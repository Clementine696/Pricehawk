import io
from typing import Optional

from fastapi import APIRouter, HTTPException, Response, Depends, Query
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from database import get_db
from routers.deps import get_current_user, _parse_search_input

router = APIRouter()


@router.get("/api/products/export")
def export_products(
    user: dict = Depends(get_current_user),
    search: Optional[str] = Query(None),
    category: Optional[str] = Query(None),
    brand: Optional[str] = Query(None),
    match_status: Optional[str] = Query(None),
    price_status: Optional[str] = Query(None),
):
    with get_db() as conn:
        with conn.cursor() as cur:
            where_parts = ["p.retailer_id = 'cfw'"]
            params = []

            if search:
                for token in _parse_search_input(search):
                    where_parts.append("(p.name ILIKE %s OR p.name_en ILIKE %s OR p.sku ILIKE %s OR p.barcode ILIKE %s)")
                    pattern = f"%{token}%"
                    params.extend([pattern, pattern, pattern, pattern])

            if category:
                ids = [c.strip() for c in category.split(',') if c.strip()]
                if ids:
                    where_parts.append(f"p.category_id IN ({','.join(['%s'] * len(ids))})")
                    params.extend(ids)

            if brand:
                brands = [b.strip() for b in brand.split(',') if b.strip()]
                if brands:
                    where_parts.append(f"p.brand IN ({','.join(['%s'] * len(brands))})")
                    params.extend(brands)

            if match_status == 'no_match':
                where_parts.append("NOT EXISTS (SELECT 1 FROM product_matches pm WHERE pm.cfw_product_id = p.id AND (pm.is_verified = FALSE OR pm.is_same = TRUE))")
            elif match_status == 'unverified':
                where_parts.append("EXISTS (SELECT 1 FROM product_matches pm WHERE pm.cfw_product_id = p.id AND pm.is_verified = FALSE)")
            elif match_status == 'verified':
                where_parts.append("EXISTS (SELECT 1 FROM product_matches pm WHERE pm.cfw_product_id = p.id AND pm.is_verified = TRUE AND pm.is_same = TRUE)")

            if price_status == 'no_match':
                where_parts.append("NOT EXISTS (SELECT 1 FROM product_matches pm WHERE pm.cfw_product_id = p.id AND pm.is_verified = TRUE AND pm.is_same = TRUE)")
            elif price_status in ['lower', 'higher', 'same']:
                cmp = {'lower': '<', 'higher': '>', 'same': '='}[price_status]
                where_parts.append(f"""
                    EXISTS (SELECT 1 FROM product_matches pm JOIN products mp ON pm.makro_product_id = mp.id
                    WHERE pm.cfw_product_id = p.id AND pm.is_verified = TRUE AND pm.is_same = TRUE
                    AND p.current_price IS NOT NULL AND mp.current_price IS NOT NULL
                    AND p.current_price {cmp} mp.current_price)
                """)

            cur.execute(f"""
                SELECT p.sku, p.barcode, p.name, p.brand, c.category_name,
                    p.current_price as cfw_price,
                    (SELECT mp.current_price FROM product_matches pm JOIN products mp ON pm.makro_product_id = mp.id
                     WHERE pm.cfw_product_id = p.id AND pm.is_verified = TRUE AND pm.is_same = TRUE LIMIT 1) as makro_price,
                    (SELECT mp.name FROM product_matches pm JOIN products mp ON pm.makro_product_id = mp.id
                     WHERE pm.cfw_product_id = p.id AND pm.is_verified = TRUE AND pm.is_same = TRUE LIMIT 1) as makro_name
                FROM products p
                LEFT JOIN categories c ON p.retailer_id = c.retailer_id AND p.category_id = c.category_id
                WHERE {" AND ".join(where_parts)}
                ORDER BY p.sku
            """, params)
            rows = cur.fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"

    header_fill = PatternFill(start_color="0E7490", end_color="0E7490", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    headers = ["SKU", "Barcode", "CFW Name", "Brand", "Category", "CFW Price", "Makro Price", "Diff (%)", "Makro Name"]
    col_widths = [14, 16, 45, 20, 25, 14, 14, 10, 45]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(bottom=Side(style='thin', color='CCCCCC'))
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 22

    green_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

    for row_idx, row in enumerate(rows, 2):
        cfw_price = row["cfw_price"]
        makro_price = row["makro_price"]
        if cfw_price and makro_price:
            diff = ((makro_price - cfw_price) / cfw_price) * 100
            diff_str = f"{'+' if diff > 0 else ''}{diff:.1f}%"
        else:
            diff = None
            diff_str = "—"

        values = [row["sku"], row["barcode"], row["name"], row["brand"],
                  row["category_name"], cfw_price, makro_price, diff_str, row["makro_name"]]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = Alignment(vertical='center')
            if col in (6, 7) and value is not None:
                cell.number_format = '#,##0.00'
            if diff is not None:
                cell.fill = green_fill if diff < 0 else red_fill
        ws.row_dimensions[row_idx].height = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=products_export.xlsx"}
    )


@router.get("/api/products")
def get_products(
    user: dict = Depends(get_current_user),
    retailer: Optional[str] = Query(None),
    category: Optional[str] = Query(None),
    brand: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    match_status: Optional[str] = Query(None),
    price_status: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=500)
):
    with get_db() as conn:
        with conn.cursor() as cur:
            where_parts = ["p.retailer_id = 'cfw'"]
            params = []

            if retailer:
                where_parts[0] = "p.retailer_id = %s"
                params.append(retailer)

            if category:
                ids = [c.strip() for c in category.split(',') if c.strip()]
                if ids:
                    where_parts.append(f"p.category_id IN ({','.join(['%s'] * len(ids))})")
                    params.extend(ids)

            if brand:
                brands = [b.strip() for b in brand.split(',') if b.strip()]
                if brands:
                    where_parts.append(f"p.brand IN ({','.join(['%s'] * len(brands))})")
                    params.extend(brands)

            if search:
                for token in _parse_search_input(search):
                    where_parts.append("(p.name ILIKE %s OR p.name_en ILIKE %s OR p.sku ILIKE %s OR p.barcode ILIKE %s)")
                    pattern = f"%{token}%"
                    params.extend([pattern, pattern, pattern, pattern])

            if match_status == 'no_match':
                where_parts.append("NOT EXISTS (SELECT 1 FROM product_matches pm WHERE (pm.cfw_product_id = p.id OR pm.makro_product_id = p.id) AND (pm.is_verified = FALSE OR pm.is_same = TRUE))")
            elif match_status == 'unverified':
                where_parts.append("EXISTS (SELECT 1 FROM product_matches pm WHERE (pm.cfw_product_id = p.id OR pm.makro_product_id = p.id) AND pm.is_verified = FALSE)")
            elif match_status == 'verified':
                where_parts.append("EXISTS (SELECT 1 FROM product_matches pm WHERE (pm.cfw_product_id = p.id OR pm.makro_product_id = p.id) AND pm.is_verified = TRUE AND pm.is_same = TRUE)")

            if price_status == 'no_match':
                where_parts.append("NOT EXISTS (SELECT 1 FROM product_matches pm WHERE (pm.cfw_product_id = p.id OR pm.makro_product_id = p.id) AND pm.is_verified = TRUE AND pm.is_same = TRUE)")
            elif price_status in ['lower', 'higher', 'same']:
                cmp = {'lower': '<', 'higher': '>', 'same': '='}[price_status]
                where_parts.append(f"""
                    EXISTS (SELECT 1 FROM product_matches pm
                    JOIN products p2 ON (CASE WHEN pm.cfw_product_id = p.id THEN pm.makro_product_id = p2.id WHEN pm.makro_product_id = p.id THEN pm.cfw_product_id = p2.id END)
                    WHERE (pm.cfw_product_id = p.id OR pm.makro_product_id = p.id)
                    AND pm.is_verified = TRUE AND pm.is_same = TRUE
                    AND p.current_price IS NOT NULL AND p2.current_price IS NOT NULL
                    AND p.current_price {cmp} p2.current_price)
                """)

            where_clause = " AND ".join(where_parts)

            cur.execute(f"SELECT COUNT(*) as count FROM products p WHERE {where_clause}", params)
            total = cur.fetchone()["count"]

            offset = (page - 1) * page_size
            cur.execute(f"""
                SELECT p.id, p.retailer_id, r.name as retailer_name, p.sku, p.barcode,
                    p.name, p.name_en, p.brand, p.category_id, c.category_name,
                    p.dept_id, p.sub_dept_id, p.class_id, p.sub_class_id,
                    p.current_price, p.step_prices, p.url, p.image_url, p.is_active,
                    p.created_at, p.updated_at,
                    CASE WHEN p.retailer_id = 'cfw' THEN (SELECT COUNT(*) FROM product_matches pm WHERE pm.cfw_product_id = p.id)
                         WHEN p.retailer_id = 'makro' THEN (SELECT COUNT(*) FROM product_matches pm WHERE pm.makro_product_id = p.id) END as match_count,
                    CASE WHEN p.retailer_id = 'cfw' THEN (SELECT COUNT(*) FROM product_matches pm WHERE pm.cfw_product_id = p.id AND pm.is_verified = TRUE AND pm.is_same = TRUE)
                         WHEN p.retailer_id = 'makro' THEN (SELECT COUNT(*) FROM product_matches pm WHERE pm.makro_product_id = p.id AND pm.is_verified = TRUE AND pm.is_same = TRUE) END as verified_match_count,
                    CASE WHEN p.retailer_id = 'cfw' THEN (SELECT COUNT(*) FROM product_matches pm WHERE pm.cfw_product_id = p.id AND pm.is_verified = FALSE)
                         WHEN p.retailer_id = 'makro' THEN (SELECT COUNT(*) FROM product_matches pm WHERE pm.makro_product_id = p.id AND pm.is_verified = FALSE) END as unverified_count,
                    CASE WHEN p.retailer_id = 'cfw' THEN COALESCE(
                            (SELECT mp.current_price FROM product_matches pm JOIN products mp ON pm.makro_product_id = mp.id WHERE pm.cfw_product_id = p.id AND pm.is_verified = TRUE AND pm.is_same = TRUE LIMIT 1),
                            (SELECT mp.current_price FROM product_matches pm JOIN products mp ON pm.makro_product_id = mp.id WHERE pm.cfw_product_id = p.id AND pm.is_verified = FALSE ORDER BY pm.match_score DESC LIMIT 1))
                         WHEN p.retailer_id = 'makro' THEN COALESCE(
                            (SELECT cp.current_price FROM product_matches pm JOIN products cp ON pm.cfw_product_id = cp.id WHERE pm.makro_product_id = p.id AND pm.is_verified = TRUE AND pm.is_same = TRUE LIMIT 1),
                            (SELECT cp.current_price FROM product_matches pm JOIN products cp ON pm.cfw_product_id = cp.id WHERE pm.makro_product_id = p.id AND pm.is_verified = FALSE ORDER BY pm.match_score DESC LIMIT 1)) END as matched_price
                FROM products p
                JOIN retailers r ON p.retailer_id = r.retailer_id
                LEFT JOIN categories c ON p.retailer_id = c.retailer_id AND p.category_id = c.category_id
                WHERE {where_clause}
                ORDER BY p.id
                LIMIT %s OFFSET %s
            """, params + [page_size, offset])
            products = cur.fetchall()

            cur.execute("SELECT DISTINCT r.retailer_id, r.name FROM retailers r JOIN products p ON r.retailer_id = p.retailer_id ORDER BY r.name")
            retailers = cur.fetchall()

            cur.execute("SELECT DISTINCT c.category_id, c.category_name, c.retailer_id FROM categories c JOIN products p ON c.retailer_id = p.retailer_id AND c.category_id = p.category_id ORDER BY c.category_name")
            categories = cur.fetchall()

            cur.execute("SELECT DISTINCT brand FROM products WHERE brand IS NOT NULL AND brand != '' ORDER BY brand")
            brands = [row["brand"] for row in cur.fetchall()]

            return {
                "products": products,
                "total": total,
                "retailers": retailers,
                "categories": categories,
                "brands": brands,
                "pagination": {"page": page, "page_size": page_size, "total": total, "total_pages": (total + page_size - 1) // page_size}
            }


@router.get("/api/products/{sku}/price-history")
def get_price_history(
    sku: str,
    user: dict = Depends(get_current_user),
    days: int = Query(30, ge=1, le=365),
    start_date: str = Query(None),
    end_date: str = Query(None)
):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT id, name, retailer_id FROM products WHERE sku = %s AND retailer_id = 'cfw'", (sku,))
            base = cur.fetchone()
            if not base:
                raise HTTPException(status_code=404, detail="Product not found")

            if start_date and end_date:
                date_filter = "AND recorded_at >= %s::date AND recorded_at < %s::date + interval '1 day'"
                date_params_extra = (start_date, end_date)
            else:
                date_filter = "AND recorded_at >= NOW() - make_interval(days => %s)"
                date_params_extra = (days,)

            base_id = base["id"]

            cur.execute(f"SELECT price, recorded_at FROM price_history WHERE product_id = %s {date_filter} ORDER BY recorded_at ASC",
                        (base_id,) + date_params_extra)
            base_history = cur.fetchall()

            cur.execute("""
                SELECT p.id, p.name, p.retailer_id FROM product_matches pm
                JOIN products p ON (CASE WHEN pm.cfw_product_id = %s THEN pm.makro_product_id ELSE pm.cfw_product_id END = p.id)
                WHERE (pm.cfw_product_id = %s OR pm.makro_product_id = %s) AND pm.is_verified = TRUE
            """, (base_id, base_id, base_id))
            matched = cur.fetchall()

            matched_products = []
            for mp in matched:
                cur.execute(f"SELECT price, recorded_at FROM price_history WHERE product_id = %s {date_filter} ORDER BY recorded_at ASC",
                            (mp["id"],) + date_params_extra)
                mp_history = cur.fetchall()
                matched_products.append({
                    "product_id": mp["id"],
                    "name": mp["name"],
                    "retailer": mp["retailer_id"],
                    "history": [{"price": float(h["price"]) if h["price"] else None, "date": h["recorded_at"].isoformat()} for h in mp_history]
                })

            return {
                "base_product": {
                    "product_id": base_id,
                    "name": base["name"],
                    "retailer": base["retailer_id"],
                    "history": [{"price": float(h["price"]) if h["price"] else None, "date": h["recorded_at"].isoformat()} for h in base_history]
                },
                "matched_products": matched_products
            }


@router.get("/api/products/{sku}")
def get_product_detail(sku: str, user: dict = Depends(get_current_user)):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT p.id, p.retailer_id, p.sku, p.barcode, p.name, p.brand, p.category_id,
                       c.category_name, p.current_price, p.step_prices, p.url, p.image_url,
                       p.is_active, p.created_at, p.updated_at
                FROM products p
                LEFT JOIN categories c ON p.retailer_id = c.retailer_id AND p.category_id = c.category_id
                WHERE p.sku = %s
            """, (sku,))
            product = cur.fetchone()
            if not product:
                raise HTTPException(status_code=404, detail="Product not found")

            cur.execute("""
                SELECT pm.match_id, pm.match_score, pm.is_verified, pm.is_same, pm.verified_at,
                       p2.id as matched_product_id, p2.retailer_id as matched_retailer_id,
                       r2.name as matched_retailer_name, p2.sku as matched_sku,
                       p2.barcode as matched_barcode, p2.name as matched_name,
                       p2.brand as matched_brand, c2.category_name as matched_category_name,
                       p2.current_price as matched_price, p2.step_prices as matched_step_prices,
                       p2.url as matched_url, p2.image_url as matched_image,
                       p2.updated_at as matched_updated_at
                FROM product_matches pm
                JOIN products p2 ON (CASE WHEN pm.cfw_product_id = %s THEN pm.makro_product_id = p2.id WHEN pm.makro_product_id = %s THEN pm.cfw_product_id = p2.id END)
                JOIN retailers r2 ON p2.retailer_id = r2.retailer_id
                LEFT JOIN categories c2 ON p2.retailer_id = c2.retailer_id AND p2.category_id = c2.category_id
                WHERE pm.cfw_product_id = %s OR pm.makro_product_id = %s
                ORDER BY pm.is_verified DESC, pm.match_score DESC
            """, (product['id'], product['id'], product['id'], product['id']))
            matches_rows = cur.fetchall()

            matches = [{
                "match_id": row["match_id"],
                "is_same": row["is_same"],
                "verified_by_user": bool(row["is_verified"]),
                "confidence_score": float(row["match_score"]) if row["match_score"] else None,
                "reason": None,
                "match_type": "auto",
                "product": {
                    "product_id": row["matched_product_id"],
                    "sku": row["matched_sku"],
                    "name": row["matched_name"],
                    "brand": row["matched_brand"],
                    "category": row["matched_category_name"],
                    "current_price": float(row["matched_price"]) if row["matched_price"] else None,
                    "original_price": None,
                    "step_prices": row["matched_step_prices"] if row["matched_step_prices"] else [],
                    "link": row["matched_url"],
                    "image": row["matched_image"],
                    "retailer_id": row["matched_retailer_id"],
                    "retailer_name": row["matched_retailer_name"],
                    "last_updated_at": row["matched_updated_at"].isoformat() if row["matched_updated_at"] else None,
                    "scrape_fail_count": 0,
                }
            } for row in matches_rows]

            return {
                "product": {
                    "id": product["id"],
                    "retailer_id": product["retailer_id"],
                    "sku": product["sku"],
                    "barcode": product["barcode"],
                    "name": product["name"],
                    "brand": product["brand"],
                    "category_id": product["category_id"],
                    "category_name": product["category_name"],
                    "current_price": float(product["current_price"]) if product["current_price"] else None,
                    "step_prices": product["step_prices"] if product["step_prices"] else [],
                    "url": product["url"],
                    "image": product["image_url"],
                    "is_active": product["is_active"],
                    "created_at": product["created_at"].isoformat() if product["created_at"] else None,
                    "updated_at": product["updated_at"].isoformat() if product["updated_at"] else None,
                },
                "matches": matches,
                "total_matches": len(matches)
            }
