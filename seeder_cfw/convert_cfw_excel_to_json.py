#!/usr/bin/env python3
"""
CFW Excel to JSON Converter
Converts temp/Product_template_CFW.xlsx to temp/cfw_products.json for inspection
before database insertion.

Merges:
- POS PRICE from "Price" sheet
- step_prices from "Promotion Step Price" sheet

Usage:
    python temp/convert_cfw_excel_to_json.py
"""

import json
import sys
from pathlib import Path
from collections import defaultdict
from openpyxl import load_workbook
from typing import List, Dict, Any

# File paths
EXCEL_FILE = Path(__file__).parent / "Product_template_CFW.xlsx"
OUTPUT_FILE = Path(__file__).parent / "cfw_products.json"

# Large max value in Excel means unlimited
UNLIMITED_THRESHOLD = 99999990


def excel_cell_to_python(value):
    """Convert Excel cell value to Python type"""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, bool):
        return value
    return str(value).strip() if isinstance(value, str) else value


def read_excel_sheet(sheet) -> List[Dict[str, Any]]:
    """Read a single Excel sheet and return list of dicts"""
    rows = []

    # Get header row
    headers = []
    for cell in sheet[1]:
        header = excel_cell_to_python(cell.value)
        headers.append(header)

    # Read data rows
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
        row_dict = {}
        for col_idx, cell in enumerate(row):
            header = headers[col_idx] if col_idx < len(headers) else f"col_{col_idx}"
            value = excel_cell_to_python(cell.value)
            if header:  # Only include if header exists
                row_dict[header] = value

        # Only add non-empty rows
        if any(v is not None for v in row_dict.values()):
            rows.append(row_dict)

    return rows


def build_step_prices_map(step_price_rows: List[Dict]) -> Dict[str, List]:
    """
    Build a map of SKU -> step_prices list.
    
    Format: [[min_qty, max_qty, discount], ...]
    - min_qty: minimum quantity to buy
    - max_qty: maximum quantity (null if unlimited)
    - discount: amount subtracted from base price
    
    Example: [[2, 4, 11], [5, null, 16]]
    Means: Buy 2-4 get -11 baht, Buy 5+ get -16 baht
    """
    step_prices_map = defaultdict(list)
    
    for row in step_price_rows:
        sku = row.get('SKCODE')
        if not sku:
            continue
        
        # Normalize SKU to string
        sku = str(sku).strip()
        
        min_qty = row.get('MINVALUE')
        max_qty = row.get('MAXVALUE')
        discount = row.get('BVALUE')
        
        # Skip if no discount value
        if discount is None:
            continue
        
        # Convert max_qty to null if unlimited
        if max_qty is not None and max_qty >= UNLIMITED_THRESHOLD:
            max_qty = None
        
        # Convert to int where applicable
        min_qty = int(min_qty) if min_qty is not None else None
        max_qty = int(max_qty) if max_qty is not None else None
        discount = int(discount) if isinstance(discount, float) and discount.is_integer() else discount
        
        step_prices_map[sku].append([min_qty, max_qty, discount])
    
    # Sort each SKU's step prices by min_qty
    for sku in step_prices_map:
        step_prices_map[sku].sort(key=lambda x: x[0] if x[0] else 0)
    
    return dict(step_prices_map)


def build_price_map(price_rows: List[Dict]) -> Dict[str, float]:
    """
    Build a map of SKU -> POS PRICE from Price sheet.
    Price sheet columns: STCODE, SKU, POS PRICE
    """
    price_map = {}
    
    for row in price_rows:
        sku = row.get('SKU')
        pos_price = row.get('POS PRICE')
        
        if sku is None or pos_price is None:
            continue
        
        sku_str = str(sku).strip()
        price_map[sku_str] = float(pos_price) if pos_price else None
    
    return price_map


def convert_excel_to_json():
    """Convert CFW Excel file to JSON with merged step prices and POS prices"""

    if not EXCEL_FILE.exists():
        print(f"[ERROR] {EXCEL_FILE} not found")
        sys.exit(1)

    print(f"[INFO] Reading Excel file: {EXCEL_FILE}")

    try:
        wb = load_workbook(EXCEL_FILE)
    except Exception as e:
        print(f"[ERROR] Error loading Excel: {e}")
        sys.exit(1)

    print(f"[INFO] Found {len(wb.sheetnames)} sheets: {wb.sheetnames}")

    # Read Products sheet
    print(f"\n[INFO] Reading Products sheet...")
    products_sheet = wb['Products']
    products = read_excel_sheet(products_sheet)
    print(f"   [OK] Found {len(products)} products")

    # Read Price sheet
    price_map = {}
    if 'Price' in wb.sheetnames:
        print(f"\n[INFO] Reading Price sheet...")
        price_sheet = wb['Price']
        price_rows = read_excel_sheet(price_sheet)
        print(f"   [OK] Found {len(price_rows)} price rows")
        
        price_map = build_price_map(price_rows)
        print(f"   [OK] Built prices for {len(price_map)} SKUs")
    else:
        print(f"\n[WARN] 'Price' sheet not found")

    # Read Promotion Step Price sheet
    step_prices_map = {}
    if 'Promotion Step Price' in wb.sheetnames:
        print(f"\n[INFO] Reading Promotion Step Price sheet...")
        step_price_sheet = wb['Promotion Step Price']
        step_price_rows = read_excel_sheet(step_price_sheet)
        print(f"   [OK] Found {len(step_price_rows)} step price rows")
        
        step_prices_map = build_step_prices_map(step_price_rows)
        print(f"   [OK] Built step prices for {len(step_prices_map)} SKUs")
    else:
        print(f"\n[WARN] 'Promotion Step Price' sheet not found")

    # Merge prices and step prices into products
    print(f"\n[INFO] Merging prices into products...")
    products_with_price = 0
    products_with_steps = 0
    for product in products:
        sku = product.get('SKU')
        if sku:
            sku_str = str(sku).strip()
            
            # Merge POS PRICE
            product['POS PRICE'] = price_map.get(sku_str)
            if product['POS PRICE'] is not None:
                products_with_price += 1
            
            # Merge step prices
            product['step_prices'] = step_prices_map.get(sku_str, [])
            if product['step_prices']:
                products_with_steps += 1
    
    print(f"   [OK] {products_with_price} products have POS PRICE")
    print(f"   [OK] {products_with_steps} products have step prices")

    # Write JSON output (flat array, no wrapper)
    print(f"\n[INFO] Writing JSON to: {OUTPUT_FILE}")
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        json.dump(products, f, ensure_ascii=False, indent=2, default=str)

    # Print summary statistics
    print("\n" + "="*60)
    print("CONVERSION SUMMARY")
    print("="*60)
    print(f"\nProducts: {len(products)} total")
    print(f"Products with POS PRICE: {products_with_price}")
    print(f"Products with step prices: {products_with_steps}")
    
    # Show sample products with prices
    print(f"\nSample products with POS PRICE:")
    samples_shown = 0
    for product in products:
        if product.get('POS PRICE') and samples_shown < 3:
            sku = product.get('SKU')
            name = product.get('Product Name (EN)', product.get('Product Name (TH)', 'N/A'))
            price = product.get('POS PRICE')
            steps = product.get('step_prices')
            try:
                print(f"  SKU {sku}: {name[:50]}...")
                print(f"    POS PRICE: {price}, step_prices: {steps}")
            except UnicodeEncodeError:
                print(f"  SKU {sku}: [non-ASCII name]")
                print(f"    POS PRICE: {price}, step_prices: {steps}")
            samples_shown += 1

    print(f"\n[SUCCESS] Conversion complete!")
    print(f"[OUTPUT] Output file: {OUTPUT_FILE}")
    print(f"[SIZE] File size: {OUTPUT_FILE.stat().st_size / 1024:.1f} KB")

    return OUTPUT_FILE


if __name__ == "__main__":
    convert_excel_to_json()
