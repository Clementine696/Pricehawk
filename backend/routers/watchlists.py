import io
from typing import Optional

from fastapi import APIRouter, HTTPException, Response, Depends
from pydantic import BaseModel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from database import get_db
from routers.deps import get_current_user

router = APIRouter()


class WatchlistCreateRequest(BaseModel):
    name: str
    description: Optional[str] = None


@router.get("/api/watchlists")
def get_watchlists(user: dict = Depends(get_current_user)):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT w.id, w.name, w.description, w.created_at, COUNT(wp.id) as product_count
                FROM watchlists w LEFT JOIN watchlist_products wp ON w.id = wp.watchlist_id
                GROUP BY w.id, w.name, w.description, w.created_at ORDER BY w.created_at DESC
            """)
            return {"watchlists": [dict(r) for r in cur.fetchall()]}


@router.post("/api/watchlists")
def create_watchlist(data: WatchlistCreateRequest, user: dict = Depends(get_current_user)):
    with get_db() as conn:
        with conn.cursor() as cur:
            try:
                cur.execute(
                    "INSERT INTO watchlists (name, description) VALUES (%s, %s) RETURNING id, name, description, created_at",
                    (data.name.strip(), data.description)
                )
                row = cur.fetchone()
                conn.commit()
                return dict(row)
            except Exception:
                raise HTTPException(status_code=409, detail=f"Watchlist '{data.name}' already exists")


@router.delete("/api/watchlists/{watchlist_id}")
def delete_watchlist(watchlist_id: int, user: dict = Depends(get_current_user)):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM watchlists WHERE id = %s RETURNING id", (watchlist_id,))
            if not cur.fetchone():
                raise HTTPException(status_code=404, detail="Watchlist not found")
            conn.commit()
    return {"success": True}


@router.get("/api/watchlists/{watchlist_id}/products")
def get_watchlist_products(watchlist_id: int, user: dict = Depends(get_current_user)):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT id, name FROM watchlists WHERE id = %s", (watchlist_id,))
            watchlist = cur.fetchone()
            if not watchlist:
                raise HTTPException(status_code=404, detail="Watchlist not found")

            cur.execute("""
                SELECT p.id, p.sku, p.name, p.brand, p.current_price, p.image_url, p.url,
                       c.category_name, wp.added_at,
                       (SELECT mp.current_price FROM product_matches pm JOIN products mp ON pm.makro_product_id = mp.id
                        WHERE pm.cfw_product_id = p.id AND pm.is_verified = TRUE AND pm.is_same = TRUE LIMIT 1) as makro_price,
                       (SELECT mp.name FROM product_matches pm JOIN products mp ON pm.makro_product_id = mp.id
                        WHERE pm.cfw_product_id = p.id AND pm.is_verified = TRUE AND pm.is_same = TRUE LIMIT 1) as makro_name
                FROM watchlist_products wp
                JOIN products p ON wp.product_id = p.id
                LEFT JOIN categories c ON p.retailer_id = c.retailer_id AND p.category_id = c.category_id
                WHERE wp.watchlist_id = %s ORDER BY wp.added_at DESC
            """, (watchlist_id,))
            return {"watchlist": dict(watchlist), "products": [dict(r) for r in cur.fetchall()]}


@router.get("/api/watchlists/{watchlist_id}/export")
def export_watchlist(watchlist_id: int, user: dict = Depends(get_current_user)):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT id, name FROM watchlists WHERE id = %s", (watchlist_id,))
            watchlist = cur.fetchone()
            if not watchlist:
                raise HTTPException(status_code=404, detail="Watchlist not found")

            cur.execute("""
                SELECT p.sku, p.name, p.brand, c.category_name,
                       p.current_price as cfw_price,
                       (SELECT mp.current_price FROM product_matches pm JOIN products mp ON pm.makro_product_id = mp.id
                        WHERE pm.cfw_product_id = p.id AND pm.is_verified = TRUE AND pm.is_same = TRUE LIMIT 1) as makro_price,
                       (SELECT mp.name FROM product_matches pm JOIN products mp ON pm.makro_product_id = mp.id
                        WHERE pm.cfw_product_id = p.id AND pm.is_verified = TRUE AND pm.is_same = TRUE LIMIT 1) as makro_name
                FROM watchlist_products wp
                JOIN products p ON wp.product_id = p.id
                LEFT JOIN categories c ON p.retailer_id = c.retailer_id AND p.category_id = c.category_id
                WHERE wp.watchlist_id = %s ORDER BY wp.added_at DESC
            """, (watchlist_id,))
            rows = cur.fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = watchlist["name"][:31]

    header_fill = PatternFill(start_color="0E7490", end_color="0E7490", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    headers = ["SKU", "CFW Name", "Brand", "Category", "CFW Price", "Makro Price", "Diff (%)", "Makro Name"]
    col_widths = [14, 45, 20, 25, 14, 14, 10, 45]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 22

    green_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

    for row_idx, row in enumerate(rows, 2):
        cfw_price = row["cfw_price"]
        makro_price = row["makro_price"]
        if cfw_price and makro_price:
            diff = ((makro_price - cfw_price) / cfw_price) * 100
            diff_str = f"{'+' if diff > 0 else ''}{diff:.1f}%"
        else:
            diff = None
            diff_str = "—"

        values = [row["sku"], row["name"], row["brand"], row["category_name"],
                  cfw_price, makro_price, diff_str, row["makro_name"]]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = Alignment(vertical='center')
            if col in (5, 6) and value is not None:
                cell.number_format = '#,##0.00'
            if diff is not None:
                cell.fill = green_fill if diff < 0 else red_fill
        ws.row_dimensions[row_idx].height = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_name = watchlist["name"].replace(" ", "_")
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=watchlist_{safe_name}.xlsx"}
    )


@router.post("/api/watchlists/{watchlist_id}/products")
def add_watchlist_product(watchlist_id: int, body: dict, user: dict = Depends(get_current_user)):
    sku = (body.get("sku") or "").strip()
    if not sku:
        raise HTTPException(status_code=400, detail="sku is required")
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT id FROM watchlists WHERE id = %s", (watchlist_id,))
            if not cur.fetchone():
                raise HTTPException(status_code=404, detail="Watchlist not found")
            cur.execute("SELECT id, name FROM products WHERE sku = %s AND retailer_id = 'cfw'", (sku,))
            product = cur.fetchone()
            if not product:
                raise HTTPException(status_code=404, detail=f"CFW product with SKU '{sku}' not found")
            try:
                cur.execute("INSERT INTO watchlist_products (watchlist_id, product_id) VALUES (%s, %s)", (watchlist_id, product["id"]))
                conn.commit()
            except Exception:
                raise HTTPException(status_code=409, detail="Product already in this watchlist")
    return {"success": True, "product_id": product["id"], "name": product["name"]}


@router.delete("/api/watchlists/{watchlist_id}/products/{product_id}")
def remove_watchlist_product(watchlist_id: int, product_id: int, user: dict = Depends(get_current_user)):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM watchlist_products WHERE watchlist_id = %s AND product_id = %s RETURNING id",
                        (watchlist_id, product_id))
            if not cur.fetchone():
                raise HTTPException(status_code=404, detail="Product not in watchlist")
            conn.commit()
    return {"success": True}
