"""
Email Service for Price Change Alerts

This service handles sending HTML emails with price change notifications.
Uses SMTP for email delivery.
"""

import os
import smtplib
import logging
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from typing import List, Dict, Optional
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

logger = logging.getLogger(__name__)


class EmailService:
    """Service for sending price alert emails via SMTP"""

    def __init__(self):
        """Initialize email service with SMTP configuration from environment"""
        self.smtp_host = os.getenv('SMTP_HOST', 'smtp.gmail.com')
        self.smtp_port = int(os.getenv('SMTP_PORT', '587'))
        self.smtp_user = os.getenv('SMTP_USER')
        self.smtp_password = os.getenv('SMTP_PASSWORD')
        self.from_name = os.getenv('SMTP_FROM_NAME', 'PriceHawk Alerts')
        self.from_email = self.smtp_user
        self.frontend_url = os.getenv('FRONTEND_URL', 'https://pricehawk-ruddy.vercel.app')

        if not self.smtp_user or not self.smtp_password:
            logger.warning("SMTP credentials not configured. Email sending will fail.")

    def send_price_alert_email(
        self,
        to_emails: List[str],
        products: List[Dict],
        status_changes: List[Dict],
        period_start: datetime,
        period_end: datetime
    ) -> Dict[str, any]:
        """
        Send price change alert email to multiple recipients

        Args:
            to_emails: List of recipient email addresses
            products: List of product dictionaries with price change info
            status_changes: List of product dictionaries with status change info
            period_start: Start timestamp of detection period
            period_end: End timestamp of detection period

        Returns:
            Dict with 'success' (bool), 'sent_count' (int), 'failed' (list of emails)
        """
        if not to_emails:
            logger.warning("No recipient emails provided")
            return {'success': False, 'sent_count': 0, 'failed': []}

        # Build email content
        total_changes = len(products) + len(status_changes)

        if total_changes == 0:
            subject = "Price Alert: No Changes Today"
            logger.info("No products or status changes - sending 'No changes' notification")
        else:
            subject = f"Price Alert: {total_changes} Changes Detected"

        html_body = self._build_html_email(products, status_changes, period_start, period_end)
        plain_body = self._build_plain_text_email(products, status_changes, period_start, period_end)

        # Generate Excel attachments
        attachments = []
        if products:
            price_excel = self._generate_price_excel(products)
            attachments.append(('price_changes.xlsx', price_excel))
        if status_changes:
            status_excel = self._generate_status_excel(status_changes)
            attachments.append(('status_changes.xlsx', status_excel))

        # Send to each recipient
        sent_count = 0
        failed = []

        for to_email in to_emails:
            try:
                self._send_email(to_email, subject, html_body, plain_body, attachments)
                sent_count += 1
                logger.info(f"Alert email sent successfully to {to_email}")
            except Exception as e:
                logger.error(f"Failed to send alert email to {to_email}: {e}")
                failed.append(to_email)

        success = sent_count == len(to_emails)
        return {
            'success': success,
            'sent_count': sent_count,
            'failed': failed
        }

    def send_test_email(self, to_email: str) -> bool:
        """
        Send a test email to verify SMTP configuration

        Args:
            to_email: Recipient email address

        Returns:
            True if sent successfully, False otherwise
        """
        subject = "PriceHawk Test Email"
        html_body = self._build_test_html()
        plain_body = "This is a test email from PriceHawk Alert System.\n\nIf you received this, your email configuration is working correctly!"

        try:
            self._send_email(to_email, subject, html_body, plain_body)
            logger.info(f"Test email sent successfully to {to_email}")
            return True
        except Exception as e:
            logger.error(f"Failed to send test email to {to_email}: {e}")
            return False

    def _send_email(
        self,
        to_email: str,
        subject: str,
        html_body: str,
        plain_body: str,
        attachments: List[tuple] = None
    ) -> None:
        """
        Internal method to send email via SMTP

        Args:
            to_email: Recipient email
            subject: Email subject
            html_body: HTML version of email
            plain_body: Plain text version of email
            attachments: List of (filename, bytes_data) tuples

        Raises:
            Exception if email sending fails
        """
        # Create message
        msg = MIMEMultipart('mixed')
        msg['From'] = f"{self.from_name} <{self.from_email}>"
        msg['To'] = to_email
        msg['Subject'] = subject

        # Create alternative part for text/html
        msg_alternative = MIMEMultipart('alternative')
        msg.attach(msg_alternative)

        # Attach both plain text and HTML versions
        part1 = MIMEText(plain_body, 'plain', 'utf-8')
        part2 = MIMEText(html_body, 'html', 'utf-8')
        msg_alternative.attach(part1)
        msg_alternative.attach(part2)

        # Attach Excel files if provided
        if attachments:
            for filename, file_data in attachments:
                part = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                part.set_payload(file_data)
                encoders.encode_base64(part)
                part.add_header('Content-Disposition', f'attachment; filename={filename}')
                msg.attach(part)

        # Send via SMTP
        with smtplib.SMTP(self.smtp_host, self.smtp_port) as server:
            server.starttls()
            server.login(self.smtp_user, self.smtp_password)
            server.send_message(msg)

    def _build_html_email(
        self,
        products: List[Dict],
        status_changes: List[Dict],
        period_start: datetime,
        period_end: datetime
    ) -> str:
        """Build HTML email body with product table and status changes"""

        # Format dates
        start_str = period_start.strftime('%B %d, %Y at %H:%M')
        end_str = period_end.strftime('%B %d, %Y at %H:%M')

        total_changes = len(products) + len(status_changes)

        # Build status changes section
        status_section = ""
        if status_changes:
            status_section = self._build_status_changes_section(status_changes)

        # Limit to top 15 products (sorted by price change percentage)
        limited_products = self._limit_and_sort_products(products)
        more_count = len(products) - len(limited_products) if len(products) > 15 else 0

        # Build price change rows
        price_section = ""
        if products:
            product_rows = ""
            for product in limited_products:
                product_rows += self._build_product_row(product)

            # More products footer
            more_footer = ""
            if more_count > 0:
                more_footer = f"""
                <tr>
                    <td colspan="3" style="padding: 20px; text-align: center; background-color: #f3f4f6; font-style: italic; color: #6b7280;">
                        ... and {more_count} more product{'s' if more_count > 1 else ''} (see attached Excel file for complete list)
                    </td>
                </tr>
                """

            price_section = f"""
                            <!-- Price Changes Header -->
                            <tr>
                                <td style="padding: 20px 20px 10px; background-color: #ffffff;">
                                    <h2 style="margin: 0; font-size: 20px; color: #111827; border-bottom: 2px solid #06b6d4; padding-bottom: 10px;">
                                        💰 Price Changes ({len(products)})
                                    </h2>
                                </td>
                            </tr>

                            <!-- Price Change Products -->
                            {product_rows}
                            {more_footer}
            """

        # Build "No Changes" section if no products and no status changes
        no_changes_section = ""
        if not products and not status_changes:
            no_changes_section = f"""
                            <!-- No Changes Section -->
                            <tr>
                                <td style="padding: 40px 20px; background-color: #ffffff; text-align: center;">
                                    <div style="background-color: #f0f9ff; padding: 30px; border-radius: 8px; border: 2px solid #06b6d4;">
                                        <h2 style="margin: 0 0 15px; font-size: 24px; color: #0891b2;">
                                            ✅ No Product Changes Today
                                        </h2>
                                        <p style="margin: 0; font-size: 16px; color: #374151; line-height: 1.6;">
                                            All monitored products maintained their prices during this period.<br>
                                            We'll continue monitoring and notify you of any changes.
                                        </p>
                                    </div>
                                </td>
                            </tr>
            """

        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Price Change Alert</title>
        </head>
        <body style="margin: 0; padding: 0; font-family: Arial, sans-serif; background-color: #f3f4f6;">
            <table role="presentation" style="width: 100%; border-collapse: collapse;">
                <tr>
                    <td style="padding: 20px 0;">
                        <table role="presentation" style="width: 100%; max-width: 800px; margin: 0 auto; background-color: #ffffff; border-radius: 8px; overflow: hidden; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
                            <!-- Header -->
                            <tr>
                                <td style="background-color: #ffffff; padding: 30px 20px; text-align: center;">
                                    <h1 style="margin: 0; color: #000000; font-size: 28px; font-weight: bold;">
                                        <img src="data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAA+YAAAF6CAYAAACUW0I2AAAQAElEQVR4Aey9CZxdRbU9vOve7iQMQcF5AGef4/OvH8ooEpxxQiBRkBASlElBxae+50QriPickSlhSMjAkCgKSJgJyBhAfSriPPKcHipjgKT73vrW2qfq3HO7Ow1pOj3cXudXddbea++qU7Xq3NNdfdKdmumQAlJACkgBKSAFpIAUkAJSQApIASkgBcZMgVHamI/Z/HRhKSAFpIAUkAJSQApIASkgBaSAFJAC41qBztqYj2upNTgpIAWkgBSQAlJACkgBKSAFpIAUkAIDFdDGfKAmD8soQQpIASkgBaSAFJACUkAKSAEpIAWkwEgpoI35SCk58v2oRykgBaSAFJACUkAKSAEpIAWkgBSYBApoYz4JFnnoKSoqBaSAFJACUkAKSAEpIAWkgBSQAmOpgDbmY6n+ZLq25ioFpIAUkAJSQApIASkgBaSAFJACgyqgjfmgsoicqApo3FJACkgBKSAFpIAUkAJSQApIgYmmgDbmE23FNN7xoIDGIAWkgBSQAlJACkgBKSAFpIAUGDEFtDEfMSnVkRQYaQXUnxSQAlJACkgBKSAFpIAUkAKTQQFtzCfDKmuOUmAoBRSTAlJACkgBKSAFpIAUkAJSYEwV0MZ8TOXXxaXA5FFAM5UCUkAKSAEpIAWkgBSQAlJgcAW0MR9cF7FSQApMTAU0aikgBaSAFJACUkAKSAEpMOEU0MZ8wi2ZBiwFpMDYK6ARSAEpIAWkgBSQAlJACkiBkVNAG/OR01I9SQEpIAVGVgH1JgWkgBSQAlJACkgBKTApFNDGfFIssyYpBaSAFFi/AopIASkgBaSAFJACUkAKjK0C2piPrf66uhSQAlJgsiigeUoBKSAFpIAUkAJSQAqsRwFtzNcjjGgpIAWkgBSYiApozFJACkgBKSAFpIAUmHgKaGM+8dZMI5YCUkAKSIGxVkDXlwJSQApIASkgBaTACCqgjfkIiqmupIAUkAJSQAqMpALqSwpIASkgBaSAFJgcCmhjPjnWWbOUAlJACkgBKbA+BcRLASkgBaSAFJACY6yANuZjvAC6vBSQAlJACkiByaGAZikFpIAUkAJSQAqsTwFtzNenjHgpIAWkgBSQAlJg4imgEUsBKSAFpIAUmIAKaGM+ARdNQ5YCUkAKSAEpIAXGVgFdXQpIASkgBaTASCqgjflIqqm+pIAUkAJSQApIASkwcgqoJykgBaSAFJgkCmhjPkkWWtOUAlJACkgBKSAFpMDgCoiVAlJACkiBsVZAG/OxXgFdXwpIASkgBaSAFJACk0EBzVEKSAEpIAXWq4A25uuVRgEpIAWkgBSQAlJACkiBiaaAxisFpIAUmIgKaGM+EVdNY5YCUkAKSAEpIAWkgBQYSwV0bSkgBaTAiCqgjfmIyqnOpIAUkAJSQApIASkgBaTASCmgfqSAFJgsCmhjPllWWvOUAlJACkgBKSAFpIAUkAKDKSBOCkiBMVdAG/MxXwINQApIASkgBaSAFJACUkAKdL4CmqEUkALrV0Ab8/Vro4gUkAJSQApIASkgBaSAFJACE0sBjVYKTEgFtDGfkMumQUsBKSAFpIAUkAJSQApIASkwdgroylJgZBXQxnxk9VRvUkAKSAEpIAWkgBSQAlJACkiBkVFAvUwaBbQxnzRLrYlKASkgBaSAFJACUkAKSAEpIAUGKiBm7BXQxnzs10AjkAJSQApIASkgBaSAFJACUkAKdLoCmt8QCmhjPoQ4CkkBKSAFpIAUkAJSQApIASkgBaTARFJgYo5VG/OJuW4atRSQAlJACkgBKSAFpIAUkAJSQAqMlQIjfF1tzEdYUHUnBaSAFJACUkAKSAEpIAWkgBSQAlJgQxRY38Z8Q/pQrhSQAlJACkgBKSAFpIAUkAJSQApIASkwTAXGeGM+zFGrmRSQAlJACkgBKSAFpIAUkAJSQApIgQ5RYHJszDtksTQNKSAFpIAUkAJSQApIASkgBaSAFOg8BbQxH8E1VVdSQApIASkgBaSAFJACUkAKSAEpIAU2VAFtzDdUsbHP1wikgBSQAlJACkgBKSAFpIAUkAJSoIMU0Ma8gxZzZKei3qSAFJACUkAKSAEpIAWkgBSQAlJgNBTQxnw0VNY11q+AIlJACkgBKSAFpIAUkAJSQApIgUmugDbmk/wGmCzT1zylgBSQAlJACkgBKSAFpIAUkALjVQFtzMfrymhcE1EBjVkKSAEpIAWkgBSQAlJACkgBKbDBCmhjvsGSqYEUGGsFdH0pIAWkgBSQAlJACkgBKSAFOkkBbcw7aTU1FykwkgqoLykgBaSAFJACUkAKSAEpIAVGRQFtzEdFZl1ECkiB9SkgXgpIASkgBaSAFJACUkAKTHYFtDGf7HeA5i8FJocCmqUUkAJSQApIASkgBaSAFBi3CmhjPm6XRgOTAlJg4imgEUsBKSAFpIAUkAJSQApIgQ1XQBvzDddMLaSAFJACY6uAri4FpIAUkAJSQApIASnQUQpoY95Ry6nJSAEpIAVGTgH1JAWkgBSQAlJACkgBKTA6CmhjPjo66ypSQApIASkwuAJipYAUkAJSQApIASkw6RXQxnzS3wISQApIASkwGRTQHKWAFJACUkAKSAEpMH4V0MZ8/K6NRiYFpIAUkAITTQGNVwpIASkgBaSAFJACw1BAG/NhiKYmUkAKSAEpIAXGUgFdWwpIASkgBaSAFOgsBbQx76z11GykgBSQAlJACoyUAupHCkgBKSAFpIAUGCUFtDEfJaF1GSkgBaSAFJACUmAwBcRJASkgBaSAFJAC2pjrHpACUkAKSAEpIAU6XwHNUApIASkgBaTAOFZAG/NxvDgamhSQAlJACkgBKTCxFNBopYAUkAIbqkDs6anduHTem29aNm/x6mXzfrt66dz/Xb1s7rU3LjngyJvPnrv1hvan/ImpgDbmE3PdNGopIAWkgBSQAlJg8iqgmUsBKdAhCkSzsPp5dxxZC7Y4mM02i882s6fFaDuHUPtisxFPu3X5QY8Bp9LhCmhj3uELrOlJASkgBaSAFJACUmB4CqiVFJACG1MBbL7DzWcd+I5gsQfXeTw25bnAZYldZuENvWt7v7x8+cy66ehoBbQx7+jl1eSkgBSQAlJACkgBKTDOFdDwpMAkVeCWc+Y+PcT4CUx/s7wjj3AGljh3m97puwzkxXSSAtqYd9Jqai5SQApIASkgBaSAFJACgyogUgqMOwWa9vVocVszbMeLMsgQowWzujWbH7l1/kGbDpIgqkMU0Ma8QxZS05ACUkAKSAEpIAWkgBQYcwU0ACnwsApgg91989J5H8GWe08kh7Qvh9m/RN+zk4W17drN12ITT0+1ExXQxrwTV1VzkgJSQApIASkgBaSAFOhgBTS1iaxA76bNnbDR/ojvumGgDJxOBIvCQIIn1pphd/qqnamANuadua6alRSQAlJACkgBKSAFpIAUeHQKqPWIK/DrlYdPrdeax1iITxnyTXm6ctqU0wvR4mz9EThK0ZlVG/POXFfNSgpIASkgBaSAFJACUkAKTAgFJssgb51/0Kb/uuuBE7DB3mnITXnajSdI8kTs5e2p26zddI9ECDpMAW3MO2xBNR0pIAWkgBSQAlJACkgBKSAFBigwpkRPT0+tb7O+Q82ac9e7KR/4z9fTmKOxDZ1m0z5CVO08BbQx77w11YykgBSQAlJACkgBKSAFpIAUGBMFBr/oW557xw4hxA9ig13HNnuQpBbbspgGD4UWK8xXXL9oznNoq3aWAtqYd9Z6ajZSQApIASkgBaSAFJACUkAKjCMFbj3roMc3Q/PL2JRvjY31wJFFsCgMJKCJCg8FhheaIVjdauGVZjo6TYFap01I85ECUkAKSAEpIAWkgBSQAlJACowHBW6df1B3X+w9C5vy7bmxHjAmbsoT2R6Hh5JCaF5YeOtet9h8JbyAutGLLjB6CmhjPnpa60pSQApIASkgBaSAFJACUkAKTBIFuClvbLruQ9hB71LZY1dm32JbFsPwUGixtsxo2MejO3vOdafP25yxDqmaBhTQxhwiqEgBKSAFpIAUkAJSQApIASkgBUZSgb7pfdtZqH0Em+mpA/oFmV+DtzbezIKHQou1ZSYLgLfmT6119W3BuOqGKDC+c7UxH9/ro9FJASkgBaSAFJACUkAKSAEpMMEUuPHsA54Zon0L77ifNHDo2F0nsmVZE6/Df5836wy3YslKYDE81Zq16cxRHYcKDHNI2pgPUzg1kwJSQApIASkgBaSAFJACUkAK9FfgpqXvfXpo2PwY48BN+frflP/Ygn3I0pH34OX/k9YimPG4Roib0VDtHAU2dGPeOTPXTKSAFJACUkAKSAEpIAWkgBSQAiOsQLDG4dHCawd0y015ImPCAmKzHmsfvuuuLS7FG/a+VixZCYpcM7jT6rHxhOwLO0OBcbox7wxxNQspIAWkgBSQAlJACkgBKSAFJocCsaendvOyOe/B/vvDwazePmtspxPRskjEB3D+wKv2X3jN7kd8c63F8Bv4KCkrAQgvhRutEWwbJ3TqGAUm98a8Y5ZRE5ECUkAKSAEpIAWkgBSQAlJgLBW45dl3vKUZaydZsO62cURsp1HIJYBJKzYQWhqnPrAIRC5/9nfi9JhCTLVwi3Mt2JaJFnSIAtqYj8JC6hJSQApIASkgBaSAFJACUkAKdK4Cq8953/ObteYXMMMtUFsFO+/sFFvq7JkFi7+xGD+346wVD1o+QvyLm/2SC7c4GyDG8BjP06ljFNDGvGOW0jQTKSAFpIAUkAJSQApIASkgBUZZgRuWH7hVs7f3a7jsi1FbJWIHnbyWRQJe0+5o1mz/HecsxhtyckUNFu7nxrvwijOyYRTnHAtm+uNvUKWTijbmnbSaozIXXUQKSAEpIAWkgBSQAlJACkgBKnDr/IM2DWsbC0KwN9Mv66Cbcm6uI96Uh3uixQ/u+J7FN5f5yWjmnXfyo2Nx7hfyiE6do4A25p2zlp01E81GCkgBKSAFpIAUkAJSQAqMYwWwKe/u3az3w7UQ3maGrbmlY9BNeY5ZjE07ZYtNHrwoMe3QbLnFdrw4t2/KyVUSW01kTWAFtDGfwIunoT96BdSDFJACUkAKSAEpIAWkgBQYjgKNzRt7BLMPYZs8pWw/6KYcGdxZF3BhzZpfesmsFevKNhUDm7PpdJlqbNNyaKHmSBi0PRJUJqgCWPsJOnINWwpMHAU0UikgBaSAFJACUkAKSIEOUuC6xbO3xcZ5Cab0eNSiDLopL0LcY+Od+v9M3+RZM7ebs/ifiR0AMdgT89bbg4XjJq7nSCpYuNsdnTpGAW3MO2YpNREpIAWkgBSQAlJACkgBKSAFNrYCty494KX1WtfCGOPU8lrr3ZRjG42CvD81Y+1DL5nV83Bvup+cN+DczKNdKkUnfsapaU1tzJMynQLamHfKSmoeUmC0FNB1pIAUkAJSQApIASkwSRVYefybp/Za7WvB7EWlBBE75eS0W/BQsNGOIcZPrt36Gden+75xjQAAEABJREFUtEEB3QRsuJ/tQW/nFk6F42c/kar9H84qHaSANuYdtJiaihToJAU0FykgBaSAFJACUkAKjCcFfrx49mZbbvmk7waLr8W4in0UdtOwveQ9szs8FURvsPDJ7ecsXjpjRk8f6fXVG5bMfgJyNx/qTTnbIqfZjLHtv1kjrzqxFShuqIk9B41eCkgBKTBcBdROCkgBKSAFpIAUkAIPq8Btyw/b/IHQdXQI4U1l8qCbcu7GUYvSsGBL77Hm8WWbIQw02ebhNuVs3ozN+0K9cT9t1c5RQBvzzllLzUQKSIFxq4AGJgWkgBSQAlJACkxkBdasXfueEOzQcg7r3ZQjAztsnLEnt5/V+8In3rj/kjX0H7Y2wvNbOUUnfvZTK4Jx/LOrVn+gxcjqBAW0Me+EVdQcpIAUkAJUQFUKSAEpIAWkgBQYUQVW9ezadfOyA/aMofl1dDwNlb8y7sBTvz0zYmRZ42+aU2zPV81d9Dd6j6SGEF9Y5BW9+tlPBVucQQT7e98D9sg2+0UjnSeAAtqYT4BF0hClgBSQAuNJAY1FCkgBKSAFpMBkUAAvxcO05zx790YMJ2O+Q2zKsVnmv0EnIBHwJ7iH7vTuM38L9xGVW+cf1I02z8s7e/SRzUp7Zy3E8Pda1xT9U/aKMp1gamPeCauoOUgBKSAFOk8BzUgKSAEpIAWkwJgqcOtZ856Ft9jHB7Mn+kCwU3fEqdgiw8BumucMsJvYOR+7w5zFV8B+xKVv+n1bxRCfzgbet5/o5doimhb/csUd2+iNeZamQ1Ab8w5ZSE1DCkgBKSAFhqOA2kgBKSAFpIAUGKjArUsPeGkjxgsQeQYq3l63NsYtyyOIFQh+Leoxa3/7p9MLZgPOvd1PQvbWaF/2Bz8VZ92G1YhN+0VPT0/TCZ06RgFtzDtmKTURKSAFpIAUGLcKaGBSQApIASkwYRS4Zdm8Z/daOAUDfjEqNsrRgad2Cx4KedQG3qyfNOXBqcfM6Ll6yP8WDbkDSp81ngbyKZU373BZWhfwF/bRmla32xlR7SwFtDHvrPXUbKSAFJACUmASK6CpSwEpIAWkwKNT4LrT501vxHgsNtk7ek++G3ZrqD0zEy5vWvO4bQ9e0EtnQ2uMtZ1wge72drF0W5ata8Tun5UBGR2jgDbmHbOUmogUkAJSQApIgVFRQBeRAlJACnSkAqvP3P9xXVOaC7Apn+kTjK3tcLsFD8VzcIJ5Y7NpH9hp/yX/B3d4JcY3tTdEr4nwYZRu/OmuByz6ewoJOkgBbcw7aDE1FSkgBaSAFJACnaOAZiIFpIAUGD0Fbp1/0KbWVT/OQngX9sA1891wcX34hYFX2m60CLo/7+qrvWunAx75X2Bno2q94eSZT8MPA17R4loXqAwD4YjhxfNCyAMBpdIxCmhj3jFLqYlIASkgBaSAFJACG6yAGkgBKTDpFbht+cwpfZv1fjzGuD/EwLY3AorSsgq/uiVG7M94U/6hV81b+L8pOiyI06buh4bYm+NcuQD6J5EqvBj7wrrmeYkQdJgC2ph32IJqOlJACkgBKSAFpMD4U0AjkgJSYHwqgDfl3fev27wHo/uYWZiS35RjG9xviwwGxdKBN9n34fX1e/+86QNXgqpE4G1A+fXKw6fGYNyYo1WrG/RvrQEkvhZu3/Hgc3+HRJUOVEAb8w5cVE1JCkgBKSAFpIAUmJQKaNJSQApsgAJx+cx67+a9h6LJh1Gn5U057EpJm+IEDGAjfb/V7OAdZy+6ZNasFQ1yw61///u/Xo439VtXduGt/bh3Wr2wrQrVRI/r1CkKaGPeKSupeUgBKSAFpIAUkAJSYFQU0EWkwMRXAJvhcMu6zeaFaJ/BbNo25a2tcLISIM9oBgtfsKkPfJf+o6kcg1l4NcawmaVj0DflRezBaM3rClPnTlRAG/NOXFXNSQpIASkgBaSAFJACE10BjV8KbCQFsPkNq5fNPbRp9g1c4nH5TTk33azgUJKVAETalNvnH/+46V/ZcdaKB8k9mrp62X7TQ4i7WLAu9oNxEVKtXBgMYn9qhHAbTJUOVUAb8w5dWE1LCkgBKSAFpIAUkAJS4OEVUMbkU+DmpXPfbRY+i7pJ3pRb25E2xQkYgrkO9eQd9j/z08/b/ZtryT3a2uxtPhUb7h3YD/ompNrPgxst3v6Ezfr0++VJoU4Ebcw7cVU1JykgBaSAFJACUkAKSIHxpIDGMk4UuGnxAe+2EL6C4Tw+b8qx7/W34eBQ6LUAlseChe+YNXvgpwRYj7I0YtgbXTwOm3Pzi8CpGIWXrxbj914ya8U6J3XqSAW0Me/IZdWkpIAUkAJSQApIASkgBSafAprx+hTA/jasXjrvnbEWlkaLT8mb8vZ8ZJFIQNMCtszRrr9j2pr9d9p/yf85NwKnlce/eWoI4cO+KS/7q14YW/TkYryN2roafjBQJsroQAW0Me/ARdWUpIAUkAJSQApIASkgBaTARlNgAnZ849K5b8EG96vBrF7dlKe9L2aUrAQgWJoxxiv7ao2Zs0b4bfX0zbZ6Ly61FS9SVHiF4efWhh18COe++rCz7vKATh2rgDbmHbu0mpgUkAJSQApIASkgBaSAFJi4CozUyG9cfMDbasFOQ3/PzJtybHfxKhyMF3owEsDyGNxf1qO9/9Wzl/6N3EjVa5ftu6XVbF+/iHeKKzkWp7ZNebS+Rm/z+CKicycroI15J6+u5iYFpIAUkAJSQApIASkgBSapAj09PbUbl8x7e6iHRdgEPylvytvliIVb2RvDRLHbmt311203Z/GvgqF1kTUi57jWXoWxvKDojJcqLJ7bNuUgQs1unNL1wO0wVTpcAW3MO3yBNT0pIAWkgBSQAlJACkgBKTAZFXjT8/74tlCL38C2eitshF0CboNZ3UHAsUVk5kfI33/nfc74i8dH5FR0go13wOV2h7dl//0+YqBZkEEw68OAVu50xyvWFK7OnayANuadvLqamxSQAlJACkgBKSAFpIAUmIQK3LzsgD3xpvtcbGzb/vl6S4q0+U1APpkPWCMcsMOcxf9DbqTrz1bM7K5beBf6xfBwTmWQTTkj/2rUmleGnp4mnUdUlTRhFdDGfMIunQYuBaSAFJACUkAKSAEpIAWkQH8Frl86d+9mtBOw2Z2KN9/9w/AjKkoCWNi/4xziL2Mj7LTj3EU/xa65EkVshMpda7r/K1p8UrU7jDO57ZfEGH74y3WP2Sg/IEgXHDao4cgroI35yGuqHqWAFJACUkAKSAEpIAWkgBQYAwVuWnbg7C4L/GNp5X+Jxu0uazGcZCUgRxOb4D/UmnboTnMXbbSN8PWL3vUc/ATgI7xmruvblDOOHxJ8/eCDF/TSnqR1Uk1bG/NJtdyarBSQAlJACkgBKSAFpIAU6DwF4vKZ9ZuWzJlrFs/AG2lsyos5ctNdWDwnL0GFubtZtznbz1m8itzGqPxDdM1mF8Zn03P/Q23KsYH/wS4HnXVpzhVuTAXGR9/amI+PddAopIAUkAJSQApIASkgBaSAFBiGAtjghpvXTX9ftNoXLcYuVO+lsv+Gn7wEILD35Tn8Eu322uk9Z36f3saqr93mV8/EDwzenvvHmJNZGVBmYlwbLHwxuYJOUeBh5qGN+cMIpLAUkAJSQApIASkgBaSAFJAC41OBGHtqq8864FMxxq+EEJ+wIZtybInvrls47PLfPePqUZjd64LZi3idoTbl6acFPwqheT1zVSePAiO1MZ88immmUkAKSAEpIAWkgBSQAlJACoy5Ar9eefjUm5b98T8shv+0YJsadtocVAKaqMlLAMLTsEm+w5r2ru32X3hVT09Pk/zGqqtOnLk5+j4Stf5wm3K8VY8h2sqd5p3zV+SrTCIFJtjGfBKtjKYqBaSAFJACUkAKSAEpIAWkwKAK3LD8w5v84677jsUG+1NmEZvy6HnF2U2ckpcARLEpD/brRmgcuNMBZ15GbmPX+rTuI3CNf3v4TTmyLNwfrXtpCD5UEqqTRAFtzAdbaHFSQApIASkgBaSAFJACUkAKjEsFfrx49ma1dXfNDzHwLfT0vIWt7L8x7uQlAJHT1jXrcc7Os5deTm5j12sW7vs8bLI/9Mg25Xxf3jzlNe9b8vuNPS71P/4U0MZ8DNdEl5YCUkAKSAEpIAWkgBSQAlLgkStw6/yDHv9gret4s/Bu45F2vJX9N9jkJQBhVryBvs1ifPVO+y6+0UbhwFi7Q4xzMcStistVB1Qw/GlBwWJTbvaH2LX22BQRTDIFtDHv/AXXDKWAFJACUkAKSAEpIAWkwIRX4PplBz6jsVnveZjIHLPYjU02TFh+zqdim8sNb5WJMf622Qjv23HO4pszv7Hxnvr9Tw1We4+Z1fuPEpxTxWhxjtY0CyfPmPvdu03HpFRAG/NJuewbY9LqUwpIASkgBaSAFJACUkAKbBwFblhy4HNrzeaZ2MK+Gleo5403fLi5JC8BWZgxBvtdX7S9dp676CZyo1FjT0+t2+zj+IHANr4D739RDsw5GMSa/SrE5oU0VSenAtqYT851n7iz1silgBSQAlJACkgBKSAFJpUCN501b+cQGstqNdvFN7mx2MwW5yxF8hKQhYk9uV1fb8Z37jJn8U/IjVb9/tN/tTPGOg914CU5MGdhOCKraef9/THNXyVXMAkV0MZ8Ei66pvzwCihDCkgBKSAFpIAUkAJSYOwVuGnpnDdYs7k8WHgV3j4XvymOYbW2tHAGeX1exMOvrBFmbz9n8U+ZNVr1wvkHbRqCfdqCTR1wTQwMBXRxhsHR3xO7pn1p1qwVDfqqk1OB2uSctmYtBcaFAhqEFJACUkAKSAEpIAWkwCAK3HrrQd03LZm3l1ntdLPwFOOR9rIJyKAmLwEIbHRjH06X1bq6XrPD3EV/AFeJwtvI5TH1+9+Od+CvGnAZjAIFdHGGgTTrCxY+PmPuIv1uuQsyeU/amE/etdfMJ40CmqgUkAJSQApIASkgBSaOAjcs//AmjV/0fixaXIhRP5271w35Q2/Y6J5fq3cfuP2+p/0d7Ue1rFrV0xVD80hcdAvUVsFeHAV+cYZRlBBuqK0LywtH58msgDbmk3n1NXcpMJIKqC8pIAWkgBSQAlJACjxKBVYef/jUeu892JTbUSHYdO8u7WUTOOWbdVoVMgbrs2AXx2lbzd5+v9P+l+HRrDFa6P7dr/4LPxh4Zdt1oxkKqOIMoyjRHrQYTn/1YWfdVRA6T2YFtDGfzKuvuUuBCaiAhiwFpIAUkAJSQAp0pgI3nz136y23vP/kGOOnMcNu385itwvb2re0yUuQ4g9aM5xodz6w146zvvYgudGu15yxz/aN2Pxk23UxRhRQxRlGtdxc6zX9JfaqIpPY1sZ8Ei++pi4FpMB6FVBACkgBKSAFpIAUGEUFrj1z9jbNhi2uBZuDy9aLTTkslPYtbfISIGw0sZn/2n2h8ckdj1wxJpvy606fN70Ww4dDCFM5Jq8YGArM4gyjWhqxKzYjpnUAABAASURBVH5Ub8urkkxuWxvzyb3+mr0UkAJjqoAuLgWkgBSQAlJgciuwqmfXrhsXz92uu6vrUiixK7awxf4EBnzfdBOLOpAE80/UD+40Z/En37j/kjVF3uifY+3BXS3EN5VXxqBQ4BZnGK2CnyJYbJ6y69xzb2mRsia7AsWNP9lV0PylgBSQAp2sgOYmBaSAFJACUmAcKhB7emqbPveZs0LNVmAH/oJiiNjIRlQ4xRmGl+QlIAXz7hDsg3+e9sAp9MeqXnvSvlvGph2D65e/E4+xwS3OMFoFcwtmvwqNeFKLlCUFzLQx110gBaSAFJACI6KAOpECUkAKSAEpsCEKrH7uHYc0LSxGm61RUbCRRYGBfTrPuQ4ko8W76r2NXXeYfeayWbNWrMuZo43cZ4dp8Xhc999RK/8CP43ZyXRCMizs4cOyXQ5Z8XPYKlKgVEAb81IKGVJACkgBKTABFNAQpYAUkAJSYIIrcMOZ+z9t9bK53zRrnhjM6sV0sJFFoZ2AJmryEoDgpv0WC/W3bnfg0h/TH6uKfXa4+vR93tGM4R0+BowRBWZxhtEqSKYTg/2sVm/wDf+AJKSEK+e/6zlXz581A/iGa07e6+UXzn/bpmyn2vkKaGPe+WusGUoBKSAFpMAGK6AGUkAKSAEpsDEUuObsuVuHWn0ZNqEHt/rHHhWFfgKaqMlLAIJlFV4577PDfgtvpDOW9bqT931sPdiHQ4ib86cFxTCLc9u4MNnkPxCsb/9d5q24M/klXHf626dfvWDWUfUQV2Lzvrxm8ZxmqJ0/PUxduOrU9zy9TJTRsQpoY96xS6uJSQEpIAWkwLhXQAOUAlJACkwSBeLymfVbzj5g16kNuyUEew2m3Y2Kgo0sCgxLQBM1eQlA9MYYz5o+7YE37XTAmb9FH60IgqNdenp6as2pkT9c2AU/KAjFYIpz21jSphyRZohh/mve++0ftcXhrDpx5uZ9jWnHYU6fRt7zQT3eLG5pwbaOMcyyxrobVp387meCV+lgBbQx7+DF1dSkgBSQAlJAClABVSkgBaTAWCqwatWuXavXTT+k0fDfJ39SayzYhqLQT0ATNXkJQPRGi1+Z0j3tgy8Zw98nxzjKMmObX+4azD7FnyYUwyzOZQKNWHDF2X7TZ+F00tV6/cmznxi7wgnIORQVL8pxTu2YB88iN+i1xnGXLp69GTnVzlRAG/POXFfNSgpIASkgBaTAaCug60kBKSAFBijwY2wmp/3lmafgDfBXEUx/5A2WYcuJkixCqv3IYPc3LRw45cFpn9l23wX/SEljCt+fv99TQtO+gClsVoy2OLcNKm2uK5Gvz3jf2bdXc/j74+vq647Fm/HZ4IOlNrC9tNrCis0dpzzwwAs9oFNHKqCNeUcuqyYlBaSAFJACUqBTFdC8pIAUmCgK3HrWAS94sNa1KMRwAMY8BTUVbjQLE1Zh+Dl5CUD9IcZw8Lrf/eHsbQ9e0At/XJTY1XcQBvLyYpjFGX6rxIIrzk5fGJ/7wlMDtt7u4bTqa3s8dnp906PxA4sDgtmAPVmrLSwUs/AU/IDipaajYxUYcBN07Ew1MSkgBaSAFJACUkAKPFIFlCcFpMCwFVjO3ydfMmevvmbtPGxG90JHddRUsMtEoZOAJmryEoC4zkJjr8t+98xzZvRc3Qd/XJRrT99nZ/yg4YPRYjc21QPHNHBT/qvQ6PvIjBk95RxuOG3mVmGzKV9pNuOH0EHd35SndvDxIp5n1li9RFeI8XVkVTtTAW3MO3NdNSspIAWkgBSQAlJgAiigIUqBTlMAm/Ipz1y32byG1RZiV/lC7DdDa46tjSasFp23oiCR3ETgikZ3/V07zF76w56eHvqgxr7cdOo+T8J8zo9mWw46GgTJI06wEGxdDHbSLgd/+zdO4LRq4QHT1jbj52EWb8pTG/heclto58XJfArhNUiHRJkQdpIC2ph30mpqLlJACkgBKSAFpIAUGKiAGCkwKgpcf86c52zTO/2b0cKJFmx6+0VjudGEVQklr4D7ACf09tX23XmfM/5SSRpzc9WJMzdfG+yzeFO+VTmR6qiwY6aL8ROKGsN11qgtwwbd6cvnz9wm9D14erDaQSCKP/RmrQNccmChJKcFIT718vl7PbdFyOokBbQx76TV1FykgBSQAlJACkgBKTBmCujCk1mBG5Yf8NKuRu2cEON7safsbtcCDAq5BDRRk1fAvXgV/B933b3Fx3aZt/BOBMdVqXXXd8P49h10UINsypH7r1p93ftmHHy2/8G678/f8yndIZwcY9wX093gTTna8F+8h9AMbxp0DCInvALamE/4JdQEpIAUkAJSQApIASkwiRTQVMeVAvyr6zctO3B2vTdch/3ptthA9ttfgEHhoBPQRE0eXkHHYL+rBdt3+/3PXLD7Ed9ci+C4KpeeOvNZVgsnYKj9/hUAholJ45z/MT5NvlB/sBHjR18977zfkeCb8kaon44Z706fO2zHdALfslpO4tgdTQRQsDPfhZ5q5ynQ74PTeRPUjKSAFJACUkAKSAEpIAWkwIYqoPyHV+Bnyw/c6qF695eDxZOwP91iYAvsJFHIJ6CJmjy8PrZgl9Stvud2s8+8CIFxV64/efYTp4X6qdgeV/6rtzRMTJpWmg1NryGEC0P3uvMQDtiU79wVwrfMwpuNB0hCrq22sFAyn7GgcEYhFyy8hKjaeQrUOm9KmpEUkAJSQApIASkgBaSAFJgQCkzYQV531pwXr+m12/D29xDsrzcfOBHsJFHIJ6CJmjxAjOHbT5+61V7bzT7jxwiMu7J8+cx6X1fvoRjYrqjtJWICYIozjFRAN3pj/PSMud+9++qTZr6sHsK5CL0SFXv79uyWBwvFcyqngsIZpUI/g39AruLL7BAFtDHvkIXUNKSAFJACUkAKSAEpIAWkwOAKjBz765WHT1191oHvnRLrF0drPqV9z5ivAxaFXgKaqIUXov01BvvoDvsvmrX1rK89iMC4LE+9r3uXUAuHYXCV/+4NHnbfOLf/83USZmsA+216f/OOqxfMnG3d4apg9lRwI7UpZ1eh+dCax9FQ7SwFtDHvrPXUbKSAFJACUkAKSAEpIAWkwEZR4OblBzz5X3evWYi35F/DW/Kt0/60dS23sPlGoZmAJmrhYaP6+xjCvltMfeD4ELBFR2Q8lutP2/cZjdg8A/N8Ytv40qSL2bQimEkzxuZp1/zlhcvXTQ8fxhyPR7T4b9VSG/heWm1hoThZORUUzigVOpkxNNY1N02OoIMU0Ma8gxZTU5ECUkAKSAEpIAWkgBSQAiOtQOzpqd1yzuyXxN7aeXhNvA9e/24+6J4RQRZevz0eQcc+8Bc313Ztu8PsRVe/ZNaKdfCHVTZ2o1ULD3hsX2ieiOs8E7VV0ga7fW4Ig8Cm/MfNhp31mqfcjs28fR4TfiwikApBN4pTy4OFUrB+pte6Aj2nqyfqaBam9U0zHR2ngDbmHbekmpAUkAJSQApIASkgBaSAFBgZBfhX12953h8Pbja6v4ced8BOEy/MYQ0o2EmikE5AExVetDUhhuNCfd3cHd97+r9AjutSi2vfb2a7tQ2ytWVuo7EBhx/vj7Xa8no9fN6CzQFRlNSmcKhcxYIsFQ9v2+2iGOwuz6rEcg55p3FqrK1pD9cSpmMsLWrHLKUmIgWkgBSQAlJACkgBKSAFRk6BGHtqD9W6vxAtfBW9PsM3hxHWgAIShXQCmqju/SWGsPdDWz/rs9vve/bfQY7r8v0z9nlNbMZPWbRNyoGmDbbPpiRhkAgR0XgJ4Ei4rwVbFLCFUZwRKwx0zJIcCxZ6gzUPb1jzOPzEo68ayzmuO53UyZSuxkN0VTtLAW3MO2s9NRspIAWkgBSQAlJACkgBKfCoFPB/ur74wJfcfNafVuIN8OHobJpvDtPGEH6lgEQhkYAmauw1C9c0+8IeO85edMmMGT38p+w2no9rTnv3y7Ep/04ww3zTSNMGu31uiDmBU7RmsLAXrCeADaiQCp4bxanlwUIpWJyD/asZm0eGf/vXgnq9FrEpnwq2XyENqmwXYqM+9UEwG1aUPe4V0MZ83C+RBigFpIAUkAJSQApIASkgBUZPgZuf+8fdm/XmObjiG1F9o5n2p+62TtgtotBPQBM13hcsHtddn7L3TnMX3QJi3JdVC2c+OYTwBQw0/cE2WGnS7XMjj4pdNM+odcSLDTkcS21osiJGQIWFAsMLzL+ji/et7X3g1Bkzru6LzcZmCPT7o279N+WGn5PEuEnX2rtsnB4a1vAV0MZ8+NqppRSQAlJACkgBKSAFpIAU6BgFbp1/UPfqpfMOiSF81yy82PxIm0O3qydsLVHIJEhmvA/Ge7fbb/FR2+674B+wx33h/1ceGrWPxhiLH0T4iItZFWcnipMTfnK/ZcF9hJty/KTjtxbqu+92yLfO2/2Ii9eipdWs61lm1o2aStK97QJoibfsOx94ATVOeZMSOnLStY6clSYlBaSAFJACUkAKSAEpIAWkwCNW4NazDnp8c/PGl/BK9hvBrF40xOaw38Yw83jb62Yl3IvslaEZd95h9pnLA16Ze8I4P/GHEU+4u/7hYOGD5VDTBrsytyLkhJ/cb1lwUxtYXloxWChO4gTz2hhr++528Lk/gluWGJsvKh2IizzswltMYZGNtxe2zhtfgdG9gjbmo6u3riYFpIAUkAJSQApIASkgBcaVAjecuf/TGs3GedhYH4aBTUHFphAe94HuVE8gUcgkcBPZ36g345zt5yz+CYkJUsN9tXvfUavFT2K8deyHfd6w3SSW1SfrJ6daFtxHtilvotNltdC952sPXXFzwJXQslpeXjhQkkbbBVpEbNrN9FQ7SIE0FW3MkxACKSAFpIAUkAJSQApIASkw2RS4+ey5O9bqXZdbiK/G3NM/pU6bQxDtBbtFFHIJuL28I0Sbs+PsMz+63ZzF/2RsotSrTp/5/Fqwz5mF4v8cx87ZcJRzg+3FCT+1XLdwemSb8vuh7wmNWnj/jIPPHvDP+y+fP/Mx6GY7iulX8RP6LksiAM2mXVXSMjpKgY29Me8osTQZKSAFpIAUkAJSQApIASnQCQpgIxhuWXbgu2MznBmCvbCcEwIopdsysCtEoZ+A5g0h1vbdbtqas+hMpHrlafs+ox7rZ2LMLzROKE2aJrhWccJPzrUsuKkNLC+tGCwUksHs97UYDnto7QMfe/3BK+4h17/Wzd4SQpzqTfxUzUgEAOVBe2jN6mpUduco0CEb885ZEM1ECkgBKSAFpIAUkAJSQApsTAX4x85uOXvekU2Ly3Cd56IWBRtNbP4Ku+0MFoUUAZvNpllYeMfUNbtsP+eM68KsFQ2bQAffUHeFOB9DxltqnH1nzvfVtCuVk00xsu7SYIVWhFxbMVgoif8zfvAxc9dDVyzJf+Qt8e0QG/t6Ez9VQ4kAoDDwP2/86GVraKh2ngLamG/ImipXCkgBKSAFpIAUkAJSQApMYAVuWnpAy81KAAAQAElEQVT4Ftv0bt4To/WYWWUvECtbUETKgi0hCl2HYH9vWPMTD03d5IhZE2xDzjnwhxJT6vVPWYy70QcW4OfKySfrJydbFlyIh3NZWjFYKAissRDP7Q5hx90OW/ED+OstV5zwzuc3o20/UPyiI/LJYh/X8DSseuQxW9sHjt7VPoQ6FDK23tqzq/fxgcGQ3BD18J4thjXuSdSo8mGcRLMe51PV8KSAFJACUkAKSAEpIAWkwEgr8Ivz500PtuaUmoWPoe/NUYuCjSZKYbedsSVEIVVA+FFfo7H3ut/+6SszZp10P/mJVp98X/0T2Oy+H7Xb0qSLuVVm4oSfnGxZcFMbWF5aMVgoFmxdCPHjm2427aBXH7ziT540xCl0GX9AML09hR2BAaDAMHRrD8YYhv+H3/psH6uFVda0VVZDbaAG1P4Y4ipcDJWIaqjkiFZbZSU2YaM/y0gbNaIaKtErrkE0q/zVedMxiAK1QThRk0MBzXJjK7D4x5vZuTe/zc5dfZSdvbrHzr3pv+yc1QejzgS/y8a+vPqXAlJACkgBKSAFpEBWIEYL965pzsama69oVvzldQ9Gg+9W+wksCjlAA+2u7+sNe796zpLrZvRc3Ud+IlW+Kb9mwbvnQYejMJdNLM0ac2ufhhN+cr5lwUVjnMvSisGKRo1+0hXDy2cc/O0Tt99v2b1l4nqMVSfO3DzG8FoMpW09PN27dMtPWKU/Iffn7gznFCK6QKeYPK5nBOMReCKfEQRSjQmkQ/IBrllGxo0HCObRJML1H3gQeSG2JzKuOqQC2pgPKY+Cj16BSdbDubfsZOfcdI6du/r3NuWh+/FgugDPoh48u44CHosH2imoyy02r7FzVv8f6lfs7BvSf49hOqSAFJACUkAKSAEpsFEU+OUF87AJrL0Znbc2gdiAoYDqX7DDQiELeAB7q1OmrgvvfPW8M35HbiLWJ99bf1Os2Wfx/Vjd0qQxt/apOOEn51sW3NQGlpdWzK2HQJ5Sr4U9djlkxe2wH1FpxsbzMZ7dWsneF75V9NKiaUW7rREf+D3NYdUmWmEhvWdumnmpjJaNhAQMzGlHEJ6f0ANOoFMgaBjmtOHgdahXRlDWx5PqUApoYz6UOopNHAXGeqTLb3o93oJfY7F5nYXau/AMe2bxcMoPK2BBYKR4etG1+ASL8Ujk/9DOvvE6O/em3RFUkQJSQApIASkgBaTAiCtw95raptgrPafsGI5/O1IS2QCLkrxmCOGY5pQ1H33FvIV3Jm7CweWnvZsvQRaHaE/3jSlm0JoiHBYn/EQP38o5FCdoVRjFuS2rab+JsfnGh9at+chr3rdigzbOzZp9ED1uhYqSegWgwK+WaKFm3x3yD8hV0wezQwj4vhOR9H0oXIuF65rQbxFFgHFYJU0d0HzwfCSW+TDYn+d7A7MuxFWGVKA2ZFRBKSAF2hQY4Cy/+f/ZuatvtVi7DA+pXXzv7Q8hZOKZZCQytp5qThsPPqsYD2Ena9pFds6NP8UmfVeGVKWAFJACUkAKSAEpMFIKNPrWRgsxvbeM/l3JwL4jvp0pWFh/CiHM226/hcftOGvFgwU78c7fP3Wff59i9i3MBxtgnDGF4gwjFyf85EzLgsvv6wC5lLEY/xliWBQbzde99pDzvr+hm+bLT97r5fg2cN+i39QrAKWgyjOYaH+px+7vltRwDVzQF5jIeTmyMxi4jPH7Vr8z4BMTlDQNtivRcHhDIEqZD8Npohtm6c5Dlsp6FNDGfD3CiJYCD6vAubfsbjFebxb+P3/I+UPKjMBnmaMbhoMPJTyc3CeCYkJJwyAd7SUW7Ao764b/YoaqFJACUkAKSAEpIAVGQoHNzdaYhV/hexcWG3jgexEU8oCfNprNA7ApPxPfnsAlO/HqqoUHTIvW/EiI9sw86QGTccJPxqNlwfNNKDCVViyuthDm3Pk3O/i1Hzjvjyn8iOHSL71hsxDjZ9AA75FTrwAUULnQY4Vfsy/NeP+KEfpje2lFAa1vS9N1MuHzRgLpkBHjYJw+0cCTIjKPNpE02xOZ5/kIYqY4qwyhgDbmQ4ijkBRYrwLn3vJhbMYvMgubAs34UPKHj7W5lvkSDQefWoC2fDy9nHasW6gda2ff9Oh/MsrLrLcqIAWkgBSQAlJACkwWBV62/5I1oWkX4NuNQd5+g0XBdyH8TeSf9tqUN+48Z/Gqia5NPa6ZbhbeFi36ngdTtLbDCT853bLgcnMJyAWxGILdZ7H5uYfWrnnNbgevuGhWz4p1Ob4h2LXZprvgG8bXoBbN2HlhDXKOfw7NsGSQwIZRkRPCCrd9/4kucG0rv0+txHNeRuZBALqD55t5N4bD89AgIyjTG3OqMGT1m3TIDAWlgBRoV2D5LQeaNb9akHjo8CnEZ50jWH8IEVH59KJPZJwUkc1oE/0ZCIPIvDLf3mFn33iOxegRpk/IqkFLASkgBaSAFJAC40KBxtT7VmAgX0et/MVwfA9SlIfwHcfpD9bX7rbL7AV/Rc7EL721rc3ilpwIpkhoVSf85FzLghvbPIP3j2Dh9Bhsxm6HfPuoDf1n6+ixLKt6erpiDAeD8HGx8winvRRMjIYflNTO/m33nyrr1Z75iL0QAr6nRHqAJAXw2vy21AmEHQuiSCiG0aIxoCKMwIB8NAGNM/JhMO75uB5JvTGnCkNWbcyHlEdBKdBPgeW3vB0PtQVWPJXMHNPDh083Pnv8IWR4KKEyjrARGc/IPMNBZNwfXvAZZ3ui59u77NybjmNEdWgFFJUCUkAKSAEpIAWGVoC/K775lPt7LIQ9QojfxTcr94Zod5vZ5TGEmZtP2fTIGfue/Q/4HVFqobaWE4k8VasTfnK2ZcH178OARemzYN/Ct2mvn7omHrHb+1b8oKCHf24+4cdz0WfxB39xYZRKZ/RYQRXwz75o5x188A96wTz6golgzQ3XdzAefp2QfKB//5kwQSsfRJnvBnrICBPhojkMp4k0gAirDK3A8Dfml/422mW/i+YIu8TfFNylCS8ZDME5T/x1NNqXJLyYiFrirxCnX0HGLoZPXElELfGX0RhbWUHauV70i2i02/AX/7KLfvEnu+iXP7eLfnEL6lV20c8vRN7ZsE+17/3ia/CPgj3XVv5iJ1v56ycMLauiHanAt255tsXmuXg4FZ8bf8bkh03CBMVDiSqAyHl8GPrDHgRoQ0dlHn3QxVORRiII0T5m59y0vekYDwpoDFLgkStw/He+bMd/J9rx56ECv0FELfHb4OlXkLFvwCd+nYj69W9FIzcYksv1ayui0a7iN5a/6ZEPeIJkfmnZJfals6L997I4AMnl+sWl0XPWi0uiMfbFhMcNhuCcJy6ORvs44BeXPn2CqKVhSoEBCrxk1op127/njFXbvWfRO7ffb9Fjtpu9aMvtZy96w477LfzeS2addP+ABhOYuL/W99cY7A9WPfi9lX8DVpDuFia+DSu9O9FuZa0eXrvr+1bMmnHQiv/Z8chH/wfwLj/xnS+0EI7H5bo5hPJqINoKAiiG7wgvX/2Pl61uiw3XCfgmFMXYa+rceOAimDgt1BRoy0MCfYDnZWQ/aGFENjMcRMbb8p1AUOXhFCg2GA+XNVicmwzeUVnrjFwczycBo8yjj9VK4GlwrTRSgGA4SoTBPPZDRMhviowIu5mR/fnNAILowcrJ+0FHGRmKxn9Kwn/q8gK426LOMAtvtRjfbWbvxfU+ZBZ64J9hzXidNfv+z77383uxWb/Rvnf7SXbhz99nK3/2ctPR2Qo04ufNwjTe9sYj8hRwopEwgWWkwbAjDPLsoIqgGSZtpcEEBAiGo2k9pmMSKaCpdo4C+Bznz3X+PJcIg1+nql+PigcBpo92CMMwY9x4gBiQTz5V5vWPN1KsowA6cD6DzZd8roxTzxIRcH0SWuonY5lHPukP8DCxNBhHHypSQAqMewUe39jyPnzsz8FA+1DxLT3P/oGmwSeEo5/8d7Dtl8HisbUY9p52v+39mvcu/z4eDa0Gnji806qv7fFYq9mno0X/XrK9U3qs6BuAwrH+va9un+zp6WmCffSFvWAy7LjtcRbZNZ9rNBISqE6JMDycsH8HoNmL0zR4HQhvGckVK0BLdT0KDH9jTrENq1BdJLh+HUcG6BFBtOWDpw+6uDlo5DzEcr80E41NsZE2P5jvhjd3i/0Vhlm+CYjGo5qPDskzn+jhFE9Ayq9HI+dlJFfkTcf9yreYh1qwBdYIP7Tv/TzahbffiA37sfa9219nq34/jemqHaDAObe+wkLgD2rMfP1xHxkPIgjeHwy4Sx8xgN+gGRkHbUTmGQ4i42xPxE1lfl/mAHKMgfhGvTWnFqojqoA62/gKDPg845L8eAP86wzj/vnn55xkRtjVPLjF8wTxtnwPFKfMZyRb56nTahImzzMjpwl5CEVlHgjG+Rx1lz6iAJyNtOtqPDwBbkZwzEsuAhUCpooUkALjXoFtD17Q29Vb/28M9AZ+i1V8juGh8KMNYMFeOd6Iffne6+6d9v+95qBvfeo1hyz//ki8IWfnufZNre0eou1tuDBKpoEVDyYKOLNmsOPefOh329/2e2SYpxq2fet9HuKq/HrEwRUPRlwED0DQMKyk2R60EwPykVnmw2Dc872BmX7HHAINXWpDh4eIlmIjx0UnorJgLcwXlQYXA+iQ0XCAgGvMY3siVx10sdjmkGmjUc2zdHg+bB9PRnaMgOeDY78EVuZlZJw+kRybEVnJl4gAuiu+iQIJF+dUGIDp+QgE2x55/K+uLrf7H3zQLvz5lXbh7R+zlT/WP3uDTBO2hOYnsK7F8LHMxvvReHD9QbTWH/ctfcQAxryMfh/mfCuO5Brb57x8P3o+0xDwPP9vNUioSoEJpcCkHqx/nv0D3JIhuwM+9zmQUunSzHmOeB6Q835pVCrjdImMs/2keGOOSXO+gPKxSdsoAAzqwQCxf57LmfI8HwRd5jmiPQto87jhYACgIgWkwIRQ4NWHnXXXlPuab8L3cUfim7SLMegfNi3eADwnWPOIaI2X7vrXF++82yHfOu+NH12yBo8K/8QjPmLlqhPe9uJg8Ut4FHWvt3MEUPI1V3fXu/mH+rL/6LGJV+b++MJViG3PORCgzZ9zNOBjsP1dj7MdA46Gg/kAltTMICKbm4HIeXpjbg93DH9jTpGz6NDcxSe6wctykUAwzxcFXDUfLml8QGAhz9sB2QxMdvEhsjKP7cuA4VhfPnjmeT7SqsXHA4LIeEZQRp9YrYwb+uO4GHdMCaCL8dNPAUI135q7wf2iNbruwFv06+2C2z9k5//iqWyhOkEUWLiK//LhLcb155B93WE4csFpJEzA289Iu4FcRxI5gRz87PI+g1vcT26Y8XqMG44C32xLrvs3eCpSQAoMVGCcMvw88wMMJHCURLgP//WNiWiQnw+ObAiOzwdAW2GcBJFxNp8Ub8wxac4X0FaoAwlH6EbMeY4UyBN4wuOXPvPgMp5ceGagyy1N1QAAEABJREFUkWDFwUBh6SwFpMDEUIBvv19z0PKvP7DplHf3xvjO0Gt73fmXOGfXg771zd0O+vbPQs8I/XPxQeS47Jtvf2qz1nUGvhV8amyL02MFCUCBwRIbIcSTrv37i0f2L+OXuz480HgxAMZkA59vCPB5yYDn0Tej689BPh9bDc0YYJ7hICJ94Nc3xPTGHCIMXcolGjptkGgI+Dk81HfxEc/IxYFrGclz8aqIZh4mlgYTQBAMR4kwQBtvAr9JEGN/BCLCBCsRBvPa8j25ODkPk+gNYTMf4DcRsVqZx3hGxjieEnE97yehAxIycmBwrcAdAV+zWvNPduHPltr5P3uu6Rj/Cmy2+RuwbtPK+8PXE8N29IWGkzAB8v3ZZaWRA4bDGxaYab+/QFXz/b4jh8o8gNW73klQlQJSYKwU2NDr8vPODzCQwOYlwvDPeULG/OsJDeaTh83nA8CIA/IZSJVxts9IuukOrU6qEAbT4dQ2RA+28ny0JRoJ2BnZX1U/hrEMHiaWBgNspyoFxk6BW+cf1H3tsn23/OHyuU+4dtmhW9Ifu9FMjCvjIx5332/Zva8/eMWfZrx/xd9m9axYt7FHvmrhAdNq9fon8GjZ1h8jg10QAZQUiUi179/1hCnLRux3y1PPFqFAfo7xMcaLZsx8RvKGhCrCNXZBNAZoJCTwOiXC4HPW85mHYB+qypAKDH9jHmPduCjU2kU3o2s8sBZpV0IPFQTzmECE65scIhedfIlIp888mr6oMBy9gRmvZzzgMw9Q9pfzMjKtWp0HQTQ2hF32B7ss7BgO8xjPCCo3K5B56IdxEu7SRyIAZyPd0oMJVsd032O18Gu74Laz8AZdb0FtHB/B3oz1smIdzQr0dbRiXbHQ611/5PH+6d8BaDQum5f3L/MG5CMz51tzV3gqUkAKTBgF8Hzg57p4cGDU8PPnmc8Nft4zImr0iazMQ7o/H9wHwXg1n3yu5A0NMpKvuUOrkyqEwHQ4tf560EeoKEyr6OEufUQBOBvlKh7EhsMT4GYExbzkImDm/ZMwHVJgzBRYfc68lzU2b3yx26Zd2Nsbrp1iay9qTu/77xuXzH3FmA1KFx5UgcZD987Bl4A5eGpU9lzwQHoDmChu8hkD+5dd9SmzZ81agRegiR5J4HPTAi9lBB8GXCfK55sTZkzAgIwHkTTbE9lwQD4SmQfwr1uMe743MP2OOYUZutaGDg8RbXtjnlYhAdfKuJhucDEQcMhoOEDAtf55xgMBhIvmMOAa87i4JRoOBBCGYeaLb1YgeEPA8639YB4ZIuMZybVVtKfPOAdC1/NBsnsAaeN1DEeZ54nm9IA8EpU4OwhhHwuN2/EG/cv6Y3E2Po8Yn1muZ1o+c8JwJKJcf1Bt9wniEZzn04Dv604ONbnG9gwbCLYnMg8pXkAn99/d10kKSIEJooB/sDHWCubP82Cfe//8I52FeY4wyHt+6oc+Y9XKOH0i42hmD/PGnOkTsHJmZtV5cr6GIyNMsyLNyjz4jAPS89S/Fy3yqGsKOMB3tOKAawYio+mQAmOjwM3LDtwlNux83Lwfwh25E0bBlzs7xBg/iFv9uzcvnfNacCrjQIGrTtrz7daMX8NjY/PWcOBlByZK8tx6oB7i0a89dMWfEznygJsE944Zbp4Bz0EfQg4kTNDKB5Hzohvm/eHsBWHvl9fxMAjPA3qCTkMpMPyNedsb83QJ19xXAQQRRF4Md+kjBMDZfJGL1bNiURFgHrwWDQK0x7nIZYBJCCBMq6R5vZzn6NHWiXF6RMYzkqNP9Jo6ZpwDpcu4oycYaR+X8UiB/vkcGIbpeWxPv2iIRggUzWqgP2L3P/ATu+B2/qV3xFTGkQJPwfqYcf2K9TI/sHy+rkYjBRIU+cwCwTAJbw+f+YSMjNMnMo98RvKpG6fNnmLF77yTVZUCUmDcK+AfbIwS2P/zzK8XoP05wudD/twj2/iBL/NhMO753gDhhFY5GKdL9Hw44+ONOQYyogWCoL/qPDlfUG2FcRJExksESd/1hu0IPRnPujPOq4BmBmlfJ3cYcEMnKTCqCty4+L0vjBa/jtvyGbgwAGcW3Lso9LeOFj7zk2WHbkladewUuPykPbdvxubXo9kmg44CAZQUKiys4YXrGr0XJnLkIeAKKMYHGi/JO8ZwOJKAvd7nIeI5LyP7YRMiwm4SGed1iOyPz1OiJ+g0lALD35hnkSl6uQi8FIkKlnnkkZiAa1isUT+CbqW5sT2alchY0RAWAv3zDYTfDAmR1Va8P7arxJnPpIy02Q9x0HwGUNGNeR4N9MdxOcDPyDhcz8vIPPJEzzNL7vPwhf9GvD0/1nSMHwWiPblYH19ALFEamrvtC1jkIZ5pEjmP9xf5vO4ZGSdPZD75jOTRnbuOIDbreiJNVSkgBSaCAvxg43PLzzWBQy4Rhj8XEjLGvIyg3eTXocLA8wf90Wc75yqnzGdkaOP8Y0j2PIY1CZPnmdH6DYk89SwRcerG5sTWgxUBlDLPE8zDkLvE0mDcdEiBUVVg9fL3PqvW1VyMi74ctVVwL/M2zUS0sOMDce0rsy8cfQVWHj/zCXj0fD2G8MzW1SNM1gKSVTg4o/yiK9Q/vvsRF98Le+OUJrotn3OwOYiQsO35hgB5TMJpRxCgje2JBp98Rrroyl1HELg3zfO9gZl+x5zKDFmHvzGPEV/uKTr6d9GJqCyg8d0DLVQuBggujhFBAYrfPYDNRWV7Yo6XiHhq3soHxziBlXHH0jBjf7we0XjwgkTUzGcE5fmOPKXKOE0i+8lIrtJdMU8QjBsRCZ5PRGUBXeTR4ThBtOWDT3TK+y+74Lb/se/89DmIqIy1AiE8VNyeWDeOxYELRocIom096SMG8PXMyPsDtBHZzHAQGWd7Ii/E+4fIPKQYkXmGg3mNoEcbpFCRAhNCgQGfZ4y6+nlmnJ9rIkLGz7ulo5rnFAjmteV7oDhlPiPZEPzJQrNjaozFnOIgehSRNFXEqSfzHEFTjkTDM9L+nDYeKdCWDz7RRV4w8+czeBUpMEoKfP+MuU+w3uYi3Hrbtl0S9ypvzzbOrMtCc+d+nNxRUuCSr87cakq97zo8L7bDeqUnUmWVYKKk0RQWHkv3NPriu177gfP+mAIbB0IIvp/ig4+XhosxGl2MF8jhesCKAz5dOkS43p7IhmxPLDpglnfjBu5NYzwjyS6eVIdSoDZUcMhYCPVycVx0ZHPRAL5GxlUjkdABviOTYMA15rE9kQ1B+6oSGc/IeDXP0sE4TV98GI6poeeDY78EVsYzMk6fSI7NiKzkS0SA12GeIwO5koDt+SmP1yMNl8Oma6WRAgQGiP3z6Ft4mdVrP8Xb87mmY4wViH8vls8Xxm9Pc8JwpAWsrn/bfYK4NwNyvb0dCLponV1je9BmCLC98XACBhA0DDPm7bPzX0yHFJACE0MB/zzzA4zPcR5xdvl5dhrE+vLYJuc5egM8hxIynivjtInsD93i6/QgiUyawDWEYk7t8ywmVEQK2ygATObx+UukLqDoEgpMeZ6PDugyz9GzoDeRRBVpb7yqnqUAFVh5/OFTp04NX4C9I2qr4B7F3drySwtsA8HSlzFaCqw6ceaTu6b2nYYVeH7rmvCyAxMledmKDzWbdsy0u++9PQU2LvhjDNcm8jZx5CVhgLb8HMwI2p+TGcmzXYmGwxsCUXKeP2/hV/P0WomCDFmHvzGP1rAserkIvFZeHCIC5eIhVs2HawgXX+1o5HwGUJOLbyqszGP78u4wHGjHPFglzevlPEcGK5VxukTGM5KjT6xWxjkAXodxx5SAyxfjp58ChLZ8EDmP7cuBsg0CCDO9RYMA7f0G28Sa8Qy74DY+kNlAdWwU+L9ifXxhzNfLcLiL9XIiYYIiHzluZGSDnEAOfnZ5n8H1dff+EOf9wjjMshuz39NVlQJSYKIowA82P8hAAodNhPvwX9+YiAb5+eDIhuD4fAC0FcZJEBln8xBSAwY6pMb+b8wxr8GmGSkAYo6QgZjzHFPcEjLO5y9dxh3RngXNi+czHQaIHVE1iXGsAG7JsNVWa76KIe6P2nrfiMDgdyFYFGva9chXGUUFVvXs2tUIfV8zC2+18uBiJAcmSsvJVgznT6ndc8qMnqtHcduKBxoHA/DHH9ENDioFcI/ZoM9DxPl89HxvaJbzDAfCdAd+fUOsdQfDURlMgeFvzIPV/YsU16RcBMNBAuCrAnQXCVWE62FiaTABBAHNnHYEAdp4E/hNQpJEQoTLe4N0zsvItGp1HgTRG8LO/WYEVRbmkc/IAK9TYr8BuIuEjAYDrlXRrwuemIBhukUeGyDg1zUe/2kX/OzzNFTHQoH482J9uC64foJyvdzAehETFPnILY0cIFfpINN+fzFWEvh4IY9upokWbnXQSQpIgQmiAD7H+TnQ//Psn3vEM/qM4Gd0Hg4RYET/uoCOiOSqlXF/DiFOZCwOlsjABK4D3phjvoNNs78eSCu/WfR8EtQh4aD5iDPsy5INIniVDVBAqRuqwKpVPV23nH3gobjdD0XbbtSixJg/3YVfnnGTosBd3dy0S98rQIjRKvzn631PeOxSLM27sTpprYrF8DHARHET39wlNP6E8cbHTWnOm/H+q+/P5EZF/3rA5xdGk8AylkYiCJbyMsI1PieJxgQaCQkcfIkwIIh5PvMQHMUfPeBqE7IMf2PON+ZG0TFvF52IygK6cuOBAeFrkhBQfnHkYrOfEpFO3/Nh+6JmZEPYvB7Ach7p/nnuM2Dth/OgiGwP08r+6OSaBsA8xjMynLt1ZB4Mxo2IBM8norKAbukxWD6S2D7nsf0APeInTG/OIdQYlBC+2385hl7PUIQBbgy6nmke6XZo+zwMyEcu8wDo73IHnaSAFJggCvBBwA8wkUMG0qXpz336IPxzDzIjzPK5wzz3Ux79ah5jrOT96xDyiORCwAVodFCN/H4W84mYJ6eXEZTRJ3pF3DB9xh1BMp5oeEYaz1UrjhRoy0ck0Z7H9r4w4FXGnwIdNKLN//KnXbGP+i8zv+nMD9ybvB3dbjuBLcpdSPn8ZbdvfXdbWM5GU+D8L759ete0vi+ahT2tPLAY2YaJkryWBeKHdWu+b9uDL3wA9uiUEPr9jjkuyyHhMdn+fHMCQSDjsPyxB/cRf7+KG9Fv3YzsQ2/MqcKQdfgbcy4OV4lI0XmZ6uKZB8AmdECCI2jG4RqR/ZRoOBDIebiHPMy4XwcBR6QxAJeW5TxHtDcEyjzPKE6M0yIynpFcW0V7+ozn63g+SHYPIG28juEo89gOCQkQ8XvdPA98RsbZAdFpGFXMecYDAYRh/add+NNDgCqjqcDer7wRy/FXLhewsp4cBBeG65MRXNt9Ah5hKxvCZ0cEw0FknPcPkXlsT2QeUrwUeU2r937XfZ2kgBSYIArENM4KFp9ns8E+9/75t+JgHi3PQ/uM5Kp59FkZz8h4bk+uk2qeF+eb50nkHDPStpRY5sFnHJHRvD8AABAASURBVFA+XiGreR6NFHCA72jFAdeYl9F0TGYFNvbcb11y0FOaZktxnaejFgX3rt9+hVc5g0UBEZvRDtt+9qLv9fT0oDkYlY2qwKpVu3ZtNr22EEszDxd6pG/KDY+kXuxR37/bBy4cnd8rx+DKgov7N7JDPt8i0lNCAstIg2FHN5CbEWbO43WcBgGBjPmm4+EUGP7GPMa6UWSKDs2NhyMJOkQQeTHcpY8YAGdjc7853PCEwjUcyR36JzPoiHlI9y+ycIt8N8x4U1i/w8cDjsh4RlDt+aljxg390fV8JMLF2UgPGHD/fA7M89EB29MvGpoRQRNKmu3b8q39aNrJduFtL2sn5W1UBUKIuK8WGtcvrxcvmNfJaKRAgnI9aTDsCIPry3zmZQTNMN3ifqoQnoeLEUO40Gbtcic8FSkgBSaMAvw8c7BAfo5pEuHiuWLl557Pl9aDwIwB5hkOPjcYzwjK6BOrlXH6RMbZvkaHZEdVzgyPS0CeJ7H/FPPUiYyXiET6rjdsRywI40YExzi6pwvPCiRhODLCVJECI6zAzWfP3bpR6z3PzJ5Udo17c/C7DiwK8u7F2/WP7LT/onNw6xYMSJWNp8CVJ+z5jMbtWy6PMeyFq6T9VEV6mCgIsbQsPGH+FqPtsdsHzl+NSCsAZ6OXgCujGB9ovDIGYzwcSdAhgmjLSz4AD15jc8fCMCOymeEgMo/tiXy+4qYs8hFXGVKBdCMNmTN4MISGi0zRy0VgKokK5sUoEbEB+RWiX3NjOw8jQERzv25G0G5mNBh+MyT0YOXk/aGjjAwxv4q02Q8x52Ukh+YE3mvmeSRwPRIO8DMyDtfzMjKPPNHzzLLrho8HAUdrPziOpn3Xzv/FdNMxegqE+n/jm+h7/IK+jrAcsU5cx7yA/Vwjn/O4nowzv4qM0ycyn/GM5A0HMcTPwVKRAlJgQinADzY/wEACx14iDH8uJGTMP/80qvmIk+JzYUC+B4oTvz70jzdJFuEOOheCcGr951udJOPUs0QEPT8h9YRpGcs8dp/0B3iYWBqMmw4pMOIK/Hrl4VNj0z5modb6b9Fwz/rtN+BqYFESPb97y6nzk72BoPQNVeDak/bdshma/Ovrb2u1bS0GHzstr2UhF99H1j75hg+cvxL26JcYgqH4Pirg8hxaRstGQgInUiIMz084aD76RBhnM14H926JhkO/Yw4Rhi7D35i3vTHHKvhipYvB9UV3NwW4OIZAcrHJMbqex8Xj4pNgvETzMN1WPjgniKieTywNM/bH6xGNB65LYM18RnI5r5Lm12Ms52UkV80rx83rI+CQkcmocIuJwB40H7y3AzLu4ykJkOiALiyGLdgzzXpPMh2jp8CsbfEwtWP8gliOodcTCVwvgOdlxMJ5eyLjdIiM+/2ViMHX/xx79w4/ZIaqFJACE0iBAZ9njJ2fe4B/nWHcP/98EJDMCLuaB7d4niDelu+B4pT5jGQnxRtzCEUdOV/IQygqeH/eZgTLvOTCM4ZdV+ORAtSPgeSmr7tIqBJwVaTACCqA2y7cc9cDbzerHYJnQ5d3DZJ3ndttJ7AooJrRwsVdD0z55LZvWzB6v6uMC29w6YAGWI5wyYlvf+5D8aEfmIXXmVmxTv6QgMcSW0+UqoVQb2zYyV3/uGsx7LErmISVz7dQDBHgBp+PPhcnzHKe4cC86OLeNEfmDcg37wZnIBowzusRSSa1aKoOrkBtcPoRsBTZxUZuibBZsBbmq0aDiwt0yGg4QMA15rE90RfZcCCAMN1MG41qnqWDeTR9PDAc0b7MB+cdEVEZBxiR/WUkx2ZEVvIlIsDreD5IuDinwgBMz0fA3YQAo08sjUQQOC4i4xmZR5/IOLrG3W3uGo6cF8J+dsFPXgNGZbQUeNervox1uL1YFi4EL0zEglXXv+0+QRxhQ0Nfx4ygjQeRcbYnMs72jBUXghUfstj8TxgqUkAKTDQF/POcP+hp8Nkd8LnPgUoezZzn6A8KPE4SMp4r47SJvC67a5DotMqJYU7VeXK+oMrHJm0+T4nMY4DYP89lTP15Pgi6zHNkB6igzeOGgwGAihQYQQVWLz3gdQ2LC/DhLrYvuAf9thtwDbAooJFhl2y5RXjXtgcv6IWvYmYbU4SrTtnjlV1WPwtr9KzWdYrFcB8mipvISehwP87H3HPXlKNm9Izmf4uGq1ZLxC3jjy+MkgjXH2twjUZGPi/pEz3PLLtusB0JR8PhDYEoZT4Mp4lumOmNOQQaugx/Y86/yu5f5HABaM614xq1bkQuAgK+aETmZYTNArfIp5HzGUBNrrE9w7wAr0csLoQkBJgHq6Q9HzwJz2ewUhmnS2Q8Izn6xGplnNfjdRh3TAm8DK/jbgoQ2vJB5Dy293wn0AqIMNNbNAjQrgvz4SLR3QJBMO7jqp1iy5fXnddplBToewfWq/ipNNehXDisCwL93crCYXxskPLgWVs+eIbLDsysWP+mmb3L3rPTH4EqUkAKTDgF+MHG59uABI6fCPfhv74xEQ34vC/zaYDj8wHQVphHgsg4m4eQGjDQITXGYk4RE+T0AEbsPz3GyTmiCTHnObKhJ/CExzV95sFlPLnwzEAjwYqDgcLSWQqMhALXL5rznFqtdhz6eiwqbrXo3w243XbCvYdCChm/DiF89AXvOOM++qobV4ErT9lzl2YzfAfy89cM0sXgJYsL1vJaFsK91rSv9jXXfXFWz4p18Meu4IYpLo4HGocI4LgHPt8Q4POSAc+jj5YA3JxmgYYHrDjg06VDhDvw6xuCxY+cYKisT4Hhb8wpOleTWC4CL0Oigu4ioYpwjT6xNBJBqDQ3Lj7zHGkwWEHm0y0RBm+mtny2SdV52ESOH6bfPFWknSvz+vfH6zHuiOt5PwkdEMhoMOBaFQfkmzHsNA1eLyOaG48SYbA/H5e9wKa+6IMMq46SArN2/A3Waa5fjevAdQJhGbE8Vbfkc9x4eEMYwDIfBlwr8+iDCPEg22eHC0yHFJACE1QBfI7z5xofa59EiTD4vPfnOfMYraDz4IgAIw7IZyBVxvkAypjojgPI5nPiPDdED7bzfLQmGgnYGdlfVT+GuRwZPS8TbKcqBR69AjH21Lq66h/ERvvl3hvuTd5lbredwKKQAvzTmvX3b7/fwtH/q94cwCSqty2fOeXyk985LzYjvxd7KqbOJwIAq4CzF5gobvrGNVmAtUg+5e6+KcftfsTFa+FvYNkY6RiRP+fQNwedXGt7viFAPudlBG18ThKNCTQSEgxHiTBwL5vnMw8xvTGHCEOX4W/MY/WvslN8XAiAs5ljWgTjAcLdhADfDBO52N6ACU6gAZAurDIvLy45LjKR7ZiH9AF51XzPTSfnYRPZHqaV/dHJlR3DZh7jGUEZr1ci80AwzoC79JEAwNlItz6ongA3o5nH2d7zwfN6pS6GAwHQMNAOZ7jlfC322MpfbwFWZbQUeNerlmMhjiwuFwFYEK4fF9Jd+gWNPLMNW89Kvh1p++x4uumQAlJgAiuA58F6n+d4YPD5wOcHkbPMSBth8+Y0QOS8jKDaCnk2yMjgpPsdcwrGibNSN/hZD3fpIwbA2SiXP6eNhyfAzQiOeclFwMzXh4TpkAIjosAtZ/1pfwt2oBluLtyrg99dYFEMBzbw9+Ne/OAOc864Aq7KRlTghq/O3ORv/2wcHZr2ZSzNY1qXSotBAiYKLdSWBWct2ny5a7N7PjbryBUPwh/7UvwvQxgHHmwcKsC/PBFp4BbEvVXEcTZDgHmGgwi3sv8wG5BvRXNAkYcGEKHIA6k35hBh6DL8jXnbX2XnauFCCbi2ZlgMNxI6IMHRcMCAa/3zjAcCCBfNYcA15nFxSzQcCCAMw4w3B9wC3cDNQbT2g3lkiOwvI7m2mjpmnAOh6/lIyt06MgCuzKOPQAJEMA6eM5GQUPaLONujWTF++MYEErSBdGkS4ZrnuzHdeh8q3uAyrjo6Crxru6+ZhbeYGb9Amq9HuZ5Yl7xOBgMuEqy4EeDnPDAM0zVfz0TEeCdyd8KbclyDnKoUkAITVwF/AGD4FUyPgX6fe+QgwK8zsLzALRAGeX9OpH7oe7ByYpwukXE0s0nxV9kxac4X4N8MEr1SABilHvCZB/DnLkIFZiKhA3R2ZBIqXDMQGU2HFHj0Cqw+Z97LcEvxr6lvynsX9iCdgkVhANCMVjul+4GpeEFARnVjKXDFCe983IPTGudYM34Mum/Zug687MBESV7LMgvN2AwnN+K6o2fMvfohGy9HjMH4POSDL2BQeciOIDIybvCJCbJrNHJedMPwPStqKmU+DA8TaQRLGYIhFBj+xjxW35inK0D71uKkRfBFQ8DdjP3zwXPxy8VG3POJMBhmPN9MzEPIiAgbDyLzeL2c58hgpTJOl8h4RnL0iV7ZIQzG83UYJ83rIET6Yefr42YyGrI9/aIhSHQE2t0SYYD2fj0fadXC8TDu6Iah/RGmY/QVeNcrV1psvgr6r/b1glEuL5aRrvNcJga4nlw3BhjPyDj9YPwjb9+w3uZL8Kb8BtMhBaRAByjADzinAeTnnCYRrvF5QMzPByKfC8whMo+25yExIzk+T4jVyjh9IuNsPynemGPSnC+grVAHEkTGSwRJ3/WG7Zj0zbozTv1AM4O0P8/dYcANnaTAsBVYfeb+jwuNcDTuqyl8Fgx+V4FF4UUSXNlnvV/VH3ujIhuvXnHSXv8e6mFxM8a3Jt3TxSoeTJQBPB4Zd2NNj5ne1/2J8fPP19MwAx6CKMYHGgePwRoPRxJ0iCDa8pIPwNyMzR0Lw4zIZoaDyDy2J/L5yucpEeFRKBP6EsPfmGeRKXq5CNSCRAXLPPJITMA1LNaoH0G30tzYHs1KZKxoCAuB/vkGwm+GhMhqK94f21XizGdSRtrshzhoPgOo6MY8jwb647gc4GdkHK7nZWQeeaLnmWXXDR8HAo7Wfvh4QGVkw2jPtgt++lawKqOtwLt3+LnN2m57LMNr8ZBaBQRgEFg+Lq+RyOvO9STPQBVj/D/kfdFq67axfXf8kO2/E3zTIQWkQEcowAcAP/BAAudUIgx/LiRkjM+HjKDd5PO+MPB8YT8IsJ1zlRPzyGdkaFL8VfYh9KCe1MMRgrg+Cfl8hmkZyzz0V+abeRiytwzGTYcUeHQKdNX3idZ8I281v70G9AYWhXQB8TeNrtoBu8xe+ldyqiOvQOzpqV150jtnmjUubDbj7rhCZZ9UrAI4PIe9uFm1QNwTLHz87idO+dyO4+Wfr2NQZWnCKp9zsDmlkLB80CWCwJuzRBien3DQfPSFMM5mvI4/b0EQDUcfaseUjTORyg23gRdwkSk22rn4RFQW0K0b1VcRbkYkMM72RF90GpW4waeLVIbp8qeJjuQK459m4bPoGBXYjEDUiBrss0afGGFnpM0aakdb006wWlhiwS5AvQbX+T3QyoPjo0MMHA8GRCQHl1BU8GzIPEewzEs0PCONcVpxpEBbPiIpnwEqAAAQAElEQVSJ9jy2x4CsaGhGZNxwEHl9ticyj/kxvh9RlbFSYNZ2V9m7ttsNS/Vs1COxLFcCMRosWF6nggAX7kT8WtTPWMO2tX12eJLts91/2qxd7jQdUkAKdJYCfD7jw27l59/8MY8zkM8HPCD8eQ50MiMchHEu8grDjP215VvryHxGRuo8dVpNwuR5ZuQ0K/JBODAgGDci3YywWeAWeXTYL4i2fPCJLvIQ9/UEryIFhqnALUvm/ZtZ+DxupSm8vWzAARaFdIK7GjF8cOd9zvgLOdWRV2Dl8e/Z4qon/+T90Ps0i2Gb9iuAzQRMlOS1LBB31Zr28ev+8f9OmzVrxfj8kWgN27625xueZ5wCwJ9v/PpiJYEpIUAXVkmzPejB85FY5sNgf57vDcz0O+YQaOhS6x9+xH4pNlq46ERUFqyF8YtgaxWTiwDXBuBEiTRSgMB2xApd5LcR/7A3P6/H3vxvqMC3EFGHQsZYd3/+Z+yt/3Y42u5vu7/gHai72ltf8GxgsFDfGj8E2AnDPwTDOBm4Gv4DQNyDuH4wIGpZSMAJxBRHwyKfPCpo608wfX15bfmGA4TnwyTCNV6PuhuIAt9k5/9iOjJUxlKBWdv9Hhv0r9m7t3sdMFiz6wnWF59vtfp2WO5X2rpNNkfsibbP9rugHm3v2f4HYzlcXVsKSIGNrEDxfMZF+OAGsOCxjeeBmT/HDQeI9eUhWuZ5furH8xmsVMbpEhlHt/jhH5kOq5wYplSdJ+cLynUleq3kMZDzGUsykjZLeY4I0GV/jlYcoM3jhoMBgIoUGIYCt5510OObtXgSvrfcwm+rAX2ARSGdAKk2vzZtzSpyqiOvwBWnvONF3fUHTo3N+FU8E/qtS14FXBcmCgyWNus3MTbnve6I8xf09PQ0GR2XtYmh+eMrmj/O2p5zCIBOAbMioYTsOsF2JBwNhzcEoqAbaGjmz1vDASLn9cFVGVKB4W/MI77cZ9GhebEIvFZeHCICvhhExFr5cFBAF7tcGjkfPEty+TQyhnkBtieSiNFZpo5o3f15/2tvfeEN9pYXzre3vuAw4Pb25DWPxTX2xWVvypeHb/CtGD8gBzhuBogcb4ThIyXSADLOJsTsZmQ/THOEQZ65JcIAXejiBqJEgPW+nmfVcaTAvtv+A5vvX9usV96Mt+K32v4vWzOORqehSAEpsNEV4PMZz+38vOf1kls8xxPBrxd87jOPFJFfP2gTvRs09DyQGWGWhXl0iIwj3UJgS7KdU2P6+h8xQU4PgHkOnB/jZB0hA5H55BzZkE5Cxo154BgnDReekW7/em86pMAGKxBjDH1x3XvwUX81b6+BHYBFIZ+At95P7B9bfm7HWePkL3tzcB1UL58/86WxWT8/1Gwm1qUr615MseLBRCloJCbDsEB/rYf6vm84/MLzrfWQgDkOSwjpiQbgZAA+FaIbHHMKxIzg2Cy5PkX6nu8NmeA0jAJJsz2ReZ6PqN6YQ4Shy/A35sHqrj5Fz4tFxB1qfjAAwwGBKsL1NGJpMAEEAc2cdgQB2rioXGRyxSKTdW/9pxGKbLttLzboZ6PuYKH+csz7Yu/ZR4DxcTyW0AGBjOThWhUH5Jsx7DQNnyc6IAKMR4kw2F9IyFjRkBZ/H4aoKgWkgBSQAuNCAX9gYyRAPLZhGB/zxsOf4+QR4POeXPk8zzxI5gGMyLyM5KqVPNtnZAw7AUJHVf6XP5wQ57khekDm8ochbGckvCOezNhfVT+GsQyWsTRImA4psMEK3HTO3GcEqx2G26rbBhxgUUgnMNxyf5hSq799xyO/Nj7+uy3rnOPak/bd8vKT9jg8NPpuCBafa00Lpe4+zYoHE8VZ7AESWh+eJ1c118ZdXvf+827BU6GVkjPGJWKk/pzD4Dji5BputmJuiSDkvIyejwBx0Hwzpw0Hn6d8zmYEBcV4Vh1CgeFvzCk21efiuOi4CtYKZyNdLK6lAwHmMUCEi5vZ6BZ5TpgVhBmReYaD1/EwCF4HlGWkPdr1Lc//H3vLi7ABbs62YPf6+GH4fBwxII4Pw6ULzwokYTiImBDnxUByvT1o74/t+SFgHC2MyDzDQWQe2xNBmefDCIH/fRcMFSkgBaSAFBgfCvBBnR/cHBF8ujT9OU4fROs5zkhRQRvC/vWBzGD55HNlnA0ykseP0AmdVSkMZsR5UreMoIw+0SvzICDjRgTJeKLhGWn/ums8UqAtH3yiPY/t/eszeBUpsAEK4LYKoWGfw8/Knj+wGW4yFPIJaD6A0xdf8Z7T/wRUGUEFLj3xHVuvtYfOCBaOQ7eb8yNd0R1UxYOJAo6lZVkM5/TFrjlvOvKC3zAyYSpuROODj1MJ5o81um6UzzcGEGOAeTSJpNmeaCAG5CMRNM7oDgbjnu8NTL9j7soMeaoNGR0qSLF9UZBE0QF0CQVyEbAoXFQS7sJ3ZBYMuJbjJRoOBBBmM+N14JqB8OtktLE93vripdbb9xKzcJvx4Dg54BJB+rgzYtyMcx5EuhmZx3ZVzHlojrvb6BoPtivz3ECYiGCMT7aLbn8GLBUpIAWkgBQYFwqk5zOf9z4e+G3PcZIgBvv6BppR868PbAfC88BmhFkW5tEhMo50G59/goijfBSVE0Pz6jw5X1DlDzFoW/88+MwDVJfDPA/6ZvQ4fEcrDrjGeEbTIQU2SIFw69J5bzMLs23AgZsKhXQCmmYhXvPQPWuWGr7LQ1UZAQVunX9Q91Un7fn2eqjfDln3QJeb8lnQpjsJBLwggOIm8h3xWHgQa3NC9z/vnrv7Eef9r5M4LV8+s37+F98+/Ttf2+Oxy0+cuTl90OOv8LnJOWIiBB+gTxJERg/AJybg44+u0ch5fJ4aDydoGMOex+s4jQ48D2g6Hk6B4W/Ms8gUPWvtSIKXJYJoy6OPGABn88UrVs+KGx4BNoPXokGA9jgXuQwwaYzrHi+9w2pdu+ED+lsfHyfUf74c73rHjwCmx2ZMKxAEaO/P59tvjuyfcUc3kEAkAPsag/wkFrFHXpQpBaSAFJACI6YAnsveFxCPdzeJcH0TSeQXAH/e5wCzEKBLMz/vM5LzfBqVyjhdIuNsPynemGPSnC+grVAHEkTGSwRJn7rD9K+3/ALMuCNIxqkflgGekS7yDAcDABUp8AgVuO7seU9p1OJHB6bjXkIhn4Am61+s2fzIjPevuJ+O6qNX4MqTZz7tnuY/v9iMdiY+y5t7jxAdxc3iVPFgohR0flYEuwPf83/w9YddcMSMnqv9T5nhsREuOeFtL37M39cePXVa/M6m9cYVj2msXfHYv679xMVf2+OZqYPxAQGjRTHDg42TAxgPRxJ0iCDa8pIPgHbG5o6FYUZkM8NBZB7bE6kdn6dEhFWGVmD4G/MQ8HN4qE/RAVyTQnMSvGjCvBglIjYgv0KkZt4fUo3tPIwAkZxfiAF3xva0+/PutL7anhgEPqAYIIZZfrPFm9JAgLYqFuM3v6kRrrrmeWyAgLe39oPT9jDjDAFzB54f/43s+K8aoRSQAlJgMijgD2xMFMjHNSw+5gnmz3PyCPjzmyx8Ap/roN1kXmHgywbi9Mt8DxSnzGck26RDo6NqoQynRh0y9p8iedeR6dQNCZ6f0MjDzjhoPuJMQ3PzPBokTIcUeMQK1BvNt+BW3K69Ae4lFHIJaKLGBjaPH9lh/yU/h6PyKBWAtuHyb+6xq8XGpTHGIyxE/kFnPEu9VHqPLRsmSvJL6zdm4e1dd967EI+Kkrz0m+94bS2GCyyGj4cQXouc/w89v6kZ4mdqXY3vXfnl3cfPv2SNIRgKxodhmoOFhKWRCIJhmiXCgGtsTzT4jGeki67cdQThz9uE5LBTIqiuX4Hhb8xjxM/hKTY6z4sEF54Vi+KrZuWqc3EYIM08+kQuKtsTc7zEVvNyswvKGHfCxsfxjhf8BAP5onFcPj9MzNGKA26hA90U4PzLfPCJ9ryqHghZzjMcOY/tc7/V/BiehyyVrIBQCkgBKTCWClSfz3kcfI7T9uc4HuQZyfF574hTNQ9u+fWhLd8DxSnzGcnW6NDoqFoow6lR34ycIuQkFJVpIBinru7SRxSAs5F2XY2HJ8DNCI55yUWgQsBUkQKPQIHrTp83PVj4EFK7UVPBTYVCJwFNVL9ZL6jXelfCUXkUCkDXcNk33/7UVSe98z9CvfYd+C/GZxh7F3QKBwVGLhUPJkoK0Aq92HKcV5sad37D+7/7P/lN+bUnvWXLS7/5tv8IIV6MVXs2+k57qug7GnTQBf7FvVPrZ/Oft8Mf+4LBYi4YR8BwCRlhc9R8nhKLByPIFIdV0phUEYY2A/KRCBpn9A+Dcc9HPyT1V9mpwpA13URD5gweDHhj7mIjXCJsFqyF+arR4GIAHTIaDhBwjXlsT+Sqg8ZqOk03005U84yrbePo6PsKxvugDTV+JJgnYOJD5SFc5hkPEJ4PmwjXOP2shyNi7D+Y3phTilGuupwUkAJSYFAF/PmcH9wpI7v+HCcHYn15HmYchufzCwBszwdWC+P0iYyjmU2KN+aYNOcL4JdBQlEpACzqwQCxf57LmfKMCMIhI9qzwDWPGw4mAFSkwCNQoD6l+Z9IexFqKriZUOgkoIlKL9wbY/Mb2++37F4QKo9CgatO3Pv19XrXWXhJjBdnhrfk1Nd8i5EsOCwVDyYKSVRY0dagQU9fnHrI6953wd9Berny+D2e82CjfppZOBqPlsp2s9yUW3k07aVTrbFr6Y+lEWMoHmOYW8BAYkbYDMA1IiZVoueZlTQNtivRcHhDIEqZD8NpohtmfYirDKnA8DfmfGPuX+TQPzQv19ANcI4IlIsHrpoPl2uKGx4W8sp8uCxcQ6dhEBlne6I39I6ZOT7qW//9LqvZCh8ex4dhG8fLYQ46fg4bAc+DnZEdgHZdvD1iLIw7wmDc+3UDLBFgwGjb0FLtSAU0KSkgBSacAngu87lePJ+L0eMxTtffXDDMOJ/3RA8wDQE+52kS4Rb5NEB6PrBamEefyDivU6NDsqMqZ4Yvk4A8T2L/KeapO0I3Ys5zRHtvk5Bx6k+XcUdPMNK4oBUHA4WlsxQYSoGbls59Ua0WPtLKwb2DQj8BTVR4KLjtlu64/5JrQKgMU4GVx7956hUnvfPzsdZcCUlfg481P/xFbyBQCtvPFQ8mirPps/7nZi2+7Q0fOP/Y3Y9YcScDy2fOrF96/Fv3btTiz5G7Jx7Z08gXdZBNOZLwknqzWt3+nWaRN4bnADX88kkSAOZgA59vCHgq0cxwY7bl0W8RZpbyDAcmSrf4egWfeZ4Pu/IjDHgqgygw/I25iwz1sRbUvFgEXoFEBd1NeUykD7c9v0IwXmluvI6HEfCbhEESxHFXVxiGWd6MPl4QPtwKUoci0TKUSKNsh4ZoZjxKhAHasi6MeX80EAhW/EELupOxnnXr4+2cW59j3771Fbb85tfYilt2teU/3MHOuulJk1GO4c15HLdadNWLbfHVb7Mzr/6ogOfDCQAAEABJREFUnbnqOFu06gTUhfBX2KKrVwKvgn8B7LNQ59vCq78C/1O28Kp97cwrX2WLL91sHM9ufA9twSUvtlNWvg314zZ/5RftlItOBC5C/Ra4S1CvAPdd4FLUk1G/ZCdf9J/Ave3klS8f35PbmKPDc9kCLgAkwHLXEQSf9/48R5xc2/OccZCMA4w4IJ+BVBln+4ykGzx1WoUunBLnuSF6sJnnozHRSMDOyP6q+jHMZcnoeZlgu0lQe+Y/3o45fVv73Omz7HOnfcyOPu044Dft6NPPsM+dei7s76FehXox6rfAnWmfPfUk2F+yo0/9tH12/sHw32E9p73Kjpo/qV4crDz+8Km4Qw7GrTYFiBLTfq8EcCwl//MHe+1TZFQ3XAH+xfXLvrnn26fUN7nMLPwXPsp18wP6EgEotFKteDBRMr8WxgUWwlvf9P4LVsHmt/Xh4q+//d8fs8u6461WOyPG2I3+GUp18E05Vxrrz+3w0362fGblVxlSs9GGGDDN9EBLYBlLIxEETrKKaG18ThKNARoJCZxPiTAwefN85iHYh6oypALD35hTbKPo6N9FJ6KygObNSLOoIHxNEgJ4l7N5kecEUokABjwfNq9D2pEGOF4PMO5K7L2e97BxfBx/Gq45kuCIiSA4HwaS26YH2xcdsQEq8wEsA/JBej6QJcYtCJOiLr9hK/vWD2ZhE368rbj1J7Cjddud1hV+Y037gdXqV5uFVRaaN1hX/W/IWYON+i3IW4xN+5NNx9gqMNTVl37/RbbkmvfYomu+bIuvuRIb7jvszGuihdpt5n9kxf7bLHzczN5vIRyAz8/eFuzNwBkWwtvwXNkH9SAL8UjkHI26zGJttTW678dG/W/YqK+yRVd91s5Y9TrEVKoKnHrps3wzveDiY4GX2fxL/ojNN77riLchjd+sHGfRPgadDwPOAbcXtH4j8LWo70B9D/xDgP+B+gXYK1B/iE17xEb9D6gr7ZTvfcZOuXB3O+l7WyKnw0vA/CIqEWBAujT5dSDQB0Ekl5E2aKbjvqYHGUEwntsVbOtMng0yMlLnqdMqdOCUOM/+etBnzCvzkr6uC0jGEw3PSENYK44UYL8MJBf3udH1PLZ3wjrvOAab56NPPxyb6uXYjP/Yjj7tXqt33WnNeAvmfy4m/EUrnrsfgBZzzWwW+LfAngH7TUA8C8L+sA+F/R/4Ovw5C7VTYH/XanG11cIf7XOnRvvsgp8Bl1jP/COsZ8GOyO/I8tgtH9gaE3srKm9CyAALhbcVIBV4KHAeqjXt6BlzF90NW2UDFbjmxHdsfXffPxbW6nER7sld0Bya45w/q7GUnyQqCJy9wERJZrwvWu3Tfc3NZ/P3yUku75k55bIT9vhArW7fwfcUh1q06agMpRrbXbLeoZ/oocan3XmnTYExtiWEUHw9gTwcHsAHT6SBcKGUExgrkHmwGIa2CIMADcNsQL45jTOQeUjk89TzwOqNOUQYugx/Y+4iU3RcgKIDfNFKxGI4kdAh5zMJBFwzYDXPeCBQ0jDgGvP8OvSdsHF38J+zh/Bn3I3G4fq0DIcPF+N2IiGBPpFx6lnF/h0wD12VtOd7A1yOyCBqCI/BuXPL8lu3wcb6o/btH9xgtan/xETPtRgOtxBe6g+bUhdE/H4BspA329QsbIu82fhGY/L8AMMmwLFw1ZPtzKuPxCb8+6gPYH1+hm/qllrNPoL12s0sPB03uhUH7vfy85AMri/XO2OR6Gc/kafhGJ9kwXa1aJ/BF9nLsUmPqNehfs4WXvFKm4zHgovfbfMv/p4tuOQuazZ/ZyGsgD7/BXw9dMdbLupM3TNCJLiWXCuNkjC0M6cNR6aDPQMefoASPmsxXIT4v7A5/52dctEJdtL5/45YBxYKxWlVsNQDhtNEGhmZjwoXZ7MAo//9Td/6HcwjRWQczaxBotMqJ4Y5VefJ+YLC84LnVPvnwWcewCg3sxwzkdABAUcmocI1A5HROuA4+vQ32NGnf9mOOeN6IGYWVmNWx1sIM6EPPo9heutzzDCiCBhkIBgN0I5OpADBcBAZz+vk+STsRVin/SyEb1iw67FRj3izfoF99tR9rGc+vk6jbQeUWmjOxDSe7Rr6tN0ClQtIFPeiXb+2WbvCbZ0esQIrj3/zFleetMebe0P9StxL+KFwqPywN4kLQKn0WfFgojDWCBZ/W6vV93njB77zpd2PWHbv8q/O3OTy49+2/WO2WnuZWfN4JD0btznSYJXl4Tfl3n+wrWxd79hvS/Gq30L6YCbwqaRBYn5wcyBhAstIo8x3A20ywsx5vI7TIPjcZTuEVYZWYPgb8ywyRYfmfhlHEvSIINry6CMGwNmKNUp5fjcgQNdwEOHi4W1lHhc559k4PWL8P+OAffw+AXeN4x4wfsQ9zzzMZgPni1i1UE80K/LcQJRIALK/Vb+fBm/osvxnW9n5P+2x7/x4A+uPUn7Cbw+G4JxPSLtaz/vxE4ce3CDR5bfubN/+4XXGn7hb+G/IuYNnUQ8aROqbkRx9YrUyTj/izSlxQ+qKm/e0c1f32Lm3oFaRdqpnPxze2GPMObuKtB+mnnMD3lBsyGAnQO7p1023RVfPtSXfv8Lqtb9aqH0F6/pqjHwToLU+D7ivSQSitR/leuLGZ5w+kVkZaefKODtGuuV40e1OFuzTZuFmW3TVHbbwyuPtjCvxAwF7pMfEyzvt0l1swcWnYTN+H+Q920J4C54rj7VCDyuQQhkOIgL99Us0EsxczzaieK4ZjkyzPbrxgOcjFsKzcN33W+j6sZ184Q/slO8dZvMv76AfMPqEOVGfNowCSQ/QIwvFLCTQpel59EEEILmMtHNlHm0i40i3SfHGHJPmfAFthTqQIDJeIkj6uPFhoVAo6Mq4EUExnmh4RrpYOMPBAGCilZ4TN7djz9jbjj5jmX3+jPsx/EstBP7wc0dgMT1MH7yZI+fphhkJuoaD6DSMEmFk/ZBinp/i5Kmn64084wFEmJbTFt5msXmWhbAGm/OzsUnHDwU9OiFP31+y31Mw8CNc1DTPBKBZ4KHQYm3UbOEu8xb67zHTV314BS4/eY9dp0zZ5LQYat9F9vNQKyWJC0AZyJNBAIXWWuCCUKu/4fWHnXfR8uUz61ec9I5XP2ZK7wnNULvIgr2GSV6R6OinR7gpZ260TdfGB2s0x7zy84hJGeeCj6GPx5EEPSKItrzkA/yezsh+2ITIZrSJjLM9kRcqP/9MGKN68Kc+aId9pscOQXX8VGEfUkXa66kHJX4AfgL9IHYQkLE2JPdw9b9eUFVk+DdJFpmil4vArklUsMwjj8QEXEOulZVGChAMR4kw0MzYDxEhvykcx+EphHss34xES+Ovok8cPDEBw3SNRtmunLCVR0AD0hmZ7w3JIwCwB+/CT7nLFoMbm4THYZxHWQj9qh2FBkdZSGj9sXaUWTzKLGEN7SP8jOwvok2JtPvVWu+TcI1HVr7946djQ3621WvXYrw7FY0wz8IwyzpktGDIMyvR2g/PQ/t6c0174BF4TdvTDPP1+VcR83MeWEPl/DNSR/qO0Im6hH7ItuSJXtFHibRRm7UPWKccS695qS2+9jTrbt5rtXAG1ov/BBqzw7pg+QzA5XOkERPhiLRq8fUE4cg82DkvI6iy5LyMDKRmlevh7Xw4HJe+Ehv0v+NN+lfs9Kv+jaljXx/lCOZfuCk244dhM/5La8ZrzMKB+MBsbtSjEACuFYfr0m9B3EUgIzPhEormHoCbMIFlpOHrAsIRqUVDGOwovMKadqI11/4Nm/Rv2KkrsRYITeji88IMgJg2DKMMxoO6U4eM5Kp6lPmlgfVhP/DZzvMrp9xPRoYmxRvzIfSgntTDEYJQN6TjuQOHBiAvSJlHnjojBvAwsTQYR2yilGMX7oG34hdZ1yb34XO/AtPYF/PfDJjuJ0yEugAoE6FAzpMTT5igaMcsEN4uIQAdMoCKdtQTYBmNCU6kOICFtCMM9hfCuy02L7PPLrgdb9IPtq9+dROGJ1KdGqZ8GFo8udARVtvgoQFKpmD+qPfpf+SvCmRKOIQCl8+f+ZirTt77lJrVVkLfvVGntKdDURIAFFqpVjyYKOTvqllj78d1P/GDrzvsvN+tOn7Ppz/2b+vOaTbsUtyDc7FyWzHJa2rgNi7a5pJ0wk/0kOGALlCibdrsqtUTM3YQYzB+Hjm6gGFwuBktGwkJOc8RhOcnHDQffSKMsxmvE9Ego+EYq98xP/RTH8V4vo7n31HG7+Gb8Sjz77ObFbSjjN+vc99DbKvI5/fu3O9Ukf3w+3fiYP2xr2ple/pE1qa9yBZ84RdWOYa/MY8Q26C+Q8bUM1zD7ArPE+BmBMs42xOZx0Uj5v5KRG5qZmU+OMYJ47FGfFvJ+eRxc4x5nrTLeaaEBJmGUGZs3yLMDB0wz3AQ4bb0ANE//6E6M5A8RFm7tj1IfcmgO2N/jiRQvTcSsPO4PB8B0m35TiAxIcFSHlhv/kh/zeZbPzzUYt8daIYv0uiI14Fj7M/SAdpdjgeXGVw/5DIPUOiGxFnb3kN3WBXN/TremB2D4PU5EHfpIwjwvIyMgzYi8wwHkXG2J1Ign2cOIIf5uK1oTei69Jp32eJrrrcYfmLWPNCn5RPyibtledoPpwfz2CLnOaZ+XD8GK5VxukTGM5JLzWhaHgDjHGC0J4I/0kL8hZ1x1U14i44v1GAmWjn5wqfZqZd82UL3n83CiZjn841HOU8KmoSgPoy5m/mEDgg4MgkVLvozyuVIo+zXjG6W1ePef78O6BoOIvuLcZpZOMJ6G3fYSRedad+8sO0nyjaRjgHzxeA5T0D5PKJenkeSAhBRq3lwS/3a8j1QnDKfkezYfyvIUYxwTcLkeWbkVSryuV68ARl3RAJ1ZvOc50gCsXyjtuWDZ7jMo0EC/HguPfMfb59f+HHU3+Ob0e9gqLtvFD2oZ9YNF7Gss+GgTC4XDCLzBuSnPMCAz0OMLzQLp9i9m/2v9Sz4D+tZiOeCjfvj5rPnbh1jkxtGHytm71ic4KEUtq/ImrrVD5ox4+qx2rLkoYx7vOKEdz7uipP22q/WaF7ftCb/qB7/dZ3fWa3BJ3EBKC2a9172EEDpxa14Xq0r7ryu2ff9f66789WXffMdJ/eGxs9wC++NOH8Y1OobRG7OVWtzGXDCT/RaV0tUCDFuMrUreZ4yNqdaDcPnMDA1h4wcDgiIggQ44HE2iFFOBmG6xefUcIAYkF/QOKObFOfz1PPAjsU/5j/4M3j5E/4bVzcff/8JcXwMYLjGcToaDhguA9DcAAekC8u7gTtsPaLdan1T9mdX1YoVqrobYPvgMbrqoOB6D44egJvQAQFH0JwkXHOkkQIEzpZYoQfk2Tg9gm2Bu9E4LU6jxNJIEyMwgThgnoYjB2DmPJqZzvqzX7+pGGRHwM3s4d8G19gBcnOhy34ykk/d8fLG6xiPNAAHJGRkCK4NyPOEgmacbl8vzzbk8e0fLFqVqRYAABAASURBVET8JPPxwDI04fhoFgNyi7QbnocLZHQSviNOaI6zGeNmD9ijObzbskP0BML7JdLNCNuQB9eIPm74RILhIDLO9kTmDZgnA8P/qOIqY1uWXHOwLfn+n60ZzrEQduT0zTBxnyewIMwPujSyHo6cPwNEBoF0aTKekf3RJ5KrVvL0iYyzPZEcuiMUlQFYzOO4iMxzOm6HyBnYnPOfus+GPf4L/wr9gou/ZPXu/8V0PoIBPxYVxSdkxvkh4JjnSTQcrkvKMyKIMj/FE22MI+zo/RkOEIzDYpi0sT1oMwailUjXcJQIg+NgvjX3t7r93E688HQ7+Tz+oASJE6hwHuV807gxvQF6rC+PTagDdXOkAdLzgdXCOH0i47zOpHhjjklzvgDXleiVAsCgHgwQ++e5nCkvrxNd5jmiPcuAPJLjsB696Dn2+TPOtK7uO/G9yHH4pvGZrVFyQvCoQ9bDEZzPL2PKq+pRzWPY82FkZLwtH30hXNAwqCfjjoh5gIiKMM5mHBf7c6RhmIIj3lrGL5mt+5199pS5YMd1afY1+UdF+U/Zq7PEmDEXFBhe3Ix20R+n3fcjJ3QaVIGenp7aFSft+bpavXYObo0F+EH5i/sJm9q5orhnvCSOkPhkwvsDvgWeWw/NTzT7bEZXmHJ+sOZy9H0wUrZAHFCcYfTrLLZfmmmsFdZdNnSDJ9Roaxu942BjHvnGPA2On7uIsTmSgwHXDOjzSZigpGmwXYmGwxsCUcp8GE4T3TAb7R8/HXTUNlaLK6rTMY7beGBcDGDhfZHpuo8Yh0ue6Pk0mAAkIMVpRxCgjYTrYjicAKIgjLNZ7o9o8Q9Wm/pmW9TzkPU7hv/dPi/OziN65EUzclKgfJKGAPMcQVbz4ZIu87wd8xlAzf2xPWjPY/uch5RxWWLEN44YsI+fCAMw+PgRQNh1yMj5gS7zyXOiJcJg3HVxA1EigB0hbG982cNvzLvZAdokwBdvM+pLn2g4MsIsC+OG6/E6jDumKGgft7spMFh+VzeDnjXgtPx/nmbf+tGPLIYDeJliXMxCE16PpgdoBFwOPE1eh/GM5OgTq5Vxs/uq1AbbuKy3ceT1aSRMYBndYHYmMpJDu+xyXHAxIQTcMOP4GTceTZ4mVl1yLX9//A9m4RTI8FTjtKrz5PwQQNyKAwl5vkS4xfq7Yeb51n6wPzJExjOSo0/0yg5h5DjdtjhivIyPB7YjiLZ88qjBno5xLbaFV/3IFl6+K5jxWU67ZLatjb+xUONfR7f167feeVpxPwKqehjzwQUihAR4XkbGETYiwsaDyLjrmQi2z/2SYmU8I+P0ieRqYZ7F7t/biRd8ku4G1DFO9YljDEDqAKucts8vET5PJiCPlOtHH07Oc0xxz0esWhinT2SczSfFG3NMmvMFtBXqQMIRuhFzniMF8gSecBvTZx5cxpMLz7gcSLDiYKCwxs25Z/ET7Zgz5uOb0N/g/trfHnb8nCfnQcQsmA8wuIQCUzwT2c2IC3mI6O3LAGh05HrDdITPvKJjM+Yz3XCUCMPTiDQQq5YYn2IxnGFHnXKLHXPi1tXQeLF/v/CAaRjjWzGeaZgFIBd4KBXPosX7cU+dM2vWio788Vme63CRv+u96tT3PH2XJ912agjhYsiHH3jETXg7DewTUZIAFFqpVrxoDXi3Wqx9sBnjy/ua4SYknWDB+HX8ceg3IG5YE9RUCqJ02tzEVvPLeGnkpPjQQ7V1Y/+NHN7cFyPC54tjBOBGNGhgrXmkAAQxBtxFoqPhgBHgtxqCgw8aRtENXHyfZGzuhOeb2Wi/Ma83L8AwtzRen+PjuDAMc8xEQgJjRMY5fyI6MLYnmhPIAjIP1qDz9HwGKzX3F+M91qzvbgt6/lGJlmattDbU8ItiVBhbOVa4Vh204WCcCVVkHn2i59NIBAHNnHYE4WEiDZIZaY+juvy2zTHubcpF4iKA4PStik5gPsQEDNMt8jg/BNgeYDxKhOHhhIwVDWEhEOwR/hPtqchHKdcR/bEfB+8HnyUgUtrKoPnI8HZAy0ZCQu63xHVMHFi//cNn4JuKa/HT0P/n3XgGOqAOJBxJ5nEBOR6AZbTB8tkmVc/jF8PkDwd4PbZzxPU4L16X2M+1zJdoOLxhgWU+DKeB7If5nC9dZJoN/6PqzUfztOy6Gbbk2t/gBjoDl31Gno6jzwusrwPQi08cFjAUwLRWPngS1APhtuL9IJ6RwZyXkRzbE3NeRnK5ohvzPBppIA7wHa044BrzYvx/Fmur7IwrltjCVY+18XKcctnz7NRL+SsDi6Hhk9ufR/0GmXUoEfGIWs43GwkdkJDRdUB+FXFRo0/0PLPsuuHrgoCj4UB/OHvxccDKyIaeV+ZvaiEcYyd97/d2wkX8FwxIHi9lfePg/DD+rAfT6DrC4Px8vswjWUHnwREBmDs+VojTZzty1Urer4N+iYx15Lf8nB8mx/lSh4yg2gp56lAiop6f0FI/Gcs88tSZeajJNc8D72jj4/jS4s3smEVHW3ff7y2Eg4rPOwbMeXKEHG6J4GlbQgLzMjI2IJ+EJ5g3q7htBPsxHkxISD3pZqx2wHx2y9QSYTifkLFqZT/0Q9jW+uo/tp75b6A7nupfp9ozMc3dOG0rD3go2a2YP53S3Xtt5oUtBfjX1h//j8bHmn0PXY8fYMxDBFs6WBXxwKWSSABK4ggVrzDvAtQsNM7Gg7T1L8hAtrKT0yJooeLaOLcVT/WT06U10OBT6B+N++qj/b7Yx9V28h8N4PPFESXA/Qo5mNWPoJvzMnJu/BwSy4ZMBEHI3TiC6P95Hk0FDvn0Qgz7ZcUwMT6OKQF4eBifGwkTFPkI0+D4SyRX6aDMh+E0kUZG5leq64Z4CO+00475eSXSZtbavA1xysGiUeAgiKgscNMq00MFgbGYT86MUHzxMBwMIF6KA4oJpGnyOh4GweuQy0h7PNXNu15WTAMDxnA5DR8e3JYeKcB5MSG5bXr4/HKAPaADujSJcFv5ID0f6OURbjrx40JPLwaM4aWOHXABxyKDw0RCclJgwPgRRjPP8/HkPPDsgC5NYrObZ3qt+q1bnm0WrkP7ZxmPMgNG7s+RQb8QDaQD6HI8RM7H83I7xFkYzxjCCLwxR//sj9fz+dHHABwyMgFEHg/zSBFBu0lEems9QQzIZ6Y/TWmM37p81ebYkJ9izXgVFuY5lufJ+WBadI1H23wTwThNYlscBNtnnZnDjkC7yXVlPCNJ+sRqZZw+kfGM5Hg9old2DILxfB3PRxA0zkYa87Pi8Pz9LDZ+jg36WwtuDM8LLvmg1eOvcD/tWI6i//g9wHHD4Dw9Dt8RXImwS90R58QdIIQj46hw++kBMiUkKLuhwf6J7A+ZRmSe4SCyPx8XfOYNyAcf4zOt1rwJb88/DG+cF58QxkgEtM03mnF+Pt8Up2/pQNhIM06KyHhGctVKng0yMjYp3phDKOrC+WakzfunqgfSjHFHTzCjvp5nOFIg65dcD+c8tncC6WNdjj1jb1vX+L1Z81NmYVN87gEYKMfv4zSzjMaDEwIyDvBpME4azUiZIwl6RBCeTwTXlo84/aIjBFlSHk2EvT+2B23MG5CPRMYBDz9+JqF6ftgS+ZfaZ+ePq39BE/rCezHtyg9q4aFg1F5aZsSP2+PCbfc9+x8e0KlUYNXCA6ZNm7LZuTGEHpDboKJELDdgQEmKAlAq0YoHE4V33+ODxVcgaVPUojAAq4DiDNdvVUc/4dqO6cQ0VvZYodwseTdalNX+1r1J73reSnna6JxCwGQ4tlDMEeDTINIINFLcRwSfLm0iXF8I4qD5SGQeoMhDIj+v3i9I/HgF541fDvvM+3CRA4zX5XgwDPhmjiAykuD4HM2sLZ95TAQa0XAA6cLi9J1me9AuKNuXASahlvk0wlxbcOwqsOstw9+YB44CF3EA8hIJfEzFaMF6ghUuEpLrBFxHb1AGrJhcAcbr5DxO3pDniPh4K83mXsbxcj4YJsGHmMfvRAoQ6BMZZ7sqWg6wBwTo0iTCNc93A3IRGWR9hJvO2jr2hAYZgNQ19wuXw0MC+ueZREZcb6g8hM2YD4N57Ci5Ttf6/Y75+T98qoU6fmocn+5xw1Hmw0A3xgDHR2R/lg6E3QowGM9Ikj6RlXwLh78xx2WKy9PwDnHCAL1/It2MsDleuEb0hmwHgmA4iHCN7YkGwscN9HzkeBn+R9Wbb+wT35KvrfMngAcXl0rjJ3AeRM4vY/95kmfDEmF4PtENRImA3B9N6paRutEnkqtW8vQd0Q+69S8a5OASisoArP55pHOeIwnkcR4cj9mT4V2IzfnpwNEvp1/2VLwlv8xC+LpfnOOnDhlJ+rhpsAaeDPnmH3C6nk83JTowAM6IINgf55vcTFtp5IDhYD6AJdNsD9rKfMMBgnFYTjuC4HhIOJJEHoEVYYKF8FU78cLzxvd/r5bHXUGOn25VD58nAo4+OzO4xsPz0CAjuWoefVbGMzLO9sN8Y85uxm/lxDA6zjfPkwiq/FzTzgKWeWjHPABvY0+BrOZ5NFLAAb6jFQdcY15GG8OjZ+Fj/ffIY1iBUTzBOD8YBWKA9Mt5wmfMKycEg3EKQCzzyKN6esozIgiC5yOeEbQxnpE8faLnm2XX8nUMAV7PeHhDGmYeN2vHFC/zrXUw36+D/prxGDtq/jdawbGzVi59zxZWi3y7mwaBOaAkhyNOJvZGZn/F7vCsRAiSAlju0HjovoNijK8DNQUVBXpVdASRSiIBKIkjVDyYKCRRWxYc/9JHLNjiTL+yUHBxbZz7lSYa4wdidhr4P5UtSwNsKolCJ/Evb/7Xdr2JHjuIMfjnjJMMGEYaIF0zEO4DnUiYgGGnaeS86IZBD9RUynwYHibSAKaUjQqH9LwKXwcWWPmcwNXy5YmGcWTkhKp57uf8fnlsl+NIya6xPftzpMFgRtjoBmezWviCnXrsIreHOA3/u31fjFCsBcCv4ZgHQwTRlkcfmQCczRxTHidLgq7hIDLO9kTGOWki85Ay/krYBzcDhoUB+/iBHC/BkYYHkMMCP7slwgDtwvp8mVepWQ9HJiI/6+H5kQ+LSoOHMSPbI4fI9iWCow9oK4zzekTG2ZzDYJIjCTpEEMzzfHCeTz49a0F5adg5wKdazjMcTENz15M4qH7MY2LC3D8RlGWknauPJz787+Dn/P7Iy/l4EHDMRMIEHK7leIkw/PpEtGfCYPnkcweMM9XwdcBxHJ6WXvuf/pY8hKf7fetD5MDzPEEkd+j1XF8+eOqx3vVk/7gA49SXCMoy0s6VcUN/SLcch+thRwboEUE4VJDjYDhjtT/zvHnYnN9op12ylaeNxum0y15rjeZtuNTrXV8Yjpwfx0ckl5F2rowX4zZj3OdrxYHpDFzPlOCABMci3du7LuCdAjJOmwi3GFcieD2axgANoI8HNtHjaOgILiPMsnievdP6HrzO5l/4+JIfVwbm5eMBYjpuEuFAYc2sAAAQAElEQVQO1CMHmIUEujR9nvRBZB0yMp4r82gTGUe6je835hztMConhmbVeXK+oNoK4ySIjJcIkr7fr7Adk75GBMc4LwMXnpFu/zzY2BxfWPQG67bbLdT2HzAAzo8DJQYMfMjxM46EMo8+egTgbOymmG+FQLrxIDoNo0QYvG7REFnJBxT3uRsFj7MxD82NB9HDMEqkwWClev/wHREncvwWj7Ce+SchMqblsdb9VtxKWxaDwFxQCrtQsrATGe0bL9t/yfC/Fyk667jzFQtmboGVfT0mlt6tck8Lb0ApdYTk1WDiScFEoYXasuCUC1KwxbnKu42eK5HUJtwD4+t48/52/PDgW4hP91wY4GG6ASw8NyyurYXwq9DTM/bfyHHXx89N/vxBbB+jY3SzGDmItrzkAzyekf14KxD9mhefewYRCIhDT3obtR7Ys5VZ89vGcQ0Yvxlp48Hh+HhoYHwMEBgjOg2DyLxy/E4gC4gwmzFcIAjPYxhxQKvgh6jzj/1Ey1+/xSVaf3SoiF+cg0ASoBgUbDcqWOZxkEhM4GlwrTRSgGA4SoTBPPZDRMhcBTfGz+nCX+yNcT25mA4GimH7TUkCrlXRx+8JTlddJ8qbyRta25F1yMh+cwdsF2pt/x9eW9uq08sOQCSwEsF5PwkBbcXzMK4SEYVraTpWGokgcHwlwqj+VfZv//BzCL+6bIcw/NJ1w8eDANF48IJEVI6DbkZe3/Oq+cjLpci7P7vDQl6PDR1xneqA+7nm4zEcCHBcANwnRrpE9kOe6AEaiSAYj+F/VNl6o9Ul152K6X+hHHZpcOCYR4KSpgHaSjQcThQ4aD7jCFA/ZLWVYj3NHJlnkLUfgipLzsvIQErHPODhOm4kdEBCRqOBtIzsh/klIhZte6t13WILrng+vI1bTrvsYDxnrrAQim8GQxpfibg8dSNNhNtWPI/z84QiBNcNx8wndEAgIxPhEiiDmQfMMvZznfdxIOBoOCodcDx0M7Ifz6vmo0kunocGIbzEGvE6+/r5T8qh8YMYH+dBgTANH1eJMDi/PA8PMp8GEGFaxnhhFPc3fbZzrnLKfEaGGjx1Wk3C5Hlm7D9N8q4785Oe1M1d+EaDjRIOmo84w0z3fBokwI92OXbRN/FD0Etx2afgcw/oVzh+nx/GR2SYwy0RPPXI83AXCRnLPBigLedl9Dyz/5+9N4G37Kqr/Nd5VQkEgoiIQ6s4AMrcKvRfWxESBDMnTMERkjDFBBIFcQxDCUG7WwWZCVNEUFtQgYQkBFBABYd26HZAQGm1VcRZQMhUdc9/fdfZv33Pfe9Vhpd6NZDan7PP+v3WWnvv397n3lfv1quh0gRZxwJo8AtUU/MEqcdZYR+IEd0aF2nQQeZpCDfvzEM+x6X/HO16xYuQD0Qf33j6jmEc/X0gq3tvvojoy7BFoz62tlj7RbTDffUEjtq547Zm/IMa3/1a5fESrfZ+jnbMlcZDOfRF5L6MnPSX6MRO9zmf2DPPFCh/Dtf7Fhq+4fjvedv3Hfc9l/2JP2x/vYU72GpYdffMgV8Xn1rs2fNhmw78NfrN4yuH4LdZai/c9P3JBigbtNEgxoNyPp+AFGtHBzzA+DNA2o1hG/sRize7LP+QyGt4+XydLDSVcjtGcNawgcfneESQsi2wD9FCODCaznyF8ZvvaBvXqN/VK5/3aMIb07f+3X6KdDWpobAt6bTtykQMTgtNoTMeZFc8NDCbsd7RcRvWD9eU0HWQtWHxQ6Ku1OuNBTU1pz6AKe77bIYGRcc3P4+M8gT4iEGny/MwMfePC/44Mc7r7/3vmDcbz4PQ04n5ghDurJcCHQdNxA+aW/HXwIaAms/WDN/Z/o75m/+QfwnzmUKP0HyGSvt5sB7riIYBdGd+0ujOGRhfCXDupIbp3NY+QbjlznqskwmY2ATrs4+k5BYNU/2O4w/hxIjPUacZb3ryE2AAMYELgoOnv+G3P0tv+I13ut4nsO3pXCmv6m4YcP1BdHenHqeMEy0EQejwN3QezMeI8gXbPHn+iLOOTgqiF8K1YYS9AHQKYZ34rZYviGCuHuCK3/yor9BOPc3R9l2vfseLfe6vEPXVKr2sFgDowTKBEMZeN3k2pj5f0uIbBiwENTWnObdCOejz2uK0jmnyrRCTgfkcdR/jsUFQP8i8eOjohejJh6/SkcP79KLLp1+U0Q+GTn0b6m+FUTd6YehsPFGGEaGDEBv8EaYbvvX6LeIn5n4BsW9OYXZ8eb3xuuFcgjbgs53UmSaEkBvoCVb8S7rPx3MwvV+v5138Fr/fnyLq37AwdZukbvRCU6t+fOv3R26jwXcpiE9uoAnmQ2hptm8658F6S8JjLOBz1GnGm97cb2P3O2C++AcLvsgNKxc6BIjeUefpR1/x7Uj7u//mVUfdxWvePXv0NhznWoYtmuA3r/70p/45hsO3lRO47uo9C79uruMceawrYpLpAO3hCjPdGk/i0BeR+zJywrQzmGmzEFOlftn+u+NLF3v0iDsd9YXHnnj+Wz/ABL930ZOO8Hdlj1hXBNKS8sBpruGfjzj6uhv3vXlm2Mbbhr9j7h1Sp4FaxfspOwjhQozojjrNgzG9ud/G7nfAfPFngPx9kQ3bdJ3zzJ/SoAemLpbw8sr6TtryCiKYqw1RHwL0it9E9yfwIKNpB9MyTv11WQwPwfiaV2l/peHIkxLdyNvajfRttLE4m5kXNS82VULEoCl13tIQTickaALApsAZvcGng6i97c+eJw33mx6K3Fz4ZvWzr+kgFCDfzOfhkyGBlvM67H4HnD8TBa1lvrUb9+Y/4sianIFSf541r7Ks7w3NJwE9NDBDNKcSgtxAE8ybukw5jVy/KTDqFQqBIDcjwxx1mvGmBTHfp1rrfgfo8WeADYUOLfsuTfrW/4653DJtn3AimLfv0wZkg+SgEJ0cNC0aiM54EJ19oOHruPW3aqbYl7c3vv9zpOt+3eU9hHKNnt0bqfpDOAeBQnRykIHZZyc8hy9Sg+o8ggxAABGNpITohcxHDsLNOzw5iM54EM7TAVNHcIQvdTvHZyC1Mns/MLAJAAZwot+vxz3ku+Pfjturrpy+SWdu6gPprB90AM8+QDjqAtOtg+ip2zk+Q36RQYsfggQ00f3mnKrRIiAHM5/cTKA7KlqMNy0hEDQE5NbRQeppaKkdfCIxDxGYaZpv1F20Y/ev+8P5nZAPis4+5Po4lyqo0qoffW8+xpQvyIZNxm+cX+jkIDrrLIbv0gvetGvv/Rcn7adAd/Bm9V/YpYwHZ/0nWgzepP5zu/QT67qGu7JNzffJfiHb8RBKHIAkfJw/uN4Xf/MJNBEo1NScKrrcMBj2x7XrDZ+l5/3Mb3ip0zSv38TyavWgz/cJXfuNGcLB3Be/ueyvsPkEWgDmPnLTQi9EJwfR5Qais17qMBG0hg+gmwYUn6MgAx13v+O6Si+EL99Cr9IzLuJDMuyN6PvGsnOH7uJt3Hm+rbYDL9CiCa4bF/r1Y8590+E/xu6TWX/d6hp9fBzGv6jHuapPB8gZt6jJs8yhr408TBMmmO7QzBfMbeypP8O+eTGMD995q1s/5sSnXvqW+5/9Sv+GQUz6p6v/7pGu8d7Tr0vLuXqUIDcGvP2hZ7/r4wQHvPNzf79QUzfoTShIZQ5SsjGn0LDBqg+jBcYzNP4E6r68P+U2823XT8zPedbp0tB+GOL1qIf1g8p2e10JzAW9j7lvxe95LKt8haZF6+igfKvn4Wc+nrC3/xaNKTbrW/9un8XZDMW4JvaSmhOwVBPwISS1MYju7nT1tEygW8o0TvNNIgjBeiDz4TkY+mUf4F8F/pG+j9TvgkHqZf9Oo5Ov1G8hPm+kEN1098Nbhgb6eWRejBhAVOPOaz9EdIO9/vE35sEMUl8hHDk47+icf5b1esFmcJq6kzZhM/+aJ/6lP/we275q8megUyPDHGW/Tvt+ITwMSawvmg2Z3zGIXmhK5OC8o2u4eX+U3ctmyiAFEzRsQLmCToB7g2DShk47cLpyHtRv2kZf/r1Z3w/49cZ33l7X7nmnt/WfVfWlblcWpGCChoAl+6XQJkAIxoMR5GbBsoPlMfC8yheMuryhk4HohXDkYHqbuHTSFd0mLz8t7LjqWvHDu5eP8ejUz3yFw/AXWuw+xc7tuV595c954umbdAcrV+oxA1Z9oCkVElfHJ28IRGcfTiMHIchAE/jiNxe/edM5t0J0ywItiwaiMx7kfBkP4sNDRy9EJwfhComrR3cSzMTSMHy5P5xfoee/8SgdFI262gEA1ASGdgByDgOBc4GYjOyLEHQ6fT0kMBm/cX7hIwfRp+m+09SzPe2zpfHZiQtH8rVne95naw0cjO6D+5r7oGeLuFDk7h09Ho35CuV5yMHRc8gdZL7RY0G8YM1DTmfcemQe5ujo+aTpg9foDfZ9tnPRrKGTBq2D+OGCHk+shujCZxId2qkzQat8HbW9jf+b/Mjd7/O6D1DqcUEgqxYSV+/1N59Be63fQvd7gprPtDNN49ZNUGlhPwcTGW+cBkog8zNfMIEELzf82B0upzERG0iAOOvMQwpm/MxHjibdVjvGX5rC/Xd3Sf/VnT+GnUVdWdDPbsLl/T/GYfg1l7u0LLVbfPQNT3vTVdLa/5SGdX+ysR2XwZeWbZY59NWkZRSipRNM9zmf2D9xsPJRx7+4Y1i773Hnve0RJ5z3tvc89Ow3+QOW2Xa98b895PZ+zT5/kFY+Q3ns5EiQG/meHWvD6wkOij6Mw1SHgRIN3oumtyWE3EALfkELISm5RJrXtF/AwYmQQHxyA233ryuCji9+aVt+Yn72rnt65p/JOg4mdAF7rd8my91XB0DdlirdWD+GDLTLSOpoo8/ktN+H65U/fuP+erGH1LVWwU3GLOqqXNtyE8wCMcOkzYeR3GkeFtiDJgCz4WIdfEECxELiA9jf9oELNQ7PX3m41M+LoVAOUu4MBeEcbICNVASML0SXW0cHGd7Q0rQ+wfiXOvFr/4noBvvK3zH3hJwvBXja/iIbza+fqPtiFGUyrGMPZjoG0sI9i8+1zWcnNwTWaQiYtc7d3UTqKDTFPABIPRmODgmG8LGAcLOOf7gZ//gbU9W0Qa9HHSnYxLpUxXeUm32+u0B1mrpCzyZg36Sibf2tyuh90i/6vdvo2qPe7rm+VtSbfTtL3YUUDNGwgQoJIpsAPSznUGg60xbGj9EE5xHf7EYd8IVI5HMkZh6wfIVw1VkmPgKvRyEB58FmdCp8IPPE1wyD/sW/VB+vJxz/r9qO9qp3vNzLfYdYt/Y5XweevJAYX8qjYIhZxxfdhqA1bE69jpMKGgZsKJQDp5pjBpoHGyCTTj4GWGA90chB99QzQwbGN/dbryt+jy+Ej1/3061u/QbSA99dH/vgALyN1NPRAfXO68cXk8eFdwIaBG7wI7SOzvhCaE8DQEteT7SG+NbPt8EPgd/YINM4FcF6zELN2ABbaALWK0TXfvje9gAAEABJREFUrKUe54UOV74+JJ/d4nMBhTMpITwLdzTL+qwLUocpFXZfDBPt6SODPUDX9rcjFm9y+fcWddX6qdtLFzrsV3w2FiI4BTyPoepuCDBPoR2Tz0HGRXDSsIEKe2CCeexceV6pw2Ther+HWVVoucXnhQtNrVzwEB2dsC7zgE7b9dV61ivOa/F+AX/eOa4W8g5a2CKAPrF//NHbfPomf7M+Db3h+2eCY8fd//ntGsYf8JH9/bQfRwQGX0StzzKHvjbyME2YYLpD12vdL51P+nV76Z5h/L7hiMVxt/rXT37XQ8976x/Hs+427tq1drtbH8W/vP+FHtPVPmuC3Ep737ecf9kfVHLAceTNkzeM8r6j1JZuIOA5pDnGbwJcP4Fp0To68OGKJUG0ff0T83N3Ha21xSWe+jaS1xOtIcC6hUip20EwgpOGDTJN6YyHCNrKeQBg9ztY71+MZ+kG/lu0TLPJbevf7afIVkwO3bM79V3sYf6CDdGLlqIzPn4EgkK5OSd1tPIhkXXgCokPRL/8z79Yl37gcg3DBdM+qZeCQRdEfS11JpmefHJrAvtHaGnfZx42A0rwkPIRFs14bHCsB0q8OKfohu71x8lZL/O3iQOeONgmcXrD9dtbvtTTJgAyv3Uu8kGv83xHq/wJJiHllA+EqPmCkBlI4GkMpP08PE98hda50AvH0V+ESbbYWY+6MtzrUH/mt5C0EIOJqgcfFGg6IWj79T9/nAf4J+YXv/vWuu3Vb3edX6/UX4VLpHkQ7CPnQtD0gPOgpubU82gaJzcT6I6Ww02YnuYlcN4HOCeN3wHny/mDcIXE1dGJQfRCOE8HTN3zsQ560Gz8oDtX/PhIQBOT/2pp7SSd+c0fQdnn/dVXXujz+G6lHq8LZpEh95Vb6mkMPtu1YoOwji+686C5jo6XD8SJJ7BNK/pEuy4HXDE4aNigT0PAeFCez06B+OQGQqcu5/g2+M3jM0yvIw+I3wgXP4EeoZde4jNLfABv1EXBIGUYSQmr7kK4Zf3TsdqefaKVrxBu3uFznizAQIs1X0vN+EIHjOiMA00xHJjQOgF60Aq+RjsT9FSo3Jqw4oc2n/VBAuM00KJzUkc87tCMNw0l1kswu6GTguiFcORgOhN7InSBJtEb7UzQN1y/1H2MT6Ha/ra24/Fe95/y/F1+6sz6XrrQYfRCePYLwhUSV93o5DkHTxyEcHeadRxOaCJ+0CTzdb8D8sxr3bJccFK5WSZNfZFNbPDjM2+YfDayXnwmCx32C52EYeiFcB6euvmTFz/wmttBbXf/ndc99o5e437ufeupYZUg0ziMb3j0o990qP+TjNnLdt2OPfY9u3/zH+77qh3afazPkd9gXcjP2NdsyVnm0FfTllGIlk4w3Wf8n/o3AJ42HLnjrkfe7pOPPuG8S19y3DmX/cmxu96z14+PV3zO73y5a3p85mi3PmuC3JqSV/Qr/BJdIUs8IDgMY6rijUlVvF8KOWQX6/25NARD+QjLx/svsokNfhtN++5pHKDHnwHSzij77jaOv+DJ7kKZxnZ5XSIg6ztpy08+C8mNENQXxGeh0d6AxHjOReZFM6ITgk5zniA+/ON4o/5bNKbYrK9tRt4ojsVThN3ZFOjORbHZBAHVGgOFmEw41XqfaBYsM71Yx6lkIusUav+3d37k9rrsg8/VYveHNQwnpD65HiqhToiOJlN3IT6IhkD8TWdcZAtg5k0wGUw7UKfjR7eQc5FfQ+PbfL9x163K5vGENV9Hk0xvoExlYbnht3B9PsuK3wE+JmjDJnr8kuiWg+iimcBHCDoV40GZyD6N3W8jqUHx2VgIFz+BO7whvrVh6x/MWc/LSARyA01kftAU6zZa+EwHU3cTALmB6IwH4ycowZ5cW3+rZvjNvR15xKv9AvsmUSf7SHmuM8jkDpxKRtFAE0D3myc3LXycE1i63NANYh18wQRmQcOKvw0oXyG2eYcnD3oehmV9k059bxeCw/U+6PIFIewTaAIYFt+pxz34d2D3eX/1O07X6N8IpK7U7QWDrOT1gXkvXyHais3j4dBlgZT5CtFMK/uTWwRpxS+p0cugE/LrRdHlVjTj+7wJLBrRHS39JqgHIohoH0C3DCjzmS+E7H4SXaAXv/UbEx2wm+vL2jOkftLUjWgidRfCuTv1XTdin0pjPgKQ+RgPwrEemB5BmXfD89fU4t/E1/2TjVRqvkJSBJB5Uo+kYAjl9eF7EB8xGNlB6i5EnHXmIQXx2ZZvjuDIwXQEB93nHN1AeVayvFREw4ALCWpqToWvUPuh/fBj/kLjwL/l8fHsr/bB0uwDpMN3dIHk6KnfOVo6hAN0mQe7D97dtNinaPhNAPHDtdwgfIXo5GD8UqWqdWSB9UTLQAIpurSKTe9+LRv+rDObr3wZZl66o26z+weXg7YvWgxr/NWlHVk6yyyjpO1m9tM7jx5+paWH4XpOYNeuXYsHn3vJhx5y7psfs2Pc+UX+NMmfgPiVUfp9x/wk/ZoMN+ErYXszt9gwCXsM/+IxHzbzvmEcXj8uhqdct2fPPY87/5L7HHfepS847pw3/+OxZ73nar9qbLVrL5drWtM4uI7xXmXpAxLkVhL4Jzt37vhtgoOmLxbK+0yu1RsGRHMqmSiM4BxsgEy64qv33SQorfsd1HzxOY9hH93OfdYz/cxPzmxZJ5FvrGOif50w5XTT+ql77ktefs/DuD7Qeem2LGnz+Jhn1Jv0qhv336IxxWZ969/t1yGnmDa1a/MhtaQJKz4bGh2T08lP0AQAEQztAMxhEDjPaWDaT50P5G/74HN0zXV/618cn+Hlj8rKlENdSVpd6/eLXj4eGrknyBCwDes041f8k7PfSw9irAn0KT3svu/qvhsK6o+yMw9ekPo6miQ3rFzoqdvrohtI46GcbISsCd1vLn7z5SucT2DZztnLwkT5GJ/5Q8Tm57FEdNYDYQuJq6Nr3PrfMXc583JnhXoF11V6IfWajo96sr4JdCYCnWIjjS9BEwDPLC1yPyC3n3vf+T7n7xR1zevPfiRo0dhHNkJC4SaAWep5NPktMD5++0QzmibqNOuVLxh1eUMnA9EL4cjBeUenANYp3cvGEkQgA00EZpjCSjcu5/s+nfXQ7fmG61XvvIeXfR1lT+fnekyo6ncZ/Uo9zkD0QlN793u+lX1idjct1nE4oQnmE2gy84PuXOTxWycvH3Hmd8D4yCbiN4cPANGJwegzX3LEWS9f4UxiutS9NvzMnN7/8dCWNHo7SUCn0/OEMZH9GafCTdpA6mjykZuIz2Shw35xDiQguu0C4YIQJGCbT6A59EY7E3TOL0ETALmBHha9I0EXbHJeaUcHrMPrJGjb/ErdJoIejy/rm8Pv4Y6y7IQmigdDzm7MQwqidzRJnvkdB70eOut5WqEH0d0tzxY2sZ+uC874I3/5P971XNVfByxNfeC89/pdOLpBqdumIITj+X67zwZkA45p3DqCFBGMz0FHB7U+HiYgN72s2374rieYjrX7zGUchOP5BU8etA5SP1whcfXF4gL98Gu3/x+CHMZT2dm0bIsA+kS2Ex/f/3WP+Nl/adRhuJEncOyT3/Sxhz75LS/Zec9/+9bde4aHrWntNC1G/p2Vx40anj2MwwvGcfEaaXidP4C/chz1YnO7xmFxtl9cjxg0nka/9prxYQ89761nHP89b33pyU+9jH8oeRTtRvavu8PvnaZxeFLZ++AEuZUUNPO2I//lk3+T5GC5DYOPwpWJ94+LMvguUp+VplZ6oVkPm97HjmU+44zTQJMmSB0hhx5NmM68jI+AYR/0c57FD0efk3WYrtYhzjomWB+DyxDrB2OQLKeuBE0A5AaiMx5kPsaD8dsDdp9zfIvxt2/Kf4vmUZteW/9gXkVSzLw4is1SCA66j9zGBrE5VQ+aAMitowN8zANamg4zwb6/Xf7nd9HlH3ywLvvQU3XZn71Ol33w/+iaa//dZfI7M0dnwaoj6Pr6wzKRdIYeGHmOIWIU9DwNwYsBIajVVudQiC8T6B2rxhvI6l9lZx6sHZ2wbsrzPpyuXPGZ72jV6bKMDDTZEKC+jg7ib9gHOi9fdE8Boqce66DpledPHfgKy98xA5Y3fAut+0dFlvKNilgPY9B1UTfrgetSwZeP+tHLV4gOD+KHL4QXbetvVUZvub/+N47xF+MXKvVI4vyor6Omx0Gde6sffq6Ty0TQ45kPAE0DyEECzq1jjMsbdTBPkMBS/DN02K/yFSK0YX29BK2QgA2F1MGYwmmel+nxD3l+6H19e81bb6dhfKuXa78Z6ELY37TuxtXgYTs6ib+hYeWKj/153uzbqlPfp+fqhaeg6QEbCjE6BabhEZw2bNCnIaCejrZOAx14IuoxqFCeYIPf1rri84DC4kHTYrx0V73srf9DB6xRiPfBPgHq6OiA/a3Ujx+T0TKR0KfAjwPeAuPCzW744AuRyDt6HLEalq8QzdMDlKv4IPAbG3SawHR8QblVYNzUbz7jQNvnV9VRGK352AfzwXV0EL4h2rwzDxvpaDH+htThUIXd5/kyTj5v95YqPuqB0P5rzzjrt/11+LQsSP0EhcTVqR++EJ5yO1I3RMOA88Luc2BaK/s1EZ/UaQLWK0Tn3ETDb8Kg1APnPHpDoOigCeaLn4GQsw5POscb8h95zVkM2a7+/jeefpRLOHaaf5OaLUysPzIOeq/Tw9cWT+DYY9+z+8Tzf+VvH/LkX/lf/oB95UPPffPF3/KUNz/noee9+WnHnXfJE77lKW8587inXHL2cee99fyHnveWHz3uKZe+8lvOu+QS9/cdd96lHzzl6Zf+s1860+O4iTVc/qIT7jmMi5f6FZpfi/skCXJbmdHMNcOO8TXH7lr+sfgVw4FKRr9afQj5wubNzN+O0jqCFENHB96YGA/KOXohqdw6OvByij8DpN3Wb+71lGd8uQb9fKZhWi+TMkyGKwxvQyGiU2Dph4hBGTZLQ5AjsA/RQjgwtmHIPk/+WzT+5Iy1m3dt/bv9FOmqXJty6C7Eqe9SEEFuoIkVP7R5096MxPickgnTJkKLRm7avyhJoGgJPldX/PkuXfEh9z/fpctA9+tDNPrlH/pJXfbBV3nsLxqv0OUf+i1/CP+YLv+gv3Lu+QuX8qsaF/5Ge3isxvG+oj7qB/vyBPRWILpcV9JCdHenGza04rcn44xeXFmnEyY9AakjZDl1XQpCxK+X66a0/nfM26DU45h1mC/onIv1WIc4aCJ+0OSKvwY2BCg0aC/oYSv1o/d57YneED7zeyBoWvFraqaTUg/jyg9GmGxJCfFpx9Z/Ys4ctQ5xJjaReUGT1Fl1oZuenn8CDFPqCDllMj6yBzK+C5gQFgT7t1/87i/QMLR/3dZ1sTp1yvV0NOn0+p+nDQw32C2G5wAqnwillY/5o5vY7DxMT34H+OInMBu/cX6hk4PohXBtGGHqoh70oNn4QXeu+L0uscZL/aH8yTCEcPMAABAASURBVAm347Y46iJPe7ecr4Ng6vH6IFwhsWkgPgJy9CBEdQjHfZ/k2ZiEX25Ji28YsBC0h8tpzq1QDvq8Njhdvpw9MPMb8VkWSCo3MH4HIAM3+PFZN2Sf6KwHwhUSV1/o+/Xit96/0v2KqYd6s6FpaVKiqrsQjvMI+jb3OZ3O2fOs+CNMt+ILYbO+Aw/zvV1t4vIVosbXdM6fetCDNjAfcnzk7vGBdAvohKDT6Tk1gvHxI8AZ8RGyDnohHDk47+jkIHohnKcDps7EJtBvVP2b+T1To3P+zJP6ze/P64Kz3qlx7RFivxvWpUCT7BO90NSqH98+Pg/Wy3l4XtaTkWWIQafX//xtxGeYfB6w1/oxuaMbNvdH4ObuiYfhCQ627drz6dvc05PfenptOPKSOQ6HXKRoLvmacRx+H+5wP7ROwB/KP2tYrF04au3zqXx6po4S5OZkeZlZDBp/6MTzr/jIkt1v0fUvNAz+nOMK633K+7elvE5Fnhew34eZyYhODDqd3neN2OA3j88w+TxgNBGfyZv7d8yf+tSjtFh7qzR8tmjM6+k1kFSHcAys1+NDsF77pD4mgF7xm+j+BB5kNO0gx8Uw7/PjWls7/qb+t2iZY5Pb1j+Yp3hX5xpdlKbiNDXT6kQMLbXQ0hBOJyRoAsBhgTN6g0+6o0/l2ebdx2eL/3pl/l+y8F+8jHq2y3i2CvnHQOjS94kv1ovx0cbjXT//oNXn26fe2B9J0IWknkKE6giO5z4mgrZdhT1oBLA3H+O633OXj5Bx6Kw3JpDr/2Oddp93Im+513yFTMT0HVmYBLQQmCGSU6VuucUgMV/Vj95oC65bbp1wbAOpI4FOxXhQJrJfxBAEghYtPvOFcKwb9M3DfZfQF3s+qZvTvIzUJ1Q2wrysB02dQUv4up+gCUCTGSbGI8sC49EiECBs/a3KDFvqRxzx0x7n95nv1EU91Bk0R50ul1TolAmGaAJADqIX4mM8iK7W0AlZJ34TG3wWTGMTPgIQXyHcvMOTg/E5AQ3z5aU2MT4EEB+0lxUtGOIPdd1V3wa1Lf017/gWl/PtmTt1OApSgNenLlN+/3OfuukE5SuEZBiY3ozoG/bZjIHmcyHT61x/6+G/omH4Ca/7OA077qGzTxz03ScNwcX4Odo9fpnGtf+sxeKBGvy1dtQVxk+xjAbPN3oGGSfCiQlSR6GDJthf/Nbh4k8gwcsNRA4SmGOcYeVCXxtesMLtryT1eD8r9XtxyqUuUNb35rNV5QtmgB9HQ/Tq6MQg83laPycY+yeY7hGUeakrqecLyo3AoIbMV75C2yODPcBvApgNF+NNSxaoC2Qe0SyYJlL5guZlIf6oyxs6GYhu2973aSM+1gPxmyIFJmQCMtDrBgrh3Z2KekTDAO7n/owz3iwtzty4aquH/bEhkH1Cg30AhBP08gXNZX+FzZf9Wkja0FC0CMhB5imMX6pUrJc6LATlloFGX6Z91+STJmx692vZar5ClPJlWJ9Qfl3cTc982YOwbEffOYx39SI7MnfWTpTblE53Ex9dW+zh66fDw9chdQJ71p7pl7d/Gjqu9aeZILeVrYQZ9b7/+PjRL1sRDqYkbw9XCvK+CVKgA9OS0RvuuC4NP2K0wHjRyEF309NwB6HBBLrZPzG/+rNeLQ338fvaIL/12rwNzPjyehTA14egKfRGK/uTWyPmvhW/dcbFT+C8dI8ODUoP18sv/NAU3vz72pan4GGwmVmtomaKzqRNwIeQ1IZgDIL2qUoJmgDIDbS9Hz7zsh4Yvz0gPkKw+xNI8Wu1pR5TIHqhqb37PV/mL8Ts7nSq33HVhU8WwD5/08nLB9V9TuI3ont45sUP32ign0ev24Zx+MloW7kxD+PArOf5QLhC4ur4et0u1HbSyE5Td5Im3Ch/BnqUkWGOOIbMy3jTmbfXE8IuI7qj6VxaXr5C9Or4x53/UemW0MtkXJCCCRo26PUnwL1BMOlxnXbgNPvMxpFNmHbka+G+H6/X/wZ/p/Fbp3pYl0Kop9Ac59tSZ5rKLqIhIDfQw6fn5JxzYTw4DTRpAz5HneZ5lS+IOOvopCB6IRw5mN4mLp10RbfJy2+6X3zx28MVn/5aw44TdfYpn4balj4OL8s5MHmtX/UjDFMhyL2jk4DohXDk4Lyjy/OA6LUOHtPL89BlGtdO8IfvL9HZJzzS/Qf8YfxinX3c6n/7c+7J/6annPTXOveEP9K5J/+GPc/XOSedaDxaa+OD/fx/neUyL+uxj4lgRdMU4LDX4zw+c4UO+xWfs2AKlspHCi+3CR+gl7zlRGf7+UohXtPo7TjwPn136vOQQM4hdWMIISFMddtvHpo8PkmFmjV0UhDdw8Q4uCAECWgCH4ak5NYMvgvaC2tqMWxIb7j+aTTb6/NRFwTItM0SoB7WDyYwDRqmCQiWvXzMw3xLZYrQiYKeByxfkIExcPP+yPE5RW+pM03LQ8it0OH+vn7kca/zkueL+hysXOxP1O/60A2k8Zj2BhNOaKL7TeM3bPTbB48wzfcPfnwf0DD8puf5Yw3j3xmvcm7XZHDgy+OY3zC9ThJMvO/xYyfu6CA2kABx1pmPFKTeQjhycN7R5XnWhv73gufyPonXdBfPM30wd1CXd+Bwuvu85PL+/hodceP+O1uPPHwd+BO4eNcxt778BSc93c/u6Rq0sz1Nv9yprWck6WHG8d8HjT/x6F1vujbkQXvz+2J0cQZen96fEwhDEfX+gfYhFJ0DIF8SHuSJ8DnqNONNL/0Wb+pPzD2kX+c86zzX+R396wnrZX47ghCOqwDW94AptQHZgAM6dSVoAoAI4mM8yAR72++oM7f636JpL23rH8yrSIrum2AViBkmtWGOTvtZ9ACDBWA2XKxjekICxBniJ+3oIIfZEPu8z+fjsNHwz5G4evxeoBDeKTAN9zoJGgZsKJQDp5rjBr+EHJog9TDOAw2idXRgWqknwt/rYff+WaKb1PdcNz175mHhjp6F9b1MXvxOV67ui0Fq0LEHTQAyv9xcOLmh23rQBMDO0CAB9XSEZIKG1ENa2H2eKOPwzTq+HeOnZsxND1mPUUGvw/5YF1yXqviOcsvACbvfQWgj8+CnflI7pelxJdzu2xvff5SG4aLp+Q9ttYacH/UVoqZuB0F8BA0bqJAgsgnQw6YvjgQmTDM9tiAB59AR36xTB3ohEvkciRkPlq8QrrqXV3wErZCA86CmNuqTrvlEnXXsxyZiG+6vfuczvMZdUg7Ts37QAfvbrP7S50gcvwPQsHLVPB2tertZF9TwK1qMX+sP4ifru49/u9WtX2ef/G6dc/KDNC4eIQ1/rl5PFlJa6nBUSCHxDT6Omc+WXPGZL4SM34FpwTtc4tr/IN2/PYV4SaO34UDq6IB6qRMUzT6AN0B4J6BBIL5CuHmHr3Eg2mw6yeuJ1hD/+vk2+CHwGxtkGqciWI9ZtxkbYAtNwHqF6Jq11OO80KEffO5LbCkQnwsohJt3eBbuaJH1WRekDlMq7L4YJtrTRwZ7gK4D1y4468Ve/Nnuqxf1s69C1NTtIFh1NwTiR49BHJcz4/inxpf7ffpkLfQQ7Vl8iZ71hEHPfMIXGO+lZz7+m/SsJ95Xz3ziFxtvoz23vr1/843/AYH/SuqVnkBKHVJHZUEJrHXlBm1Q/K6jEG7e4ck7Oql5QKcrV3yeT8OpK/w+Sn7voicd4fP6Mvc19z4rK2b/MFNC+o9XffJ2/wZ1uB8aJ/CFdzj62/0S+mGqrcc4PeeeIaWHyWtw+OUde/b8WsiD87Ym8YZzxQ0q7XwR6Gx4jh4mHwr05n5J8UuKzwMK5bbVv2N+7rMf4PlexBvJ6Ik8r+9THQ6S9oVN+CLlmRSa2uhn4KBpXhtmqeSBjO8oNwwGJlrTj+mVz+M3SiH2WfcD2uJcvViPz6GD7lzeSzZJnG4ie2lo8Bd7sdfJF0KaCAmMX5YdRAYJzLGeQeWDpp6ODpIbta6FNwcy3qH6fCTVvR4hPvRCuJo2iM8BukAb4gfduUx7I0Tum/mhzZeP8Tx05rMk0LJoID7WA4fFj0Df5N7/jnmbkPlqHdZvdOZlndRD1oQVv/lGZ5+Mj98D4Wte24re6/OP38ZCBtR8QWvMB9DxeZmV+eKzEMTknnobfurj/mDleKsX61FXxnsd6sn8FpIWYjCROoz4oEBSYtD2TevHhyd9P/7EfPe4y0veeWX52m/f52gLhRc4T2rMwIYBC0F7uZwu99sIdEJwRTdxvefXdOqKz5MUOuwXOgmIXgjHemC656N+9KDJ+EH36XqYHveQD0zhNtxf9e4v9vvoArG+Wks9jsHU4zpBU5r71Bq+Fgqf7dqwTxvwRbchaC44/om0uK8/kD9S55z4h2b33XXOKW/WOSd9pTT8stJcmJdPCDpdvj5MUE9efwhxKSnh+vrh4ieYdXyk43gvveyt30W4/zp1ex/9ATgnpQDqot5COHKQjs/2nEdyE+hzP3x1eNYphMcf5Fbd8xDiQy+EY72O+EygZ14L8YPuXJZnD8SMiRW/KXLT8TE+DzCERSPLOOp095uM3zi/0MlB9EI4cjCdiZm/0CR6S52JbaUu0ZrAfAgt7XURMB7EfiD7j5z5HC//QneXT6GOqJv6Ck2JHEyf+chJ0YMmBl0hrT1Se251ez3zCfd2P1fPesLL9Own/Kp2PfH6/xj2ru/6hJ7xpPfb/1o964lna7HzDp7xqRr00bx+B2ecG+uBFmDSqZcARC+EIwfnHZ08dXviQjinPhAiQwTwaD3rFQ+dyH14v4Nu4xX+03xG506ne7bpjMvMX554/ouvIT7cD/4TuOInT37AOI78202f42c3FZwgtylv9zB+TQ7j+NFh5/i0477/HTfvhz9t3m2BDX/H3KuwgXrf5P3WCYsWSB3l9ez0ht/PmN19JmK+QlPayk/Mz931BeKHBMwjF1AoN+Y3QANTbwUD6MFJmXwmPI2/MJh0UPOZ1orfBHnfuO3CH3yjXvE8f59GfKP7jTJu/YN5Fesa85BYznsAsgeKTxCDptSGloZwGpz7RLNQPtZxKibohxdCOVR8jlS+ILqF+BFnHZ0URC+EW+keT45OfaTxm2R6A7SEIKn7yG1oIJpTqYiGABOA6IyfY/fLzQI+R52e/O/Wqffd2n8FdCsmo7eJMx/rOL++fVYB3e85PMz3PA6hJ/c82V+hJr2lWu8TzQPRCUGnyjqNoC6VAOdOalB8HlAIFz+BO7whvrOOvZpwS531vIxEIDfQROYHTbFuo4XPdJDzIAfR5QaiMx5EZzyIz5bp2vpbdRp/I+8/974v9U9IfyDu1JPIt1YodVJXUhuClqnXqUDRmgB0v3ny8s33CW95OdwEPtab+/DUfMTDcLW//nzE495r+hfE331ejHyz+kuW3ycN/0805unoiT29x8FMr8sp8h0BAGe+ltr8GD3um3/Nju27ht38Dv3yHxNiJep3Od6fS0jQEJFUK0L8AAAQAElEQVQcnPX4zRciOQWmzoYcofvgRMo5B/USnX3ifXT2SX9sx/Zd55z0KK/7Bm9ERqWxPgF1UQ9CEHK2gfW+7rev+x3XhZ59euBC2/KLaS21EavuGboMyhF1hTaRugvbLE4TxWdjIWT8BLOOTgqiMx6E83Bg6hGU9SkkqQ1BTc2p1AjmK19hdOXxqXyFDJv7GI8/mEBt4ITxOwQjO0jdhdbmF/OQg/hsW76fmQCRjmDsPufdb54rdvPUO69/b77uZ/AB7hec9b3e98Vif5QSdIFgr985Wjr7dIDOfif0b1SPz5N2frGe8fgT9cyzfkV8yLbtZl27zvp3f0D/aS0Wd9Pa4N/sHa+SBj/2qqdQEnWQriCEZn7HdeGr+tknfEcSrxMAPU/8OgVqX/bduxe39vnzGxCZ1isZpzvlOenX2jB+qCeHg4P2BP7kjacf+fYXnPgY7cz/dHSr9jT9OqTknpGkh/Frzy+xjy60OP3E86+4ef/rT2bdxtvCP+RxsdlQe3tktWzERGFewM7BBiokKJ/3nvH4pkDISVlnxccE2kIb/Zv4453EfIwGWZfpQLisQ0CPoMlvoaUqJDBtg6ZzaAJA4SB61rEFP3zhMPyWLrrwW1G2o69tedIcxiCl+DYLm0nx5E1Y8dnQaBzZY/zmC9ERwdAOQHQOCZwG2mXBsoOcbWjWK18w6vKGTgaiF8KRg+ltYnQmJkUPxiDoaWG5NWG9n3pdZnyMJ58GeoyFNqzTjDe99Ns2v0of/YvcsIc/MjZXb3y8OJKVvcwE/sVFor7M7wKgybWuoQvdBnQDaVymPWHCCU10v+n4PcB09ML5BJbtjBya8eVjfD+ouOxrA+KzsRA5foJZH/Vvs+ymhyznZTIwWETDBiokKB/1UB8bQy9EJwfxwxfCZzF/MQ1u9218TpZnmV4PCYWYmNef/Vgz7bumcfjkBloAnNV2ptdZIxjfBbi9+c3jiz++f9U4/LTGtQfojGOO0pnH3lVnHHuMzjzmO3TGMT+gs479Xp157OnuD9CZx3ypzjx2cG38tOQVXv/jciLqqvkGTS2IQAqaCHR8ln9S7g+S6NvUL3r353rmx7n7aus6mup1wPmnbhcWNDe/0MlB9EI4cnDe0Zfn4d+wWpyis48/b27Z1vi7T36M53+duy/263058nOSUq/zoDTlWm3Uj144Vz2dv0A0xvPIxOS7u158yYObsB/A62YVI2UQg06nfTaCffQHDWcDPsKpbm/HRHwmCx32Cx8JiG67QLggBAnI/IXm0FvqTLLsBaUETQDkBpbekaALNjmvtKMD1mGfQdvmV+o2EfR4fFnfHH4PdzSVRVA+eHS4eUcnB9E7miTP/I6DXg+d9Wq+ILq75dnCJg6S6xmPe5zL9zeurqfX78LZn4HtWNGEEHIzDsMntdCF2r12Z/9k/Bl6xhl/Z2HfX7vO/rSe+cQf1WLHXf16/01NhUigyxAN5HypvyMB4qyjkwatg+wTrpC4OnrW8QLDvv9gvmcc/RuoQz6YewWvOt39PBzPr1GLYTi4/j/reXmH434Cf/PRTz96HIaf8nv9qPY0HSL3jCQ9TF5j4k9CvOC2n7z2dyMczLfNfmJOvX47tY06Y2cmsjfQ1ACaN8RXyPvLskDLooHojAd5QzAeRL8p/cm7XmL7N7h7WSZ2lHk9ManBjFjeBk2tCfgQktoYnBzQk9986jKiI4NO/fVK3bes//9q3LktfzVGrW39g3kVmeI9W+G0i0YATbhefz8FrR8uxo1SR9HwN2R6wkI54GFkXPkwtF58ITT+ORIzD1i+QriaNuj16qGCSS0UMo9TzREfORifVGmC1GMhqNWWOqCGZ+jU//yXRDerMx8TdHTCul4+L0qnK1d83lBHq05X6mdfRTAPeUcH8TcsX6Fp7JWKIPVYAEXzBMPwV47epmF4rfF5xqcZH6Nh7aEaF8f6DXdscE0Plgb+6OqPeN63aBj/Wje3eflMERwcEjRsoEIC5KADeBdC6hoVNN2xBxgtAKJt/a3K6BvV3/D++9j3WMrrZSQwWzhQEHUVWnMaGezB9ejlC3p8FmzYhvVpCPLcLYzjP2kcv1/D1XfWmQ96qs58oH8izrgb0c849l0689hzpOELJD1Bw/Bnnsuhr6oj6HVST8MAwvCzOuvBz7V7e6+13U/x68Lf7LGM12V9wo4OOI+hIdq8w5N3dBJ/Q8PKFR/rMJ8epSed9LYVfX8k55x8pjT8ofdtSB1LFLnr66jVRv3ZX/lmMsMYF8o6WP417b/ffFj/ekod3NyrnkJTOYfCVrbQw5nY237R8a3XydGCHk+shpv6Y2hl4OMgGzbIcGiCFWRsCAfGTf3mMw60bX5tVk/OzybqZz6HDAe0qT/KdENnfEfTNQ+odRN2H7zraxCbU/UAQQdPe8ZZj3IxV4r62VehSbYPTEjd3sio/6nrdt5Vz3rcM8VPtmPY5tuux39Uu570Tf66+xKp1QHIraOD9fVb7hf7IpnjDfq93/jHL9MzXnE3hu+rPi6uubW38tljJpzu0zmHaLcx1Jh/JK9Rh+GgO4FLd51ym8tecPKT/RT5U6h3Mk41Jshtyts9DK89f6UcxvES7dCLj931nq3+Deo2636C6f0gv3ZdvdSxB34f8qoFVtAEG2c8uKlfCi23+GwsNHWT/lX2c3Z9l79eLP872pqnkPk8fdYDE0C6zgKeESkIF5+DYAQnDRtkmtI7Jvi4FosT9Mpd/+xB23Zt/bv9bNK7oNYckmt06ruyKR6maDH44Reaw8d4EB/jQQZi62gvOb7uN4cO0NGDPZCYL34Gyq3QYfGFpuIPivvU0YnA9fPNpvPG7DKBTyBpoWMup5OPhDpNrPjNNzo+1qvzsKSaV26T72067d7Pd7b167prXcRseOpxnvktBZ1zOU1dxFVX/BbiK8QQwkFDYGP9ns6Chzmwl8B5+WZpdM6D9YbhIxqHH5J23FuP/Nov16Pud4r7492f4f4C9zfoUV/7Lp3+X97T+yPv/26dfr+f0+n3/3E9+v4P1+n/5Wu84M27BoZTb6EJ6uv1k1szTPU7Fv4QToykjjrNeNOTnwADiAlcEGxzH398qsfrzZfPqhAOqDP7JJ/54J2mfvRpopZaaHbPoMjoppXWg+Y3OZ9vGD5h4QJdffSX+cP1T+qxx23972/x1xjOPPY1nueeWhv4cPYvUz1eM1crlPURSKV363EPPiPydt4uuvQ2/o2j88TrPevMz2UqxL9IKTr1la9Qbs0Wn1ORowchqkM4Zh5lnW/V2cdfZubAXLv3PNILf2Kq1xF1pSzXSf1dsMaFXohODsIVEldHJwbRx/FheulbvgRq2zvrbai/rbqsR4pPbtm40dfozoUPZB585GC42a34QqTyzab1+wnF4AXQV/xI5gHWkweiB03GD7pzWfZERK2bWDc8r0fT8TG+5s0IC93vAJ31QPRC4uroxCB6IZynA6bu+VJ3odn4QXeu+NFJQBPMl3HmnC7LbXoIawfTdZ0e5nN+v7I/1wnKrdCh9b/VoBP1zMd/u3Y99h+h9nvf9aTz/Dr4ZbkQVXO5CTl36i2EJAfnHZ0cRC+E43n158PEJtBZb238eiz7qu9crN3KS93e+5mmZLkpavfpQ7k9Gq874qONPAwH2Qm8+6WnH73zDnueuzaMP+lXy47+GBPktlJxmLym/GSH8ZLFtTvOPvH8K/ip+YrvoEyGYfDXAZfmnbIRp96FeHvkdUy+JCQEfHIDPSzjQXwb/M1nmHw2clbxmdzpfmOuc3d9tb8fev1kZWFHNU+hKTEvspchnTqEI2C9Hh+CdeqXCeYLmlvx22c55xJ9fLhe9eMftmtbr61/ME/xrWg2RfFOU22wiIYBC0FcDpxq2qwEckimcwggeiE664D41Bo6YepxEGwD4ze34m8Dyle4ziZ4uKDnYxjzBRGqQzie+1gP2sNU2AMIC8DefJbV/XIzEb/DQR/SuPM7HN2861brhlN/9ueFQGQvC1CmUo/crEMEbCi0Aq0Nvhgm2vbIIEHhNFD9uTtCDk1d0p+ZOtUfxu+q07/2v+tRX/2nzg/c1eumhNpfoTnOr6ViI91P0ARAbmBoB2D8CSzOcetvVU90w9fP/tY9fP4nsXy+mDKilieeBCnPw4LLjS8oNwemJeP04OT5lJR8HR2Bc4ohA5WGj4B1oIfhvRrX7qMzjvkxnX3/fftfk51x7Eu059qv9D74HXJWdW8FDKALGPQBHXnNwyxs/7XzVo+Vhs9xPZqa158CiXpIgwl8vKCW6FCUHXTA+Xa/yWZ35Mu674quH9STjn8j6QHr5536lxqH71MrS6nLBRemMOdB3+ANk88BOft1uDw/ktbR5fEgPtZZrJ3b1O0F1hMLev1aqdLUA2lib77I6A7ib/PEb25+oZOD6B7Wz6MNQ1bVgy/noul1FL/j6DOc+7rf+oqfBRoBWF5OYwIZgrpA5sEDWk7IOviCCaa6Is5u6KQg8zEehGMYmI7gAF/Wcb7eF79521R1keILamobfBN90Nx3nXW1dl53ouv532K/vf4U7nPUG3TE7nvrgsdfYc+BvcYd57ug5X9byjlTUep2EKy6G5ruV+mFCOw3yK1P6MTjy7c2fJ2JfXZdN+4+wvu4dSb0MsF+G3nFWZ6Ibzr6P/51ig7fD6YTeOd/e8jtr7ru07+icfBrUrfujzFBbivlhmmvNb/Kfkd71r7n5B++7N9WTAdzkr9jToHeiTeQXx+CcA5MS0YRNGzQaYKcgYWg3PAbuExPwx2EBhPoRv3E/Ek/yG92vZmppu7xBLyPmRhkXWgQrU1PKOorX1DT+zB+x9ELPZD5yldoWvCgPHAcz9zX/y0aFWzW1zYjbxTHYVTRrpm9UPu0e2YYfbOADyEp+UT7LujJb74mwCc3MLQDEJ31wGmgTRYsO5hNY6J8wajLW+pxCqIXmhI5OO/oauuge3rSWExPC5M1AcAAxu+gfOSb1Y9uG8MyHzk+/PA1vYZPaveOU3Xa3T8JdYP9+gw39HfMGcv64Lzf5PNgA95QoLBN6DT7nTZu0gQ+R2w/9Dj+gD+Q31OPut+lOliay0wpQQomaNig158A9wbBpMd12oHTlfPg/E3b6Gubf2K+Y3GGaj3QKyZfX/+Nev5shMKNAHOBTvsvAsybdUrAZAMpIbiml+mxDzpGZ3zT9A+4we/r/oTj/1VnffNZLocP3/7gz8IsAg4f09riBH3Xif6JPdx29+HMleevYUpZNuU44Pzr3ILm5hc6OYheCEcOzvs4vt8/KT8A/4XYvIgWn3vSqx39vbv37Q1T7w3XL45pel35vOTGOEPniZkHAkT39P7d+NORtr9TVxb0vtpqLZ3qhjNBXX4hijpF8zjqJQSdTn4Ck/Eb5xc+chDd086n21AAPgzxed4gE7g73eg3z1U+xpeP9TbUj9m9/OjlA+Et96vmCzIxBhBHIXHr5YttLzpWcJmqxgAAEABJREFUfNmnjawLF3ROTF1g9zlBR65pgxDWyk94MPUfOvvjfl1/i0v6sHr9w6c0Do/UMx/3GKFbPOAXf6x9WHuu1h8n5885BwnWVQoPBWZ/ngCEKySujk+eB0Rf7NufmK/piNu6gpX/Kk1p7UN5Ym7DJ4dHv2kP0eF+8JzA21540jfsvvWR7/QrhH+DZqef5VRcgtymvN3D8Fqa8v+7ezGef9LTL7/5fz1ymm//3Pk75lnJu2ZDhrwPwQSITcheLSQtbDrvp/jNQ+V9lmD6ZQOa8SC++K3fmJ+Y77z1GzUMX2b36sV8Gjy/C2I+A2lMpi0knNBE95veq98+6mMi5rO10unX2xAXbsd/i8bMm/WtfzBnk1Q/2xOp2JxoCMaAdztHp7GBPcBgAvCw0EETpsV6OWRIiIaWs25HB/hW/HhbD+8YzEDH+A15COC840MvRJstL3m9zNMwYEMhutP4Cjf4JWyhCVivkHnkNsgfGvQwPeIeH9a+bOyLhTt6ctZnXdDpytV9MYgyGd6xBzMdA2kh50AObuqXNAz+ncjxbv5A/hM62FrqdlHB2kjDBirsQSdq4IRFc67r5+P80e2Utv5WzfAbuo3jGfliNtg4upDC1G+ukDrzHGNAyLAEHqb4CJreoNMEkS2AolVgNM30Zn9Uj3nQk43753rcg9+qxfAgDcM/TxvyN7MaTtAZD92+3xSY7+ziK+/uffunOXUADQF8HR3k+TREm/c8HxOFDvN1zfYg+bJfrR2Dn/uSOODROLzA5yA/h+kxyIWz345abewTvRCVPOgbvEFzRJ/yu+hlb70v8vZ2v67lfbAxgMU6Oqh6QDR8hdSZ4fbBJTdRCDfv8IwvRLMdgFbqkNv1zLfBD4Hf2CDTOBXBesxCzdgAW2iC7NMCaNC8UTfzFUaDICgkbj0+84WN7gDPwh2t1Log9ZhSYfdRGPNqeh22VPGZD+rgbD/yuH/Szt3f7OL42vUx4wPEv7Tu4KC6FsMb+jFyvhTH+fNcCuHmHZ68o5P4GxpWrvj8vDrqftp18fQT7hXjVpM9t+PltTq6fSj3skttPPzT8tVDOqDZu3cds/PyF5x0zo5Rr9cw3J9ieFzg9Mx6FopbGF5rJBr+z+4948NPffrb/1fSQ+7GG847aqDCHjQC4EDm6GHi/QQKgaAhwFl0dMCZxY/P4m7367vO3XWhhuFbNvl+ReaVL8jMR13OCib0enBqCGR9k6Bh8jlIOTE4adggw0sfx1/URc97pk377Vrb8krZpHdB8RxSsM1mOofXUrFL9EJ0xoM5JQIMoNyMpI7ycJxOSGCS9QzabL7Ma1+h1rXw5kDGO1Sfj6R6KwAfeiGypwcUxOcAHSIpuR0G3wW9PI8YnBZK0Rkfv3nW6+cit+GTGhYP0an3+jUn++a6rv6OudejANYPenrWb7QzQbtgTa0JK34rjY6P8VU/PBMEZz7G137RV/z+YvnIr/l6Pfp+f+ERdR082OumJDZmgv2wj6Tk1gwbzsO0yie3+EEHe/Vb1zb+xPwNv3mCV/gCVV08P5dDqt4gnPR9kqdgTT7nSY0hGgYsBDU1p9P7eUqFH11u4Npwgc540C5n+/d6/LG/p3HPg6WBv3f+rTrr2P+t/dX2DGdyDJu/XlxEzr1hno8PCjSlaaBWWvkh8dm+ajMxjv9dTzj+4HqP7R4uSp3UP1C866T++voARTcNTK8jG+M3QsZPMOvopCB6Rz0cens7dVEwyEpGUsJehwnqgiskNi3bs8/kJtBrHNy8wzOgEA1/kFt1z0OID70QjvU64jOBnnktxA+6c1nO65a4ntOK3wJ5+RhfPkuqeeXWlst+4zcXv3F+MR85iF4IRw6mtwnRax30RseSdSDIQBMrfvONzj4Zn/rNH6zXDz7xb8U/hLp7z3/RBWftv69jN+U8+Kn5qN/MkJy3I3Bo5w+aUiFxdXzEeS74nZTPaZ6TqbyOZCJ+0OSe6+7j+z65xsV49OpE46avDLOfWvUdzg7UCbzrJQ+/49V3uO1L/XJ5qWv4Cj+wgZeR4/ay6VkobmHyGiLT/11oz5NO+/63/1GyQ+02DMPyfeHi/bbwGYi3SQ7AcnAiJDAHoCXNWTCOgRv8zWfIOujxZ4B0fT8xf/KzTvNyF/RxzJHeCmAeG6KDaMwf5Fa9/M7RSdvy0zATyY0Qfd4N/t/SNUecaXa/XmtbXo3N5qF4hmwKdOdir2wWvZBDwBfE5GAzHxLjLANinfIxXhaCGC04JVL5guZlofvimG7oRCB6IdxK93hydAohjd8k0xugxTpy674YFXqDD2KmM0FLxfjIJkAZJ92/0zoco1Pu+1val+2I9t+lZR1PnPW9cEdzTn2f3ozlC1q4Pp9lzX3TPqZ5Bk2IPveZjjDoxXrk1z426QG53cCi8/pjbQTn0ffpjTVaG/bZBEBuoO1iPCgTvM5A5rNlurb+Vp3GX899GB6rrO8CCrE7BabuuqgHPWiWOhstOYjfKBpoAuh+8+SmFT9BIwC5DXqNHvPAH3N0YK6zHvrHunbPXXTmgy/brwUMwxn9FxvONYtzPgkkzp00mMBvF1BuhQ7rKl8h/IrNB37EyDcmKAdPP//ET/gc3jHtl4Jd56bn0Uqu/RVCdz9J6+h5Hbb5DF4H8RHctrezD1aYYdY3R12hTaTuQmtcToF+HvFnwOz5xzHd0IlA5mM8CNeGEUoRDIWa5mupM+fcG8F8OT84T9RoMmjVfIXoCKDtYnzHBGoLTIjPkcDIDlJ3IeKsMx8piM+29jw9LRMg0hGM3ee8+81zxW6eequApBaCmNydCr1QB3n7kTM+rF3+gH4wlzkM70h5eT6OVjAHve552sOFj+cF8jzhOpLw4Izocx/0jt377F9mH3bsmP3x9JGVvKivVrojXyN/3m0bfwH3EoevGzyB37voSUe8/YUnHn/dnmt/0S+VJ3gAr4Z1z2zlwdnilx93Dwho/N3FMH7XKU+78uD/b9EoeLM+ejP9fWFDbTnoIynMyTgHG/Dlj1QE5fN0SguRCDk+1gntCeIzTo6N9yfv+kqNa6+/wa/jjMy8npjpMq9Jp763K4KEj0JaqkKC+Bsx95Vf4/Tfov3Mrqu1n9vWv1jkMLwpNmdI3UEIMtDEio/cmsF3cTbTyx6i/JpaS/OQkHNYBCVgc05KCDqd/AmkHLZWW+oxBaIXmlr1M6FJdHk+0vjNOfVd0FP9covBaaEp/KnbcZCBTYdigko7OsCGX8NfSYtv1Kn3/IPYt+OW/XlikHo7miM3rFzoqZs6B01lNofTiSC3Xr6gOeZjfPkK0S37usgfyvkHOBwepFfbVqpbqb8JDZbnYqJ8tX/2a1qF6ORgH9gIQLQFt+3qJ/T3TX8+XqrX47jqQu9120B9BhzQG59/lInGx3iQ+TgPcBpo4/AHeuyD+AXT8QG8zn7ox/fr6q955wN8/l+gjecxnRvF1Dnn/DhAE/Ejzjo6KYheCEcOTv1n9bgT/2kKD7b72pU+D/XzqLoLNWu1v8KZJI4pry9InxdE+ZLaMOq++ulL7oxj+7rXyeRG1iUGnU77bET2VwKcDaSEvW4T8ZksdNgvfCQguu0C4YIQJCDzF5pDb6kzyfL0AiRoAiA3EJrz7UjQBZucV9rRAetknHW7Vq7UbSaIbv9UiMQ4Url1dFA8aGnlYh4IEL2jSXLqcNj3iS6v62mFHoxBMj355IZgOHzdzBPY87eZgOPkfDn/jgRRlzd0sqB1kOcEV0hcHZ0HB6KzzmLYZx/Mvcx17r7ah3Lmp5uZrp7smPLD9wNxApe/6IRb/fNVf//ccRx+zm/5b/YrJ599+tNJkNtKeWHG3P3WH9+9OGLxbac89e379odkKyvupyR78imwNUNWDUKQgSZWfC03+DDE2yo4BRLIMLmB+BgP+tDF+w+0vOF6+tNv67ne4n67yWdH/Mb5xXzkIHqtA1frEGcdE/gEmtyr33r328c1jh/XniNO2O7/Fo2lNut5cW4m3CDHJtnMbE+k4hBEQzB2H7lPsUFsTtWDJgBy6+gAH/OAlvzwcg9aTlLIfDyM+PuAWHIrvhAS/xyJmQcsXyFcTRtkYYKGAeeFzONUc8xBxRB6noaY6vkt7V67v39S/kFtR1tjQ564gTqaY/2Ul8JNzK74zHe05lTxO+5BIwA22NFB/A2X/v/jD+XfzQwHfad+igx6H+yv9rEuFXz56lzx4ytEJwfxwxfCi7a1tyojr7e/4f338fvo9rre58kMrZDuI3fBDVTYg0Y06DSBh6mj3CCGT2vY861ObnnXqGPEuXIMqgOTmwlSR6GDJvI6agg378xD3tFJ/A0N7Xp9w4MP1ha/Kur39qXaZ6FWW3w2Fs5V02K8aB4fMK4/jyP0jUjb1ynE6/K+Blioo4PU0xANX6HphOxvCvx29XzkjAs3uxVfiFS+YJ8QRSpfodw8ve9ehzt+iIYNVEiA3JExIRwYy1cYn/mOts2vqqMwGn4H1M88DhkOqHyFIWc3eM6zo7WaB9S6CbsP3us2iM2peoCgw+1mn8CO6YN5HSfnz3MpXD8/PNwcb9DvBxc/6MGDvtL3fXKNw+gP5u1D+YYZvV7jFhqObOFh2I8n8O6Lz7z1lS886Ru02PHeUYsf9NKf456rP50EuYWvW5i8tsSfinjvziOPePyp573jL0s/ZHEcB83fD2x08G5AVdAQyNdP60ET+BgPbuq31zbfpfhsLJTbZn/H/KqjXycN99jUr1mreQqRPL1YD0wACWEERguFptgGMGEEpw0baNzxcL1q14ctHJBr69/ts1kOwXtWDsn1symDgggkoIkVv3ly0/nVn/E5JRPYmSCIz9308icLztEBevf1QGK+zM9AuRU6LL7QVPxBbq2jE4Lr55tNl/plAl/Qg+IH3bksTz4S6jSx4jff6PgG/U+deu9v0CPu8S9Wtufqf8e8TZ96HKcO6nNs8F1sK3WJFoPTQnMr+238jX6e5dcTPdOhceVcqJtyQRM5P9Dc+vMw7QNDcOcywTBC0Ony9W2C8XV+ePIAFon2+W1YPEjMv9f61ZrrIuo+chc+gbfnwKkDuwicZ15S5y11Juj4TCvNwbj4fj3m2L9Ieou7DQ/c+Pw5BJ8LQM+5OwB5fRSaEjlI55yDLQDQgwit33o4eH/X/zZH/sXG81i3AfbPVsDszzoIV0hcHR8xiG778tzGr0Pats56K+9nr8T6hmmffs5VF5ycB32b+5xO7xvrK/4I0634Qtis78DDfG9Xm7h8hajxNb3qRqcuaOYLYnZf8TsvH2H5GF8+xte8eFb8HoAefwZI5FrX0KFA9EK4Noww5yUT6EGz8YPuXJYnH4nXj6/QHHpLJ18R1g5fN+8ERv1DJuB8CXhOeT4mQLhC4ur4iEH0QjgeT15fJJ5n/fPUsM8+mC8Wi2tYIcslYE36MiEaRt3JmMpQD/ftP4FLfvrkO1/9iX/+ycU4vGXQ6td4P3w3DhEAABAASURBVIupgAS5TXm7h+E1Nepqv+dfod3Dtx133qWH/ody9jcMPg52OHhrJpzm9es0BPmSwBDawYT4OBsQ3wa/nUxv2PTXt50Is/6UZ/+AfY8Mk3k9cWFIbm3C4guRWB/Zw0inDuEIWK/Hh2Cd+td/fVgbHqNXPufdqAeqb/2DeTbrzbHJHJK34NT36eGx2do0uMFnIn6QwFi+whkt5mMdEF2tMYww9TgItoHxm1vxtwHlK1xnEzxc0PMxjPmCCNUhHM99rAftYSrsQSOADb6R/wvxyTrl3t/uGbf3OmKkuuUa1A9ViFKOYAo2C5oIzNAK21H2KbcYJOZDaKkKCTzcBrm9TI/8mkPnX7dc1u3S24ZW9mlDo7WyT/PkdR5yKx/jkdF5DpbaG8kRwtbfqp7geq7hARGzvtdJPUbIBoSirtQdQ0ttaGkIp0F83S83C+UrRM8+Q3xAZxzzMhtv4PoMlYfxGzQ/j5yL3HxuvufK83EUhPe5la/QMtMAKl8hJMNA6YN67HGfmsKD8J7ahk/5F2v1fYi23EB4qOzPQbDp8/OwlAud1yWI7uPL/BG1vR/MWU9ZcFqNe6WppxF780X2ALYXP4HJ+I3zC50cRPewvs82DFlVD76ci/zlxob4HUef4dzX/dZX/B6fcUZ4y0mDJkwLgrpA5hHNgmUisY7TCRNMdUWc3fCRgszHeBCOYWA6ggN8rAeu98XffAJNBAo9nsupossNg+HwdfNOYIf8E2dPUceZ50Nuoj+nHLzJ2RWf+UKk7ifxeCDPa+aDHse7RNoHtyN2DHzPxk9UZ7N5vZYl8m3UeJs/+tnH3KbRh2EbT+CNbzx9x9tfePJ3HDHoz/yWf7K/gNxpvpwfx5QmyG3K2z1Mey355fJjn3+bOz/1pB+4/GNNPvRh9Oa8MZ+LpreHd0xuCFHow0sOll7IQE8TPSi3DDT6Kl/en8673/H8J+ZP2fVgjcN/V3weX2hb/3WLWEzoAF0zH3Stb9qOdkWQul/ZrhqtBHJrxNL3XL38uW+wcECvrX+3z2GwGQ6j7W3aKwR7Ai3gQ0hKbs3gu6BXT8sCPrmBTvNwwHoY4DTQJgv4HHWa9agLIog46+ikIHohHDk47+isxzrowWbw8lP95E0AVvwmysd46kJnCGgZ0GJ8qY4c7uqflL8s0v66ZX9eDKS+QlMiB+cdXd5Q6i5sBqcbzyPGiWa+lmYE/nH8N+086oeSHyo36qbWYG2oYYPlY4bADDKgEM55paMDp+2gEKU6L9G26SfmGh/M7NP7zAW4jKwLyfrZCEkEl1doDr2lziQPt0FK0ARAbiD6fJ+MZ/5huMCOg+fan5W85h38Vy38/SqpzkMclNyMnJsjjil0zs88RPyIs45OCqIXwpGDGqafUiU+aG//IOrdtP5ZzdGdBzkXx4wziBSeOGgCROdcwWj6emD7utfleVEQ67IQGNoBiJ56nOPDA1IvMYgvSGAyfuP8QicH0ddNN70/Y/DN8+DLOqTkoDuX041+BPeal/HlYz32wXy2CMQnN3C9Dz+85X7VfEEGYABxFBK3Xr7Y9qJjxZd6bGRduKBz4tTtoPscoyPXtEEIa+UnvKX2C1/9pXrORV+n577mVD3nVU/Sc1/1TD3n1S91/2Xn7zb+tn70VX/g+E/0o6/8sPGvjB81/rPxE8ZRe8YP5Pj6sTrgnHkOnH/E2Q2eFEQvhCMH5x19/tw9vdM7zC03J15cM1zrl8LsNzlZYJoxUW7koz4x6o5Eh/v2nMD7n3/6UVe84ORjbvf3V73Bj/1iPxf/Rkh/AFm0ZwlyC1+3ifFo6cMahsee8LQrLrz/2a+cfvOoTIc69k99fqOxYYPPSgITyK0JOQoLSQst48v7LQKEe9MdIWc+xpvOryPxW6yfmJ//7DtrMb7RjGXPgx5/BkjkWtfQZR1E9zDSuEx7ooQTmsCHIT5yywbfBd19U8H/U6+48Fk6CFp/RDe5Fg6FzbDJbNozgNNunSAUWEja0BAb2AMMJgAPCx00YVqsl0OGhGhomTLU0QG+FT/e1sM7BjPQMX5DPpyA844PvRBttryWCysP2ctnnkL0+E0UZl3n4KBL/IvTvXTavZ+i4+/1r55k/1yL9q+ys6/U0evxNlxoUuP6ajb12xS/kf0yXyE8+RyZljw4vE6n3f2TjDxkOnVTbLBvxIyJdanqHDrKzT7ffdDq9OCBoY2cF0Jed2pt62/VNsFG+Pnf/nxp8O8me2HWZ90sTy6XZ6QOhyrsvhiVho0gWHzDBhleOsg8oIb/p8c88C0Mv6X1ab/DAzh2Lc/DNAfTDg4wo44O8rpoiDbvzENeSBy/A9Dgr0/31iuvfI9eecXm/aLGg+mXv0cdiW+gv/wy++0Bb0x/mf34Ol76Hu/383IuDlyvq76e/bIv9gvaOfkd5Bg9zqHQC/GRg3D0l735fsD29BTiqY2tHHV0QB0r9dhnt9+AUngpKLfk1gtNrVzwHFwhou0AtOT1RGuIb/36G/wQ+I0NMo1TEazHLNSMDbCFJmC9QnTNWupxXugw57CCSaZbfC6gcGKXd3gW7miJ9VkXpA5TKuy+GCba00cGe4Cuz9y266Lb6Hmv/Vo997WP0nNe80P+8P1q91/Vc1/9QT3n1f9mHLXQX0k7flvj4q3ScJGP+Tl+Vue68z8dHGP8Oh/X15i/l4bhbsYvte8LNY7+gDrczuhUUxsmsM/DfNA8hzyfxhfAE3d0go/xoNOVK742nwtQfHbsevnn+X6zr8URw7Veon3/4nXajIlyg5iCYbH4T2SH+74/gbc//9S7fnzHVa8aB73Js3+b+5F+IRmW1/QUnCfIzcnyCuPXkF8i7xsXe77jpKde/no/29BL12dANA7ek3c5fz+0VP0N0gigfIUeLR8M6eZ+SRknKT4PKJTbbneuPcObrd8x86DDgX4G5v34PA5u3tEZUIhWtmBfGEUizXxazhcfuftkIHi/XvHc7f+Tyqx0I/rWv9tns2yKTXJIwbYih8HhtVTlK0RnPBgfwboJSOVWviA+c6xn0Gbzla9Q61p4cyDjHarPR1K9FYAPvRC5laEgPgfoEEnJbTT4Lmi/KjS1GBwuftFbv7dOuddpeti9P2Bi/17975i3em5q/St+l96myT45L29ObBy+oyKT9l+Ux+HVZg+tK881G3PdoImV8yC3ZMiG5+dhWhwAw+QG4mM8yLlt8Nvn74K479O+GO+a55B6WiEBFxKcrwbhnDrnflOk0z5J8Hk8+0BI6jyI7u4064JavMbMLfca9ZUc0/I8fFCbPf+cu48JRC80pUyg1YY+MVL8knLektHfGC/4twUe5Me0EYdh4gdZG92dq+EI5zzouKN14tGY8Q1lT+YphHec+RqiZ5y1IPOPR7tOTa8rCve5TIQ5X6SG6dysjybYJ1whcXV0YhC9EM7DNe74SsLt6VnAU4MG9uFyiW6wfnwMo14GgOvrh6+OnvlroAX8Bmhg6uiO8KMXmuo+1hU+B+gIScltNPgu6Dwn0WJwWmgOH+NB5mM9cBo4GbA76nT3m4zfOL/QyUH0QjhyMJ2JvTA66yUlt2jwXdAuWFOLwWmhWXwttSBlfgh9ZrQfe+2d/EH7eH8Qv8D4y7rwtX+pHTs+5Z9m/b60eJMG/bhfp4/3Zh/s5/NVGobPNkqD3HwO/TxCmDOadqDu4/xNh9jgtxPd4HUkdHJQboUO+4VOwjrohXC1DnF8JjqaxK/hcx3d7GvnYs0/MR/8wZwCpukS5UbeA29t+HKYw33fnMAbTz99x5UvPe1L3v7Ck5+jHYs/86zfOQz63Lzmppup6epPIUFuk9DuMMM4/oc/2L9cO/SQk5/+Dr/2m/iZBsMw+MXoXfG+mCDH5XTz96cFDsjW7uP9ZHpzv43d78DLZT3QkviJ+bm7Lnb4tSvzmei+zJ8FYN09j+/RZR4dhKt5TZNOvfzO0ElLD5roOPylFjtPs/Ogubb+wZzNcqpsLofkPXmvvudZKYcGEUNLnbc0hNMg83S/3CyUj3WcCj3rWAjKzYJTB1L5guZlofu0bOhkIHoh3Er3eHJ06iON3yTTG6DFOnLrvhgVer3Pb3zzL9aw+y465d7fplPv9ac6UO1WtTD1Oqb+7M950Nz6+k3JGxAb734l9b0h48mM5QuaY75Gi3lGvV+P+uoDdwbaQpvXn+GN4Dz6Pr3RRmvapwSiF6LLDbRdjAdlgvMH47cn19bfqhm+2W0Y7yLWRQNZtxAu9RDQXRf1oAfNxQ+6y3r8RlLy8hWiI4PojF/s/OXYb7E3fnOEzftgOI86FyjOLegb5865BRP4/QZam/tI6eUrhCs7MeuA6Iz38vlFrxAtfggS0MSK33yjpQoKrdW8hEUz3tMo6yeQN6KkcsNnUHzoJnIukOSgu2nf1X3dL0838znNhZ56PJD5DNkvIvZRdybcns4CzDzDrG+OukKboC4VWuNyCig+Gwsh4yeYdXRSEJ3x48i/Rnyst3+spdaHCYfN0Fx4cO1Yr32s1sA9M1wcqz3ua+7gYjxWC8eFY8uDjkdZX4do8KCsE3f02slnKMfVB03/JsngDS736e35IjdMl3WC7nOObvB5oPj1AhTRMMB5WzP4vvQlxxD20Lvt+tnP0/Mu/g5d+JpX+kP4h7Vn/EcNwxVajBcaH+H3xZdJtT/QGwY4sDmajq8QPeNMxCdVqjp/WeD8RbMPoEd3sIJN737rdeFjPbD0jpi8TgD0POVL6lzaJx/MFzv2fHrU2P/bycycG4v3gMTVjj7XhIdvN+ME3r3rmJ1XvvBh97v9A6/+0fG6PVeM0gXufNxr71Fns/l7liC3mVpDxj/fs9B3f9kXHf29J55/xTUrhs+0ZLGQeD/4Fam8HzS1HI2JQvQymJ6ngi9fve9i0NS638GKz/nCP4wbdKZoTgFRD/MUQpKD6c2ITg6iQ4NwWYeAHkGZl7paqkKCyf9J7d5x/IH6b9G0l7b17/ZzGN4lmzNk/iAEGWhixUduzeC7OJvpbQFRfk2tpf5FQt3Hw+CQJ0IC8ckNzDQOyhe0Nr9SjwkQvdCUyMF0zwOiyxOTogcR3E1P9Tuuutb7w+s37ThDV3/W5+nke52vk+/7f50fHFfqdSlg9ucNBs2BhpULX85j7muOG3MejC/f2vALbeShA94220/BtY8QTWiQx156RwfsHz++QtMrfvgi4mM1fzEF9m1vPzH3pNTF8w4656Ku1EFCISbQqS8puTWD74Le+H7Q1OJ3yPj4TQzDX+vMBxxavzHjLezba7hLP7eBg/G5TISXcU7qKI/B6fT1MIEUv1Yb5wsDohfCkYPzji7PxzrowWYwvfF5NkPAhmD5nfdC4ZyjE4JOp/obwXqErF+YepyA0T0waA40rFzlK5yLrJd6ID0P65QvaQyI7osv8W2brlrHyLqsAjrdeB4lYLKBlLDXbaLOoRC9Oj5iEN12rY1/pKee/p6b37+9zQGk6EeQAAAQAElEQVS6f7/7U93B/d2l6a99zffJftn7vKOTg+gdTZJv9vrI66TpnJ8fgzNBr74fdOi0C1/9/+l5r/1v/jD+p9q5+x80Ln5OGp6ocbybOIfs0xsNampOp/0mMGdEd5Rjcxq9owPOV0Y8IDlpMIEV0BAddGde6PjITVCXw5ULHSLoAWD5CtGro2edNp+B1L8Jccey3By89pNXf1pa/CNzMHXOhWQZTNkk3s3luOhQh2838QQuf9EJt7rsp0955LV3uN2vjsPuXx0X4w/7Wd7L00yfY6Yzdrq8QpEmyI2sdzOLcVy86XY79Y2//x9X/MK9H/2ma7v4mRpwWn4hyoeXl2m9IoM+EdFAEyu+lhvEwELmYQjIMGIQnfFg/AQII88Ml6chdxif9UJT2uv72WL5GO5hZqQghNxAE/gQkpJbMvguaA3DKQfyv0XTXhqPaC/SDdA5NO+WTRpUmICxEMbuI7exQWxO1YMmAHLr6AAf84CW/DRzD1pOUsh8PIz4+4BYciu+EBL/HImZByxfIVxNG2RhgoaB8SoNw9ttfbo0fKVOuec3uf+sHv0lV+lgaSt/x9xFsT+DpvrVUetafN5vR+tOJ7/jHmQiJc2bUm42Ms6g4PAek4feRf1UHRwcETRsoEIC5KAD+JXzWB0ufOiF8ctt629VD97LtfCHQhagrkJbndbySiA3dAC0gedHnS2NzbR60IQGnSYo3zi+S7f0NugrOEZxnnUuITgYE5wfYUcHfL2K3zravMOTd3QSf0PDyhWf5ylEdApMZXi9BA0DNhRidArEpghOGzboNAH1dLR1GujAE1GHQYXyBBv8ttYVnwcUFg+aFuNF8zwBI/MZ/KEEZurD2p2nYLrv2zuFZEGtL0dVd6Fo+BsyjBAdlInU3zDc7IZvvb7u34yeuQ/h0Pun+s32C18dnddXRws5n4acp0MVdh/z+zk0iOx08hEg6OBuF77m/nrea5/v/tfS2u/4GH5QGu6Z75tqQ4Ocej+FTu3jbh6I4KBhgxouAs6zEH0+Aefp6VUoDCGUBUgdhQ6aYL748UHOOjzpHG/Q73niBz3YoDXtk38A7phz7/VpDcPHmLJvexl4Me8yosNRd/7Nl3/HZzs6fN3IE/CH8Ttd8dMnP+DKF5/67DXt/N87hvFNo8YH+ohvL2mtjta505459rnn7lvo3Jz0a1xo/AsNi6eOnzjizAeef8U/7dqlbfnpR1/xYAn4O+br3w+DixvdVUFDgMPt6AAf48FN/Z7HNt+l+Gws1LoGDzXHG3o/r9c9vVgPTJAJuSlp/PILIgajY65x8Vi9/LnvJTzY+ta/22ez7Jq9cqjBtj0OiYeZtAkrfgvk5WN8/Caw17y2FZ1voixDCV2txe+Y+Qw5deYjB8P1gZY9AH6uk+Ob2bIeXPkK4eY+ChzW/tH+yzUMuzToQf4QfhudfI8T3H/K8Z8z5KDre65bffbsjyJ9POI8ghDug7sgCk3ED5pb8eMzH7+RVIX2kjv1ef2LHvGf/8TMoXdRf/ZH6W1D13ce3Z/Ag4wMc5RpnPo8xDH5BWoMIU2EFNyOXzOGr8h6zL/X+tVaK7j7yF1ng5X6ma82Fn3mYzan07prv0t6i+0Xv+NudUwr58f7KYeSg0oUnYjzRy+EIwfpnHewBQB6EKE6hGPm4XkVmhL+IDd81NEw4DyI7u40z7NQDmq+8hWy4czfiWkCUkfIDM9+Pc00LwEGEJN75m/IfOSgKRUSV0cnBtEzXZsPGMcvRd6Wznp9Y20F1pe8PQfoY8PIFJTAeqH1hMYN/gjTreYphN3B7TOt+xzYUu2zEG52fNMBmkAXaEPOD3TnMj35SJjXxIrffKMnn/U8T/MH4/Vjrz1ZF178Xg1r/8tlPlUa7qze2kZW9uf9NDo2p9M+yZqw4jff6PgGBnTConNSR15fHPsNv581NdZhvkJYcnDe0clB9EI4L5+6iKsAdAqhLvx7dJvIN/Pmb/kWi4X+n5dpv0CzwHLSLJsUfvxC7dxzp6SHb9d7Ale84Pgvu/KFp1wwaOebh7XhEv90fJfP+O4elKdrdMrdfXRfZiTLbBPNL4Or/Vxeu3NcPOzE73n7i0/ZdemnM+iWc1vL+9EHkYPi/cA55WQdkE9COxELppOATjMexLfBbyc+w+Sz0Qeu+CCdA/S5L7kJfCv+CNz8tt5Ej9/ybFobTfiyPesGnXPhG3WhXvG815MejH31w9lNqTCH4d1mk4VtAqdSBKkwqYWg3Bw4FToPAcxDlpsFy6RFi2DuU2v4CFOPg6DHd7+5TAS6oxsEMl8hHMNAOnxHC1lnvNrU7/vF9kZpeK77adq958466e6f7w/gJ+mke/yoTrrnr+tQaDuOaL+QtGLZ797Ow9uXcgBS0ARp/PJ7wLlhOmYEEtA88yK0tIYb34vrkOzelrwBpbWNrezThkYLn9Mg50AOossNRGc8iM65WvLB5j7h1t+qbZLN4HZiPbllfReQeoymKBOYegQJH0JS+4Jyc+BUMqIXkpKD6IXRd/+BbsltMXx+joFzIchz9wEFOZgIBBLnThpM4JcFqCU6lIcDKl8hZLMTSs2InufjnHUN/vqmtPghyEAT3W/OaaYBCQozn/WOjttwMX7uUxMA25IGTaSehnDMF/SNeQwCmS9IYJJxhpULnfEguqdd3ef4xSv+fZmwnrLgctZKUw+0ib35IqM7iP+G9jnzeZj2OP+Mu9iYN1XnQZrzM9eOx5EvBAC0UH5TvByACdHJQHyOma+lzvw+4w4xR+KDpF/4M4/Uha/9A+/nUr/c/FPF2T56ia1+zsFGdbTBdt832aeFDGtoEDlIUMh85CA6k4HorMN5ogcREUB3fAbF5yDY9O43X1fphfDly7A+oRUT5YPGt2O4tYV9cnlK/vHe3e3g+pwsMyVefwr+0+4da583hYfv60/g9y663xGXv+gRX3zli0554bDjiD8aB13os/1Gv5zuMHn7OZqamClY8rA9S5AbNNaF5/v7hYZvPelplz/x+Kdd+ad+WSwNcd0ibj4C9umt+0Dy62AQzoFpyciJFa5Lo+cFbiEotww0+jI9DXcQGkxgsdChad8lP4ipDhM1X6Fo5gOgx3e/yfKZdtaumY9CWqoJP6KLnvvMZjwoYevf7XMYORzva9qssmkOQTROyQI+hKTk1gy+C3r6YgYRw5TKraXTw3LOvKwHMnDUVdLwXqd8wJtwUEPzcjyAo/X1COcef8P4PWbUr0r6ea/7UxrGp7ug79SgY7RY+2J/8D7KPwW/v06+57fq5Ls/y/0SPew+f6NDsa1dy6F7exy0N8BzGkwVmhI5OO/owmcSneFOnQnaE2pqTdirf/jI5DsE7xv2C1H79X5m6Ybz8AtW00FJYBvm11vSyc8Ekup8RVtw29f9qKzLrDynWg+EC1IgCei68Ak0h95oZ4Je1t8EQG6gh2U9kHPYveOvrdxyr3E4euU8OE/OZTpIn4sPinNz1GnOv3xBxFlHJwXRC+HIwXlHZz0QnfW8bCxBCDLQBL74zcVv3nSeeyG6ZYGWRQPRGQ+yIcaD+PDQ0QvRyUG4QuLq0Z0EM7FUPlJ4uQVNgOipx7ml6Ro+Z8LtuLNOFswxZYWW3vDzx+gRqbuQ+RyzD8PKhQ8CRGf4Doj90PfrEmzMC873yX5NrVzoEEGfG1i+YJsnr0Mb0YXPMTqyU2eCXn2AOjja8y7+al148W/5+5VfckFfIwplHzdYvzeGL35J+OVm2ncpuO4AKi2scwMzvgsSEzA/8wQTSPByw4/dIcOB/n7o/rDLGzwZmPGeAIQrJK6Oj/VAdNtJtRj7P39b1q3i7uGIP/HO/MF8OQPLTRkLTpHvn71jWHy18fA1O4F3/fSpn3/lS079tn+59otetUPX/bFP7HzL/qGB705899WD/lKZgiVv00T1YEX7u2Ecf+Laa3fe/+TvvewSW1ZE57ecaxj9cmW7Bk7BkIMDE6A1IS9kC0kLm877KX7zULyx8BGD0IwH8cWPGMKBEd3R9L5vefkK0avHP/PVOuimV78+m+h+G5hv8t9F3/2MJ5o5aK+tfzBnkzls723abDsTH4Yp8ZDkltSGOTqNDPYAgwnAw0IHTZgW6+WQIU0M+n864W7H6MSvPEYnuO8V7z55TpwjcesnrcOT7/4QnXT379TJ93i6TrzHT+mke/y88b065av+jpU/c3r7dYlzzXP0OQe9Q845qc/Z6cq1qd+O+I2qoCFQ8xZm2uFjuA/JnvpdeZANEjRs0I+hByV4HOcAgEVzrkwz99dziHfrb9UM3/x2tFg3dbRCAi6kUAQMbrjej2Q7wDSKHwK/sUGnCUyLeR73wP6v2eqW2IbFUTmHOg9Q6w6MlLPp6CCvi4Zo8865khcSx+8ANKxc+OA7WqUOT7/352lDdCP1AnPMwBgEPU9DsB5CUG6ZwOgrdcyw+zxf91uvK36PL4Qvn2nBw80RnRxE2/aeQryK0dtwIHV0QB0r9dgnmjG8Y9AgcIMfoXV0DrwQ+jPs75izJcnnJjf2eVPOg2HxeyxY8xQy3/z84re3MD4/l6D5A3n92OvuqOf9zKtcwh+6nK83StTPvgrlRrkGtiWxEbk1BOI3BRomn4OMi8FJwwaZZr1e4+cTpA6Ge+BmftNWlfnkFr+NhaZWLniIjk5Yl3lApytXfG0+6orPjmHYZz8x/6fbfupvFho/6llzLcvwumGWt3Fce+gyu2VHu3btWvMH8jN2r+kKLfQafzg7wye2/Dv4TqYT6gFPcEYtecieJchNfvxXa9DFa0fuPOY/PnH0s077wUv6c2LMLbIv2HV7IzTwGc0+v3F2TQA49TlGNgH2gc67z/OTGuQH4OeqjqJloANjdIegofvIl28klKnDZx0W8HjYBtASvNwaAsxTaCW+YXihzn3GPUgPxr717/bZLIfAoQzedbBt0Wl7yksCvftNM777CDCA1spHWL5g01kP7XDf+glcU/8jRDt3zrfOnfNtdBbIsUOQgSZW/OYbnefO+OnVn1Q1r9yWvn9wdmhe3n7bmOtvG1o5DxsaHd/8PDxCm50H4z1sc7/c8tXUuC+v4bb5opl6WsEBFxKcrwXhnDrnflOkU90k+Dx+/vzZf9FYIo+H7vNnD/uij/JPzD3RdB5anmMIaSKUlnN3BOY8faCgKc19ag1fC4XPdjFtcXk+TvBFtyForqPj+BhonQkCzoPo7k7z/B1OCNEMDfo0BMwPMl/G2I+PGHSa1yWIb4PfRnyGyWdj7QMufoJZRycF0QvhPDx1v+DNy28K4fdZZwEKBpnUSEpYdRTCUR9Ix2d79pncBPrcD18dnnMthL9F/MS8nQv75XzAdPPz80jqAw3GIOQ8f9GaMI5OZr5Gx5f5IWw5UBd/j3wcP+h6npAS8rwdgdRXaErkYHqrG52cFD0I4e5te14HXE2I30LSwqYznvepzEOB+IhBaMaD+Db4bUQ35HWOTg7CZ/7K8QAAEABJREFUFRJXRyfO/J64EM5prz8+Ex1tyHyLffbB/NGPftMe132FZzZwp1MQuK4v9KDfu/SUffL329fNfEikfgzDr7785C9654tOe+x//Zzf/6A/kP+Mn87X+HndZuXEetIDXjnTHkPlNuW+9yxBbp8w/b5xHE/63X///55w/JMv+YtH77oF/Ivr3vQNXmv+2OcHoXqf+gHkcEGCvD84wxBS+eRWNOMjm9jgbz5D3hDo8WeARK7WPFzQ6FAgeiFcOkYH8AwoNCX8QW7Vy+8cnZR1nAocx6O0GH5JZ+7aZ18HmHpfdT+hLU7FZvMQPT6HBLpzcQjT7p3lFDSlFloawmmQeSaDpmahfKzjVOhZx0JQh9vNOYH2A3NxrnLLOfugO5pz6ru/ZnL3uQPxW7g+n2XNfTxfhsMXSofmf0vR6yeQG+iNcR59n+SWDOIcCtHJQYbJDURnPIie13cJ9uTa+ls1wze9jbcV66KBrFsIl3oI6K0e9Ko/fmvxWS80JTkvXyF6oyVd5X7Lvgb5/H0EnAvnxXl2NM+5AXTOHV8wgd+XIGIhcevlK4ResfEgTKKzDinrF1qCFvWIFkFa8Utq9DLohFygosutaManjiKssRApYUcH1MMEQcQMJBC0aJnPfCFc95O0jp512ryGfNOA7OFiwh2Lo7UtLQt45hlmfVPUFdpE6i60xuUUUHw2FkLGTzDr6KQgOuNvET8x96bZr6E/V2JxAA76eTjHZ+DlYCUvU8Xn8y2M7jyoqTkVeqH2RdvCHM+7+EUah0td+Of20eyPJOgCwb5P52jpbMgBusyD3QfvblrsUzT8JoD44VpuEL5CdHIwfqlS1TqywHqiZSCBFF1axaZ3v5YNf9aZzVe+DDOPe70POr599xNzllnTzp/NtKKlAIJlNxV9GG//qY/d7lFL4ZYT8YH8XS857Ud27157y0Ljq6ThbkobeZKJcvNZBWfsKtWz2HrmYBhG/pz2u8ZRT9xzmx3fcuJTr/g1/2R+EePh23QCCx9Hf1+Y8rn57i8n3P0GSW7M+TdsoEKC8vmwGdkmmMLysc6Kz0L32+rUdyk+GwvlNvepGdEtCUSHBuE8HJh6BMXHPlqqwgS6p269+/k6CNvWv9vPYXiXHIYhewtCkIEmVnzk1gy+azqb5uPwIEjlBuJjPIjOwwDx2XL42gcnwPkyDcj5djRJbli50Dl/EL2eE6Z6TsT1nPDFbzJ+D4jP+aF4uXy2k9KzjyIaNqjt54tV+Wr/TICvEJ0c7AMbAWQxfzEN7uNbno/nBKu+1GEuWAWAJvD1usnLV4jPce1jXZpvmj1MGrZpQzp02rh2q+V5+KA4/zq37MIHZToh6HTyJ5Di12rL8zEFohea2rvf82X+QszuTvP6dTihCeYTaDLzg+5c5DeqfsxeMH5izwdkXvPErBPdedBkocN+la+wCw4yrcc7XKmfeaDBaO22NmzTT7JSiBcxsq6jfkypuxGpB4N9UDmPBC7fPHr8TSdvcgd0EhDdw7QD4jOtszHvab5P9mtq5UKHANE7miTPg3Ac9Lmi17mjs4xpHNB+EAmX2NL9Af/9Nf9JF178RxqG87Ic9SWY3Xr9LhzdIOrHEoQgAU3g7z5yawbfNY1rPtEskBKCTnMOHR0w3zTQrpYbVr5urei2cdV8jF/xI846OmnQRpD64QqJq6OzHoje19ERZdkX+IDHv/73Pc9f5TwcrFxe05ep6e5f+r7XP8l18aZuIdeVL37YCf5A/n5/IH+u32r397aPdPe17z6U++w/MozDKZ/8otsef9LTLn/jKWff4v61dZ/njbj41Mf7Ie8L++uVGGyvUT8kRSe3EAAdGHzWkYP4RLNgmaiGT+97GAu8/yKQt546HIPohaZEDs47OjmI7mnlZaEmhCADLeBDSEpuzeC7oCWdo3Oe8TDjQXXxiLZWEIfCIbPJbNrTgG23Kuy+ZmwQufsJmgDIraODyA0tTS+GBIdvWz6B9iPzPB9P0tExL2Yfd95UTleu+PxAOlp1ev3P0wbmy+vFgVOPOnSvqj/o/bCvOoB1qeDLV+eKH18hOjmIH74QXrStv1UZvXkfP6X+HFnIBTRQYQJGQxhvlN/zZJyRYYZKRUAubdOHIB06bVhcI86T8yhUHRjbsEBK2NFBXkcN0eadecg7Oom/oWHlio91PF9ed1ad+t6+zBbfMGBDIUanwDQ8gtOGDZbbMkE9EEFbp4EOPBH1GFSozfy21hWfBxQWD5oW40XzPAEj6xo2fH0btuuDOYVkQa0vR1V3oWj4GzKMEB2UidTfMNzshm+9fov4ifn1nAevL84l6LPK+TTkPB2qsPs8X/dLkfNYGh9C+6/x98n37Hi3l71Pf92yj/UVUD98IXrqdhCc128i6QxtY9vAhDE4bdjAdSy/PrAexDgKaILHMK8HGJR6TMl5Jm4IFB00wXzxMxBy1uFJ53iDfs8TP+jBBt/39TVqGN+12aRjyOmeUPqK37j4276mxZ+xcPmLTvisK1902qOufNGpvyYtLvdG7+xH7wfsKNfN/lDOV7Y/9Q/JX+/f6PjW667d+TXHP/Wyy/JXCzL/4dumJ8CPRNa/H3gqeYlW0BDwQ8vbNmgCH+PBCATmo7cVSQnjs14IN+/w5HO8offzet3T9zISZEJuShq/tPy66ZiLcRgWw8V60rNn/6Uk4oHtW/9un82yKTbHoQbbZvJQIMhBEyt+8+SmfVoS43moMoG9oyKT5lAtm/HVA8eHry2dQP875m00z4OQ8+d5BCHcc9wQjvtzIrcQKJzp3QfXdML4CQ7h7u1ML0z20DaU87OQtLDpTid/ApNGfI7qmJavbwucfxcw2S++mhLvyz58qte11/prPddF2H3krqvBSv3TG9bu0gtNcTm1/yjCW3Qf9an+mDlXzgUiz5+TCUHgx8RBO4zPfKEpdb+TZvP5OvFFjh503i8IJ8yjNh9oSviD3PBZpy4Zkxaiuzt1gUIOEvR5JdIaHj3zZyJNzROQkoBOUz/IwA1+GzN/Q3Ry0JQKiaujE4PoWScLSAETa2uJtK8b67GPaaFpdi+XoNdjIj7YWRmmYXIeUyDhq3Fa14ovRL5F/MTcB8W5sN/Z8eX1JhOcR9AGfLaTOtOEEHID1/uXdJ8vz9P8/rj+20W39/P3hxt9ZZaj/gTzG3U7Z5/ohaZEDqbjW78/cosG36UgPrmBJpgPoaXZvumcR+YvwUPKR1g04/fqtxGfwfuUmC/+DJhyrTQv2waUrxBbhjW9CkWXBWjmx7ePu2f/NU+9/Ct6XivLpoblYqaPXuwej9u1S1v//ns53UEX7dq1a+2KFz/smJ3Dzrf6Nyte42M/NkV648HcxtVT6VoPlnqo3NpIXTVq+MUdGh+mtcVpX/Lx2z6Bn5Cf9oOXfDKGw7frP4E1v+zywvQrdrTVL9octlO/sSTyJSHJAj65gU6n96lzfBv85vEZJp8HsF58JtfWjvGwDznKchO2AeUrjMjtenTmRfYyOKcO4QhYr8eHYN2FeHufrbU9P092sPS1LReSzXpzbDKH6Jmc+t4OO4LThgEbgqZ9GpyJguZBCPTCGb3Bp8PtZp3ArdaN7s/TD4Dnicz5dzRPPH9O+KBBtPghSEATzLvp88RziHZvSzkHuW22TxsaLXxOg5wDOYguNxCdcwLR6zzx2TK9obb+Vs0Um91G/8Sc9dCyvgtIPUa4BoSKzwQ+6ipfUG4OLKt8haZTP4jecbi9Lvq9W/ZPzdfWrqpjyrnlufuAgnLjwAxcnDtpMIGPFbTY/Y493HepfIVya3ZHvpoRPc/TOfMY8oupHdAShNxAT9D9ppxGBgkKGYi9o73k6IwH8aMXotuWNGgi9TSEiz+BxDxyA5kvSGCOcYaVC53xILqn3bDP61ZG7LuE9ZQFl3NWmnqgTezNFxndQfw3tM+Zz8PEz5VMfWZdbMw7qvMgzfmZa8fjyBcCAFoovyleDsCE6GQgPsfM11Jnfr9xh5gj8Tb3Pbd6u18+983rdV7/yrKtLnQ2BPb6vZ/u3cQXvw1lCzafF87Gk1oADUWLgBxknkJ8cgPRqx506rKUeYO+4TMoPgdBBjrufsd1lV4IX74M6xNaMVE+6PJZuf7rpqnX7dYfeeq/zigv6cvhdHfQr2HUDpdx3AM+9+F36ORnQDD6A/mvv/aEO33DHf/3hWsaLxs1HONtfZb7yqMmGUO2W096kFdS1FC5XeX8L32+P3ntnuu+6sTvfdu3Hfe9l7/txPOv+Mi9D/+jbj6am3Dl75jj97kORh+qgo4JTAvMU2hCg04TMK6j3DLQ6Kv7HYQGE0i7F9doXPtOu8Rw0fr7c+bL/Ih08wHQ83S/yfKZdtaumY99tDTrxQeBFYQYvlFnP/O5MAdD3/p3+6N/uc/heBt9b445BCBoIYcGmpz7neaQylfIGaGBHjb9YtQIxncf/8YD/OG+pRNY8NtmHpnn05DzJQdNqZC4Orr8YPJ8CpvolC+6UxaD00KzzNdSLUjMHYpX9unCg7Whhg3qZeoDsJFrg2DSE3TagdPJn0DiiEyLNmzD631t+HTeX8zPc631QLhgFQC6LnwCbUBvtDNBL+tvAiA30MOyHsgB3erqL7dyy73GwT8x98HUeXCenMt0kD4XC5YdzI7VRPmCUZe3PB+nIHqhKZGD844u1mnzGkhjMT0tTNaE7jfHfOTlK5xPwDBb+7ZW/AzAAGJyRzdMrxPz5KwDV0hcPbqToP0OVT5SeLigCRA9yzpHq76GUMm+RNbJghuOc9ona1nP8kbhhzNSLyHodPITmIzfOL/wkYPomY4A8jOoj+3r4egNsj2DwPVbRIcL+tzA8gUZGAM3Px9yfE7RW+pMMm2DpoYwRdt6f97PPN/rfn1/7hvqX7c6ugcs/dZTd2HVDVrofnTnBoYDEzZfEZUW5o2NaILzSt7mYQLmJw0msBkEjB7maHasJkz3+iPObsxDCrJeIRw5OO/ovQ5P7OlJ/ZNcJ3PjzY+HI271Ec/yRxwByyw3ZbauCL4Nw3+59VFH3Df0Z8Dtjc8//ah33vH/POmq/zjyXaMWP+y9L3/T3dtdbnHkeOZpi5emHjnwQ/oXG35hHIcn7thx7QNOeOplP/Cw73vH35g7fG31BIbBx8pgg88474fC/nQaMRbaz7CW+vlK5PF7HtGM6ISg0+l93Ij4He90f/mzft/jfzS60yA664FwhcTV0eWJQfRaB9106iKuuvDFb3KvfgZ6omG8QE+64AF2HvBr6x/MB+3IIbQ9sfecRQK5IRSwaccYoJ3GBvagCQDWjg7w5VAJEI3D4Bvx4b6lE1g7YpFxnGuei885aJYXc9JNjnhTv8fEb9z0eXoe9D6/fWssYjwUL28nZQfZGEHDBv0YelACI/E3LJpzDd0Jv71MkMa6Da/3xfgfYt08FxZq6/FokjoXAQU0XO9HwtYRH0TDBpkGmgBknrU9X8WwW2wfd39KnIoaNW0AABAASURBVEOdB6h1B0bKAXV0kOfTEG3emY+8kDh+B6Bh5cIH39EqdXh6XhbarJ74Nb0+0fHPMQPbBA2QQxMwvqPnmQQCKXVIHeUJNvi1bPG7gEKU+B2YFrzDFUSHB9Gqr427K9y3mEI8pdHbcSB1/P/ZexMAza6yzP+91R0CCIgLM8wMo6MDI46OoOOCyJKwZYEkgBIHIfvSSYCwKALZaBMQF4T5h+xAAggixgWSkLCFgCjjuIzLuIGAzujoyAhuLFm66/5/v/eec7/7VVV3kkpVd1X1ObnnPu/ynHPe89x7v67bX3cHwzrm6oEXNjDj2CIQ4jK+idLNq2dFw70DNLZQr7/+u0+3V3HpFo1P9UDu4YdAiI4LA9gVV+STl8bliORpGIj1ba+55hCWe7Hlh3W5bLBu1s3SFTHHI3kQK5rAFXIex6fDPKLgPBWNLeMbSELk8Ik7F3CesEkomHVgV5xOIN9pSWc4kUDGCxqbdufRHxEn+QWBuSN51DMiWdzo1/7X00NPesutXd+9x+n5cGShJUcmhhdTzPvuWux+ZAljXd31mvyWNxz18AcedPuv9rH4X/kNj/nfbGCjs3X7WOKW1CxarX4x/hkNXx+Ldzzq3l/5xZOPfPEN73jKCz74194VZVCD1SqQvx6gpFejQD5/Kf6SgG7lVZTn8ySOAyUSEKxrRIx8Pguaq7/KXvLKndF1/8NpQTORuJQfk+a6DqhoimUFwxGsE7aCQs5HTAQGHkaOSwKOyOCFhXfF6Ttn/7cLMvvjWP2LeW6Szbg5RUosWyAc4+6NETDPvjOMO/7iOAtANAFUnqbrGE7UIOh6QDvugQLj3zH3wqCr+lbd1beEcwXSs+tZEnN8WCWcPMfX62q8zguthjU3bb9TPSC4b2CZHrlpEua1Rdy552Gqn5zsi3le01PX/d9cN69PKSSBghKnqxnAH6+7PjxCDh/2qVPjBRPgJZqn4+a6XfdIvAP32Nb90/g8qGunFAi10vU3n+mS10+ewRyoMevmqyePYcN1qkED2PIyj59IbETsWYE4rAMt5vJDOK8/5oDw6rjkk6ho3PFiyCMnmtcUDWddJbCMT1weUO4jliWQPIIVMcfD+XRE8xWNuZ71+FcL9Ne8uwD1uc+cG19Xu9ZR0Zj1iXZ50HOf6RMwP+Ubr92461Q0zm+hC1uro4Mbcp9L9dA3l10eAspLXQiaL2G8MMwNFEMriTk+mRJOnuO9Xwiv27Hz7Q+IfuFn87pT/oium+uzckXMzFc0bv2isYrajhfNJ3IyP+5Pn155iRSQfJHcHJ+B+iMv/nd03W+i03UwryJ8USzEj+KfGV0cF/3iM2KheyLxQ6Pvhh79GZHzM0LsXKfOS0wfmDvkGYAW5isaYzjraQGZmCBh+cB6HLu23et6FvvbZXNbRvT8N5c54qNvfPYT5iKbyLnl9U9/4AcvPvq0O2LhZsr+TvbtX5DELEfuudjsfIlbErMol/R25vjjWOxeu21X94gjXvTelxz+kvd/6tCTPnJrITdYCwX8w7KIHcGDovwAlydd9AfnAhGVF7SRjyHNgfk84cuDkoeuhuuYr2jMb8xFe7/tuTnMfPoMXMo37jpi8li4ojH5iZ5qZx5NwXyiATrDh31i57wE6nx9/69jYddbzezPvvoXczfrptjT7EO1bEURBrUJJCEGl0RxM4Cb6DwDIYZGovJcBzfMz8SL1u6hAuNHqEIzV+qM0CMSw+Wc93Cof9jkk9gbj3QkH0Oe17cMq+HYrK3uIzfiJkpgbp/umxwQ8iqqg77osKCJ5h0vmq/3uTwow7H6R3UYv8K57/88XNeU2FNARWO4wtBLoeatK10IiTIwcCPAsIkEhJFPXJ9wyOvju+KAbgf9hTIoT2io/4hBS6FADnXXTUyD51IkN0ygMeuVV9FMpWu7jmje8XldIFQ0hxuVl0hgjh9RwzMjJ4ihyR+sked4wjEGgkbAYVgZTiSwVz0k0XM+x1c+sRwHTg95uc/CA+Z/3SLQ3/GP0yFrZ1NfTjZBlrOcsK4ME8i6K+aACNywJQ9iRWPJ15h087qieccfcH/HHJ3UILsCYIx64FddKi2R+HBBIBNItyIhD9zwglSMdWz32vVqHvCHhHW7jJh1U5horKK2+REpUN889PE+Nx8GMMy7X3HkGaczPCovkUAOE7svM99NfDN6XvTdSdEtHBb97m+L7ff62rjg1C5eedrXx/mnfk9ccPoxccFpO/AviPNP+2n8K/DfHjt3vBv/w7HztI+Mve/9Y61sNyI6FmKZATWCeEHM8ZAXxEXrNzGiDvMkiBNeuvjm1qHzrfk/8GX8T81Nncv1VjsXJtD1i4sv+7U3H33/+cTG9pC5u/nyp3/3ru392/suLqbaf71sd7lnMnksyY650eCWij/iUr6w7xaPPuLF1//oYS+9/s9z6B5PLbFqBfLvmOeDEFEgbHk5CFTkBo1KIDx1M1553BCRLQNp1WHBRc1xwQTJAwfGcL7s/D8h9dJIHuMrmk2+hr2MM58uvnnAm8fQsE5anDIROa+J4kZFDZaDEHzAROS8BMxHHBlnnv+i2I9t9T/tK0qwC/Yi5B5wc5PplMQcD0IJJwV34GuUhGBSzDCGmOJq4I8LSmz9HimQ14cZRG/OEYnpA3OHefUXzU8vh5fH65QDSkJe8gkmn3jy8DfjQfluJ0vPfdRAwQIpQ82PiDHVw4lW4huvE5jPxRbzvLanhU/NPtRYaOn1odzh+XRV8tY1rT/55JJXUR720voNy3O8aL6Lw+OaWx4o+4DsJx36D+j/eWVNndVTXYYAkiCUumGN4dSPuIHkm5x087qi+YrG9MWIz7PeRyO6j7I+GCC9o/d0MSr2Qz5E+Inkan4RW36OI9/D60TjBfuCuR5xx2Y3Dj/jK2C3wNrGp2OqXeLjeviu67zWM0Xt7Kxn/SMPP+tmTvHMZ3421qV5vZwYXPF6miPRkfe6hmgMJKzFdYowv/L1jLGZ1xGTj7ONvuWOIsx0n+536T7NGxPNj0hQP/XGTlRv5xWJmS8uXgRhnpsYmonBWvPza3/OP0r5/JzXejXErId1RWMVtWuXZ6Gieei6mV5WP4E5nj5MgHMM45ig6z4TEa+Phf4pcf4p940LTj2Sl+tXxwWnvAX/A3HBjv8Z55zg3wWGtsqDZXK9rIc5Emsh+PUwrp1IXnSfxipq127eiUXzdZ2aXwfsF+5zNdP+Fb3cLkteTDPhiWK6/rt27b7PY/Q2Q+9/4VnbPvCGY562e1d/I9U/NaK797LdkYixLcmOuYwv4v4Npxcc8aIbvvXwF95whf+YG0MJc94Ix1aswbc+n4d8LtggjxHn0C03bAxIYo5XfCDzFYeBjCFQr5yIO/y6RSoI+PyJutN+yc7X4v56mHc9kUD64rSb1xflMW24jrFEAzoiAXkS0tUnB3AOw7mPNJIwuEFb7F8fZ52z3/5Ep5eIKlZxKIoiu8myJ93ITQbNhCBC2CuffI4DpTMs3UQChMPxorFhobTaaZUK3KGgjC2Q+uIGco8PUz8KbmboySc+ImHcYRz2aOREka7XSzcRI/k5QWzKZv0Wnsh+3FdulMASN4wTHhDDvPwpEpZmOHlpSCAhhG31j6qjV+6Ln4m8DK7jQiJMIIobaQTNgCBCGMcRw02aOBqVt0K+8sTt234AxoF7+EOvOqinOOqnJASUUXNEDJ/L5JM3N+3G9UfESX5BgOf7M3H6EYfE6YcdEjtE+h7xyIGzY4rapZ+5FJ8Kn9iZE5SjPyK59MGzjjoktPeG5pb1ow8JY89bCYllvKD22I85JNIu+AKRLqrNunSvE9fN51pwjREx8voUNCevIuE06/X0/ljGT8Zwkrc0f0B8Y45Q7ntQYXZWD/UckZQ86DwHOBqAuiboc73m+CRKOJJHPjHWp912++zbGuvI5ShAdEXrn6J27cmHWNE4rqAMkXUbYD4DCfgVg4bL2R9Sb4jF7glx3in/Ps4/9SVx7mkfzPian7YvZFnOW+uuaGzajetPUT30RXPTbjz3mRuMXKfuL9an3f/2e32x7/tf4f5ipXwBXb5QT2qIfhWM02/ZuXP6B3yHzAY7v/eSI7/+A5+97bXdQvdLlPY1dJTthVmfc9nZLOP9lB7RRa7GH3LRL9jex6OOfNENl2TiAD7t0633PBQceUG4EFzEmD0XSwK6EkbE8Bo7XhwHEh95ERkOWvIgViQU9e+Ya9e+sO2ENCuvYgYnpxqvaIrpcz0xDYPWAwo+axUJWaYwYCZwCxbIaRa7a+P0nbN/xBDWvjoWVr2Qm7V6xUiRmMlNAYbzomvn7knM8UnoE06e40ceOSdwXk1R3sg3aEBsfdUKLCroZHR1U2/0TSx53LxO6ZZE8kmkW1FCBjAKCitfz8xA3HwH271beoz8NNgvWHcv4vKLeChTzjt9HsImYR2+Me+2fyrXc+E9Xk/Xt1soOPL0qavAXP3ON32e3U/lMUWmzTM8FvsTDR2wve8/lXqoq3qkLmkgSUVM88CgM3F9dTVWUbv3RDcPOF2YN84wQ/gPS2ynfaOA+g8XYrae10PP62S+orGoFwpnysPN53UZPxPDqc5T0egB8Y05QqmL+53Il3qpp3okQpAHXRcvBjQQNJEJ5viz8DhfXk/ia33s/IX7Uc/zxmlLOcNzX6LWX8wZSMSzbvMVCYW+mF3e0v3pkwQ4R3TxSU5P5tvwo+KVp9wS6952P3CUs9Zd0bXn6jdANw8MulC4fuXhDtcpCZwImGdjuU7lkVmP4zt3XHVHLPS/EtF9XrVj2gz0noagFuUc2T/4T44eIhvzfNNlR3/TQXHQu7q+ez6aH2SVvGALs+5mRm9Jdsz1X+Kl46LFxe1H/be//8+vefKLbvjf45BmrLcCw/xdx8XxgvhcEOIGHJ4LbA190efFkChdW2QY90AYzudsGT8yzBlkgHnvedHgSr8F9YbzP82SzxvmZYEp3zEkB1hhPuclPNSTLE4GCizNMz2FkfSQR8D1nCBdfXPdQ6O74yqtfd15Rla5ZG6WXbCHQUzmweVc9pwJ3IIJEBIJVxESiYuKb77iJBzmq3jmo7V7pMDBS0aP15MLkDqTV39gkJu4ttfBgK68iuaSb0BHJOC8I594CWNt3oNtReoQtLKhuX1CKOGQh5uoDvqi+aCJ5h0vmldXUuVBwjKx+keVCVY+nvOoPwvXC1quzzpZD0jIMoWhZyJCnol04SUGDQM3AjRfUVdfNF/RvPvsuu+Ln73loXGgtm7hM8oTVY8Rg6ZggIe66yamwe0hklRHIA/11ai8isYKHfMr46r3fQvYjn2hQF4fL8zsAniZ87rn9bEI8nviZdo8RvLLPMknNj3M64vmGRa7DWy17sbY03Sf7pdQ6ipmn/BMVL65IqPhyAsSNPkkEioS9sCNOV6sfTv4i0cw6QPpw2EdWrXuisbmeiGad0OiehgWR64BHPOVl0jM/XXx43zAQ6iuAAAQAElEQVRD/k1x/skfIrKvjq+dyUp9WYeoQQlz9eN7WL/xisb0Ez0xXsiJmafyDFde5tfn1H/jZz/Wd/Hry2enlhKsFuXcq1vor/jQG5/9L0tqw8Af/sKz7vXBi49+5rbF7ne5Tb6HwvK1ijc7zMlRN5OhJdkh94WIxff0By18x2EvvGHnES9+91/s3LlzHb5tyALaaW8KLCJ7J6GP4fGoGDQSuDEkIioS5vqPbhrcuDMMWg6MCMyRj5FhMY1Y8RtzhsRlOy+Lrvtwvk/q5/wadsYniMyz0vNMWMrQJzwLL25UTENmCThf5VV0vq57Tpxx7nEy92Vf/U/7iuZmsnhKruimcCORTctThMzrkwQ4h+GRN/JjaMnHdHzyCbhe5ZFqxxoooL5OI6pvRWP64rSb98JxOcJ8YiHU65RuSeyNn7xNeMp9Undi2We9L5e4w/0Nt+ZHNMYEIx8Dd+CnEaO+YePDVFjz3v+3nNLrNF7PyfpZrwzqC+LyEomNfGwP0rP6K98Evbj5oVt5jnf+fuEVMA7MY3H37ypn6lb1GALogVDqhqVMGVb/yks0OenmdUXzFY3pi/Y+Dhda3xcKcB3rBVzxeloDibw+YF5oY4zz+mmKuMPzo0Ew+eD0kKcvmne6A+IbczbtfoG5Qx0MJKKbWHmJCpQETzyG+vJwzRcXL4bLYiBoFTHX8ui7+b9vbL3OL2Y9rCsaq6hduzwLFc1D180022KDaQ5IQJ6E5PXH81J+biHsO1hY+Np8PFzReihr7j43Xrt5bTH3R+GisYratcvL/RUeoFvT64WHHvqRXQt9dz5rfTHXcN2sJb1xu4PHue8ftNDf8bJbrjnx3ngb4rhl5yHb//qzt/9Iv9BdScH3qUUtee0ebqWahOhWRzdiN5fzo7v7OPF+t2477oizrvvEJNfM/aHA+NbHlfFiAVy24F6N2cUsibxnIaRbEZoD8nnLhAF6yWOZzvkcTzjnTT7J/K0dcKWjv+PEiO6fw1b52rU7nxOL5pcsn+sktyTkJZ/gHvkWWPnwPIobw/gr46yX79Mvj8ZLZC13q7tJ1Z/sSTcUIWwmwAR2OUXcpImjIYGAwLAMJxIgHK6XIhk0ILa+egXKV+bq6oUbkRnVGdmHmxJ/eoy8JEQUGHE0SkLI+YPGddMHcDbvUetPrBsqWGCUYTRqwm3nQAywhtUVN6b8eh3CtvpH1dF77F33sfwwc/16neq6YtbjaAsFR55+FpzDyRSs8YIFcpqkExCdR8yBcXK846MH6B+v3vYxZY9RD/QZAiiDQLpYMSKG1yX55M1Nu3H9itrJxxCBPLpu+Dtd6bTT+irgdeK6eV0FFxsRw+vi9RLNyauYcRwRCFFeRWPTbtzxFc0dEN+YFx3d77RXHUYkmfoVDMZhRsWRZ5zrViDTuANPw0SsR/u+uUmtJ5djPdFCrF9SRe3akw+xonFcwdsiHB825qvgPAvdcXHeqT9raJ/3xf5Bs7Koy3qm9U8LMq4/Ik7yCwJzR/IQYESyuJzX/Xjcae/8g24xrolcL0+55sxKl183h0jXd8f3t37pP5fofoVbLn3W/XZ9zQPfRWUXUf/X1mLu1ks5gxn3k7fe2h311Bdd/0uPedl1wwsXwXbsRwXyOx6eMy5sPndep+LG0oDxyquYfBLiivyIDAfN5y6fT/k5IPb4jTn0uOzVfxnd4guijjM27cazDuYTzZVpB9e4wYJCrk9MBAYeRo5LAk7BAlExuuA96D6xe/u1kPbZsfqf9nOTVO/mFCux1E2YT5viCATMu1sRl82G7sDLQMQQiBDlBc11Mk3AdQhFRe3WV6fAbbeVcegaCKzOiYTVt4TxwvBwnYJWEnP8WTh5js+733nJOYHDNEXCmpu2Z/1uxB2IBOb00CcHLNODcKykh+P3yA9afpqCa3ws+mLOwq5f6/L6lW3NVjOAN/L0GUfIYcM+dWq8YAK8RPN03Nnzj+8Eu3r/JVudA6uf9qS/Zft/NtMDodS/Pj9VjdQdRzRfkVAwQSxt5mss+TjqDgzH4n+KKz7wnYPdzuurgMJzXcfrhK/rol6nvD4ERGMVtQnnMHnpEzCvLxqbduMOqGjugPjGvOjifud0IT7VI92iPyDddH5+pZME3IoE5RWXRETObyDWtvn3yyN4MXPBMrXXUbeinwu5PvmKmPn5UdF48h1IUB8YjlK3eQO6Xbwjzj357br7pS/Ed9/1+kuFc/Wzz9wHaDrBAE7yCIxIbE4P/HU8FvuFS6Nf/ExdolRVXW6nWQTra7qF/qL9/b9Pe/+lx/zbO+KOqxe7/uldH+M7wl19Ke+iu4Pb9OZtff/Yw8++/txj2gv57HpvBGsj/X/MV9Lj0gvfGov9DcPnbCX0g9GLXXluQKP1eS6uIQgzMF+GZTB5BCoGxjgvjDm+PPLRPzJ2nPt6sut1zM07PnRz0bviWDxPn3uK3BSDerpHYm4Gr2ACiUTCDsQN0XlGDBqJynMd3DCf65BIjNbuiQLlC/NQ16Clzgg9IjFczuUeR/d0RBJ745GOnBdDnte3DKvh2Kyt7iM34iZKYG6f7pscEPIqqoO+6LCgieYdL5rP+7sm4OSx+kc1h+/p1O0e/h5crk8BFeXjCkMv9Ziv9VtnCUfWHTQDgH7lVXQ+06J5x4uZ7x4Tb/3oCx154PX+V0Mdqh6JQUuhQA51101Mg+dSJKd+wrRXXkVzla7teguLL0mzndZZgSr8BOtzkNfH5Qnkda9ojI7LOSJ5jK8YtOSD08O8vmje8QfEN+Zs2v0C488j2qEAGKMe+PKA8bFB1kieRkkk4CfG0HBDXsVY43bvL33HMGMuMJjWrSVm3RQkGquobX5Exuubh75HPRRA3kL34w7db50XuLAOyp5HA1TlPoC5Q36tv+ZHlOnGwaU8w5VHer2PT8b9P91H97Msu9gvXWxSR82Bj7/9jnv/hP9bsqX0feHfdPnT/13X9z/HPfPM1b2U93f0ff9Tu7fvfu6TX3TDr+2Lmtsad1OBxcXI5yyfn+DnCLoHN18Ed2pF8/oiYaG6c7zxPs6BkW3kY2RY1ACTcCenrjuN9f5+xirj8nkmKrquYZEQfM+lZyJCnoniRkUNyxEzXxJC+jHo4viRFy+KM885IvZBW/1P+ykGu7BoIGtNNBARbs5Nz/EgmAbm+QZKQjApZhhDdD5FEp1XTuv3XIG8Pkwjqu+IxPSBucO8+ovmuTy6yanXKZ2SkCchXQj6QFI24yn3UQrPfdRAwQKz25RA5aVe+FWPiuZLuHwasEAJCHgRfJgmrvHpOY/lw6//A34hjqj1WU/QEmsBIgGvX4jmK2J74EZuXKfytenFHdbBl+d6ovMZ6rqfiLd+7D9oHmCdH2AQaKke6kI4tRDVV/0rLzGzs5N5PdF8RWP64qw/O6666btmbrPWRwEvnDODXkdNEXevz8Pc9WeA12/v15PHD17ODyYfZxt9yx3szz1VPXTdr7FpN68vmh+RoH5+/mAnckHMV93N57zm6aQRGMPDhLiGfXc8eNlsWQ9RMethXZFQVNSuXV5QqGgeum6mCc/Xb4DP/nNO/uPM74/ThVc9PqK7T5V/fB5q/bGkGTeUSP2i+zRWUbt28wogmp/qUTnriDt2XHXHvRYPvmyxi8/OLWM9JWBJxeTy9PxMvnjiRz4fPwCFDY6ZdTfed8lR37JtN99UBr9J3sf4qXFXvilHWmmf6KJ78uEvvP68pz7vxv+77gW3BVanAHdYPmf5XDBFR/dIrHejSICbMEZe8QFu1Awnmg8bCYdpiri5jugDzk0y8CXcSb9sp/fP6ctYWQ9R0fnqOoSGMgzoiCwsz0S6+uQAzmF4qMdAEgY3aMVdVv9i97Y4/Zx/BWNdDy/R6hZQlBSb4eMmsIfdYrhZQYSwVz75HAdKZ1i6iQQIh+NFY66bGNFgtQqUr8zV1SlGxPFmRva8KXHnjuRxIUYkixvJxx6NEhC8XiNiyJe6WXutP5H9uL+67yVuGK+8qqt8eRXN64vyjVc0HrbVP6qO3nvvro29Xk9Hl0JGnj4FF4iKo1ECBcawBsNixKBlQLx3dIvviLe9/ytw9u9xzS2P2ocFfCDUNWWogrk6AV3NETHyPipobtqdR39EnOQXBJYcvxxvet9XL4ltLPfSGz4Ql91wzsYq6u5Uw3UMrpfPteDQETHy+hQ0J68i4TTr9XSeZfxkDCd5S/MHxDfmCOW+BxVmZ/VQzxFJyYM+/PqmQUxdE/S5XnN8EiUcySOfGGvbFvrlz6F15HIUILqi9U9Ru/bkQ6xoHFdQhsi6DTDfEPhQ7M/Wx/ANlOVYR627orFpN64/RfXQF81Nu3H3OSJJtw/sq+P7zvzZz/It8g62+OVcc1LnXCmz+H37WHzFr13+zG9I/j44vf8NRz28i4WruD3+o3JFaf3UMTZfcM1+idQbu1j8gaecff2vYrdjIyuwSHFLnwduzuFiVqOgYGJEDO8Bx4vcMMPbLPGR5/x0j+RBrGhsl6e70C/b+YtM//NzzDpPRZNMD28oIw2D1gMKPlcVCVmmMGAmcAsWyGmcV2PENL42YuGdDFjXY/U/7fc9v9x3gxgpEnXicg73MiSC5mZIKI6J4uYvioST5/hUiYD5ymN0Dc/4BuX1nLT3Sd96iyzeMX/t8/qwTfX3eiTie6TSBnREAskXic3xS75eON2Vrueigxi7GQ+2nfdt1u4GCexND9IDPw1GgQ7DqjLN7m8SKQ2obnLEbh3v9wX/frcbKHW5/nT5rMEAhjTrGZEYw+bqN1835rCl88k3n8j45Iv0Pr4z4j4fjWtumf0vgwjv0+Mtt5wbsfhr8ZYPz/9DTOtVxClP+Wv0u3mQAcHUK9caBQryGUk0r/6i0YraDBeSp6FvPtFA7Qaw+3gIl+JXYufO+c8DUvv96LnnL7v+PbEQT0aAV/Ny/sdxxXu4P/Z7ZXevAPVH5BgucGQr8ud1Mt8TEDM5ve4ZYPvk0wTlzfEzMZxqvKLRrptMaGAL9J57w230K+gxt1vy6i4vkUHKUcJ4YRiBY2glMccnU8IDzwUMEF/Lo+/4oc8JnV+ku4yu9YiEwvrFuS6RQPIgViQ0z5dX8gFG/zdS9kt/3ev8l76Hb8UsyyJq3RWNrbRf8+ZE8xWNua183nScmIB595su/nr+euqyS/qv/vW33EDomnzeMTwsRcx7KusbvOHcfdvtCwtX3XLLIdsHf/3O+VLeLXwIVR4di4o0rHXXXsr7f0T+s3f3u84+7IXv/UNGzraF044NqEDXdcN9yBX3auHm44Kb96L+LMAGSMjDGsPer4RX5kMc+RjOl/wcEHF37uhbu+cx22dzHYyse+l86ZMs02NxsC7noV4SuoCh4Q43oCeSsD4T6eqTA3LdiuYJA4+PHeeep7lefWHVE3fdtkEkZnBTFu+mcFOMqIGCCRASJWHghjzHiw4kPIphXl80P8frMhpbqu3DzSz4F00mHVEInwAAEABJREFU6+XNjaQVTeEKXpZQ/7CVC5IAoaIp3FjGS8IQNl/cWMiLGZuyuY+97hNC3ac83BBTyJIQgiaaT93x5Y3SmDAG9ut4vz/7MbwYdsP/PzLrYb2yrDD0TERYp/tIF15i0DBwI0DzFXX1RfMVzec+x0Bk043+P8e2hY/HO27++ozty9PVN7+Kz7VXRXTbYrG/Lt720X3zrUUfb2HdiFGXoCkY4NEhjG5iGnxMiiRTR9ADmhCVV9FgoWtGFOKQf1z8q0ddGRutXXHDldF1Rw+6WFz/zbG48Ftx2fWXxZvfc38jm6Ln9VHvyQWoburvLgjsiZdp8xjJL/Mkn9j0MK8vmmcY+pUBJrZI78rn4fw+h83N7VYBCMvLzyF8dSGkKwxIPB2RCRIqZoLnTTQxRe016l2Ub8xZt045Lodh3e5DrPkRyWubd0OiPMOiuewGMMzLi+4LePvn+ML9d7DwV9EjSllhXW4/USPQvSDmeNR8RRN1n0kfJyRDoPIMy1vPX09Zcemxc+dOfuDafRnxP6Gn8mL2Ps9zJ0JU2j9x959+zc+8/23HrdufILvpvx79Td3CwjUs/m+mRfVTh+S8O2b/PBa7U/7b566/5sizb6r/mrDs1jeyAj0PAHcXD1aEiDtg0Ahw88UYwPfiFxjDGo4bMWg5EOQY+RgZFtOIuKvfmDNNvHnn5/n1y/+Fml6Ez3HWU+YDyEe2Mn3akYmIkR+53SjhSCNoJTDl5fyknM+4mHyNwo+4KM54xbp9cbP6F3MvSi261iq6KfY0qiDPTeWeICQmIQyPvBxX8kGrPMcTTp7rVR6UdqxSgekw9dUX1beiMX1x2s174fL6cGESCwE3r1O6JbE3fvI24Sn3Sd2JZZ/1vlziLtOj8hgeo44BjYFz8xFTf8JYHIv0dT3eEnU90aUSawEiBe7tepJmI4zUqHxcj+Lmh6hpdRjnzwAsUB4W6W+O3Qf9Zrz1lsfornt/64e+jW/Ifzu64NtyV6OQrvvq2LXr+rj0lvsZWdf+lff7pegW/nmmn6vN6RHURjrrCoyIjnwsaXl9iInmKxLaM995+lPjjTe9RdqG6Jdff1WE/wBM3W/EpP4z49aFT8bl1x0bm6KlvlQKsh2MvHyBe+fPQxmQ15GRiQ7E9voCc4d5A6J5h49/W9TEVulujL1M9+l+Cc0d5g0koptYeYllHj5wpOX18MIYNp+YmTA8XLigmQDW8ugX7r1sOus1KGY9rCsaq6hduzwLFc1D180025+vn0Af/zJz++XU/0iVfYYUTFl5Hax/aV3uy5hovqIxfXHazSuAaJ7pdaeUfWUf+tf/6U/Y589YwrAmlnUNzngmis2ZAylO2valLz6HACbnNTx++ZJnfM227d2llPDd1JUzsySm53SH08T1lRzXH0beF7v7Hzj8Rdf/0s6doT9w23njKzC+9XFL9ZQLcNFjeC4MBE0kwc0RJtLVj9DNzxGfp9nACBPygiZCH55jfHnJx74735hDj8svvIn13qyZ8+U6LOB8gG7mXM910imJPdYPaeRrVD5xj+Lmeqad1/XEumC/7efixJ3r8ic7x0tkLXer1yItetyEMxiYYLoQpoibexNHQwIBYTI8XIfwgBomK2q3vjoFpn/HHD3V2ZtO/b2ZKy6dfOQlIaLAiKPRBQ/TkK/zVmS52Myt1p+4RIAlbizVI2w5EAMc+Ri4MeXX6xC21T+qjr7Tfsftv8jl+dywPIVQTn4oZcDRBsCl158Q4zzn5Y7kM76iwyZu5tMnIYatGiDhnC+x/xcRCx+Lt370p+PK376vzDXv/v3qa27+r3wT+/usW/5XRdQxq/9b4j6L71jzdZdOeOyjvxyx+AtR163YxdBGxMj7ouCQnZ3z+uBWxMzrCD1Rf9rl1fkW44S46qb9979Nsq6Lb3xAXH7DuyNfyoNG4danHomEPLp4cPTdu/j2/Ia4+MaHGNq4fe5+GspkW2lM9R/3J98smHlsEQhRXkVj0248yjjR3G5PW60XAd3v3dHDYclHDzEMYFd0PnUbkRxyZlocjTqO/FodXf+FZVNZh+tWdP2sG2ZFzPFIHgMqmsAV3FY4Pmyl/q7bP8/Oj73xVdTzb5aWE7XuipY67cb1R8RRB7cj4s4dyUOAEcnict7nR7dz5+KhZ1zrC8a7oqPYfnkJQ2g4l+z9u7577fuu+v7/WPw1Af+XbPeLns/ZeGLXx+SHi35+/jk3HU7dB768e9uzD3/xDf9jnty8TaHAolXmAxNRYMTRKAmBBzXDFbkDwudJDAkaBYWgjYjBrR7Jl0fu7nxjDj2P+37phSz/v3MejBFNlmkNR7Be2AoKuT4xERh4GDkuCTgFC+Q0NZ/jSCRCnU3wdXHwHVcbWes+eSDv5tRZpMUyLkUX6R6EYyy+BMZN4pt3vJg8DQkieVXR1ay8xJJ3PXOtr16BrfP/MV+9BqsdmbfheIMyCwHvz3rfen+aJpzPgf54n0OvPM3Kc/we+RLz01RjffpJh94aC91rhzIppNY1rmYAxzqzfn14hHSHferUeMEEeInm6bj5kijihhOYD5po3HVEC+riR+LeX/iLeNtHXhzv+NhXwbrnx5s//Ai+Ib8ytt3r/0TXTf43baUA1w8KSLc7Ovwj7vd81b3PsHv76yP328HLhcFyZD3YovdTRUJhnbGkma+h5OM4LTAczo8lL/P4ifGcuOp9vxtXvvd7yO7b45LrviW27/4dFj1mdn+UutRluk/C8JCrfypjPsEL+gb+X78pvAWLVg3qai7V35jXQbTLg556pE/AfB1nbNqNq1NFcwfEN+ZFF/erPmJ24lM90kXQxCSEaW6kGFpJVP2KO7v9COT84DBg7c59/0/DZNQ3GJTFOrrWI1pIrg+hIuZe748pz/Hz/Cfq7tP+Y298Mhs7NwbhI5v70xCtt6IxfXHazesjT5ivaKzqpJ08AiMSlA/sr6OL7c+Pxe7jS9d3C+gyhAdnsCPuv313/+H3X/LMx7ENNlPDq8OPv+5Z9/nil7tXs9aj6u3gTP3UGQKeS8/slxhzRXfbrT/wjBe/+x9KosFmU2CB1z5upAhuJe8zIC+9qJHPR00EjYQulmmH5ecNYe6HiGX8yDBnkIHmXU80eHe/MXfMa1/7xdjWnZDrBgs7n2iuzktYd+isqyGYTzRATx6Bis5T5yMcc3wC+uPGGR8MJKxF+Blxxnk70l7DE1dolbPVYqlxEIt5ZsXiZGKG6UJIJDxurgYqmpvwXAc3kq8BL0WM1u6JAgsK6wTomQCqq+FEgsoNcPNxJs85vA4G9sbLcfIx5CU/eEjpJcwvTFoENstR6rRqthWpQ9BKYG6fEEo45OEmqoO+aD5oonnHi+bVX5QHZThW/6gO4+/Cefttl7Ds3y17nnNoKdQ6rStdCk6UgIEbTBDZ8CuvovkSDnnTfRoP2ogYyRfTeBDTvC527f58vO2jN8Tbbjkx3v4bD2DEXTuu/tUHxdUfPibe+uGfims+/OnYFr/HfP7DQ/cGY9ZYT2f5Ps+Nqz/0A6bWrZ/+hD+K4Fvg1CVo7hvwsB7dxDR4nkSTFbVLr7yKhudoe9yn8z4yYuE34qob3xqXv/9fOHRd+5XX35dvyV8b2xb+MLruobnWWHdnPRnCKAgQ5hzwCff+aYqficuu+7245LrviA3XqvATtH7d3KcFE8jrXtEYHZdzlH3OMGjJB6eH8+mL5h1/QHxjzqbdLzB+fmmHAmCMeuDLA8bn3uuQPI2SSMBPZLwHbsirGGvd6jfmucAwuXVriVk3BYnGKmqbH5Hx+uah71EPBei6b4hXXX2MQ/dJ3/nmf80D+/PDWtQ5GBEdherOoYGAXhBzPOQN9c/y7ldC0plPeynPcOWZ3w/9kNPf+bmFbfFjlPn5ujw2JmcOt4Wz9PgX27fFxR+67BkPX5q4u/4X73XHichwCuvkK9KwpOfJTBPXV3JcVOuvuN9XLLzssJd+4IsTZjM3mwJ93+Xzxg0QHcVzcTnzHHkmkD5ovhKWuGG88vo0okwAcox8jEyLGiDpVR2X7PxILCy8Icf6XLuu04kGnV7MnokIee6juFFRI/klMOXN8clPeY6r+aCRjsX+ijjt/G/BW7Nj9T/tpxhUZdFAVpRoQE8kMMfTJwdwjtyjm0yj8mNoxc1fVJJPoIqX/IHWzqtU4KC8MDxL6OoUuuo7IkF9YO4wr/6ieYfn9YGVaAC7Xld5ySeWfPLyNus//kb5bofdxIA1ULBA3T4Cz3h1/w6UV1E99MVxYAkIYVvnb8xd4thDvwD8pGWBMWAtQKTAZdczYuBFDCgvaKJ8TI/i7v153hOfuLqoX87Ft6R9d03svvUf4y233BJv/cjb4i0f+fF46y074xr6Wz/yavy34n+Q/J9g/2Ms7P5sLHTvjsV4aXTdNzod6GwRXQwt0UJ1RQIJFRfeFlffvL4vft2uV8a0INcPmkgZg35pRFQ9YtJ6ifii+YqE9sxnPoclHyIuZ47ueHT727jyprfEFTd9L4G1PXwhv+yGF8Ri90km/uEIFrbeoInW44VKJFYRczwqT4zuEUzxO3Hpda+Ln94A/8u9sUj2lTaoztoi7nA9SyD3VxPGIOhquj/zFY3pi9NuXl807/htBrZad2PsabpP90to7jBvQDQ/IkF97y/MSFRv5xUJmi8uXnBvxcADki+uYe+6/7dsNus1KGY9FCQaq6hduzwLFc1D180025qvn4C8iP8aO68s/yJ8rF/beeXDo9v90YjuqyOb66cxlIU7Pg/WZf0lPYJxnUQGiJVX0Xzt5hVAND/Vo3L2MVJGf8hp136QZV9Pv8OSBgHwVjh67zVIbOERC9FdcculRz54BdqdhhjffeDipz2F+S5mSn8zs4xh8mIlzLnpfCm6/qf/+cH3+dHHnHLdPyennTavAvnXKLyuPj9sA+AcPiaz+7DmK0YEN+74fHIDRY4r+bAR0NUUcef5GTC7+n6vf3wZgz81zMt8dR2Cs3p0SoKbPkykuyc+cfdTeUFLvohh2rz7F+WRCpG0EN3itXHizuX/Rkisri2sbhijapEWXYsTs0ryFUdeIRbI9MjXKAlhMjwcn2kSorkUJ412WrUC079jziTqDAQyDzc9Tt7U4PRIHhdiRJK4wzjs0ciJIl2vl24ihvzYxK3Wn8h+3FdulMASN4wTHhDDvPwpEpZmOHlpSCAhhG31j6qjl/Y9+gctXMryfztfj+xSyHjd9Ut9wDx/Eii0ubxpA4nOXQ1wRT7x5IvyJ73rDqHe40i/Anwl+Eru33Mi+uPxnxQRD8eefbNu/d7XFSHA8wxN6Di5TsEE/AHvQ/I98cYPrd8/mHTKEZ+IWPw51uEo62KxL88Rte6KsaQZNzQiTu63IDB3JM91coNDCjeNxIyfEF18PK648X/GlTf+aFx103dlfrWny298Qlzx3p+KxYW/Yd6Lmebf0DlY0HqAqBisb/0jQpseyWNAxdQLWMoAABAASURBVCH34viK2z4Rl1x/5ODu7zP1Wb83GtvJakbEcH9z9cuXBZLWCvODEdzfkb7jYkmTZ7yi6QPiG3OEct/ud9rVIXUnn0hSXrroGxrEKq7IJy9NevI0DBBf06P742XTWU8ux3qiBOufonbtyYdY0Tiu4PZjrn4STMv99O9i+0HXx85L1+8fufyxNx0V3cLvsP5DuYEjah2uH7QRMdxfV5DU3GHcwBTvlO8+mU8BCjjF/u633f7Fn6ayK3hR3m1p2ZcURW6Qq8T5rvNxu7uD3/nBS5/xzSV0l+HmK4/55ti2/WdY5y5/U87XAV9igZ887Pk3vPzYY6/dkp8k7O/AOriokc8Pd199HipGNQoK3DAZTiTAsHC8GPjGK+qq5ogYS5/PXRJW2V//+i/zEnx85PpZwPB8sEyWERrOXVDI9YmJwMDDyOFJwClYIKep+RHTgDvBGf+b4153/H8k1+RY/U/7uUmqssYUiXpwOUduam73JOb4gZgMJIwR4fjkEyBMIMNh0yfMLx4RYthGQ6f11SiwmBdkNrK6qTf6JpY07rILknwSyasoPwMYBQUvXOIszPWsEYKb7GC7d0uPkZ8GmwXr7kVc9Ahlynmnz0PYJPhpqr3O3X+ErOteXh/H2WoWijded33qKjBXf24kE7mdcD/FZYYIhmUiMWijkeHMu06GGej4uYJIEGYgfAzcYX0NoskHp4fz6YvmKxorwzSZECBg3kKYPpI/hDk/JLZ17wHX79jW7Rwm7wbwnPVgiB3xioSyPtFuvYnFEJJPkGGcy2EC03mCREVCIT/Rkzzy6t913wr8JP03eUH/Qlx5001x5Y3n88J+PC/aj4vLb/p3jgj/Qb1L3/fQuPz6747L3/s08i+Ly294c1z+3l+Py2/oo+tv5nq9NCIeQOdg/h7wEHHJR4gsFllPTcTQsl5M0XxFQsnv49+wznvjkut+MS5976q+YXKqNenW5z4iNzRM2Q8w7JP4tP498XIIA51vjp+J4VTjFY0eEN+YF13cL3IKQyeunuqRSDT1E+keyZenIxKY4xMv4fHzIa8n8bU8Dr7Pnw7Tsf5gTJajgBq2/pofkby2dZuvaExfzC6PiczP6dE/Krbf+w/iwjev7Z+MefWV/youfOPlEYvXsRzf0rJ+1gMSyJI86SZimLc+0VhF7drNa4vmKxpjezFeH+ZzHfOJEOQD+/vwfzG20N12IeXevFItvXuw/LlkH5T/+K7rLr7lmqc/cC61d6db3BUv5/Nm8kdu+/kRc+7gdH33tsXY9ZNdZzHz9OZtUgW6ruc+oHjuPC8zFzevLm4+N/qzwIyHNYZ9nvbIh+i8QK7jfMnPARH520ImV9kvvei/Me9rfBByhpwfq0yPxVEKEJbmk2cCWt2Q9QUJw3N8AoRTF/MOEQmnKZp3fBenx2nnfn/G7+FpYdXju47f5aOqWVFD7U5IOCy+blpcxiNQeW6q8gnnRKL5iuanvFC9aG3VCuya/3FNOdW3ovOq/4heCB2RRMIETeFGmAiaSMD5xus/C0e/sPp7j2n268G2Yq/7hFC2nzzcRHXIcQTMB03EDXUSg4DXgVQ+BxW7jkQ66396zve9JaL7eJYbtbk8BWYZInHrLOHIuoNWAwUFJxIZJk03NBwvDoHIJk/DdZJPYBmPBGFpIS+wRHkVCc0dxg2IycMRgenyEWVieSZEeYZZNob2PXH1zb8wmOtwPukpf8a1fzV9NnnWgZtoIRRkXYT4Rcrz0AmnUXkVDTpMzF6I5pftsxATCi91IVD5EV/Buocz1YX0t0Z0H42u/3NewvvYtftzsW33n0W38N8Zdj37+InoupMj4tEgLlY4L/NV1K3hRAI9+a6gMetM5GQcCBHagBoEHQfk0QW/UO7+RFx63VnUy2QZ3benrMelS32uXt1afxDYEy/55jGSX+ZJPrHpYV5fNM+wWNQxuKW6O4twa3WfYtCKPFgcAy153j+VT0ZXGLDwQmSChIrJmty3+hLENewvOfbLzPYXZSFMjrpMrbsiqfmjEM27IVE9DIsj2QCO+cpLJNbHN0QXvxYXvfld8eNX/0ciqz92vumr48I3/UzsXvg0058Rrpd1sH6iU6OvYCcsDDyskY898rHrUfMVjVdeTjtOSIZA5RmW13Va5PbvceiO6/9uYWG3fzz3M9NKeHOauw2GHPvAoPyu7/sn7f5SvO2Wy58+/GYo8T0djOr4hv1U8j/Etch95/wExgPSaEPCvZ0Crj74Af/0Yn8DYZZr1hZQYCHKXZDIDTWgOyPBxY8xgM/9sNTNvONMJAYtB4IcZVj4mGWYQOXdk2/MmTqPv1vYya/nfxJMC2bIMgfDcyYiXN9EcaNiGkErgSlvjk++h5Z8Dfyar+FE4pnur4lTf/Qhhu5JX1j14L7fFm4mi2GWihaNy0PN2WJLIkF/CHOOwB15OY6AvKCJuCm6aN71xBzYy4DYjlUpsLA4fAVbZRTVt6KT6ovTbl79Vd98YiHU65RuSeyJv1n/jrl7y31iJJZ91vtyiTvc33BrfkRjTDDyMXAHfhoRVd+g8Ssx5313LHSn5PpZr8tSX1DXnq6nFNKz+ivfBL24e36e4eT8In2O78QE1IPU3JH1EBHNVyQU+mJ2xos1rzuXJ+kyK+1XnvzMj7xnxTUfXr9/CXzbvV5F/Z9ZWs6gn4VQkHVRztzh/gyI5isa0xen3Xwwn2ieaXWTQni4nnolIU9CuhD0geRVNO8QUZ62aH6OnwGyIuBhvmLWw0DRWEXt2uU7PFGDROXpGo94ALpdGpdd/3G+sf9WGPv4yEJYE2Q7GCmX8lBXJHqhs24J8MIGDvXDJ4478DXIJx+cHnN8eAyLzfxZO93bvO3Oii5ln3emh0KrT+UlDtMw0TC7+eThmjfN9HhheOR5vWJd2q/PzZr1EBGzHgoSCUVF7drlBQWL5qHrZprwfP0E5ElIHn7EAu6xsbv/o7jo6t+Mi954bvgt+qve9PU5x0qnnW96KLwnxEVvenH82FVvw/597rnPQX0J9+t9mI9lWcB6UreOFIc+YSzyeQYJmLYu8yU8gnEd0XxFY/ritJu3ANE80+tSl9aUud/sQ09/9+91/cKxFPB/ECDypXlZdSUAuBW48Lqjdi92V3/4sqO/SX9P/UMXH/U9/eLiaxmwjeGA5wl74ro2Lkf80pd3b//hQ0/6yK0TZjO3ggL8ts6wDR40rzTATRH5XKQRtJLImw1CuhVJy/N5EoeBBEseaww7nrD3dSSf5D39xpwp4tqdt/MMP2dcx1hdR7smXD9ILKsfEuGsy/zIJ+6RfAzHV17WXxPkHKerOfLi/rHtoGsN3ZO+sOrBtUiLtriKFpuTGsBIgDBF3KSJoyGBgMCwDCcSIByu5+aNpYhptNOqFTh4GKmu6jkiYXVGdm58nCXHyEtCRIERR6MkhJw/aFxIfQBn8x61/sS6oYIFRhlGoybcdg7EAGtYXXFjyq/XIWyrf1Qdfbf7Dz36T/nQ+vHIeoJmoYJIodbrdSVUYcCSz3HyIADVDY304YlhqwZIOOepmHziI8qfdOtInRggmlqKxhwvrsQ3bneZ5Gkwn4Uk4CdKouOGvL7/mbjmQ35rHGvZci7/93V9f7rLpO/6Gnurv+anqK0ejhf1p33pfMmDUDELcMMlkIBf0TxuTFHd9MXkRVQ3jayDRGLQcgKQI+uZoAOTN+WTr0fyGV/RePIxCIdxzIKPisXF/xmXvPsnDO27noWwHMg2MCJGxLBe6xTDBk9I/czjmAdClFfR2LQbr+NEc1vyb4aii3tzv3dHD4cln8FiGMCu6HzqNiI5L4c0MXkaBsit9dEv3Dg3pXXkcqwnun7WDasi5ngkH2JFE7iC2wrHh435KjiPrmis8qP/rogFfoOw/3j03V/w4t3TPxsXvul/gL8F/j96z0v4n5G/mZ8XXhddd1z08W1Ow68fgZ8wYnRRAgAL6RLJcCIB65jWb7x249oj4iS/IDB3JM91mJfCch3cOc4GcJ5w1rW/s7h78XmU9jnLnC+JaAnMrCHAi/Shu/vuLTde/awHDZH5840XH/GAvlt4BdEH0DmWzDDnjs4ndu1aPO8ZL373PzCgHVtNgUU3tOR5KG6MD0gJCN6QU/Q28bkSV+RHZDhoyYNYkVCsxTfmznPFRb/LOhdYnu6AXZpBImy6Sz8fKMfUjG8giZHDJm4GHG8iMWgSACcowyL3Z4xA3z8qTj/3NXqr7av/aT+LtAiWrkXh4kUk1uKDRiDdggAf4jHjZSBiCESIyY/Zh7fruQ6hqKjd+uoUuO22Mk6h0V99QySsviWMF4a5EDG0kpjjkynh5Dnem9aBxkec8rA364FMuc+s3w0SmNNDnySQvKkehGMlPRy/R37Q8tMU3IfH7XdcxGp/ROdwn4JIodbrPggNQJxw7jcD+GICiUTJdNzZ849feZqV5/zyvI/2qh8DzCc/B0Tox5Jm3pBovqKxMkxzrN98kGD6SD5ZXM4RiSZC+rvimg88DGvtj1Of4g+778iJsx4sMethfZFQDAXFXJNXA/Kgz9P6ISsv8xASCY+Irf45kLyYgACJ5um4CoHFURMFC4zTaDi/6HyMCFFe0ETny7rw5S3jE5cHDPcRA5IPGku+xqSb1xW7hZfFJe/5dFyyr/7XatZlwaJFgLqaWY8+gVp3xcxzIp37xEw0X8cZm3bjqSfziea2edpq3f2xJ/e7VA99UsMhDwHlqUe6+mQBzmF42f07x48hnXwmyPlBwmt+3HH7+2IoKLJZh+tWHJ8HslkH6GG+onF90VhFbceL5hM5mXc7roM7LG9ARySRfDEeFF337UzzndHF16YwHXECQSCy4TtMW8Qd7tsSWMYnnvMXNK8vEoqK2rWb1875WaCiMdysSzt5BEYkuNJ8hPf38VUHPcjflDmXOibfUrsxIhy5BXB21Fz3qO1fvuPDH7z0GY+Bw2ZnjIMWDj6c7R5qhJd4Ydbr8IyMWf8qxQlPe8l7P5Phdtp6Cvi3SLlRIrhVvAeA2eNLgBtmeH5MBA0kjDELO55wBpbxYY58DPPJzwERa/GNOUvkcdmFF0XX/XbaZfq0c0NYLE8+y4yaTyRR0YT1JTIm6xXpzqMvmjckMjxN0XkcL8ob+C+PHec8MTmrOC2sYswwZFg8rDE/dINmkYC1xZCIqGjRtfjkEaiYA/ATg0aiuq6DG87j+BGjtXuiwILCOoFCg7rqOyKx1L1i4am/12lvvBwnH0Ne8ss8JRyLmSC4yY5af+pg7SWQ23G/xFJHkS6PcIjqUNFhQRPNO14073gx+XDyWP2jmsNXc/Jb28XdRzH0n6LWY53WlXVTcGLQMHAjeUHDr7yK5ks45E33aTxoI2IkX0yDpAjU+TSzHgzR+SoSmjuMG0hkHqZd9rll3rrEpbzkm6AzPCovugfE4rYb4+03lm8kYm3b7fc6g19VPhHW47qJaRAWXa6idumVV9F8SyquAAAQAElEQVTwHM0NETRf9Uz9iImA4YjCS2SCOX5EDc8M+fDCBnYivYYdTzhmA4ONRLpBkwdE8iQSqPUMBZmNIBy2yqtobOTrlG7e8eKQ7+L2W/khtOTXFdyHC0zQ+nWzHnMEsq6Kxui4nCOSx4CKQUs+OD3M64vmHX9AfGPOpt0vMD7X2qEAGKMe+PIAbwcyeftF8tC3YubxE2NouGG+YqxD23nq55n1o/ThsG4tMeumINFYRW3zI1Kgvnnoe9RDAeZ4TJB80H0Kic6nswRxw3xF59MXnSdoovm6jnnrIlWETysyjzmHDiQ28rHrIS/XYYGaH1ES8QSReeSbTxff3Abr37njqju2/e3n39zFwllddHeM+lCupcdcI1j8zPXxrX2/+LYPXn7U40o4rrnmkHvzuv3qxT7uD9bwgLPhqFiz3RcWoz/78LNv+M2B1M5bUoG+5/bq2Bo3QQGccrsRIBwBcmeMuMTNeOX1aUSZAOQY+RiZFjVA0mt6LHTH5XxOn4Yn1yHgc+8+ijtuS4N0iJkvBCH9GLbj+Clv5JP3kG++8hINkOy7t8fpO/nNS+y7eaz+p/0++OWeqqwByHUTDeiJBPKiicSyaJHuQbjsHq/yMT2Km7+oVJ7jFU1xem4uea2vToFtKSbyKzRT9KChEYnpA3OH+UH/CPMMi7w+EQMaCJpIYuQTSj5xwnib86D8YZ+Un/uogYIF6m2KwDHwSdT9G8DNhOg8FceBJSAErdtP9/vxj//z6Prvpyy+sqfQZdfT2ugepHO/2gyIcZ8xhM07XjSvHqI8KCHW/Yrykp8GaTHmm3kjovNVNKYvTrv5uk7N12kTXdgBIoGECWa9NQ86XxcPjdvu9e7YuXP1n6dMteLxvEO/wCft01n2C5Y9fB5SD4Go9U8HWo++aL6iMX1x2s0H883tsxAIDxdOPwm4FYk5X3HxIvStK3JghGg+aKJh1xPlJZ+cPEE0ry1mnoGJBCtijkflVRwTGHUdTArnTCB58eXYtu3oeMmxvgQRX++DdXMJkO2kKeIO19MIgdwfqA6GRF3trJsBFY0lX2PSzeuK5h3fdQw0uIV6Xz4Pezbo9oAQl27RvDHR/IgE9b0PMefuj0CuOl9iEsLwwAuaCWB9jteO01qvjmi9FY3pi9Nu3kJF85bJdpKSaEBPJDDH0ycHcA6nGfY7CTgsaGKGMUbEcL5hIKTiA8N9nsYQ5xzyGB42MdMYI2qYnPScHz+RvOg+CUVF7drNx4TH9Lr8mkawkjYOHrrzI7s+99nuHX3XX8St6T8GCCytz02U2MQk8g2xuPDu97/hmB+85dJn3e8hX3zguQx+6HANydZjbszgIB3G4s/uPrhb8V+Ir0MbbhEFxueC/dQnIZHbgNBwzxCY4xUfyHzFfKAcRGDJ8OG5N0eCmyzH6a5lv2Tnn3Kf/8hQBuvk3KL1VCTo+sXFixkfHhNkwHzQxAxjiOYdLw4DIZEgne6IGMnL9IMjbv95rLt9rP4HSWpKkUVqGYtLwzpMgFkkhBGJ4SZNHA35BAQoGU4kQDgcLxpTnG71pecU7TQooK5aI+L4MCJ7PlS4c0fyuBAjksWN5GOPRgkIeb3MQXQcEJv5j1daf9lOrLRf8+5bND8ihvE5PSIfI2mGYzQkVn7Q9IH9cTznsR9i2ZfTI7x+Fjpi3Hn9bCMsX9SYQ8Y7nyBWXsXkO4CA92XyJifryPQkX3kVJ/RIPgMqmsMVXD6CedIomAChYuaDZkAQzXeHxtc95nVE1v7YcdifRtf9FybuQfTuY8RY0jrrITYitjoYFnHnjuSV+XLfZHE5s47nHIhRMAFCRTJ12ICZIFqwQFTUyDoIJEIdBmI4r3FM68KNFfkxa8mDWHGWmdRvkHkTwG7hh+LMp/2h7t3tq+NTn/twnyyfc4yIoQ5z9cuXBZLWCvODwb6Mk3BcxiYnecYrTlJbymT7uZ+6z4oZnJyMp+4OUDdyqU/BMI5dcUU+eWkMj+RpGIj1aeeccENEN9yf1pPLsZ4YNOsH9vjrs/kcVwYUUIaYq58E0+Y8FYNGmDP3medMYBQsME6j4XoVzQ8LMYaJsg7MivIyL9E8OQ/dRAznSz55Y9NuXH+Kd8pnnuSLDAYiNu7Pj8fuvPb2bYvbXh/94qvYGl+AUfN4ZPHpkUsl05lZD+wi3nTH4u0XL/aLO+p34QOH82w4I2ZZ5vpEd8e21xy14/ovwWrHVldg6fPATcMNwa6rUVAwMSKG95DjxdU8z7tYZq2PKy58HWX+emQ9QaNOzulyc49oLOvGSJSnUbDAwIejYToxDYITXJFvnoTrRjwxTnvFjzDobh2r/3Tq+22RxUZEvUjUErZEi9MRCWSRIjEg9EXUzPFiEJA+Ilx9wjM+MfMZ0G59VQrs9gJMRlY39UbwxJLHLb9KEyiJ5JNItyLpeh0rmvd6JZIXofMtIM4mPaw/92f9ZUN702Pkp8Eg0GFYOQ1u3s6iAZ8nUd3kiP1iHZGRfX467jE/HdFdnXVmPZZDwQUyjjvcJxolkYCfGEPDTV6iodHIsNPPzbeiHo6jq7vDEzWIJR+cHub1RfMVjZVhmmMB5i0k64aQODAMDzz9khj4L4w333yi0TXvpz75vazJNyCst6x+6qsLkk4z68HSTz72hMZcBDiSR6IioZCf6CknwCiYIJ8QwDkikURFA3U+wrqz25lAzg9mImgM1MUaeY4nnHUu40M0Dwz3CUT95BGsiDke5tPpL4qzjnp3mvvqlPW4Qeqsa+pqW5f5isaiy3OeprwaWMbPxHCq81Q0ekD8q+wIpS7udyJz3j/eZ+qRCEEedF28GNBA0EQmmOPPwuN8eaMSX69jMS7IqUs5w32ekQjrj6VNIjHrNl+R0Dxf3tL96UMEOEckyguaSMD5TBQ3t0849XC9WYAxJORhjWHHE16ZD3HkYzhf8nNAhH4saeYNieYrGsthzKNdCzAfJAwPfK1kbMTToc+79gtPev57fryL/oe7/GPtVjkrObdjKPssntuNuB+Rk0g9iD47CFZn9kqekduR5BmH/fB7/jK9dtr6CuQN1OXjGFz8vG9wM6A/C6AFiXrviLizzyMCy/gMIcyZ6TDMu55ocLunNe99bFsc/kh7Ts26ouC6iQbo1p/7w04kYH0hEpvjM5AwGzFB9yBAWKsO36seXffTcerL/3Py7+Jp9S/mWTzVUeNYFG6um5gJ3IIJJBIJVxESiYvu0nzFSTjMV/HMR2v3SIGDUszZFOP15ALUlPrLSCSu7XVQf115Fc2tyEtC1GEjbuYfFu90nxDKtnPjuInqVgUwHzTRfOqPb15dNZOvIWH1j6ozrEl/7vedwhcNvNRQT9Zd0dkJ4Ib1Z934olDRvL4oL/c5BiKbrkbVI9EBJkSToK6m+YrOpy8am3bj+qJ5x4vGmE4YugkseVk3vjxAl8zwGW39GSgJQb/rr4mrb3508tb6dOph/kufP5Gft67fl8Irul7WgWH9xisSsjxh6IVo3oSYfLIiYDhcJ2zyWU+eieJmmnBoVKz5ESNM64bjp7woEwlBGxHDOpKfA0hWxDQOhGg4UYOg44C5w3zEO+N5x1wwF98XTtbDflKAsmB1rSvLJrAnnkMqLzEHcB8WNF+7eW3R+Zh2z78JKnGzdjdG7dN9ul9CU5kjJjwTlR+0Kl9i4SWfgK7zJcL1IByZD5oJYL2O8074FS7wjbPlWC/rKbhsXeLG3N90n4YdZy67AYwpL/nEcn8VCy8LIJFuQaCGQ0NfdJ6KyY+obrhe1kEiMWg5EOQgzDkGXsSAJT/yY9bqfBXNVF4OGyckQ6DyDFcemY1+fHn77jdS40XcC3eAw8F2BsPzxJmY8E3O+p5zu5Dj4sNecP2fzsjN2tIK9FxxnwOfVxE3Et01Rt4roPmaWOKGcceNGLQcCHKMfIwMi2nEmv2r7Cwzd1zyqj+nnDOGGOtZv899IlGXL+GAGNlKYMqb45N3XPI18Gve8bqJGKblpS4GM6ARsbDwC3Hyj95/cO78vPqf9nt+uXczrt2xUEWLxh0+GEhkkSLBKR/XPYy8HFd45up8jiecPMePvPJ3zOS2fvcVWPSfZmSY+gL5w7766ovGKmrXbt4Ll9eHC5NYkrh5ndItiT3xN+s//ubecp8YiWWf431Z4jWccWKJDhgTBPGrq064g35pRKi/+aDtr79jztJzx0Ff9YPRd78W1lXrK+VGYiYYUlDAG/gY0306fkyQc4IV+U5MIvnyJt35dEXzFY3pi9kZL9a87lyepMtkPdiJBOb4xumE8zo53vxYNwnnjXh3XP3+fxvr0U477BXRdZeM6y9dI+shKNb6REKM8zzf5YV19xHygMANW6IBHZHAyCeWfOKEh3qIVd00nYh0mqI8x4vyHC/KSxIn88Dd+jxyPsflfAweEds4QH2/GA+6dfI76hncR6cskLVAdcAat219hKkvIuuWkIGIAM0HTcQddSEUydeYdHm6ovmcTsPgfurrsWxffv3v2aDbA+5Uj0DAyrcmx+WF0HEC0HzysM0bZhheGM7rFDYT4jr2bQftYPYv01mW9bKeggb1xWkf6y88YKgbUu7DAHbum8DIJ1bnI4wXwzj5k0B1K+Y8QSOQ48FhIDHGOT8wu28nefm6MGfTEJjjm5x059MVc7x8BxDUB+YOedYjmoeuu1H/jvlc7Tj+0fIvb7/jZ2Kh44Wjv9VtuAVSHDOr6mdkybfh3DtQy2FOTnHNfTr67k2j34ytr0DX1QfG6x/5PHhTZFQjaCIBbzgJ6eqTAnJgTpMJgh4kdDVF3OG5L4HkY6/PN+ZMzHHZhVdS24fpOBSwx/qH9MiLUrBAqrrL65fAvHKqLtquk2HydZ8VzUd8Yxy0/a2Deefn1b+Yd7EtN2Ux1GKNuZk0gmaiAoR0CwJJE0dDAgGBYRlOJEA43KSbN+ZCXWc0vXZahQILBy3mKHUd9MRFUuQeb8ZRb1L1WJFPMseBUY2CQs5vrs6Pvem/MWcPbCdW2q9x9y0uzYctExhg5akrbkz56m8+aP0Gud+P/dbb4+AvPy267vfy+ac0L68wYIfpRgoWiIoamSYgwh7ncQLCgrREDXUYMQfMTqkbE1U0k3yMipjh+KBVXkVC48E0kTyNUkgCfmIMDTfkic5joSMG2+kfFP22G+LK6++Lt/bHqU95QUR3aT6nsaRlHcQqYiYv67dgA5MuT51GJCct+dju0/1VzDiEisZxY4rL+BGmM6zheiMGLScAObKOCY48Fsxx5KZH8hlf0VzlEY6Mxy/xTfmz4thjl/x9Tcn7omchLASyDYyIETGsN+skH7YJZpyYCIS4jG+idPMKXdGw35CIW6nXX//dZ9Ej7/OlezQ/1QO5k1cxNBxUcEU+edN5WaohEl/P42XP/SvKe5Hlh3XV9d2v61bUrj15ECsaxxVynqh1FxScp6LEZXwDSYgcPnHnAs4TNgkFsw7sitMJ5DttxB9l1QfARQAAEABJREFUGFokj/EVjU27cf0Rceo8Iu7ckbwynwK4Hm5slF9P54pd2fHl/Mln/srVfd+dDuPv6BxuAvCYmPziY2TW7zTXveOIF133idmAZm15BfLXg/IgFIiKo1ECQnATTRE3fK7EMKFRUFDAETF68skHza3H3zF33tr7xeMjun8KG8vPf94TLGW4rVipfvM5LmgY1i8vkdAwEAMi6XRzf4RGHomRb5zexzPi9FecgHWnx+pfzHNRF2eNWhQuXkQiRUdtBNItCIxiDbuCKMEEphPoarqO4UQNgq4HtOMeKHDbbWWwQqOr+lbd1beEk0R69oFfEnN8WCWcPMfX62q8zguthjU3bb9TPSC4b2CZHrlpEua1Rdy552Gqn5zsw++jpLm/T8c++R9j8YuPoYxfpIeXd9hn0MYNYbOxdCsS8sCd7bcE5GmKc3kCy/SAQFj6MI8+geQRrYg5Hnm/4onmKxIa6tewM48B84nEki/SPVgu96tdb+g5fia+Lbbf5+K01uN02lOez7SvoHNkQeDkyHqKv7T+DLtPDHmZx08kNiJ23V9FaDGXh5PLm8CuvIqGzVc07ngxTDgGNK8p4g7XtQSW8YnLAwYeA+o+jCVfI/u74/lH/0Ba++1EfXvcLxux3j3VTzplMm/94lK+8drNO6CicX4LXdhaXWHYkftcqoc+qeGQh/7yUhei5ksYLwzf6fM88nv4zJfXM9a/veLEq6Lrfna4z12OQqxfs6J27g9DNF6REOM9l854LfOJnJIv0j3c3rg/+QSSL0KY45PXTz550hEg4bCJuMvrr4l4Z/SLJ0bOzwDR+SoSCn1x2s3r5zQsUNEY7ng9k0dgRAgrzUd4ox//8HcHvYudPI+9fXas1X3jCH4bjjk7DBbP3MRlikx8InYffGla7XTgKOAflp17Hti6Nwc3V94Y+XyMAZIkdLHGx9zxhFfmQxz5GM6X/BwQsZ7fmLN0XPHq/xNdz5cXOCyfnx+J+B5ZBoGKgWF9iRCyXpHuhvVF84ZEhqcpMjw/v0R5y/gw5QHRd5fFjvMeprm3vrC35F5zdXGLyU3BHhfHtvgwkIQYXPziZgA3ccoLG4nKcx3ccIJch0RitHZPFDi4DkZPzdQZoUckiMs5n71Q/7DJJ7E3HulIPoY8r28ZVsOxWVvdR27ETZTA3D7dNzkg5FVUB33RYUETzTteNJ/3d03AyWP1j2oOX+vT8Yd9MZ772Gcx7Q+7nci6g2bdgL6JdNmYCNRwaEz3aT5oI2IkX0yDpAiM82KrGxCi81U0Nu3G9ROZh2nzw9QYrjB0E1hLeYYrL9EAvBAJCLUuXMKfjt3xUzLWrZ922E8w91N4QP8enD+sf6ke1jWysuAIeWPdEAw7Lmi4EQaCJhKY48/CkTzyIwYN32FYY9jxhGMMBI3AlEckkkc8SNR6rNOcnbAQlVfR4MiPd8Xzjn6Gof3b3YcVTND6dbNucwSy7orG6Lic4072GWNzPh3R+Ry/n/6cgGWsX3djzD7dp/slND7X2rGUhy8PGG8nr0PyNEoiAT8xJ+IxEwkQjuTHOrbJ1OeccHz4J5SybtfPAqinoFR1GJG4fvIJisBwMF7DvAKI5g0zLCqmkUROJIwnH7ci4ZBX0bi+mPyI6kZdJ0i4Xtj6T3KtTtWKzGPNYU48v08oecjLdSbz1XlzGHGJS3mGK8/8JurH7rz29ic9793X7toeh1L2p9w+WCA3rTv0ietL+RAs5zG3+4LDX3LtPvpfRpa1G+x/BRYXI8bnIni+6B55X/CAVMw7C18sEBU1Km98njLgTGHaYeE6GWZg8sDYB+2yi97Gxm7I9S3EZa2jYmDoi5nHT7Q2EsWNaf3moyQEqSL0qLxEAyYrYssDoov7Rr/7Ws299dX/tF9Fdu3ZoqxlAKibmONBNA3IoEigBgoKROtwPrhj5LnpMRGtrYUCeX2YSFTfEYnpA3OHeS+IaN7rtex6OqIk5CWfWPKJVz6hTXdQvtvJunMfNVCwwOw2JVB5df9OQDgqmtcXx4ElIISND1Nho/XjHvs6SvIHhc+6HT4McT0snA0JE3fvz/Oe+MTVRf2ca9rz/iIgmq9IKPTFaTcfzGddNY+blEQTeiKBhAlah+mK0/nGeePjcevB3xWnP+mTSV3P02mHfTC67d9OOb+fy2Q9WKL7q0go9MVpNx+T/eV+C4Hw8utZCAkQEisfn0LC+cKGb15TxF1+/U2aKJj1YItZLwMTiVXEHI/Kq1gTfX8WL+X/pbr7F6f7K5WwrZQp6zZGIPcHZsIY43Q1k6dPIHkEK2KOhzwd0Tx0/9KZoa3V3Rg7mu7T/RKaO8wbEM2PSFA/71fsxKKv+ju9+UTzdNLzzwOxfXXc3h8aCwsfH56fLCTC+mJJc39BXjS/1/q7iDme45gP4BxOM+x3EnC+oIkZxhgRw/mGgZCKD8zqhj/k/wF4ZuzcMfzvuWrY8XN8ppke5vUTIYru01hF7drNs9C4fl2n5jcZssX+yDPe88fs59ls5WNsDeCY7mPOnXOGyym36/9H7N79Ic3WDzAFfOsbnwv2zmPEObiXYnaDeN+QmOMVH0hexWEgYwk4DKt+nHKfxpAmwc2b42IftX7bKaz3ubAAlg/XT4yhUS55bI2SEIjsvX75kkD5wDzfAHnXA+aOQc9HxOmv+P/m4kscL9GS0F10c1GqsgYgKqbhHAbAkacPsUDScGM0SkII2ogY8pxHJDWImUY7rVqB8pW5ujrHiDjePMieDxXu3JE8LsSIZHHHyzgaOUGk612rm4ghPzZxq/Unsh/3lRslsMQN44QHxDAvf4qEpRlOXhoSSAhhW/2j6uh17cc99iNx8LZvpuyfj3EjFl7qB8awhv6IQcvAgGWYaeYjRsD70UAioenhfejwRA2SlVeR0HhUXkUTZdi4XhqsKyZAqGgdjqnoPMlLQhC+Ok5+0vfFWY9d/i12rFM79Yn/K04/7JHMvjOyHqwRsdUhy2MfuHNH8ohXNIkruK2IHBhRMV0IFYOGy7l8LGcCt2CBOjw0rGfEoE0msA7diiOPiXIc9OmRPAZUjPhTXmD+Uzz/mMuntP1rU5/7UFC2kbWMiOG+ZvWTlg/M8eEZcp5l/EwMpzpPRaMHxDfm6KMu7nfa1SF1JJ9IUl666qxBTF0T9ImP4wjiZlocDXnk9tWx86R/iIPv8ySWf9/467L7WLq+dRuvaD7rxki0bo2CCfgVoSmTMGAmcAsWoI7Z8+56BkTzw0DGOC+BPiKyHkKBH/3fR7fwhLjg9D8ykt2wRofR13GgsWk3rz/FO+UzT/JFBgOcN/XxlBdc99u7+9t/qF/sr2Ejsyd8sree6zBxy/Uy0i9G3930Tw95wD8yth0HmgJ+x7P0eeCxC2+N4flEkRIQTIyIIc/x4or8MhyI5EGsaGy9/465a9iv2MmXRf3p+XlJ2YnGKUdwWxGZiKi4xM148kn0aUR5kAYknPNUdJ7kEUiENj1GHbqz4/RznjpNTe3V/7Tf93wYuDjT5WIi3YPwXPFzxUIwb9Giu3K8mDzyI2KrhbyRTyzzvVGd1lejwOId89defZ0n9UbaRAN03L1fTwgjvxoFhbxezOOhDz02/b/K7kYmG0r92Jhh7+fEkic86JcGQdA8Vr3t80ODcPIcPyYkkeg2+P1+7KM/H8c99tkRi8dEdP8vav3u0/0kxtDYTuYTDY1Ghr1d7lQP53Ooujs8UYOg6wFzh3kDovmKxsowzbEA8xbiOsknW3mJJojVfUb/uei6o3kp53dpje+HftphPxa7Fr+BlT+c+mEM5VFwjwNwLocBzHGf+oXgfklFujVeMIFEoiQ6brhQRQeO85b8yMfI+UF5pEPUDZroPI4Xc16NmoDjYb6i8+l33Zvivvf/9jjzaX9oasN063Mf7rMW5Xa0h7qRj0DyDLpfkU6Y85AfjAh5dVwsaTVe0XTXTSY0sAV6Xz4PewRyexXd2txuyQcB84kQki/SPUgjsBZ9Jf4sPPAcII/4vjxecuyX4/b/9dTouhX+7YpSj/vM/eGL1ldRu96H8oJ9QAvziUmIIDzsM2glMcefhZPn+DovqXAChwVNdD7Hi/lSHo+NC079XbKzQ55e8iBWNJbza0y6eV3R/IBGwuWzrrA5cZnPRLr4G/3XU0u/k37k2Tf9Vdx269ld11/K9m/PS1DG+FJezAHc92AFMvxz3y9+7Nhjr+Vn+GjtQFOg67rh5wOeA+8L3Lx3cPO50Z8FUIeEPKwxzA3HfUSExDL+EObMdCWffOYxuN3TPuqXveqXI7p3jnUHLcugLkwK5EzA+tyQYfeTSMqBpEeeocrTrjzHV57jHSdPjl1eIoZ5+X28Pc566YMNL+0LSwN32e+6bXlxLcZFEsto1o6ogYIJJBKDhoEb8hwvuhnCKYJovqL5OV5nNlpbpQIL/kWTydh6s1Q0VRVO9EIYFAkkTNAUbnidwpaECOer19V8Ccem/1fZ3UjQRDY2t099UkAE+YrqoC8SDpto3vGi+bzPTWYAA9ws/4rscY+/Lg7e9nC28ZbxOa77ZRvEMxwauc8qQAxNV6vqkTg3kCz+lEckKq+isWk3ri+6ruNFY0wnDN0ElrysG18eoEtmVn8GTMR1sX33f4yTnnh95vfn6cwj/iJOO+yJ0fU/xOfzX426WNNd3mchJuT+GC0SGHUhhBslHBr6oroYHzGihsPxU15IJCAEbUSM1L0gqSJ8WuE8WgP+YfTdk+Kso06Lkw691fCG6u4j2Eewz1pYda0/wwT2xHNM5SXmAOQoaL5289qi8zEt98EKREmbuPM2ktXP7zNDU5kjdY8IeSZEdQlaVSVRoYiFSCChonE6bmQ+aBKAfX3s3LkY55z4Qp7vM+aXLvW4v+k+Ddf95gADGFNe8on1dI/Ewsv9Eki3IFDDoaEvOk/F5EdUN1yv738r+t3fMfdNeZQ28jGcb+APybn6h1CZL0YMWuU5PpiHUCQSqPMZlnePfj2NDdMOe+kHvrhrcdePRt89j6I+RY89v5Sjw0D4m27xvr+l2foBqEDPA+BzwJ0yPB7cF/pAGKhoXl+s+YrGmWbg54DgFyR6OSovnztjBJKPva++MWepPA5aOCO67q/dRlBGYhpmS8A6TehWdFvGxeRrSACFOjyRAOGQV/fpPFEa6bTqfGL0D4w7tv98xpecVv9i3sfucPKeGV204lhMCWSRENKtyBgP3OFiaiRhcM0Vlx8qIkw7r+uJBvryO+bR2qoUWFwYrn2v0Mwgqm9FQqEvTrv54II4zHxiIRBedgH3xF90cBm32SD3SdGJVYCCBeptukyPWWKYYORjzM1nmgBhrOAHMZzYHM1vz5/7uJMitj+M7b4jolsEgUAONpQ7ETVA7ydSIeoGTcw0Rt4qFclNj7y/CIjyKhIKfTE748Wa153Lk3S9LBQ7kcAc3zidMBvBWPi16PqnxMlPPCaOP+yzBDbOcdoR74yDDmtrffkAABAASURBVH5YLPYXRBfD3+Ncul+rdX8Qhs9ZNpa6mKDjDvvEnuqRfGLO5/jKq2iedIjOFzTR/Bw/AyRFwMN8xXH+ktc3N+ufpe4z4rO/84h4/lE3z8IbzbL+IoBgeWKGMUT1zf3hq5sccU4PgvrJw66IOR7mdUTzOZ2GwS3U+/Lrf88G3R4Q4tItmjeWiNBi5SU6MAmeuN315eGaLy5eBGEIMTQTg7VfzuecfGVsP+jfsvbPRf1NChyeB84UWvdpmbgEIxINBE0kIM9EuvhkdIUBM5FuGKhuRe/bsBFQr/TLPMk3Tr7vXx/9Xz86dp75F3jLD2gZ7DEcnqiR0dnJuJ7oehWN6YvTbn6sg/mYXpfPbZwpcQPad7Ekvjm/7fAXvudN0S/8F17K/2BumPvNa5JGWvxi/Pr2j77NqXRgOV1X7n3A2wLIG0NMQzlKYnx+iDmshPNzUD/5OVBChjEGNOx4UV7yye7Lb8xZLt6w0/912qnh+mP9JGpdoVESAinLHcIETBtwvJgJSSRIa+XnLm5uPHlEK2KOR9UjMQc8Pk4/55VjvhgLBe8+dP6TMlTl3EDWKqYRNBMVSKRbEEiaOBoSCAgMy3AiAcLhJt2MMcWZ/mKUsXa6WwosHOTfNInUddAzIlFAcGTPmw137vA6JC8JEQVGHI2SEJIftDov5qb/xpw9sJ1Yab/G3be4NB+2TGCAlaeuuDHle7+bD9pm/B3+4x79qTj+cc+NXbu+JbruF7wNItiQ+xz3GzQDgATSgrREDXUYUd6k5zyMr2gq+RgVMcPxQau8ioTGg2kieRqlkAT8xBhaH7dEdE+Mk5/w2DjpSR+Mjdr89njH4RfFwsLDouvevsfnWZ2qHnWfFaMaBROqHqB5YYphIIlheOpmwPVMJAZNPuCRdWBUHHnMN/Iz//rYte1h8byjr4ydfItIaOMePaVRv0IIeG5LiNwn+YoZxK+YcRwRWJlvovTkMb6iYb8hEbdSr7/+u0/vi4pL92g8dVd4dYGQ/IJhHLviinzy0hgeydMwEPu3+f85P/ek50Qsfi+F/Do9wvpzf9QnBs1yAWWIIB62goK8iqaW8Q0kIXL4xJ0LOE/YJBTsut+KxXhivPL0l/Cc7vm7Mqd3yEr1G6/dvPaIOK7reBF37kge9YxIFjc246+nlL6347AXvvt3ti92T4Hzy/QvhPvEmBxfiOgvO/KF771qEotmH2AK5K8H+cDEXp9nbyBpS9H7yudJXDpB8iMyHLTkQaxIKPb8KWB2ffrlF97ENt6Yk1PODC3YQMECWb9hjfxcIZHoyExggISZNyL3FzQCyatIaHokz3GTfN/vjFNf8dgpbfUv5uPiTJeLiXQP1uQDQKt0Ar1mQSB/SBSHXZGUkAFsUBdr5Lme6xirqN366hS47bYyTqHVuyJh9S0uXgTp2fUsCa+HieKOl1HD8WLNjxjDNDkf9mY9sn437gZEAnN66JMDcsNTPQjHSno4fo/8oA2/j4Kx+Y6TDv3TOO6xPxjbuq+OWNzBBn5zfK5x4m7rgVDKHrTUTZ9A6kysIuZ4yNMRzVc0xnBh6MyT9VQkmvzgh5ruqugWvj1OfuIT6B+OzdJOecpfx2mHHRf94kPZ2qvpfxn5fLKBqsOIxDoEKdsfePjJBzNeEa4H7sDTSQJGwQJ1ePKcfxaAywTysMZw1lMCld91nyZyThy0+OA482kvibOP9HfCCW30g/2NG7NW/HG/GO4v90s80wW1SXO9kE2DQOVVJDR3GHdARZPbPG21vhc91HPcrjz0rHqkqw8B4BzKhcAxtCTgViQqr7gkInJ+A7Ex2jkn//c496THUNcx3Ga/A1Im9WWdlFgRk0SeI/XAhDbwsd0nEIkmdEQCyReJOV8J53z6LBzDwIgBfwt4alxw2nfHztPu/LMy5w+mY2Ln0xcJRUXt2s1rQ898RWOUyURaQCYmSHil+QhvheNJL7rub7vbbju+i/4M+q/wDfqno4tPosL1fdedfvut216+n/bZlt0oCviHZfP54UHhxuD+4PmgONw08vmoCeISdDVFeY4Xfe6X8SHKA/Jzxnzyc0DEvv7G3DrsCwsvAf7K7Qz7xLN+A9aXSCzrFenm9UXzhsRxfwTcluNFecv4cEY+hvnk54AI/a7/uTjz5V8VpS0UvPvgZFkEQ10E0BUGdFGKiIIJ+ImyMHCj5kcMGgnSOY/r4Ib5XIdEYrR2TxQo/yh7qGvQUmeEHpEYLue8h6PyEknsjUd64GPI80Jy2YQ6PDZrq/vIjbiJEpjbp/smB4S8ilMBHBY00bzjxeRr1AScPFb/qObwjXB6zmP/Po4/5Cq+Rf+e6BYfxof2qyO6388bzO0GbUSMlEFMg6QIqCNhreiKIfq5UDGTk5Nx3UTmcZh8Y7jC0E1gjbzuo3inxP3u8y/i5CfsiJMO/T38zXnsOPLTcdrh59G/LqJ7Mvq/Hf2+OOqvHm5fDFrqYgA7RAJdRWK4NRyjUfJhg6CrKeKG48WoAZMEdDVHxMg6En+eb92exMv4Q+Oso14Tpx3zt1I3T2d/WewE2Za3cUz1mO032XmSp5E8xlc0lnyNSTevK5p3/G4DW627MfY03af7JcR97bn0pTx8eUDqLwtZI2qgYAKJxBgabsirGBusnXPidXHeSd8Z3cI3RdddRHWfjKw/C8b1MAB2InFxqkcJRw4MWgkICjZFhoe8iuYj/P8HPy5fyM8/7ca4qy3rgDyHOTEfTwVJj4c81xOt38SIOp2nCPNTnuHKi63Z/Hvnh519wzsOe+ENzzzihTc89PCzb/gm8Ogjzr7+nce87Lp/3pq7rrtqeKcK+M9Ljc8F7Pp4JfKAVPS5CXyxQHVDo/LG5ykDkW3kY2RY1ACTsB9Ol+38AmU/1+0ExogaUz3SDz536MYtW77xinUbovnKSzTA2OSLdHlAZB4jsfBSv+4hsat7C5k8Vv/T/lO+sYun/PuhHwau2B/aRcZF+uH0w+jiXH9YF+mLk35EscXs/6GLEf/Dw3MH7bQ6BY769j+Lp39bN9ef8YjBT8QWV+yP7ML4M0W6uGL/9i4yLtK/n/5M+oDXrq7w/TjqB7/7uXHsd3fxg0v795SYSP8v9B+ki3P9UV2kL076s4stZv/eLkbUpv/Qox6+H3e+9ksfd+in4vjHnxfHP+6RsW37v+OH6bNZ5IPjZ5mfWR0RP7T8EDORSGx6mNcXzVc0pi9Ou/lg4pwfNFfAcPk09h9xOyVuX3hQnPSEQ+hXx7GP/rLULdNPP/xDseOI4+L0w++H9odF310SXfdXkbqUXaYuBvRFAkv1K2EZjAfmAoOcRMd5Hc80mRivTwZggeYj/hn+L8dCd2rcdsdX8TL+7I39d8gpfU/H2c/4kTj7md2K/YUlLmb//i5G1F7SX/QD5ImJd6W/+FldvIj+kmPft6fyNm38pT90eLz0h7rsP/ockC7urb/suV2YT8QW5/pxXaQv0l9OfxldXKn7R8k3ooDnnPBJvkG/gP5NEbsfwXPJS3r/u8Hjlc9d1lye0x70OQRixCREjPw0IgzIC5qY4f5LhN/Hb7SeHYvbv4oX8hPoH4Nx14+dZ/x2vPL0LnbS53AHsb31M4b8j4F77Wd2Yf5CkS5eeOYZd73AxmwKLFFgM7tvOP+n4+Lzurj4/KG/AVyxX9BFxkX6JfQ30MW5/souLlmhX1piYvadXYhv2Pkb+02+yy/6aFxxUTffX1V8cdKvLLaY/dVdjKhd+lVL8ce7uGov/Y0l98bXdDHfj6m6rP7FvM7QsCnQFGgK3BMFnvN9/ytOOOQNcfzjnxInPL6LvuMb9e5Mfoi8nB/6fpsXR2bnp0B/iMSaO8YfJif5yqs4HZB8fqocMT4bfdzIOheCT41tu+4XJx16dL6M7zj076ZDt6y948gPxI7DX8BL+r+NhYXvjMX+7IjO/z/uH4DBdaCjLwKhE676Ear64uIR91x4AzEqjKiR4+AlBi0n+H2mfw3rP55vxh8QZz3t++OMp705XvyMf4DQjqZAU2A1Cpx72h/wgn5BnHfKd8Qd9/rKiIWjeVB/gqneS/8LH8f8fO3w6vOYj6M+fSBgQOj7z2D472tcQ/gcntfH8BL+FXH+qUfE+ae/IXae1J5VBGpHU2CtFGjzHJgKtBfzA/O6t103BTauAic87jfj+MddEccfchb9u3hp7/hB8BujX3hSdP2Z0XU/ww+X72EDv8EPh/8zuvjz6HnBju5LxDjAPvzjzp8m/vsR3cf44fOXI+JS+nn0U3j5PDz6bV8XJz7hX/IS/lTwleCNcfxhXyR/4B6nHfY7ccaRb4gdh58cpx/xiNhxRIfWj0KQYyO6F2LzQ333FnR9P/5vRsQfo///Bj8PAvxU33Vo2PObGt1fEvCP0/4+/N8g/z78N4GvxD85YuGJ0W97OC/iHf2RcdZTz4kznvqr0VpToCmw9grsfO4/xXknXc9L+ivoT4vzTv6GOO+ULrrtD+Hz8DsiFg6NWBx63z8mFvr/FF33dbH73l8Z58O74NR/Hxec+gTsk3kZfw328A/OrX2lbcamQFNg3ynQVtpgCrQX8w12QVo5TYGmwAoKHP/4P48TH3dzHH/oFXH8438kTjj06XHCId8bJx7ybeA3gv+S/hVx4qFdnHjIV/Ct94PjxEMfCj4yTjzkcbx0f3+ceOjz6a/GvjpOfsL74+TH++IYrd2JAjue+t95Qb+WfnHseOor4owjToozjjwc/B7wW2LHkV8Pfg3YxRlP7cD7gQ+KM478ujjjqd9Ef2Sc+dTvBY8AT+Ob8At5Cb8mzjzyw3HWEZ+4k9VbuinQFFhPBc474f/wkv27cf5JH4nzTxv6Baf+epx76h/GeSf/ZfhCv57rt7mbAk2BA0CBtsW7qkB7Mb+rSjVeU6Ap0BRoCjQFmgJNgaZAU6Ap0BRoCmw8BbZARe3FfAtcxLaFpkBToCnQFGgKNAWaAk2BpkBToCnQFFhfBdZz9vZivp7qtrmbAk2BpkBToCnQFGgKNAWaAk2BpkBToClwJwpMXszvhNnSTYGmQFOgKdAUaAo0BZoCTYGmQFOgKdAUaAqsuQL7/sV8zbfQJmwKNAWaAk2BpkBToCnQFGgKNAWaAk2BpsDmVWDLvphv3kvSKm8KNAWaAk2BpkBToCnQFGgKNAWaAk2BA0mB9mJ+z652G90UaAo0BZoCTYGmQFOgKdAUaAo0BZoCTYF7pEB7Mb9H8u2rwW2dpkBToCnQFGgKNAWaAk2BpkBToCnQFNiqCrQX8616ZVezrzamKdAUaAo0BZoCTYGmQFOgKdAUaAo0Bfa5Au3FfJ9L3hZsCjQFmgJNgaZAU6Ap0BRoCjQFmgJNgabATIH2Yj7TollbS4G2m6ZAU6Ap0BRoCjQFmgJNgaZAU6ApsCkUaC/mm+IytSI3rgKtsqZAU6Ap0BRoCjQFmgJNgaZAU6ApcM8UaC/m90y/NropsG8UaKs0BZqEDDR9AAAHbElEQVQCTYGmQFOgKdAUaAo0BZoCW1aB9mK+ZS9t21hT4O4r0EY0BZoCTYGmQFOgKdAUaAo0BZoC+16B9mK+7zVvKzYFDnQF2v6bAk2BpkBToCnQFGgKNAWaAk2BiQLtxXwiRjObAk2BraRA20tToCnQFGgKNAWaAk2BpkBTYHMo0F7MN8d1alU2BZoCG1WBVldToCnQFGgKNAWaAk2BpkBT4B4q0F7M76GAbXhToCnQFNgXCrQ1mgJNgaZAU6Ap0BRoCjQFtq4C7cV8617btrOmQFOgKXB3FWj8pkBToCnQFGgKNAWaAk2B/aBAezHfD6K3JZsCTYGmwIGtQNt9U6Ap0BRoCjQFmgJNgabAVIH2Yj5Vo9lNgaZAU6ApsHUUaDtpCjQFmgJNgaZAU6ApsEkUaC/mm+RCtTKbAk2BpkBTYGMq0KpqCjQFmgJNgaZAU6ApcE8VaC/m91TBNr4p0BRoCjQFmgLrr0BboSnQFGgKNAWaAk2BLaxAezHfwhe3ba0p0BRoCjQFmgJ3T4HGbgo0BZoCTYGmQFNgfyjQXsz3h+ptzaZAU6Ap0BRoChzICrS9NwWaAk2BpkBToCkwp0B7MZ+TozlNgaZAU6Ap0BRoCmwVBdo+mgJNgaZAU6ApsFkUaC/mm+VKtTqbAk2BpkBToCnQFNiICrSamgJNgaZAU6ApcI8VaC/m91jCNkFToCnQFGgKNAWaAk2B9Vagzd8UaAo0BZoCW1mB9mK+la9u21tToCnQFGgKNAWaAk2Bu6NA4zYFmgJNgabAflGgvZjvF9nbok2BpkBToCnQFGgKNAUOXAXazpsCTYGmQFNgXoH2Yj6vR/OaAk2BpkBToCnQFGgKNAW2hgJtF02BpkBTYNMo0F7MN82laoU2BZoCTYGmQFOgKdAUaApsPAVaRU2BpkBT4J4r0F7M77mGbYamQFOgKdAUaAo0BZoCTYGmwPoq0GZvCjQFtrQC7cV8S1/etrmmQFOgKdAUaAo0BZoCTYGmwF1XoDGbAk2B/aNAezHfP7q3VZsCTYGmQFOgKdAUaAo0BZoCB6oCbd9NgabAEgXai/kSQZrbFGgKNAWaAk2BpkBToCnQFGgKbAUF2h6aAptHgfZivnmuVau0KdAUaAo0BZoCTYGmQFOgKdAU2GgKtHqaAmugQHsxXwMR2xRNgaZAU6Ap0BRoCjQFmgJNgaZAU2A9FWhzb20F2ov51r6+bXdNgaZAU6Ap0BRoCjQFmgJNgaZAU+CuKtB4+0mB9mK+n4RvyzYFmgJNgaZAU6Ap0BRoCjQFmgJNgQNTgbbrpQq0F/OlijS/KdAUaAo0BZoCTYGmQFOgKdAUaAo0BTa/AptoB+3FfBNdrFZqU6Ap0BRoCjQFmgJNgaZAU6Ap0BRoCmwsBdaimvZivhYqtjmaAk2BpkBToCnQFGgKNAWaAk2BpkBToCmwSgXuwov5Kmduw5oCTYGmQFOgKdAUaAo0BZoCTYGmQFOgKdAUuFMFNs6L+Z2W2ghNgaZAU6Ap0BRoCjQFmgJNgaZAU6Ap0BTYegoccC/mW+8Sth01BZoCTYGmQFOgKdAUaAo0BZoCTYGmwGZWoL2Yr8/Va7M2BZoCTYGmQFOgKdAUaAo0BZoCTYGmQFPgLinQXszvkkwbldTqago0BZoCTYGmQFOgKdAUaAo0BZoCTYHNrkB7Md/sV3Bf1N/WaAo0BZoCTYGmQFOgKdAUaAo0BZoCTYF1U6C9mK+btG3iu6tA4zcFmgJNgaZAU6Ap0BRoCjQFmgJNgQNRgfZifiBe9QN7z233TYGmQFOgKdAUaAo0BZoCTYGmQFNgQynQXsw31OVoxWwdBdpOmgJNgaZAU6Ap0BRoCjQFmgJNgabAXVOgvZjfNZ0aqymwMRVoVTUFmgJNgaZAU6Ap0BRoCjQFmgKbXoH2Yr7pL2HbQFNg/RVoKzQFmgJNgaZAU6Ap0BRoCjQFmgLrp0B7MV8/bdvMTYGmwN1ToLGbAk2BpkBToCnQFGgKNAWaAgekAu3F/IC87G3TTYEDWYG296ZAU6Ap0BRoCjQFmgJNgabAxlKgvZhvrOvRqmkKNAW2igJtH02BpkBToCnQFGgKNAWaAk2Bu6hAezG/i0I1WlOgKdAU2IgKtJqaAk2BpkBToCnQFGgKNAU2vwLtxXzzX8O2g6ZAU6ApsN4KtPmbAk2BpkBToCnQFGgKNAXWUYH2Yr6O4rapmwJNgaZAU+DuKNC4TYGmQFOgKdAUaAo0BQ5MBdqL+YF53duumwJNgabAgatA23lToCnQFGgKNAWaAk2BDaZAezHfYBekldMUaAo0BZoCW0OBtoumQFOgKdAUaAo0BZoCd1WB9mJ+V5VqvKZAU6Ap0BRoCmw8BVpFTYGmQFOgKdAUaApsAQXai/kWuIhtC02BpkBToCnQFFhfBdrsTYGmQFOgKdAUaAqspwLtxXw91W1zNwWaAk2BpkBToClw1xVozKZAU6Ap0BRoChygCrQX8wP0wrdtNwWaAk2BpkBT4EBVoO27KdAUaAo0BZoCG02B/x8AAP//jBLDvwAAAAZJREFUAwCa+QkZayMxeAAAAABJRU5ErkJggg==" alt="PriceHawk Logo" style="width: 120px; height: 50px; vertical-align: middle; margin-right: 10px;" />
                                    </h1>
                                    <p style="margin: 10px 0 0; color: #6b7280; font-size: 16px;">
                                        {total_changes} changes detected
                                    </p>
                                </td>
                            </tr>

                            <!-- Period Info -->
                            <tr>
                                <td style="padding: 20px; background-color: #f9fafb; border-bottom: 1px solid #e5e7eb;">
                                    <p style="margin: 0; font-size: 14px; color: #6b7280;">
                                        <strong>Period:</strong> {start_str} → {end_str}
                                    </p>
                                </td>
                            </tr>

                            <!-- Status Changes and Price Changes -->
                            {status_section}
                            {price_section}
                            {no_changes_section}

                            <!-- Footer -->
                            <tr>
                                <td style="padding: 30px 20px; text-align: center; background-color: #f9fafb; border-top: 1px solid #e5e7eb;">
                                    <p style="margin: 0 0 10px; font-size: 14px; color: #6b7280;">
                                        View all products and detailed price history on product detail page
                                    </p>
                                    <a href="{self.frontend_url}/products" style="display: inline-block; padding: 12px 24px; background-color: #06b6d4; color: #ffffff; text-decoration: none; border-radius: 6px; font-weight: bold;">
                                        Go to Dashboard
                                    </a>
                                    <p style="margin: 20px 0 0; font-size: 12px; color: #9ca3af;">
                                        This is an automated alert from PriceHawk
                                    </p>
                                </td>
                            </tr>
                        </table>
                    </td>
                </tr>
            </table>
        </body>
        </html>
        """

        return html

    def _build_product_row(self, product: Dict) -> str:
        """Build HTML table row for a single product"""
        # Extract product info
        name = product.get('name', 'Unknown Product')
        category = product.get('category', '')
        brand = product.get('brand', '')
        retailer = product.get('retailer_name', '')
        old_price = product.get('old_price', 0)
        new_price = product.get('new_price', 0)
        image = product.get('image', '')
        twd_product_id = product.get('twd_product_id', '')
        old_scraped_at = product.get('old_scraped_at')
        new_scraped_at = product.get('new_scraped_at')

        # Get image URL or placeholder
        image_url = image if image else 'https://via.placeholder.com/80'

        # Build product URL
        product_url = f"{self.frontend_url}/products/{twd_product_id}" if twd_product_id else "#"

        # Format timestamps
        old_time_str = ""
        new_time_str = ""
        if old_scraped_at:
            old_time_str = old_scraped_at.strftime('%d.%m.%Y %H:%M')
        if new_scraped_at:
            new_time_str = new_scraped_at.strftime('%d.%m.%Y %H:%M')

        # Calculate price change
        price_diff = new_price - old_price
        price_change_pct = ((new_price - old_price) / old_price * 100) if old_price > 0 else 0

        # Determine if price went up or down
        if price_diff > 0:
            change_color = "#ef4444"  # red
            change_arrow = "↑"
            change_text = f"+{price_change_pct:.1f}%"
        elif price_diff < 0:
            change_color = "#10b981"  # green
            change_arrow = "↓"
            change_text = f"{price_change_pct:.1f}%"
        else:
            change_color = "#6b7280"  # gray
            change_arrow = "="
            change_text = "0%"

        # Build product info text
        info_parts = []
        if brand:
            info_parts.append(f"<strong>{brand}</strong>")
        if category:
            info_parts.append(category)
        if retailer:
            info_parts.append(retailer)
        info_text = " • ".join(info_parts)

        row = f"""
        <tr>
            <td style="padding: 20px; border-bottom: 1px solid #e5e7eb;">
                <table role="presentation" style="width: 100%; border-collapse: collapse;">
                    <tr>
                        <td style="width: 80px; vertical-align: top;">
                            <img src="{image_url}" alt="{name}" style="width: 80px; height: 80px; object-fit: cover; border-radius: 8px;">
                        </td>
                        <td style="padding-left: 15px; padding-right: 20px; vertical-align: top; flex: 1;">
                            <p style="margin: 0 0 8px; font-size: 16px; font-weight: bold; color: #111827;">
                                <a href="{product_url}" style="color: #111827; text-decoration: none;">{name}</a>
                            </p>
                            <p style="margin: 0 0 10px; font-size: 14px; color: #6b7280;">
                                {info_text}
                            </p>
                            <div>
                                <a href="{product_url}" style="display: inline-block; padding: 6px 12px; background-color: #06b6d4; color: #ffffff; text-decoration: none; border-radius: 6px; font-size: 12px; font-weight: bold;">
                                    View on PriceHawk
                                </a>
                            </div>
                        </td>
                        <td style="width: 280px; text-align: right; vertical-align: top; padding-left: 20px;">
                            <p style="margin: 0 0 6px; font-size: 11px; color: #9ca3af; line-height: 1.4;">
                                {old_time_str} → {new_time_str}
                            </p>
                            <p style="margin: 0 0 8px; font-size: 20px; color: #111827; line-height: 1.4;">
                                <span style="color: #9ca3af; font-weight: bold;">฿{old_price:,.2f}</span>
                                <span style="margin: 0 6px; color: #9ca3af;">→</span>
                                <span style="color: #111827; font-weight: bold;">฿{new_price:,.2f}</span>
                            </p>
                            <p style="margin: 0; font-size: 14px; font-weight: bold; color: {change_color};">
                                {change_arrow} {change_text}
                            </p>
                        </td>
                    </tr>
                </table>
            </td>
        </tr>
        """

        return row

    def _build_status_changes_section(self, status_changes: List[Dict]) -> str:
        """Build HTML section for status changes (products going active/inactive)"""

        # Separate into inactive and active
        going_inactive = [p for p in status_changes if p['new_status'] == 'inactive']
        going_active = [p for p in status_changes if p['new_status'] == 'active']

        section = """
                            <!-- Status Changes Header -->
                            <tr>
                                <td style="padding: 20px 20px 10px; background-color: #ffffff;">
                                        <h2 style="margin: 0; font-size: 20px; color: #111827; border-bottom: 2px solid #06b6d4; padding-bottom: 10px;">
                                            📦 Products Active Changes ({total})
                                        </h2>
                                </td>
                            </tr>
        """.replace('{total}', str(len(status_changes)))

        # Going inactive section
        if going_inactive:
            # Limit to 15 products
            limited_inactive = going_inactive[:15]
            more_inactive_count = len(going_inactive) - len(limited_inactive)

            section += """
                            <tr>
                                <td style="padding: 15px 20px 10px; background-color: #fef2f2;">
                                        <p style="margin: 0; font-size: 16px; font-weight: bold; color: #991b1b;">
                                            ❌ Products Inactive ({count})
                                        </p>
                                        <p style="margin: 5px 0 0; font-size: 13px; color: #7f1d1d;">
                                            These products failed to fetch data and may be discontinued or out of stock
                                        </p>
                                </td>
                            </tr>
            """.replace('{count}', str(len(going_inactive)))

            for product in limited_inactive:
                section += self._build_status_change_row(product, 'inactive')

            # Add footer if there are more
            if more_inactive_count > 0:
                section += f"""
                            <tr>
                                <td style="padding: 10px 20px 20px; text-align: center; background-color: #fef2f2; font-style: italic; color: #991b1b;">
                                    ... and {more_inactive_count} more inactive product{'s' if more_inactive_count > 1 else ''} (see attached Excel file)
                                </td>
                            </tr>
                """

        # Going active section
        if going_active:
            # Limit to 15 products
            limited_active = going_active[:15]
            more_active_count = len(going_active) - len(limited_active)

            section += """
                            <tr>
                                <td style="padding: 15px 20px 10px; background-color: #f0fdf4;">
                                    <p style="margin: 0; font-size: 16px; font-weight: bold; color: #166534;">
                                        ✅ Products Active ({count})
                                    </p>
                                    <p style="margin: 5px 0 0; font-size: 13px; color: #14532d;">
                                        These products are now available again
                                    </p>
                                </td>
                            </tr>
            """.replace('{count}', str(len(going_active)))

            for product in limited_active:
                section += self._build_status_change_row(product, 'active')

            # Add footer if there are more
            if more_active_count > 0:
                section += f"""
                            <tr>
                                <td style="padding: 10px 20px 20px; text-align: center; background-color: #f0fdf4; font-style: italic; color: #166534;">
                                    ... and {more_active_count} more active product{'s' if more_active_count > 1 else ''} (see attached Excel file)
                                </td>
                            </tr>
                """

        return section

    def _build_status_change_row(self, product: Dict, status: str) -> str:
        """Build HTML row for status change product"""
        name = product.get('name', 'Unknown Product')
        category = product.get('category', '')
        brand = product.get('brand', '')
        retailer = product.get('retailer_name', '')
        sku = product.get('sku', '')
        image = product.get('image', '')
        fail_count = product.get('scrape_fail_count', 0)
        twd_product_id = product.get('twd_product_id', '')
        
        # Get image URL or placeholder
        image_url = image if image else 'https://via.placeholder.com/80'
        
        # Build product URL
        product_url = f"{self.frontend_url}/products/{twd_product_id}" if twd_product_id else "#"
        
        # Build product info text
        info_parts = []
        if brand:
            info_parts.append(f"<strong>{brand}</strong>")
        if category:
            info_parts.append(category)
        if retailer:
            info_parts.append(retailer)
        if sku:
            info_parts.append(f"SKU: {sku}")
        info_text = " • ".join(info_parts)
        
        # Status indicator
        if status == 'inactive':
            status_badge = f'<span style="display: inline-block; padding: 4px 12px; background-color: #fee2e2; color: #991b1b; border-radius: 12px; font-size: 12px; font-weight: bold;">❌ UNAVAILABLE ({fail_count} failures)</span>'
            bg_color = '#fef2f2'
        else:  # active
            status_badge = f'<span style="display: inline-block; padding: 4px 12px; background-color: #d1fae5; color: #166534; border-radius: 12px; font-size: 12px; font-weight: bold;">✅ BACK IN STOCK</span>'
            bg_color = '#f0fdf4'
        
        row = f"""
        <tr>
            <td style="padding: 20px; border-bottom: 1px solid #e5e7eb; background-color: {bg_color};">
                <table role="presentation" style="width: 100%; border-collapse: collapse;">
                    <tr>
                        <td style="width: 80px; vertical-align: top;">
                            <img src="{image_url}" alt="{name}" style="width: 80px; height: 80px; object-fit: cover; border-radius: 8px;">
                        </td>
                        <td style="padding-left: 15px; padding-right: 20px; vertical-align: top; flex: 1;">
                            <p style="margin: 0 0 8px; font-size: 16px; font-weight: bold; color: #111827;">
                                <a href="{product_url}" style="color: #111827; text-decoration: none;">{name}</a>
                            </p>
                            <p style="margin: 0 0 10px; font-size: 14px; color: #6b7280;">
                                {info_text}
                            </p>
                            <div>
                                <a href="{product_url}" style="display: inline-block; padding: 6px 12px; background-color: #06b6d4; color: #ffffff; text-decoration: none; border-radius: 6px; font-size: 12px; font-weight: bold;">
                                    View on PriceHawk
                                </a>
                            </div>
                        </td>
                        <td style="width: 200px; vertical-align: top;"></td>
                    </tr>
                </table>
            </td>
        </tr>
        """
        
        return row

    def _build_plain_text_email(
        self,
        products: List[Dict],
        status_changes: List[Dict],
        period_start: datetime,
        period_end: datetime
    ) -> str:
        """Build plain text email body as fallback"""
        start_str = period_start.strftime('%B %d, %Y at %H:%M')
        end_str = period_end.strftime('%B %d, %Y at %H:%M')

        total_changes = len(products) + len(status_changes)

        text = f"""
PRODUCT ALERT
=============

{total_changes} changes detected

Period: {start_str} → {end_str}

"""

        # Status changes section
        if status_changes:
            going_inactive = [p for p in status_changes if p['new_status'] == 'inactive']
            going_active = [p for p in status_changes if p['new_status'] == 'active']
            
            text += "STATUS CHANGES:\n--------------\n\n"
            
            if going_inactive:
                text += f"❌ Products Now Unavailable ({len(going_inactive)}):\n\n"
                for i, product in enumerate(going_inactive, 1):
                    name = product.get('name', 'Unknown Product')
                    brand = product.get('brand', '')
                    retailer = product.get('retailer_name', '')
                    fail_count = product.get('scrape_fail_count', 0)
                    
                    text += f"{i}. {name}\n"
                    if brand:
                        text += f"   Brand: {brand}\n"
                    if retailer:
                        text += f"   Retailer: {retailer}\n"
                    text += f"   Status: Failed {fail_count} times\n\n"
            
            if going_active:
                text += f"✅ Products Now Available ({len(going_active)}):\n\n"
                for i, product in enumerate(going_active, 1):
                    name = product.get('name', 'Unknown Product')
                    brand = product.get('brand', '')
                    retailer = product.get('retailer_name', '')
                    
                    text += f"{i}. {name}\n"
                    if brand:
                        text += f"   Brand: {brand}\n"
                    if retailer:
                        text += f"   Retailer: {retailer}\n"
                    text += f"   Status: Now available\n\n"

        # Price changes section
        if products:
            limited_products = self._limit_and_sort_products(products)
            more_count = len(products) - len(limited_products) if len(products) > 100 else 0

            text += f"\nPRICE CHANGES ({len(products)}):" + "\n" + "-" * 14 + "\n\n"

        if products:
            for i, product in enumerate(limited_products, 1):
                name = product.get('name', 'Unknown Product')
                brand = product.get('brand', '')
                category = product.get('category', '')
                old_price = product.get('old_price', 0)
                new_price = product.get('new_price', 0)

                price_diff = new_price - old_price
                change_symbol = "↑" if price_diff > 0 else "↓" if price_diff < 0 else "="

                text += f"{i}. {name}\n"
                if brand:
                    text += f"   Brand: {brand}\n"
                if category:
                    text += f"   Category: {category}\n"
                text += f"   Price: ฿{old_price:,.2f} → ฿{new_price:,.2f} {change_symbol}\n\n"

            if more_count > 0:
                text += f"... and {more_count} more products\n\n"

        # No changes section
        if not products and not status_changes:
            text += """
✅ NO PRODUCT CHANGES TODAY
===========================

All monitored products maintained their prices during this period.
We'll continue monitoring and notify you of any changes.

"""

        text += """
--
This is an automated alert from PriceHawk
View all products on your dashboard
"""

        return text

    def _build_test_html(self) -> str:
        """Build HTML for test email"""
        html = """
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <title>Test Email</title>
        </head>
        <body style="margin: 0; padding: 20px; font-family: Arial, sans-serif; background-color: #f3f4f6;">
            <div style="max-width: 600px; margin: 0 auto; background-color: #ffffff; padding: 40px; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
                <h1 style="color: #06b6d4; margin: 0 0 20px;">✅ Email Configuration Working!</h1>
                <p style="font-size: 16px; color: #374151; line-height: 1.6; margin: 0 0 20px;">
                    This is a test email from <strong>PriceHawk Alert System</strong>.
                </p>
                <p style="font-size: 16px; color: #374151; line-height: 1.6; margin: 0 0 20px;">
                    If you received this email, your SMTP configuration is set up correctly and price change alerts will be delivered successfully.
                </p>
                <p style="margin: 30px 0 0; font-size: 12px; color: #9ca3af; text-align: center;">
                    PriceHawk - Price Monitoring System
                </p>
            </div>
        </body>
        </html>
        """
        return html

    def _limit_and_sort_products(self, products: List[Dict], limit: int = 15) -> List[Dict]:
        """
        Sort products by price drop percentage and limit to top N

        Args:
            products: List of product dictionaries
            limit: Maximum number of products to return

        Returns:
            Limited and sorted list of products
        """
        # Calculate price change percentage for each product
        for product in products:
            old_price = product.get('old_price', 0)
            new_price = product.get('new_price', 0)
            if old_price > 0:
                change_pct = ((new_price - old_price) / old_price) * 100
                product['_change_pct'] = change_pct
            else:
                product['_change_pct'] = 0

        # Sort by change percentage (biggest drops first, then biggest increases)
        sorted_products = sorted(products, key=lambda p: p.get('_change_pct', 0))

        # Return top N
        return sorted_products[:limit]

    def _generate_price_excel(self, products: List[Dict]) -> bytes:
        """
        Generate Excel file for price changes with the specified format

        Columns: Product Name | SKU | Brand | Category | S-dept | Retail |
                 Old Price | Updated at | Updated Price | Updated at

        Args:
            products: List of all products with price changes

        Returns:
            Excel file as bytes
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Price Changes"

        # Define headers
        headers = ['Product Name', 'SKU', 'Brand', 'Category', 'S-dept', 'Retail',
                   'Old Price', 'Updated at', 'Updated Price', 'Updated at', 'URL']

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Define color fills
        green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # Light green
        red_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')    # Light red

        # Write product data
        for row_num, product in enumerate(products, 2):
            old_price = product.get('old_price', 0)
            new_price = product.get('new_price', 0)

            # Format timestamps
            old_time = product.get('old_scraped_at')
            new_time = product.get('new_scraped_at')
            old_time_str = old_time.strftime('%d-%m-%Y %H:%M') if old_time else ''
            new_time_str = new_time.strftime('%d-%m-%Y %H:%M') if new_time else ''

            # Determine price change direction
            price_increased = new_price > old_price

            # Row data
            row_data = [
                product.get('name', ''),
                product.get('sku', ''),
                product.get('brand', ''),
                product.get('category', ''),
                product.get('watchlist_group', ''),
                product.get('retailer_name', 'Thaiwatsadu'),
                old_price,
                old_time_str,
                new_price,
                new_time_str,
                product.get('link', '')
            ]

            # Write row
            link = product.get('link', '')
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)

                # Apply color to price cells (columns 7 and 9)
                if col_num == 7:  # Old Price
                    cell.fill = red_fill if price_increased else green_fill
                    if link:
                        cell.hyperlink = link
                        cell.font = Font(color='0000FF', underline='single')
                elif col_num == 9:  # Updated Price (New Price)
                    cell.fill = green_fill if price_increased else red_fill
                    if link:
                        cell.hyperlink = link
                        cell.font = Font(color='0000FF', underline='single')

        # Adjust column widths
        ws.column_dimensions['A'].width = 40  # Product Name
        ws.column_dimensions['B'].width = 15  # SKU
        ws.column_dimensions['C'].width = 15  # Brand
        ws.column_dimensions['D'].width = 20  # Category
        ws.column_dimensions['E'].width = 20  # S-dept
        ws.column_dimensions['F'].width = 15  # Retail
        ws.column_dimensions['G'].width = 12  # Old Price
        ws.column_dimensions['H'].width = 17  # Updated at
        ws.column_dimensions['I'].width = 12  # Updated Price
        ws.column_dimensions['J'].width = 17  # Updated at
        ws.column_dimensions['K'].width = 50  # URL

        # Save to bytes
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        return excel_buffer.getvalue()

    def _generate_status_excel(self, status_changes: List[Dict]) -> bytes:
        """
        Generate Excel file for status changes with the specified format

        Columns: Product Name | SKU | Brand | Category | S-dept | Retail |
                 New Status | Updated at

        Args:
            status_changes: List of all products with status changes

        Returns:
            Excel file as bytes
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Status Changes"

        # Define headers
        headers = ['Product Name', 'SKU', 'Brand', 'Category', 'S-dept', 'Retail',
                   'New Status', 'Updated at', 'URL']

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Define color fills
        green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # Active
        red_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')    # Inactive

        # Write product data
        for row_num, product in enumerate(status_changes, 2):
            # Determine new status
            new_status = product.get('new_status', '')
            is_active = new_status.lower() == 'active'

            # Format timestamp - use last_updated_at from product
            updated_time = product.get('last_updated_at') or product.get('detected_at') or product.get('scraped_at')
            updated_time_str = updated_time.strftime('%d-%m-%Y %H:%M') if updated_time else ''

            # Row data
            row_data = [
                product.get('name', ''),
                product.get('sku', ''),
                product.get('brand', ''),
                product.get('category', ''),
                product.get('watchlist_group', ''),
                product.get('retailer_name', 'Thaiwatsadu'),
                new_status,
                updated_time_str,
                product.get('link', '')
            ]

            # Write row
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)

                # Apply color to status cell (column 7)
                if col_num == 7:  # New Status
                    cell.fill = green_fill if is_active else red_fill

        # Adjust column widths
        ws.column_dimensions['A'].width = 40  # Product Name
        ws.column_dimensions['B'].width = 15  # SKU
        ws.column_dimensions['C'].width = 15  # Brand
        ws.column_dimensions['D'].width = 20  # Category
        ws.column_dimensions['E'].width = 20  # S-dept
        ws.column_dimensions['F'].width = 15  # Retail
        ws.column_dimensions['G'].width = 15  # New Status
        ws.column_dimensions['H'].width = 17  # Updated at
        ws.column_dimensions['I'].width = 50  # URL

        # Save to bytes
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        return excel_buffer.getvalue()
