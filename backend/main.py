"""
PriceHawk API - CFW/Makro Branch
Simplified backend for CFW (Central Food Wholesale) and Makro comparison
"""

from datetime import datetime, timedelta
from typing import Optional
from fastapi import FastAPI, HTTPException, Response, Depends, Cookie, Header, Request, Query
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import bcrypt
import secrets
import os
import logging

from database import get_user_by_username, get_db
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('user_sessions.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Helper Functions
# ---------------------------------------------------------------------------

def _parse_search_input(search: str) -> list[str]:
    """Normalise a search string (commas/newlines → spaces) and split into tokens."""
    normalised = search.replace('\n', ' ').replace('\r', ' ').replace(',', ' ')
    return [s.strip() for s in normalised.split() if s.strip()]


def _extract_bearer_token(authorization: str | None) -> str | None:
    """Return the token from an 'Authorization: Bearer <token>' header, or None."""
    if authorization and authorization.startswith("Bearer "):
        return authorization[7:]
    return None


def get_client_ip(request: Request) -> str:
    """Extract client IP from request headers or direct connection"""
    forwarded = request.headers.get("X-Forwarded-For")
    if forwarded:
        return forwarded.split(",")[0].strip()
    real_ip = request.headers.get("X-Real-IP")
    if real_ip:
        return real_ip
    return request.client.host if request.client else "unknown"


# ---------------------------------------------------------------------------
# FastAPI App Setup
# ---------------------------------------------------------------------------

app = FastAPI(title="PriceHawk CFW/Makro API")

# CORS configuration
CORS_ORIGINS = os.environ.get("CORS_ORIGINS", "http://localhost:3000")
cors_origins_list = [origin.strip() for origin in CORS_ORIGINS.split(",")]

app.add_middleware(
    CORSMiddleware,
    allow_origins=cors_origins_list,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ---------------------------------------------------------------------------
# Authentication
# ---------------------------------------------------------------------------

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()


def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode(), hashed.encode())


# Session settings
SESSION_EXPIRE_MINUTES = 10080  # 7 days
COOKIE_NAME = "session_token"
sessions: dict[str, dict] = {}


class LoginRequest(BaseModel):
    username: str
    password: str


class UserResponse(BaseModel):
    username: str


def get_current_user(
    session_token: Optional[str] = Cookie(None, alias=COOKIE_NAME),
    authorization: Optional[str] = Header(None)
) -> dict:
    """Validate session from cookie or Authorization header and return user"""
    token = _extract_bearer_token(authorization) or session_token

    if not token or token not in sessions:
        raise HTTPException(status_code=401, detail="Not authenticated")

    session = sessions[token]
    if datetime.utcnow() > session["expires"]:
        del sessions[token]
        raise HTTPException(status_code=401, detail="Session expired")

    session["last_activity"] = datetime.utcnow()
    return session["user"]


# ---------------------------------------------------------------------------
# Auth Endpoints
# ---------------------------------------------------------------------------

@app.post("/api/auth/login")
def login(data: LoginRequest, response: Response, request: Request):
    """Login and set session cookie"""
    user = get_user_by_username(data.username)

    if not user or not verify_password(data.password, user["hashed_password"]):
        logger.warning(f"Failed login attempt for username: {data.username} from IP: {get_client_ip(request)}")
        raise HTTPException(status_code=401, detail="Invalid username or password")

    # Create session token
    token = secrets.token_urlsafe(32)
    expires = datetime.utcnow() + timedelta(minutes=SESSION_EXPIRE_MINUTES)
    login_time = datetime.utcnow()

    sessions[token] = {
        "user": {"user_id": user["user_id"], "username": user["username"]},
        "expires": expires,
        "login_time": login_time,
        "last_activity": login_time,
        "ip_address": get_client_ip(request),
        "user_agent": request.headers.get("User-Agent", "unknown"),
    }

    logger.info(
        f"LOGIN | User: {user['username']} | IP: {get_client_ip(request)} | "
        f"Time: {login_time.isoformat()}"
    )

    # Set HTTP-only cookie
    is_production = os.environ.get("RAILWAY_ENVIRONMENT") or os.environ.get("PRODUCTION")

    if is_production:
        cookie_value = f"{COOKIE_NAME}={token}; HttpOnly; Secure; SameSite=None; Partitioned; Max-Age={SESSION_EXPIRE_MINUTES * 60}; Path=/"
        response.headers.append("Set-Cookie", cookie_value)
    else:
        response.set_cookie(
            key=COOKIE_NAME,
            value=token,
            httponly=True,
            max_age=SESSION_EXPIRE_MINUTES * 60,
            samesite="lax",
            secure=False,
        )

    return {"message": "Login successful", "username": user["username"], "token": token}


@app.post("/api/auth/logout")
def logout(
    response: Response,
    request: Request,
    session_token: Optional[str] = Cookie(None, alias=COOKIE_NAME),
    authorization: Optional[str] = Header(None)
):
    """Logout and clear session"""
    token = _extract_bearer_token(authorization) or session_token

    if token and token in sessions:
        session = sessions[token]
        username = session["user"]["username"]
        login_time = session.get("login_time", datetime.utcnow())
        logout_time = datetime.utcnow()
        session_duration = (logout_time - login_time).total_seconds()
        
        logger.info(
            f"LOGOUT | User: {username} | IP: {get_client_ip(request)} | "
            f"Duration: {session_duration:.0f} seconds"
        )
        
        del sessions[token]

    is_production = os.environ.get("RAILWAY_ENVIRONMENT") or os.environ.get("PRODUCTION")
    response.delete_cookie(
        key=COOKIE_NAME,
        samesite="none" if is_production else "lax",
        secure=True if is_production else False,
    )
    return {"message": "Logged out"}


@app.get("/api/auth/me", response_model=UserResponse)
def get_me(user: dict = Depends(get_current_user)):
    """Get current authenticated user"""
    return user


@app.get("/api/health")
def health_check():
    """Health check endpoint"""
    return {"status": "ok", "timestamp": datetime.utcnow().isoformat()}


# ---------------------------------------------------------------------------
# Dashboard Endpoints
# ---------------------------------------------------------------------------

@app.get("/api/dashboard/stats")
def get_dashboard_stats(user: dict = Depends(get_current_user)):
    """Get dashboard statistics for CFW/Makro"""
    with get_db() as conn:
        with conn.cursor() as cur:
            # Total products by retailer
            cur.execute("""
                SELECT 
                    retailer_id,
                    COUNT(*) as product_count,
                    COUNT(DISTINCT category_id) as category_count
                FROM products
                GROUP BY retailer_id
            """)
            retailer_stats = cur.fetchall()
            
            # Total products
            cur.execute("SELECT COUNT(*) as total FROM products")
            total_products = cur.fetchone()["total"]
            
            # Total categories
            cur.execute("SELECT COUNT(*) as total FROM categories")
            total_categories = cur.fetchone()["total"]
            
            # Total retailers
            cur.execute("SELECT COUNT(*) as total FROM retailers")
            total_retailers = cur.fetchone()["total"]
            
            # Recent price updates
            cur.execute("""
                SELECT COUNT(*) as total
                FROM price_history
                WHERE recorded_at >= NOW() - INTERVAL '7 days'
            """)
            recent_updates = cur.fetchone()["total"]
            
            # Price statistics by retailer
            cur.execute("""
                SELECT 
                    retailer_id,
                    AVG(current_price) as avg_price,
                    MIN(current_price) as min_price,
                    MAX(current_price) as max_price
                FROM products
                WHERE current_price IS NOT NULL
                GROUP BY retailer_id
            """)
            price_stats = cur.fetchall()
            
            return {
                "total_products": total_products,
                "total_retailers": total_retailers,
                "total_categories": total_categories,
                "total_matches": 0,  # Not applicable for CFW/Makro
                "pending_reviews": 0,  # Not applicable for CFW/Makro
                "recent_price_updates": recent_updates,
                "retailer_stats": [
                    {
                        "retailer_id": r["retailer_id"],
                        "product_count": r["product_count"],
                        "category_count": r["category_count"]
                    }
                    for r in retailer_stats
                ],
                "price_stats": [
                    {
                        "retailer_id": p["retailer_id"],
                        "avg_price": float(p["avg_price"]) if p["avg_price"] else None,
                        "min_price": float(p["min_price"]) if p["min_price"] else None,
                        "max_price": float(p["max_price"]) if p["max_price"] else None
                    }
                    for p in price_stats
                ]
            }


@app.get("/api/retailers")
def get_retailers(user: dict = Depends(get_current_user)):
    """Get all retailers (CFW and Makro)"""
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT 
                    r.retailer_id,
                    r.name,
                    r.domain,
                    COUNT(p.id) as product_count
                FROM retailers r
                LEFT JOIN products p ON r.retailer_id = p.retailer_id
                GROUP BY r.retailer_id, r.name, r.domain
                ORDER BY r.retailer_id
            """)
            retailers = cur.fetchall()
            
            return {
                "retailers": [
                    {
                        "retailer_id": r["retailer_id"],
                        "name": r["name"],
                        "domain": r["domain"],
                        "product_count": r["product_count"]
                    }
                    for r in retailers
                ]
            }


# ---------------------------------------------------------------------------
# Product Endpoints (from previous CFW/Makro endpoints)
# ---------------------------------------------------------------------------

@app.get("/api/products")
def get_products(
    user: dict = Depends(get_current_user),
    retailer: Optional[str] = Query(None, description="Filter by retailer: 'cfw' or 'makro'"),
    category: Optional[str] = Query(None, description="Filter by category (comma-separated category IDs)"),
    brand: Optional[str] = Query(None, description="Filter by brand (comma-separated brands)"),
    search: Optional[str] = Query(None, description="Search by name, SKU, or barcode"),
    match_status: Optional[str] = Query(None, description="Filter by match status: 'verified', 'unverified', 'no_match'"),
    price_status: Optional[str] = Query(None, description="Filter by price status: 'lower', 'higher', 'same', 'no_match'"),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=500)
):
    """Get products with filters"""
    with get_db() as conn:
        with conn.cursor() as cur:
            # Build WHERE clause
            where_parts = ["p.retailer_id = 'cfw'"]
            params = []

            if retailer:
                where_parts[0] = "p.retailer_id = %s"
                params.append(retailer)
            
            # Category filter (supports multiple categories)
            if category:
                category_ids = [c.strip() for c in category.split(',') if c.strip()]
                if category_ids:
                    placeholders = ','.join(['%s'] * len(category_ids))
                    where_parts.append(f"p.category_id IN ({placeholders})")
                    params.extend(category_ids)
            
            # Brand filter (supports multiple brands)
            if brand:
                brands = [b.strip() for b in brand.split(',') if b.strip()]
                if brands:
                    placeholders = ','.join(['%s'] * len(brands))
                    where_parts.append(f"p.brand IN ({placeholders})")
                    params.extend(brands)
            
            if search:
                search_tokens = _parse_search_input(search)
                for token in search_tokens:
                    where_parts.append(
                        "(p.name ILIKE %s OR p.name_en ILIKE %s OR p.sku ILIKE %s OR p.barcode ILIKE %s)"
                    )
                    pattern = f"%{token}%"
                    params.extend([pattern, pattern, pattern, pattern])
            
            # Add match_status filter
            if match_status:
                if match_status == 'no_match':
                    # No matches at all, OR all matches are rejected (is_same = FALSE)
                    where_parts.append("""
                        NOT EXISTS (
                            SELECT 1 FROM product_matches pm
                            WHERE (pm.cfw_product_id = p.id OR pm.makro_product_id = p.id)
                            AND (pm.is_verified = FALSE OR pm.is_same = TRUE)
                        )
                    """)
                elif match_status == 'unverified':
                    # Has at least one match not yet reviewed
                    where_parts.append("""
                        EXISTS (
                            SELECT 1 FROM product_matches pm
                            WHERE (pm.cfw_product_id = p.id OR pm.makro_product_id = p.id)
                            AND pm.is_verified = FALSE
                        )
                    """)
                elif match_status == 'verified':
                    # Has a confirmed correct match
                    where_parts.append("""
                        EXISTS (
                            SELECT 1 FROM product_matches pm
                            WHERE (pm.cfw_product_id = p.id OR pm.makro_product_id = p.id)
                            AND pm.is_verified = TRUE AND pm.is_same = TRUE
                        )
                    """)
            
            # Add price_status filter  
            if price_status:
                if price_status == 'no_match':
                    # No confirmed match = no price comparison
                    where_parts.append("""
                        NOT EXISTS (
                            SELECT 1 FROM product_matches pm
                            WHERE (pm.cfw_product_id = p.id OR pm.makro_product_id = p.id)
                            AND pm.is_verified = TRUE AND pm.is_same = TRUE
                        )
                    """)
                elif price_status in ['lower', 'higher', 'same']:
                    if price_status == 'lower':
                        comparison = '<'
                    elif price_status == 'higher':
                        comparison = '>'
                    else:
                        comparison = '='

                    where_parts.append(f"""
                        EXISTS (
                            SELECT 1 FROM product_matches pm
                            JOIN products p2 ON (
                                CASE
                                    WHEN pm.cfw_product_id = p.id THEN pm.makro_product_id = p2.id
                                    WHEN pm.makro_product_id = p.id THEN pm.cfw_product_id = p2.id
                                END
                            )
                            WHERE (pm.cfw_product_id = p.id OR pm.makro_product_id = p.id)
                            AND pm.is_verified = TRUE AND pm.is_same = TRUE
                            AND p.current_price IS NOT NULL
                            AND p2.current_price IS NOT NULL
                            AND p.current_price {comparison} p2.current_price
                        )
                    """)
            
            where_clause = " AND ".join(where_parts)
            
            # Get total count
            cur.execute(
                f"SELECT COUNT(*) as count FROM products p WHERE {where_clause}",
                params
            )
            total = cur.fetchone()["count"]
            
            # Get paginated products with match information
            offset = (page - 1) * page_size
            cur.execute(
                f"""
                SELECT 
                    p.id,
                    p.retailer_id,
                    r.name as retailer_name,
                    p.sku,
                    p.barcode,
                    p.name,
                    p.name_en,
                    p.brand,
                    p.category_id,
                    c.category_name,
                    p.dept_id,
                    p.sub_dept_id,
                    p.class_id,
                    p.sub_class_id,
                    p.current_price,
                    p.step_prices,
                    p.url,
                    p.image_url,
                    p.is_active,
                    p.created_at,
                    p.updated_at,
                    -- Match information
                    CASE 
                        WHEN p.retailer_id = 'cfw' THEN 
                            (SELECT COUNT(*) FROM product_matches pm WHERE pm.cfw_product_id = p.id)
                        WHEN p.retailer_id = 'makro' THEN 
                            (SELECT COUNT(*) FROM product_matches pm WHERE pm.makro_product_id = p.id)
                    END as match_count,
                    -- confirmed correct matches
                    CASE
                        WHEN p.retailer_id = 'cfw' THEN
                            (SELECT COUNT(*) FROM product_matches pm WHERE pm.cfw_product_id = p.id AND pm.is_verified = TRUE AND pm.is_same = TRUE)
                        WHEN p.retailer_id = 'makro' THEN
                            (SELECT COUNT(*) FROM product_matches pm WHERE pm.makro_product_id = p.id AND pm.is_verified = TRUE AND pm.is_same = TRUE)
                    END as verified_match_count,
                    -- matches not yet reviewed
                    CASE
                        WHEN p.retailer_id = 'cfw' THEN
                            (SELECT COUNT(*) FROM product_matches pm WHERE pm.cfw_product_id = p.id AND pm.is_verified = FALSE)
                        WHEN p.retailer_id = 'makro' THEN
                            (SELECT COUNT(*) FROM product_matches pm WHERE pm.makro_product_id = p.id AND pm.is_verified = FALSE)
                    END as unverified_count,
                    -- price: use confirmed correct match first, else top unverified by score
                    CASE
                        WHEN p.retailer_id = 'cfw' THEN COALESCE(
                            (SELECT mp.current_price FROM product_matches pm
                             JOIN products mp ON pm.makro_product_id = mp.id
                             WHERE pm.cfw_product_id = p.id AND pm.is_verified = TRUE AND pm.is_same = TRUE
                             LIMIT 1),
                            (SELECT mp.current_price FROM product_matches pm
                             JOIN products mp ON pm.makro_product_id = mp.id
                             WHERE pm.cfw_product_id = p.id AND pm.is_verified = FALSE
                             ORDER BY pm.match_score DESC LIMIT 1)
                        )
                        WHEN p.retailer_id = 'makro' THEN COALESCE(
                            (SELECT cp.current_price FROM product_matches pm
                             JOIN products cp ON pm.cfw_product_id = cp.id
                             WHERE pm.makro_product_id = p.id AND pm.is_verified = TRUE AND pm.is_same = TRUE
                             LIMIT 1),
                            (SELECT cp.current_price FROM product_matches pm
                             JOIN products cp ON pm.cfw_product_id = cp.id
                             WHERE pm.makro_product_id = p.id AND pm.is_verified = FALSE
                             ORDER BY pm.match_score DESC LIMIT 1)
                        )
                    END as matched_price
                FROM products p
                JOIN retailers r ON p.retailer_id = r.retailer_id
                LEFT JOIN categories c ON p.retailer_id = c.retailer_id AND p.category_id = c.category_id
                WHERE {where_clause}
                ORDER BY p.id
                LIMIT %s OFFSET %s
                """,
                params + [page_size, offset]
            )
            products = cur.fetchall()
            
            # Get all retailers for filter
            cur.execute("""
                SELECT DISTINCT r.retailer_id, r.name
                FROM retailers r
                JOIN products p ON r.retailer_id = p.retailer_id
                ORDER BY r.name
            """)
            retailers = cur.fetchall()
            
            # Get all categories for filter
            cur.execute("""
                SELECT DISTINCT c.category_id, c.category_name, c.retailer_id
                FROM categories c
                JOIN products p ON c.retailer_id = p.retailer_id AND c.category_id = p.category_id
                ORDER BY c.category_name
            """)
            categories = cur.fetchall()
            
            # Get all brands for filter
            cur.execute("""
                SELECT DISTINCT brand
                FROM products
                WHERE brand IS NOT NULL AND brand != ''
                ORDER BY brand
            """)
            brands = [row["brand"] for row in cur.fetchall()]
            
            return {
                "products": products,
                "total": total,
                "retailers": retailers,
                "categories": categories,
                "brands": brands,
                "pagination": {
                    "page": page,
                    "page_size": page_size,
                    "total": total,
                    "total_pages": (total + page_size - 1) // page_size
                }
            }


@app.get("/api/products/export")
def export_products(
    user: dict = Depends(get_current_user),
    search: Optional[str] = Query(None),
    category: Optional[str] = Query(None),
    brand: Optional[str] = Query(None),
    match_status: Optional[str] = Query(None),
    price_status: Optional[str] = Query(None),
):
    """Export CFW products with Makro matched price to Excel"""
    with get_db() as conn:
        with conn.cursor() as cur:
            where_parts = ["p.retailer_id = 'cfw'"]
            params = []

            if search:
                search_tokens = _parse_search_input(search)
                for token in search_tokens:
                    where_parts.append(
                        "(p.name ILIKE %s OR p.name_en ILIKE %s OR p.sku ILIKE %s OR p.barcode ILIKE %s)"
                    )
                    pattern = f"%{token}%"
                    params.extend([pattern, pattern, pattern, pattern])

            if category:
                category_ids = [c.strip() for c in category.split(',') if c.strip()]
                if category_ids:
                    placeholders = ','.join(['%s'] * len(category_ids))
                    where_parts.append(f"p.category_id IN ({placeholders})")
                    params.extend(category_ids)

            if brand:
                brands = [b.strip() for b in brand.split(',') if b.strip()]
                if brands:
                    placeholders = ','.join(['%s'] * len(brands))
                    where_parts.append(f"p.brand IN ({placeholders})")
                    params.extend(brands)

            if match_status == 'no_match':
                where_parts.append("""
                    NOT EXISTS (
                        SELECT 1 FROM product_matches pm
                        WHERE pm.cfw_product_id = p.id
                        AND (pm.is_verified = FALSE OR pm.is_same = TRUE)
                    )
                """)
            elif match_status == 'unverified':
                where_parts.append("""
                    EXISTS (SELECT 1 FROM product_matches pm WHERE pm.cfw_product_id = p.id AND pm.is_verified = FALSE)
                """)
            elif match_status == 'verified':
                where_parts.append("""
                    EXISTS (SELECT 1 FROM product_matches pm WHERE pm.cfw_product_id = p.id AND pm.is_verified = TRUE AND pm.is_same = TRUE)
                """)

            if price_status == 'no_match':
                where_parts.append("""
                    NOT EXISTS (SELECT 1 FROM product_matches pm WHERE pm.cfw_product_id = p.id AND pm.is_verified = TRUE AND pm.is_same = TRUE)
                """)
            elif price_status in ['lower', 'higher', 'same']:
                cmp = {'lower': '<', 'higher': '>', 'same': '='}[price_status]
                where_parts.append(f"""
                    EXISTS (
                        SELECT 1 FROM product_matches pm
                        JOIN products mp ON pm.makro_product_id = mp.id
                        WHERE pm.cfw_product_id = p.id AND pm.is_verified = TRUE AND pm.is_same = TRUE
                        AND p.current_price IS NOT NULL AND mp.current_price IS NOT NULL
                        AND p.current_price {cmp} mp.current_price
                    )
                """)

            where_clause = " AND ".join(where_parts)

            cur.execute(f"""
                SELECT
                    p.sku,
                    p.barcode,
                    p.name,
                    p.brand,
                    c.category_name,
                    p.current_price as cfw_price,
                    (SELECT mp.current_price FROM product_matches pm
                     JOIN products mp ON pm.makro_product_id = mp.id
                     WHERE pm.cfw_product_id = p.id AND pm.is_verified = TRUE AND pm.is_same = TRUE
                     LIMIT 1) as makro_price,
                    (SELECT mp.name FROM product_matches pm
                     JOIN products mp ON pm.makro_product_id = mp.id
                     WHERE pm.cfw_product_id = p.id AND pm.is_verified = TRUE AND pm.is_same = TRUE
                     LIMIT 1) as makro_name
                FROM products p
                LEFT JOIN categories c ON p.retailer_id = c.retailer_id AND p.category_id = c.category_id
                WHERE {where_clause}
                ORDER BY p.sku
            """, params)
            rows = cur.fetchall()

    # Build Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"

    header_fill = PatternFill(start_color="0E7490", end_color="0E7490", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_border = Border(bottom=Side(style='thin', color='CCCCCC'))

    headers = ["SKU", "Barcode", "CFW Name", "Brand", "Category", "CFW Price", "Makro Price", "Diff (%)", "Makro Name"]
    col_widths = [14, 16, 45, 20, 25, 14, 14, 10, 45]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = header_border
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 22

    green_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

    for row_idx, row in enumerate(rows, 2):
        cfw_price = row["cfw_price"]
        makro_price = row["makro_price"]

        if cfw_price and makro_price:
            diff = ((makro_price - cfw_price) / cfw_price) * 100
            diff_str = f"{'+' if diff > 0 else ''}{diff:.1f}%"
        else:
            diff = None
            diff_str = "—"

        values = [
            row["sku"],
            row["barcode"],
            row["name"],
            row["brand"],
            row["category_name"],
            cfw_price,
            makro_price,
            diff_str,
            row["makro_name"],
        ]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = Alignment(vertical='center')
            if col in (6, 7) and value is not None:
                cell.number_format = '#,##0.00'
            if diff is not None:
                if diff < 0:
                    cell.fill = green_fill
                elif diff > 0:
                    cell.fill = red_fill

        ws.row_dimensions[row_idx].height = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=products_export.xlsx"}
    )


@app.get("/api/products/{sku}")
def get_product_detail(
    sku: str,
    user: dict = Depends(get_current_user)
):
    """Get single product details by SKU"""
    with get_db() as conn:
        with conn.cursor() as cur:
            # Get product details by SKU
            cur.execute(
                """
                SELECT 
                    p.id,
                    p.retailer_id,
                    p.sku,
                    p.barcode,
                    p.name,
                    p.brand,
                    p.category_id,
                    c.category_name,
                    p.current_price,
                    p.step_prices,
                    p.url,
                    p.image_url,
                    p.is_active,
                    p.created_at,
                    p.updated_at
                FROM products p
                LEFT JOIN categories c ON p.retailer_id = c.retailer_id AND p.category_id = c.category_id
                WHERE p.sku = %s
                """,
                (sku,)
            )
            product = cur.fetchone()
            
            if not product:
                raise HTTPException(status_code=404, detail="Product not found")
            
            # Get all matches for this product
            cur.execute(
                """
                SELECT
                    pm.match_id,
                    pm.match_score,
                    pm.is_verified,
                    pm.is_same,
                    pm.verified_at,
                    p2.id as matched_product_id,
                    p2.retailer_id as matched_retailer_id,
                    r2.name as matched_retailer_name,
                    p2.sku as matched_sku,
                    p2.barcode as matched_barcode,
                    p2.name as matched_name,
                    p2.brand as matched_brand,
                    c2.category_name as matched_category_name,
                    p2.current_price as matched_price,
                    p2.step_prices as matched_step_prices,
                    p2.url as matched_url,
                    p2.image_url as matched_image,
                    p2.updated_at as matched_updated_at
                FROM product_matches pm
                JOIN products p2 ON (
                    CASE
                        WHEN pm.cfw_product_id = %s THEN pm.makro_product_id = p2.id
                        WHEN pm.makro_product_id = %s THEN pm.cfw_product_id = p2.id
                    END
                )
                JOIN retailers r2 ON p2.retailer_id = r2.retailer_id
                LEFT JOIN categories c2 ON p2.retailer_id = c2.retailer_id AND p2.category_id = c2.category_id
                WHERE pm.cfw_product_id = %s OR pm.makro_product_id = %s
                ORDER BY pm.is_verified DESC, pm.match_score DESC
                """,
                (product['id'], product['id'], product['id'], product['id'])
            )

            matches_rows = cur.fetchall()

            # Format matches — shape must match frontend Match interface
            matches = []
            for row in matches_rows:
                matches.append({
                    "match_id": row["match_id"],
                    "is_same": row["is_same"],
                    "verified_by_user": bool(row["is_verified"]),
                    "confidence_score": float(row["match_score"]) if row["match_score"] else None,
                    "reason": None,
                    "match_type": "auto",
                    "product": {
                        "product_id": row["matched_product_id"],
                        "sku": row["matched_sku"],
                        "name": row["matched_name"],
                        "brand": row["matched_brand"],
                        "category": row["matched_category_name"],
                        "current_price": float(row["matched_price"]) if row["matched_price"] else None,
                        "original_price": None,
                        "step_prices": row["matched_step_prices"] if row["matched_step_prices"] else [],
                        "link": row["matched_url"],
                        "image": row["matched_image"],
                        "retailer_id": row["matched_retailer_id"],
                        "retailer_name": row["matched_retailer_name"],
                        "last_updated_at": row["matched_updated_at"].isoformat() if row["matched_updated_at"] else None,
                        "scrape_fail_count": 0,
                    }
                })
            
            # Format product response
            product_data = {
                "id": product["id"],
                "retailer_id": product["retailer_id"],
                "sku": product["sku"],
                "barcode": product["barcode"],
                "name": product["name"],
                "brand": product["brand"],
                "category_id": product["category_id"],
                "category_name": product["category_name"],
                "current_price": float(product["current_price"]) if product["current_price"] else None,
                "step_prices": product["step_prices"] if product["step_prices"] else [],
                "url": product["url"],
                "image": product["image_url"],  # Frontend expects "image"
                "is_active": product["is_active"],
                "created_at": product["created_at"].isoformat() if product["created_at"] else None,
                "updated_at": product["updated_at"].isoformat() if product["updated_at"] else None
            }
            
            return {
                "product": product_data,
                "matches": matches,
                "total_matches": len(matches)
            }


@app.get("/api/products/{sku}/price-history")
def get_price_history(
    sku: str,
    user: dict = Depends(get_current_user),
    days: int = Query(30, ge=1, le=365),
    start_date: str = Query(None),
    end_date: str = Query(None)
):
    """Get price history for a product and its verified matches"""
    with get_db() as conn:
        with conn.cursor() as cur:
            # Get base product info by SKU (CFW products)
            cur.execute(
                "SELECT id, name, retailer_id FROM products WHERE sku = %s AND retailer_id = 'cfw'",
                (sku,)
            )
            base = cur.fetchone()
            if not base:
                raise HTTPException(status_code=404, detail="Product not found")

            # Build date filter
            if start_date and end_date:
                date_filter = "AND recorded_at >= %s::date AND recorded_at < %s::date + interval '1 day'"
                date_params_extra = (start_date, end_date)
            else:
                date_filter = "AND recorded_at >= NOW() - make_interval(days => %s)"
                date_params_extra = (days,)

            base_id = base["id"]

            cur.execute(
                f"""
                SELECT price, recorded_at
                FROM price_history
                WHERE product_id = %s {date_filter}
                ORDER BY recorded_at ASC
                """,
                (base_id,) + date_params_extra
            )
            base_history = cur.fetchall()

            # Get verified matched products
            cur.execute(
                """
                SELECT p.id, p.name, p.retailer_id
                FROM product_matches pm
                JOIN products p ON (
                    CASE WHEN pm.cfw_product_id = %s THEN pm.makro_product_id ELSE pm.cfw_product_id END = p.id
                )
                WHERE (pm.cfw_product_id = %s OR pm.makro_product_id = %s)
                  AND pm.is_verified = TRUE
                """,
                (base_id, base_id, base_id)
            )
            matched = cur.fetchall()

            matched_products = []
            for mp in matched:
                cur.execute(
                    f"""
                    SELECT price, recorded_at
                    FROM price_history
                    WHERE product_id = %s {date_filter}
                    ORDER BY recorded_at ASC
                    """,
                    (mp["id"],) + date_params_extra
                )
                mp_history = cur.fetchall()
                matched_products.append({
                    "product_id": mp["id"],
                    "name": mp["name"],
                    "retailer": mp["retailer_id"],
                    "history": [
                        {"price": float(h["price"]) if h["price"] else None, "date": h["recorded_at"].isoformat()}
                        for h in mp_history
                    ]
                })

            return {
                "base_product": {
                    "product_id": base["id"],
                    "name": base["name"],
                    "retailer": base["retailer_id"],
                    "history": [
                        {"price": float(h["price"]) if h["price"] else None, "date": h["recorded_at"].isoformat()}
                        for h in base_history
                    ]
                },
                "matched_products": matched_products
            }


@app.post("/api/matches/{match_id}/verify")
def verify_match(
    match_id: int,
    body: dict,
    user: dict = Depends(get_current_user)
):
    """Verify a match as correct (is_same=True) or incorrect (is_same=False)"""
    is_same = body.get("is_same")
    if is_same is None:
        raise HTTPException(status_code=400, detail="is_same is required")

    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                UPDATE product_matches
                SET is_verified = TRUE,
                    is_same = %s,
                    verified_at = NOW(),
                    updated_at = NOW()
                WHERE match_id = %s
                RETURNING match_id
                """,
                (is_same, match_id)
            )
            row = cur.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="Match not found")
            conn.commit()
    return {"match_id": match_id, "is_verified": True, "is_same": is_same}


@app.post("/api/matches/{match_id}/undo")
def undo_match(
    match_id: int,
    user: dict = Depends(get_current_user)
):
    """Reset a match back to unverified"""
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                UPDATE product_matches
                SET is_verified = FALSE,
                    is_same = NULL,
                    verified_at = NULL,
                    updated_at = NOW()
                WHERE match_id = %s
                RETURNING match_id
                """,
                (match_id,)
            )
            row = cur.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="Match not found")
            conn.commit()
    return {"match_id": match_id, "is_verified": False, "is_same": None}


@app.get("/api/price-formula")
def get_price_formula_matches(
    user: dict = Depends(get_current_user),
    search: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
):
    """List all verified matches for price formula configuration"""
    with get_db() as conn:
        with conn.cursor() as cur:
            where_parts = ["pm.is_verified = TRUE", "pm.is_same = TRUE"]
            params = []

            if search:
                where_parts.append(
                    "(cp.name ILIKE %s OR cp.sku ILIKE %s OR mp.name ILIKE %s OR mp.sku ILIKE %s)"
                )
                pattern = f"%{search}%"
                params.extend([pattern, pattern, pattern, pattern])

            where_clause = " AND ".join(where_parts)

            cur.execute(
                f"SELECT COUNT(*) as count FROM product_matches pm "
                f"JOIN products cp ON pm.cfw_product_id = cp.id "
                f"JOIN products mp ON pm.makro_product_id = mp.id "
                f"WHERE {where_clause}",
                params
            )
            total = cur.fetchone()["count"]

            offset = (page - 1) * page_size
            cur.execute(f"""
                SELECT
                    pm.match_id,
                    pm.price_formula,
                    cp.id as cfw_id, cp.sku as cfw_sku, cp.name as cfw_name,
                    cp.brand as cfw_brand, cp.current_price as cfw_price, cp.image_url as cfw_image,
                    mp.id as makro_id, mp.sku as makro_sku, mp.name as makro_name,
                    mp.current_price as makro_price, mp.image_url as makro_image
                FROM product_matches pm
                JOIN products cp ON pm.cfw_product_id = cp.id
                JOIN products mp ON pm.makro_product_id = mp.id
                WHERE {where_clause}
                ORDER BY cp.sku
                LIMIT %s OFFSET %s
            """, params + [page_size, offset])
            rows = cur.fetchall()

            return {
                "matches": [dict(r) for r in rows],
                "total": total,
                "pagination": {
                    "page": page,
                    "page_size": page_size,
                    "total_pages": (total + page_size - 1) // page_size
                }
            }


@app.patch("/api/price-formula/{match_id}")
def save_price_formula(
    match_id: int,
    body: dict,
    user: dict = Depends(get_current_user)
):
    """Save or clear price formula for a verified match"""
    formula = body.get("price_formula", None)

    # Validate formula if provided
    if formula is not None:
        formula = formula.strip() or None
        if formula:
            import re
            if not re.match(r'^[*/][0-9]+(\.[0-9]+)?(/[0-9]+(\.[0-9]+)?)?$', formula):
                raise HTTPException(status_code=400, detail="Invalid formula. Use format like *5, /2, *5/2")

    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "UPDATE product_matches SET price_formula = %s, updated_at = NOW() WHERE match_id = %s RETURNING match_id",
                (formula, match_id)
            )
            if not cur.fetchone():
                raise HTTPException(status_code=404, detail="Match not found")
            conn.commit()

    return {"match_id": match_id, "price_formula": formula}


# ---------------------------------------------------------------------------
# Price by Location Endpoints
# ---------------------------------------------------------------------------

@app.get("/api/pbl/locations")
def get_pbl_locations(user: dict = Depends(get_current_user)):
    """Get all Makro locations"""
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM makro_locations ORDER BY name")
            return {"locations": [dict(r) for r in cur.fetchall()]}


@app.post("/api/pbl/locations")
def create_pbl_location(body: dict, user: dict = Depends(get_current_user)):
    """Add a new Makro branch location"""
    name = (body.get("name") or "").strip()
    branch_code = (body.get("branch_code") or "").strip()
    region = (body.get("region") or "").strip() or None
    if not name or not branch_code:
        raise HTTPException(status_code=400, detail="name and branch_code are required")
    with get_db() as conn:
        with conn.cursor() as cur:
            try:
                cur.execute(
                    "INSERT INTO makro_locations (name, branch_code, region) VALUES (%s, %s, %s) RETURNING *",
                    (name, branch_code, region)
                )
                row = cur.fetchone()
                conn.commit()
                return dict(row)
            except Exception:
                raise HTTPException(status_code=409, detail=f"Branch code '{branch_code}' already exists")


@app.delete("/api/pbl/locations/{location_id}")
def delete_pbl_location(location_id: int, user: dict = Depends(get_current_user)):
    """Delete a Makro branch location"""
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM makro_locations WHERE id = %s RETURNING id", (location_id,))
            if not cur.fetchone():
                raise HTTPException(status_code=404, detail="Location not found")
            conn.commit()
    return {"success": True}


@app.get("/api/pbl/settings")
def get_pbl_settings(user: dict = Depends(get_current_user)):
    """Get currently monitored watchlists and locations"""
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT watchlist_id FROM pbl_monitored_watchlists")
            watchlist_ids = [r["watchlist_id"] for r in cur.fetchall()]
            cur.execute("SELECT location_id FROM pbl_monitored_locations")
            location_ids = [r["location_id"] for r in cur.fetchall()]
    return {"watchlist_ids": watchlist_ids, "location_ids": location_ids}


@app.post("/api/pbl/settings")
def save_pbl_settings(body: dict, user: dict = Depends(get_current_user)):
    """Save monitored watchlists and locations (replace all)"""
    watchlist_ids = body.get("watchlist_ids", [])
    location_ids = body.get("location_ids", [])
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM pbl_monitored_watchlists")
            for wid in watchlist_ids:
                cur.execute("INSERT INTO pbl_monitored_watchlists (watchlist_id) VALUES (%s) ON CONFLICT DO NOTHING", (wid,))
            cur.execute("DELETE FROM pbl_monitored_locations")
            for lid in location_ids:
                cur.execute("INSERT INTO pbl_monitored_locations (location_id) VALUES (%s) ON CONFLICT DO NOTHING", (lid,))
            conn.commit()
    return {"success": True, "watchlist_ids": watchlist_ids, "location_ids": location_ids}


@app.get("/api/pbl/products")
def get_pbl_products(
    user: dict = Depends(get_current_user),
    search: Optional[str] = Query(None),
    watchlist_id: Optional[int] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
):
    """Get CFW products in monitored watchlists with Makro location prices"""
    with get_db() as conn:
        with conn.cursor() as cur:
            # Get monitored locations (ordered)
            cur.execute("""
                SELECT ml.id, ml.name, ml.branch_code as postal_code
                FROM pbl_monitored_locations pml
                JOIN makro_locations ml ON pml.location_id = ml.id
                WHERE ml.is_active = TRUE
                ORDER BY ml.branch_code
            """)
            locations = [dict(r) for r in cur.fetchall()]

            # Build product query
            where_parts = ["pw.watchlist_id IN (SELECT watchlist_id FROM pbl_monitored_watchlists)"]
            params = []

            if watchlist_id:
                where_parts.append("pw.watchlist_id = %s")
                params.append(watchlist_id)

            if search:
                where_parts.append("(cfw.name ILIKE %s OR cfw.sku ILIKE %s)")
                pattern = f"%{search}%"
                params.extend([pattern, pattern])

            where_clause = " AND ".join(where_parts)

            cur.execute(
                f"SELECT COUNT(DISTINCT cfw.id) as count FROM watchlist_products pw JOIN products cfw ON pw.product_id = cfw.id WHERE {where_clause}",
                params
            )
            total = cur.fetchone()["count"]

            offset = (page - 1) * page_size
            cur.execute(f"""
                SELECT DISTINCT
                    cfw.id as cfw_id, cfw.sku as cfw_sku, cfw.name as cfw_name,
                    cfw.brand, cfw.current_price as cfw_price, cfw.image_url,
                    c.category_name,
                    mp.id as makro_id, mp.sku as makro_sku, mp.name as makro_name,
                    mp.current_price as makro_price_default,
                    pw.watchlist_id,
                    w.name as watchlist_name
                FROM watchlist_products pw
                JOIN products cfw ON pw.product_id = cfw.id
                JOIN watchlists w ON pw.watchlist_id = w.id
                LEFT JOIN categories c ON cfw.retailer_id = c.retailer_id AND cfw.category_id = c.category_id
                JOIN product_matches pm ON pm.cfw_product_id = cfw.id
                    AND pm.is_verified = TRUE AND pm.is_same = TRUE
                JOIN products mp ON pm.makro_product_id = mp.id
                WHERE {where_clause}
                ORDER BY cfw.sku
                LIMIT %s OFFSET %s
            """, params + [page_size, offset])
            rows = cur.fetchall()

            # Fetch location prices for these makro products
            if rows and locations:
                makro_ids = list({r["makro_id"] for r in rows})
                loc_ids = [l["id"] for l in locations]
                placeholders_p = ','.join(['%s'] * len(makro_ids))
                placeholders_l = ','.join(['%s'] * len(loc_ids))
                cur.execute(f"""
                    SELECT makro_product_id, location_id, price
                    FROM makro_location_prices
                    WHERE makro_product_id IN ({placeholders_p})
                      AND location_id IN ({placeholders_l})
                """, makro_ids + loc_ids)
                price_map = {}
                for pr in cur.fetchall():
                    price_map[(pr["makro_product_id"], pr["location_id"])] = pr["price"]
            else:
                price_map = {}

            # Build response rows
            products = []
            for r in rows:
                loc_prices = []
                for loc in locations:
                    price = price_map.get((r["makro_id"], loc["id"]))
                    loc_prices.append({
                        "location_id": loc["id"],
                        "name": loc["name"],
                        "postal_code": loc["postal_code"],
                        "price": float(price) if price is not None else None,
                    })
                products.append({
                    **dict(r),
                    "location_prices": loc_prices,
                })

            # Get monitored watchlists for filter dropdown
            cur.execute("""
                SELECT w.id, w.name FROM pbl_monitored_watchlists pmw
                JOIN watchlists w ON pmw.watchlist_id = w.id
                ORDER BY w.name
            """)
            watchlists = [dict(r) for r in cur.fetchall()]

            return {
                "products": products,
                "locations": locations,
                "watchlists": watchlists,
                "total": total,
                "pagination": {
                    "page": page,
                    "page_size": page_size,
                    "total_pages": (total + page_size - 1) // page_size,
                }
            }


@app.get("/api/pbl/products/{cfw_sku}")
def get_pbl_product_detail(cfw_sku: str, user: dict = Depends(get_current_user)):
    """Get location price detail for a single CFW product"""
    with get_db() as conn:
        with conn.cursor() as cur:
            # CFW product
            cur.execute("""
                SELECT p.id, p.sku, p.name, p.brand, p.current_price, p.image_url,
                       c.category_name
                FROM products p
                LEFT JOIN categories c ON p.retailer_id = c.retailer_id AND p.category_id = c.category_id
                WHERE p.sku = %s AND p.retailer_id = 'cfw'
            """, (cfw_sku,))
            cfw = cur.fetchone()
            if not cfw:
                raise HTTPException(status_code=404, detail="CFW product not found")

            # Verified Makro match
            cur.execute("""
                SELECT mp.id, mp.sku, mp.name, mp.current_price
                FROM product_matches pm
                JOIN products mp ON pm.makro_product_id = mp.id
                WHERE pm.cfw_product_id = %s AND pm.is_verified = TRUE AND pm.is_same = TRUE
                LIMIT 1
            """, (cfw["id"],))
            makro = cur.fetchone()

            # Location prices
            cur.execute("""
                SELECT
                    ml.id as location_id,
                    ml.name as branch_name,
                    ml.branch_code as postal_code,
                    mlp.price as makro_price,
                    mlp.scraped_at
                FROM makro_locations ml
                LEFT JOIN makro_location_prices mlp
                    ON mlp.location_id = ml.id
                    AND mlp.makro_product_id = %s
                WHERE ml.id IN (SELECT location_id FROM pbl_monitored_locations)
                  AND ml.is_active = TRUE
                ORDER BY ml.branch_code
            """, (makro["id"] if makro else -1,))
            locations = [dict(r) for r in cur.fetchall()]

            return {
                "cfw": dict(cfw),
                "makro": dict(makro) if makro else None,
                "locations": locations,
            }


# ---------------------------------------------------------------------------
# Price Alert Endpoints
# ---------------------------------------------------------------------------

@app.get("/api/price-alerts/settings")
def get_alert_settings(user: dict = Depends(get_current_user)):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM price_alert_settings LIMIT 1")
            row = cur.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="Settings not found")
            return dict(row)


@app.put("/api/price-alerts/settings")
def update_alert_settings(body: dict, user: dict = Depends(get_current_user)):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                UPDATE price_alert_settings SET
                    schedule_frequency = %s,
                    schedule_time = %s,
                    schedule_day = %s,
                    enabled = %s,
                    updated_at = NOW()
            """, (
                body.get("schedule_frequency", "daily"),
                body.get("schedule_time", "09:00:00"),
                body.get("schedule_day"),
                body.get("enabled", True),
            ))
            conn.commit()
    return {"success": True}


@app.get("/api/price-alerts/emails")
def get_alert_emails(user: dict = Depends(get_current_user)):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM price_alert_emails ORDER BY created_at DESC")
            return [dict(r) for r in cur.fetchall()]


@app.post("/api/price-alerts/emails")
def add_alert_email(body: dict, user: dict = Depends(get_current_user)):
    email = (body.get("email") or "").strip().lower()
    if not email:
        raise HTTPException(status_code=400, detail="Email is required")
    with get_db() as conn:
        with conn.cursor() as cur:
            try:
                cur.execute(
                    "INSERT INTO price_alert_emails (email, verified) VALUES (%s, TRUE) RETURNING *",
                    (email,)
                )
                row = cur.fetchone()
                conn.commit()
                return dict(row)
            except Exception:
                raise HTTPException(status_code=409, detail="Email already exists")


@app.delete("/api/price-alerts/emails/{email_id}")
def delete_alert_email(email_id: int, user: dict = Depends(get_current_user)):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM price_alert_emails WHERE email_id = %s RETURNING email_id", (email_id,))
            if not cur.fetchone():
                raise HTTPException(status_code=404, detail="Email not found")
            conn.commit()
    return {"success": True}


@app.get("/api/categories")
def get_categories(
    user: dict = Depends(get_current_user),
    retailer: Optional[str] = Query(None, description="Filter by retailer: 'cfw' or 'makro'")
):
    """Get all categories"""
    with get_db() as conn:
        with conn.cursor() as cur:
            where_clause = ""
            params = []
            
            if retailer:
                where_clause = "WHERE retailer_id = %s"
                params.append(retailer)
            
            cur.execute(
                f"""
                SELECT 
                    c.retailer_id,
                    c.category_id,
                    c.category_name,
                    COUNT(p.id) as product_count
                FROM categories c
                LEFT JOIN products p ON c.retailer_id = p.retailer_id AND c.category_id = p.category_id
                {where_clause}
                GROUP BY c.retailer_id, c.category_id, c.category_name
                ORDER BY c.retailer_id, c.category_name
                """,
                params
            )
            categories = cur.fetchall()
            
            return {"categories": categories}


# ---------------------------------------------------------------------------
# Watchlist Endpoints
# ---------------------------------------------------------------------------

class WatchlistCreateRequest(BaseModel):
    name: str
    description: Optional[str] = None


@app.get("/api/watchlists")
def get_watchlists(user: dict = Depends(get_current_user)):
    """List all watchlists with product count"""
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT w.id, w.name, w.description, w.created_at,
                       COUNT(wp.id) as product_count
                FROM watchlists w
                LEFT JOIN watchlist_products wp ON w.id = wp.watchlist_id
                GROUP BY w.id, w.name, w.description, w.created_at
                ORDER BY w.created_at DESC
            """)
            rows = cur.fetchall()
            return {"watchlists": [dict(r) for r in rows]}


@app.post("/api/watchlists")
def create_watchlist(data: WatchlistCreateRequest, user: dict = Depends(get_current_user)):
    """Create a new watchlist"""
    with get_db() as conn:
        with conn.cursor() as cur:
            try:
                cur.execute("""
                    INSERT INTO watchlists (name, description)
                    VALUES (%s, %s) RETURNING id, name, description, created_at
                """, (data.name.strip(), data.description))
                row = cur.fetchone()
                conn.commit()
                return dict(row)
            except Exception:
                raise HTTPException(status_code=409, detail=f"Watchlist '{data.name}' already exists")


@app.delete("/api/watchlists/{watchlist_id}")
def delete_watchlist(watchlist_id: int, user: dict = Depends(get_current_user)):
    """Delete a watchlist and all its products"""
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM watchlists WHERE id = %s RETURNING id", (watchlist_id,))
            if not cur.fetchone():
                raise HTTPException(status_code=404, detail="Watchlist not found")
            conn.commit()
    return {"success": True}


@app.get("/api/watchlists/{watchlist_id}/products")
def get_watchlist_products(watchlist_id: int, user: dict = Depends(get_current_user)):
    """Get all CFW products in a watchlist with their matched Makro price"""
    with get_db() as conn:
        with conn.cursor() as cur:
            # Verify watchlist exists
            cur.execute("SELECT id, name FROM watchlists WHERE id = %s", (watchlist_id,))
            watchlist = cur.fetchone()
            if not watchlist:
                raise HTTPException(status_code=404, detail="Watchlist not found")

            cur.execute("""
                SELECT
                    p.id, p.sku, p.name, p.brand, p.current_price,
                    p.image_url, p.url,
                    c.category_name,
                    wp.added_at,
                    -- verified Makro match price
                    (SELECT mp.current_price FROM product_matches pm
                     JOIN products mp ON pm.makro_product_id = mp.id
                     WHERE pm.cfw_product_id = p.id AND pm.is_verified = TRUE AND pm.is_same = TRUE
                     LIMIT 1) as makro_price,
                    (SELECT mp.name FROM product_matches pm
                     JOIN products mp ON pm.makro_product_id = mp.id
                     WHERE pm.cfw_product_id = p.id AND pm.is_verified = TRUE AND pm.is_same = TRUE
                     LIMIT 1) as makro_name
                FROM watchlist_products wp
                JOIN products p ON wp.product_id = p.id
                LEFT JOIN categories c ON p.retailer_id = c.retailer_id AND p.category_id = c.category_id
                WHERE wp.watchlist_id = %s
                ORDER BY wp.added_at DESC
            """, (watchlist_id,))
            products = cur.fetchall()

            return {
                "watchlist": dict(watchlist),
                "products": [dict(r) for r in products]
            }


@app.get("/api/watchlists/{watchlist_id}/export")
def export_watchlist(watchlist_id: int, user: dict = Depends(get_current_user)):
    """Export watchlist products to Excel"""
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT id, name FROM watchlists WHERE id = %s", (watchlist_id,))
            watchlist = cur.fetchone()
            if not watchlist:
                raise HTTPException(status_code=404, detail="Watchlist not found")

            cur.execute("""
                SELECT
                    p.sku, p.name, p.brand, c.category_name,
                    p.current_price as cfw_price,
                    (SELECT mp.current_price FROM product_matches pm
                     JOIN products mp ON pm.makro_product_id = mp.id
                     WHERE pm.cfw_product_id = p.id AND pm.is_verified = TRUE AND pm.is_same = TRUE
                     LIMIT 1) as makro_price,
                    (SELECT mp.name FROM product_matches pm
                     JOIN products mp ON pm.makro_product_id = mp.id
                     WHERE pm.cfw_product_id = p.id AND pm.is_verified = TRUE AND pm.is_same = TRUE
                     LIMIT 1) as makro_name
                FROM watchlist_products wp
                JOIN products p ON wp.product_id = p.id
                LEFT JOIN categories c ON p.retailer_id = c.retailer_id AND p.category_id = c.category_id
                WHERE wp.watchlist_id = %s
                ORDER BY wp.added_at DESC
            """, (watchlist_id,))
            rows = cur.fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = watchlist["name"][:31]

    header_fill = PatternFill(start_color="0E7490", end_color="0E7490", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    headers = ["SKU", "CFW Name", "Brand", "Category", "CFW Price", "Makro Price", "Diff (%)", "Makro Name"]
    col_widths = [14, 45, 20, 25, 14, 14, 10, 45]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 22

    green_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

    for row_idx, row in enumerate(rows, 2):
        cfw_price = row["cfw_price"]
        makro_price = row["makro_price"]
        if cfw_price and makro_price:
            diff = ((makro_price - cfw_price) / cfw_price) * 100
            diff_str = f"{'+' if diff > 0 else ''}{diff:.1f}%"
        else:
            diff = None
            diff_str = "—"

        values = [row["sku"], row["name"], row["brand"], row["category_name"],
                  cfw_price, makro_price, diff_str, row["makro_name"]]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = Alignment(vertical='center')
            if col in (5, 6) and value is not None:
                cell.number_format = '#,##0.00'
            if diff is not None:
                cell.fill = green_fill if diff < 0 else red_fill
        ws.row_dimensions[row_idx].height = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = watchlist["name"].replace(" ", "_")
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=watchlist_{safe_name}.xlsx"}
    )


@app.post("/api/watchlists/{watchlist_id}/products")
def add_watchlist_product(watchlist_id: int, body: dict, user: dict = Depends(get_current_user)):
    """Add a CFW product to a watchlist by SKU"""
    sku = (body.get("sku") or "").strip()
    if not sku:
        raise HTTPException(status_code=400, detail="sku is required")

    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT id FROM watchlists WHERE id = %s", (watchlist_id,))
            if not cur.fetchone():
                raise HTTPException(status_code=404, detail="Watchlist not found")

            cur.execute("SELECT id, name FROM products WHERE sku = %s AND retailer_id = 'cfw'", (sku,))
            product = cur.fetchone()
            if not product:
                raise HTTPException(status_code=404, detail=f"CFW product with SKU '{sku}' not found")

            try:
                cur.execute("""
                    INSERT INTO watchlist_products (watchlist_id, product_id)
                    VALUES (%s, %s)
                """, (watchlist_id, product["id"]))
                conn.commit()
            except Exception:
                raise HTTPException(status_code=409, detail="Product already in this watchlist")

    return {"success": True, "product_id": product["id"], "name": product["name"]}


@app.delete("/api/watchlists/{watchlist_id}/products/{product_id}")
def remove_watchlist_product(watchlist_id: int, product_id: int, user: dict = Depends(get_current_user)):
    """Remove a product from a watchlist"""
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                DELETE FROM watchlist_products
                WHERE watchlist_id = %s AND product_id = %s RETURNING id
            """, (watchlist_id, product_id))
            if not cur.fetchone():
                raise HTTPException(status_code=404, detail="Product not in watchlist")
            conn.commit()
    return {"success": True}


# ---------------------------------------------------------------------------
# Scraping Infrastructure
# ---------------------------------------------------------------------------

import subprocess
import uuid
import json as _json
import signal

try:
    import psutil
    PSUTIL_AVAILABLE = True
except ImportError:
    PSUTIL_AVAILABLE = False

BACKEND_DIR = os.path.dirname(os.path.abspath(__file__))
SCRAPER_SCRIPT = os.path.join(BACKEND_DIR, "scraper-url", "adws", "adw_ecommerce_product_scraper.py")
RESULTS_DIR = os.path.join(BACKEND_DIR, "results")


def normalize_url(url: str) -> str:
    if not url:
        return url
    return url.split('?')[0].rstrip('/')


def _is_scraper_browser(pinfo: dict) -> bool:
    name = pinfo['name'].lower() if pinfo['name'] else ''
    if not any(b in name for b in ['chrome', 'chromium']):
        return False
    cmdline = pinfo['cmdline'] if pinfo['cmdline'] else []
    cmdline_str = ' '.join(cmdline).lower()
    if 'playwright' in cmdline_str or 'crawl4ai' in cmdline_str:
        return True
    if any(flag in cmdline for flag in ['--disable-dev-shm-usage', '--no-sandbox']) and '--headless' in cmdline:
        has_user_profile = any(
            '--profile-directory' in str(arg) or
            ('--user-data-dir' in str(arg) and os.path.expanduser('~') in str(arg))
            for arg in cmdline
        )
        return not has_user_profile
    return False


def _cleanup_scraper_browsers(zombies_only: bool = False) -> int:
    if not PSUTIL_AVAILABLE:
        return 0
    try:
        killed_count = 0
        for proc in psutil.process_iter(['pid', 'name', 'cmdline']):
            try:
                pinfo = proc.info
                if not _is_scraper_browser(pinfo):
                    continue
                proc_obj = psutil.Process(pinfo['pid'])
                if zombies_only:
                    is_zombie = proc_obj.status() == psutil.STATUS_ZOMBIE
                    is_stuck = (proc_obj.cpu_percent(interval=0.1) == 0 and
                                proc_obj.create_time() < psutil.boot_time() + 300)
                    if not (is_zombie or is_stuck):
                        continue
                proc_obj.kill()
                killed_count += 1
                if not zombies_only:
                    try:
                        proc_obj.wait(timeout=2)
                    except psutil.TimeoutExpired:
                        proc_obj.kill()
            except (psutil.NoSuchProcess, psutil.AccessDenied, KeyError):
                pass
        return killed_count
    except Exception as e:
        logger.error("[CLEANUP] Error: %s", e)
        return 0


def cleanup_zombie_browser_processes() -> int:
    return _cleanup_scraper_browsers(zombies_only=True)


def cleanup_all_scraper_browsers() -> int:
    return _cleanup_scraper_browsers(zombies_only=False)


def scrape_single_url(url: str) -> dict:
    process = None
    try:
        output_file = os.path.join(RESULTS_DIR, f"scrape_{uuid.uuid4().hex}.json")
        os.makedirs(RESULTS_DIR, exist_ok=True)

        cmd = ["python", SCRAPER_SCRIPT, "--url", url, "--output-file", output_file, "--timeout", "60"]

        env = os.environ.copy()
        env["PYTHONIOENCODING"] = "utf-8"

        process = subprocess.Popen(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            cwd=BACKEND_DIR,
            env=env,
            encoding="utf-8",
            errors="replace"
        )

        try:
            stdout, stderr = process.communicate(timeout=120)
            returncode = process.returncode
        except subprocess.TimeoutExpired:
            logger.warning("[SCRAPER] TIMEOUT: %s", url)
            if PSUTIL_AVAILABLE:
                try:
                    parent = psutil.Process(process.pid)
                    for child in parent.children(recursive=True):
                        try:
                            child.kill()
                        except psutil.NoSuchProcess:
                            pass
                    parent.kill()
                    process.wait(timeout=5)
                except psutil.NoSuchProcess:
                    pass
            else:
                process.kill()
                process.wait(timeout=5)
            return {"success": False, "url": url, "error": "Scraper timed out"}
        finally:
            if process and process.poll() is None:
                try:
                    process.kill()
                    process.wait(timeout=5)
                except Exception:
                    pass

        if returncode != 0:
            logger.warning("[SCRAPER] FAILED rc=%d %s\nSTDERR: %s", returncode, url, stderr[-500:])
            return {"success": False, "url": url, "error": f"Scraper failed (exit {returncode})"}

        # Check for makro.json output file
        retailer_files = ["makro.json", "unknown.json"]
        output_dir = os.path.dirname(output_file)

        for fname in retailer_files:
            fpath = os.path.join(output_dir, fname)
            if os.path.exists(fpath):
                try:
                    with open(fpath, 'r', encoding='utf-8') as f:
                        scraped_data = _json.load(f)
                    if isinstance(scraped_data, list):
                        for item in scraped_data:
                            if normalize_url(item.get('url', '')) == normalize_url(url) or item.get('url') == url:
                                item["source_url"] = url
                                logger.info("[SCRAPER] SUCCESS: %s -> %s", url, (item.get('name') or '')[:50])
                                try:
                                    os.remove(output_file)
                                except OSError:
                                    pass
                                return {"success": True, "url": url, "data": item}
                except Exception as e:
                    logger.error("[SCRAPER] Error reading %s: %s", fpath, e)

        # Fallback: check output file directly
        if os.path.exists(output_file):
            try:
                with open(output_file, 'r', encoding='utf-8') as f:
                    scraped_data = _json.load(f)
                if isinstance(scraped_data, list) and scraped_data:
                    item = scraped_data[0]
                    item["source_url"] = url
                    os.remove(output_file)
                    return {"success": True, "url": url, "data": item}
            except Exception as e:
                logger.error("[SCRAPER] Error reading output file: %s", e)

        try:
            if os.path.exists(output_file):
                os.remove(output_file)
        except Exception:
            pass

        return {"success": False, "url": url, "error": "Scraper output not found"}

    except Exception as e:
        if process and process.poll() is None:
            try:
                process.kill()
                process.wait(timeout=5)
            except Exception:
                pass
        return {"success": False, "url": url, "error": str(e)}


# ---------------------------------------------------------------------------
# Scrape Endpoint
# ---------------------------------------------------------------------------

class ScrapeUrlRequest(BaseModel):
    urls: list[str]


@app.post("/api/scrape")
def scrape_urls(
    data: ScrapeUrlRequest,
    user: dict = Depends(get_current_user)
):
    """Scrape product URLs in parallel. Returns scraped data for each URL."""
    from concurrent.futures import ThreadPoolExecutor, as_completed

    cleanup_zombie_browser_processes()
    os.makedirs(RESULTS_DIR, exist_ok=True)

    results = []
    errors = []

    max_workers = min(len(data.urls), 4)
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_to_url = {executor.submit(scrape_single_url, url): url for url in data.urls}
        for future in as_completed(future_to_url):
            url = future_to_url[future]
            try:
                result = future.result()
                if result["success"]:
                    results.append(result["data"])
                else:
                    errors.append({"url": url, "error": result["error"]})
            except Exception as e:
                errors.append({"url": url, "error": str(e)})

    cleanup_all_scraper_browsers()

    return {
        "success": len(errors) == 0,
        "results": results,
        "errors": errors,
        "total_scraped": len(results),
        "total_errors": len(errors)
    }


# ---------------------------------------------------------------------------
# Manual Comparison Endpoints
# ---------------------------------------------------------------------------

def _fetch_makro_price(url: str) -> dict:
    """Fetch Makro product price via plain HTTP (no browser needed — uses __NEXT_DATA__ JSON)."""
    import urllib.request
    import re as _re
    import json as _json2

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36",
        "Accept-Language": "th-TH,th;q=0.9,en;q=0.8",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    }
    req = urllib.request.Request(url, headers=headers)
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            html = resp.read().decode("utf-8", errors="replace")
    except Exception as e:
        return {"success": False, "error": str(e)}

    match = _re.search(r'<script id="__NEXT_DATA__"[^>]*>(.*?)</script>', html, _re.DOTALL)
    if not match:
        return {"success": False, "error": "__NEXT_DATA__ not found"}

    try:
        data = _json2.loads(match.group(1))
    except Exception as e:
        return {"success": False, "error": f"JSON parse error: {e}"}

    p = data.get("props", {}).get("pageProps", {}).get("product")
    if not p:
        return {"success": False, "error": "product not found in __NEXT_DATA__"}

    current_price = None
    try:
        current_price = float(p.get("displayPrice") or 0) or None
    except (TypeError, ValueError):
        pass

    # Step prices from slabPriceTiers
    step_prices = []
    slab_data = p.get("slabPrices")
    if slab_data and isinstance(slab_data, dict):
        tiers = sorted(slab_data.get("slabPriceTiers") or [], key=lambda t: t.get("tier", 0))
        if tiers and current_price:
            step_prices = [[1, current_price]] + [
                [t["quantity"], float(t["priceInVat"])]
                for t in tiers
                if t.get("quantity") is not None and t.get("priceInVat") is not None
            ]

    images = p.get("imageUrls") or []

    return {
        "success": True,
        "current_price": current_price,
        "step_prices": step_prices,
        "image": images[0] if images else None,
    }


@app.post("/api/products/{sku}/rescrape")
def rescrape_product(
    sku: str,
    user: dict = Depends(get_current_user)
):
    """Rescrape verified Makro matches for a CFW product (looked up by SKU) and update their prices."""
    with get_db() as conn:
        with conn.cursor() as cur:
            # Resolve SKU → DB id
            cur.execute("SELECT id FROM products WHERE sku = %s AND retailer_id = 'cfw'", (sku,))
            cfw_row = cur.fetchone()
            if not cfw_row:
                raise HTTPException(status_code=404, detail=f"CFW product with SKU '{sku}' not found")
            product_id = cfw_row["id"]

            # Get all verified correct Makro matches for this product
            cur.execute("""
                SELECT mp.id, mp.sku, mp.url
                FROM product_matches pm
                JOIN products mp ON pm.makro_product_id = mp.id
                WHERE pm.cfw_product_id = %s
                  AND pm.is_verified = TRUE AND pm.is_same = TRUE
                  AND mp.url IS NOT NULL AND mp.url != ''
            """, (product_id,))
            makro_products = cur.fetchall()

    if not makro_products:
        return {"successful": 0, "failed": 0, "message": "No verified Makro matches with URLs to rescrape"}

    successful = 0
    failed = 0

    for mp in makro_products:
        result = _fetch_makro_price(mp["url"])
        if result["success"]:
            new_price = result.get("current_price")
            new_step_prices = _json.dumps(result.get("step_prices") or [])
            new_image = result.get("image")

            with get_db() as conn:
                with conn.cursor() as cur:
                    # Update product price
                    cur.execute("""
                        UPDATE products SET
                            current_price = %s,
                            step_prices = %s,
                            image_url = COALESCE(%s, image_url),
                            updated_at = NOW()
                        WHERE id = %s
                    """, (new_price, new_step_prices, new_image, mp["id"]))

                    # Insert price history
                    if new_price:
                        cur.execute("""
                            INSERT INTO price_history (product_id, price, step_prices)
                            VALUES (%s, %s, %s)
                        """, (mp["id"], new_price, new_step_prices))

                    conn.commit()
            successful += 1
        else:
            failed += 1

    return {
        "successful": successful,
        "failed": failed,
        "message": f"Updated {successful} Makro product{'s' if successful != 1 else ''}"
    }


@app.get("/api/products/sku/{sku}/matches")
def get_product_matches_by_sku(
    sku: str,
    user: dict = Depends(get_current_user)
):
    """Get existing verified Makro matches for a CFW product by SKU"""
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT id, name, current_price, url, image_url
                FROM products WHERE sku = %s AND retailer_id = 'cfw'
            """, (sku,))
            cfw_product = cur.fetchone()

            if not cfw_product:
                return {"found": False, "product": None, "matches": []}

            cur.execute("""
                SELECT
                    mp.id as makro_id,
                    mp.sku as makro_sku,
                    mp.name as makro_name,
                    mp.current_price as makro_price,
                    mp.url as makro_url,
                    mp.image_url as makro_image,
                    pm.match_id,
                    pm.is_verified,
                    pm.is_same
                FROM product_matches pm
                JOIN products mp ON pm.makro_product_id = mp.id
                WHERE pm.cfw_product_id = %s
                  AND pm.is_verified = TRUE AND pm.is_same = TRUE
                ORDER BY mp.name
            """, (cfw_product["id"],))

            matches = cur.fetchall()

            return {
                "found": True,
                "product": {
                    "id": cfw_product["id"],
                    "name": cfw_product["name"],
                    "price": float(cfw_product["current_price"]) if cfw_product["current_price"] else None,
                    "url": cfw_product["url"],
                    "image": cfw_product["image_url"],
                },
                "matches": [
                    {
                        "retailer_id": "makro",
                        "retailer_name": "Makro",
                        "product_id": m["makro_id"],
                        "sku": m["makro_sku"],
                        "name": m["makro_name"],
                        "price": float(m["makro_price"]) if m["makro_price"] else None,
                    }
                    for m in matches
                ]
            }


class ManualComparisonRequest(BaseModel):
    cfw_sku: str
    makro_url: str
    scraped_data: dict | None = None  # scraped Makro product data


@app.post("/api/comparison/manual")
def manual_comparison(
    data: ManualComparisonRequest,
    user: dict = Depends(get_current_user)
):
    """
    Manual comparison: given a CFW SKU + scraped Makro product data,
    upsert the Makro product and create a product_match entry.
    """
    scraped = data.scraped_data or {}

    with get_db() as conn:
        with conn.cursor() as cur:
            # Find CFW product
            cur.execute("""
                SELECT id, name, current_price FROM products
                WHERE sku = %s AND retailer_id = 'cfw'
            """, (data.cfw_sku,))
            cfw_product = cur.fetchone()
            if not cfw_product:
                raise HTTPException(status_code=404, detail=f"CFW product with SKU '{data.cfw_sku}' not found")

            # Upsert Makro product
            makro_sku = str(scraped.get('sku') or '').strip() or None
            makro_name = (scraped.get('name') or '').strip() or None
            makro_price = scraped.get('current_price')
            makro_brand = (scraped.get('brand') or '').strip() or None
            makro_image = (scraped.get('images') or [None])[0] if scraped.get('images') else None
            makro_step_prices = _json.dumps(scraped.get('step_prices') or [])

            if not makro_sku:
                raise HTTPException(status_code=400, detail="Scraped Makro product has no SKU")

            # Get or resolve Makro category
            makro_category_name = (scraped.get('category') or '').strip() or None
            makro_category_id = None
            if makro_category_name:
                cur.execute("""
                    SELECT category_id FROM categories
                    WHERE retailer_id = 'makro' AND category_name ILIKE %s
                    LIMIT 1
                """, (makro_category_name,))
                row = cur.fetchone()
                if row:
                    makro_category_id = row["category_id"]
                else:
                    # Insert new category
                    makro_category_id = makro_category_name.upper().replace(' ', '_')[:50]
                    cur.execute("""
                        INSERT INTO categories (retailer_id, category_id, category_name)
                        VALUES ('makro', %s, %s)
                        ON CONFLICT (retailer_id, category_id) DO NOTHING
                    """, (makro_category_id, makro_category_name))

            cur.execute("""
                INSERT INTO products (retailer_id, sku, name, brand, category_id, current_price, url, image_url, step_prices)
                VALUES ('makro', %s, %s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (retailer_id, sku) DO UPDATE SET
                    name = EXCLUDED.name,
                    brand = EXCLUDED.brand,
                    category_id = EXCLUDED.category_id,
                    current_price = EXCLUDED.current_price,
                    url = EXCLUDED.url,
                    image_url = EXCLUDED.image_url,
                    step_prices = EXCLUDED.step_prices,
                    updated_at = NOW()
                RETURNING id
            """, (makro_sku, makro_name, makro_brand, makro_category_id,
                  makro_price, data.makro_url, makro_image, makro_step_prices))
            makro_product_id = cur.fetchone()["id"]

            # Block if this CFW product already has a verified correct match
            cur.execute("""
                SELECT pm.match_id, mp.name as makro_name
                FROM product_matches pm
                JOIN products mp ON pm.makro_product_id = mp.id
                WHERE pm.cfw_product_id = %s
                  AND pm.is_verified = TRUE AND pm.is_same = TRUE
                LIMIT 1
            """, (cfw_product["id"],))
            existing_verified = cur.fetchone()
            if existing_verified:
                raise HTTPException(
                    status_code=409,
                    detail=f"This CFW product already has a verified match: \"{existing_verified['makro_name']}\". Undo the existing match first before adding a new one."
                )

            # Create match entry — auto-verified as correct (manually added = user intent)
            cur.execute("""
                INSERT INTO product_matches (cfw_product_id, makro_product_id, match_score, is_verified, is_same, verified_at)
                VALUES (%s, %s, 100, TRUE, TRUE, NOW())
                ON CONFLICT (cfw_product_id, makro_product_id) DO UPDATE SET
                    is_verified = TRUE,
                    is_same = TRUE,
                    verified_at = NOW(),
                    updated_at = NOW()
                RETURNING match_id
            """, (cfw_product["id"], makro_product_id))
            match_id = cur.fetchone()["match_id"]

            conn.commit()

            return {
                "success": True,
                "cfw_product": {
                    "id": cfw_product["id"],
                    "name": cfw_product["name"],
                    "price": float(cfw_product["current_price"]) if cfw_product["current_price"] else None,
                },
                "makro_product": {
                    "id": makro_product_id,
                    "sku": makro_sku,
                    "name": makro_name,
                    "price": float(makro_price) if makro_price else None,
                    "url": data.makro_url,
                    "image": makro_image,
                },
                "match_id": match_id,
                "message": "Match created. Go to the product detail page to verify."
            }


# ---------------------------------------------------------------------------
# Server Startup
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import uvicorn
    config = uvicorn.Config(
        app=app,
        host="0.0.0.0",
        port=8000,
        timeout_keep_alive=300,
        timeout_graceful_shutdown=30,
        limit_max_requests=None,
    )
    server = uvicorn.Server(config)
    server.run()
